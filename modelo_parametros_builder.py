# -*- coding: utf-8 -*-
"""
Constructor de Parametros.xlsx desde datos de entrada (Opcion A: page-driven).
================================================================================

Objetivo
--------
Permitir que el Modelo de Optimizacion LP (modelo_optimizacion_core.ejecutar_modelo)
se alimente de datos capturados en la pagina web (persistidos en BD) en lugar de
depender de que el usuario edite manualmente el archivo Excel maestro.

Estrategia: "template-based rebuild"
------------------------------------
El archivo Parametros.xlsx NO es solo datos: tiene una capa de columnas derivadas
(formulas). El solver lee el valor *cacheado* de esas formulas. Por eso, para
independizarnos del Excel hay que reproducir esa capa de derivacion en Python.

Este modulo:
  1. Parte del libro base (estructura, hojas auxiliares y rangos con nombre intactos).
  2. Reemplaza las filas de datos de las 12 hojas del modelo con los datos de entrada
     del escenario (los que el usuario edita en la pagina).
  3. RECALCULA en Python todas las columnas derivadas (porque el servidor no
     recalcula formulas de Excel).
  4. Guarda un .xlsx temporal que ejecutar_modelo() consume sin ningun cambio.

Toda la logica de derivacion fue verificada empiricamente contra los valores
cacheados del archivo real (>1000 celdas, coincidencia practicamente total; las
unicas diferencias son celdas hardcodeadas/obsoletas en la hoja original).

Columnas derivadas reproducidas:
  - SPG            = 141.5 / (API + 131.5)            (COMPRAS trunca a 3 decimales)
  - DENSIDAD       = 6.2898 / SPG
  - Iv50 / V50     = ln(ln(viscosidad + 0.8))
  - TIPO FLUJO     = catalogo 2.FLUJOS  (lookup por FLUJO)
  - CRUDO ORIGEN   = catalogo 5.CURVA   (lookup por FLUJO/PRODUCTO)
  - TIPO DE NODO   = catalogo 1.NODOS   (lookup por nombre de nodo)
  - Precio Compra  = BRENT + Spread
  - Costo modelo   = (Costo*42/TRM) si la ruta es 'CTK', si no Costo
  - Costo Operac.  = suma de componentes
  - VENTAS: ORIGEN BLEND, densidades y RELACION DCONTRAC/DENTREGA

Uso como script (auto-validacion celda a celda contra el archivo real):
    python modelo_parametros_builder.py "ruta\\Parametros.xlsx"
"""

import math
import re
import pandas as pd
from openpyxl import load_workbook

# Las 12 hojas que el solver consume, en el orden esperado.
MODEL_SHEETS = [
    '1.NODOS', '2.FLUJOS', '3.COMPRAS', '4.RUTAS_TRANSPORTE',
    '5.CURVA_DESTILACION', '6.VENTAS', '7.COSTOS_REFINACION',
    '8.LIMITES_CALIDAD', '9.FLASH_RESTR', '10.CVC',
    '11.REL_CRUDO_MEZCLA', '12.COSTOS_OPERACIONALES',
]

# Columnas DERIVADAS por hoja: las calcula el backend, NO se editan en la pagina.
# (Cualquier otra columna de estas hojas es un INPUT que el usuario captura.)
DERIVED_COLUMNS = {
    '3.COMPRAS': [
        'TIPO ', 'Precio de Compra CALCULADO  (USD/BPD)', 'SPG',
        'DENSIDAD  (BBL/TONMETRICA)', 'Iv50',
    ],
    '4.RUTAS_TRANSPORTE': [
        'CRUDO ORIGEN', 'TIPO FLUJO', 'TIPO DE NODO ORIGEN',
        'TIPO DE NODO DESTINO', 'Costo de transporte USD/BBL para modelo',
    ],
    '5.CURVA_DESTILACION': [
        'SPG', 'DENSIDAD (BBL/TONMETRICA)', 'Iv50',
        'TOTAL (DEBE SER IGUAL A 1)',
    ],
    '6.VENTAS': [
        'CRUDO ORIGEN', 'TIPO FLUJO', 'VALIDACION TIPO DE NODO VENTA',
        'ORIGEN BLEND', 'Precio de venta (USD/BPD)',
        'CRUDO: DENSIDAD  (BBL/TONMETRICA)', 'REFINADO: DENSIDAD  (BBL/TONMETRICA)',
        'RELACION  DCONTRAC/DENTREGA',
    ],
    # NOTA: 'Densidad (BBL/TONMetrica) Contractual' NO es derivada: es un INPUT editable
    # por contrato (default 6.7 cuando la venta es por densidad), porque puede llevar
    # valores reales negociados distintos del default (p. ej. 6.5, 7.4).
    '7.COSTOS_REFINACION': ['TIPO DE NODO'],
    '8.LIMITES_CALIDAD': [
        'MINIMO CALIDAD SPG', 'MAXIMO CALIDAD SPG', 'MINIMO V50', 'MAXIMO V50',
    ],
    '9.FLASH_RESTR': ['TIPO FLUJO', 'FLASH POINT C', 'BINARY COND'],
    '10.CVC': ['CRUDO ORIGEN', 'TIPO FLUJO'],
    '11.REL_CRUDO_MEZCLA': ['CRUDO ORIGEN', 'TIPO FLUJO'],
    '12.COSTOS_OPERACIONALES': ['Costo Operacional (USD/BBL)'],
    '13.PRECIOS_VENTA': ['PRECIO CALCULADO (USD/BBL)'],
}

# Constante de densidad contractual cuando la venta es por formula de densidad.
DENSIDAD_CONTRACTUAL_DEN = 6.7
# Componentes que suman el Costo Operacional total.
COSTO_OPER_COMPONENTES = [
    'Almacenamiento (USD/BBL)', 'Op. Portuaria (USD/BBL)', 'Barcazas  (USD/BBL)',
    'Pusher (USD/BBL)', 'Tasa Portuaria (USD/BBL)',
]


# ---------------------------------------------------------------------------
# Utilidades de columnas (tolerantes a espacios, saltos de linea y mayusculas)
# ---------------------------------------------------------------------------

def _norm(text):
    """Normaliza un nombre de columna: colapsa espacios/saltos y pasa a mayusculas."""
    if text is None:
        return ''
    return re.sub(r'\s+', ' ', str(text).strip()).upper()


def _find_col(df, target):
    """Devuelve el nombre real de la columna en df que corresponde a `target`."""
    nt = _norm(target)
    for col in df.columns:
        if _norm(col) == nt:
            return col
    # match flexible por prefijo (p. ej. 'Spread\n (Valor...)')
    for col in df.columns:
        if _norm(col).startswith(nt) or nt.startswith(_norm(col)):
            return col
    return None


def _to_float(value):
    try:
        if value is None or value == '':
            return None
        f = float(value)
        if math.isnan(f) or math.isinf(f):
            return None
        return f
    except (TypeError, ValueError):
        return None


def _is_blank(value):
    return value is None or (isinstance(value, float) and math.isnan(value)) or str(value).strip() == ''


# Cortes refinados que se cotizan por igual, salgan del crudo que salgan.
CORTES_REFINADOS = ('NAFTA', 'GASOIL', 'FUELOIL')
HOJA_PRECIOS = '13.PRECIOS_VENTA'


def corriente_comercial(flujo):
    """A que corriente comercial pertenece un flujo, para efectos de precio.

    NAFTA_R_TILODIRAN, NAFTA_R_BORANDA... son todos 'NAFTA': comercialmente es
    el mismo producto. Las mezclas y los crudos se cotizan uno a uno.
    """
    texto = str(flujo or '').strip()
    for corte in CORTES_REFINADOS:
        if texto.upper().startswith(corte + '_R_'):
            return corte
    return texto


def construir_precios_desde_ventas(ventas, brent):
    """Deriva la hoja de precios a partir de los precios que ya tiene 6.VENTAS.

    El ajuste sale del PRECIO REAL menos el BRENT, no de la columna de formula:
    en el archivo actual 23 de esas formulas contradicen el precio que se esta
    cobrando, asi que tomarlas como fuente meteria errores.
    """
    col_flujo = _find_col(ventas, 'FLUJO')
    col_precio = _find_col(ventas, 'Precio de venta (USD/BPD)')
    if col_flujo is None or col_precio is None:
        return pd.DataFrame(columns=['CORRIENTE COMERCIAL',
                                     'AJUSTE VS BRENT (USD/BBL)',
                                     'PRECIO CALCULADO (USD/BBL)'])

    vistos = {}
    for flujo, precio in zip(ventas[col_flujo], ventas[col_precio]):
        p = _to_float(precio)
        if p is None:
            continue
        vistos.setdefault(corriente_comercial(flujo), p)

    filas = [{'CORRIENTE COMERCIAL': c,
              'AJUSTE VS BRENT (USD/BBL)': round(p - (brent or 0.0), 6),
              'PRECIO CALCULADO (USD/BBL)': p}
             for c, p in vistos.items()]
    return pd.DataFrame(filas, columns=['CORRIENTE COMERCIAL',
                                        'AJUSTE VS BRENT (USD/BBL)',
                                        'PRECIO CALCULADO (USD/BBL)'])


def _is_den(value):
    return str(value).strip().upper() == 'DEN'


def _coerce_num(value):
    """Si `value` es un string que representa un numero (acepta coma decimal),
    devuelve el float; si no, devuelve el valor original sin tocar."""
    if not isinstance(value, str):
        return value
    s = value.strip()
    if re.fullmatch(r'-?\d+(?:[.,]\d+)?', s):
        try:
            return float(s.replace(',', '.'))
        except ValueError:
            return value
    return value


# ---------------------------------------------------------------------------
# Conversiones de calidad / densidad
# ---------------------------------------------------------------------------

def spg_from_api(api, truncate3=False):
    a = _to_float(api)
    if a is None:
        return ''
    spg = 141.5 / (a + 131.5)
    if truncate3:
        spg = math.trunc(spg * 1000) / 1000.0
    return spg


def density_from_spg(spg):
    s = _to_float(spg)
    if s is None or s == 0:
        return ''
    return 6.2898 / s


def iv50_from_visc(visc):
    v = _to_float(visc)
    if v is None:
        return ''
    try:
        return math.log(math.log(v + 0.8))
    except ValueError:
        return ''


# ---------------------------------------------------------------------------
# Catalogos (rangos con nombre del Excel reproducidos como diccionarios)
# ---------------------------------------------------------------------------

def construir_catalogos(sheets, economics):
    """Construye los mapas de lookup que las formulas del Excel usan."""
    nodos = sheets['1.NODOS']
    flujos = sheets['2.FLUJOS']
    curva = sheets['5.CURVA_DESTILACION']
    compras = sheets['3.COMPRAS']
    limites = sheets['8.LIMITES_CALIDAD']

    def col(df, name):
        return _find_col(df, name)

    nmap = dict(zip(
        nodos[col(nodos, 'NOMBRE DEL NODO')].astype(str),
        nodos[col(nodos, 'TIPO DE NODO')].astype(str),
    ))
    fmap = dict(zip(
        flujos[col(flujos, 'FLUJO')].astype(str),
        flujos[col(flujos, 'TIPO DE FLUJO')].astype(str),
    ))
    # REFINADOS = 5.CURVA_DESTILACION!A:B  -> PRODUCTO -> CRUDO ORIGEN
    rmap = dict(zip(
        curva[col(curva, 'PRODUCTO')].astype(str),
        curva[col(curva, 'CRUDO ORIGEN')].astype(str),
    ))
    # FLASH POINT C de VENTAS/FLASH se toma de 3.COMPRAS
    flash_map = dict(zip(
        compras[col(compras, 'CRUDO O PRODUCTO')].astype(str),
        compras[col(compras, 'FLASH POINT C')],
    ))
    # ORIGEN BLEND = VLOOKUP mezcla -> CENTRO DE BLENDING en 8.LIMITES_CALIDAD
    mezcla_centro = dict(zip(
        limites[col(limites, 'MEZCLA')].astype(str),
        limites[col(limites, 'CENTRO DE BLENDING')].astype(str),
    ))
    # CRUDO: DENSIDAD de VENTAS = VLOOKUP a 3.COMPRAS columna 7 (Spread, tal cual el Excel)
    compras_col7 = compras.columns[6]
    crudo_dens_map = dict(zip(
        compras[col(compras, 'CRUDO O PRODUCTO')].astype(str),
        compras[compras_col7],
    ))
    # REFINADO: DENSIDAD de VENTAS = VLOOKUP a 5.CURVA columna 6 (SPG, tal cual el Excel)
    curva_col6 = curva.columns[5]
    refi_dens_map = dict(zip(
        curva[col(curva, 'PRODUCTO')].astype(str),
        curva[curva_col6],
    ))

    return {
        'nodos': nmap, 'flujos': fmap, 'refinados': rmap,
        'flash': flash_map, 'mezcla_centro': mezcla_centro,
        'crudo_dens': crudo_dens_map, 'refi_dens': refi_dens_map,
        'BRENT': _to_float(economics.get('BRENT')) or 0.0,
        'TRM': _to_float(economics.get('TRM')) or 0.0,
    }


# ---------------------------------------------------------------------------
# Motor de derivacion: completa las columnas calculadas de cada hoja
# ---------------------------------------------------------------------------

def recompute_derived(sheets, economics):
    """Recalcula in-place las columnas derivadas de todas las hojas del modelo."""
    cat = construir_catalogos(sheets, economics)
    BRENT, TRM = cat['BRENT'], cat['TRM']

    def C(df, name):
        return _find_col(df, name)

    # --- 3.COMPRAS ---
    if '3.COMPRAS' in sheets:
        c = sheets['3.COMPRAS']
        flujo_col = C(c, 'CRUDO O PRODUCTO')
        api_col = C(c, 'CALIDAD API')
        visc_col = C(c, 'Viscosidad Cst')
        spread_col = C(c, 'Spread')
        
        tipo_col = C(c, 'TIPO ')
        spg_col = C(c, 'SPG')
        dens_col = C(c, 'DENSIDAD  (BBL/TONMETRICA)')
        iv50_col = C(c, 'Iv50')
        precio_col = C(c, 'Precio de Compra CALCULADO  (USD/BPD)')
        
        if flujo_col and tipo_col:
            c[tipo_col] = [cat['flujos'].get(str(x), '') for x in c[flujo_col]]
        if api_col and spg_col:
            c[spg_col] = [spg_from_api(x, truncate3=True) for x in c[api_col]]
        if api_col and dens_col:
            c[dens_col] = [density_from_spg(spg_from_api(x, truncate3=True)) for x in c[api_col]]
        if visc_col and iv50_col:
            c[iv50_col] = [iv50_from_visc(x) for x in c[visc_col]]
        if flujo_col and spread_col and precio_col:
            c[precio_col] = [
                (BRENT + (_to_float(s) or 0.0)) if not _is_blank(f) else 0.0
                for f, s in zip(c[flujo_col], c[spread_col])
            ]

    # --- 4.RUTAS_TRANSPORTE ---
    if '4.RUTAS_TRANSPORTE' in sheets:
        r = sheets['4.RUTAS_TRANSPORTE']
        r_flujo = C(r, 'FLUJO'); r_orig = C(r, 'ORIGEN'); r_dest = C(r, 'DESTINO')
        r_ruta = C(r, 'RUTA DE TRANSPORTE'); r_costo = C(r, 'Costo de Transporte')
        
        orig_col = C(r, 'CRUDO ORIGEN')
        tipo_col = C(r, 'TIPO FLUJO')
        nodo_orig_col = C(r, 'TIPO DE NODO ORIGEN')
        nodo_dest_col = C(r, 'TIPO DE NODO DESTINO')
        costo_mod_col = C(r, 'Costo de transporte USD/BBL para modelo')
        
        if r_flujo and orig_col:
            r[orig_col] = [cat['refinados'].get(str(x), '') for x in r[r_flujo]]
        if r_flujo and tipo_col:
            r[tipo_col] = [cat['flujos'].get(str(x), '') for x in r[r_flujo]]
        if r_orig and nodo_orig_col:
            r[nodo_orig_col] = [cat['nodos'].get(str(x), '') for x in r[r_orig]]
        if r_dest and nodo_dest_col:
            r[nodo_dest_col] = [cat['nodos'].get(str(x), '') for x in r[r_dest]]
        if r_ruta and r_costo and costo_mod_col:
            costo_modelo = []
            for ruta, costo in zip(r[r_ruta], r[r_costo]):
                cv = _to_float(costo) or 0.0
                if str(ruta).strip() == 'CTK' and TRM:
                    costo_modelo.append(cv * 42.0 / TRM)
                else:
                    costo_modelo.append(cv)
            r[costo_mod_col] = costo_modelo

    # --- 5.CURVA_DESTILACION ---
    if '5.CURVA_DESTILACION' in sheets:
        cu = sheets['5.CURVA_DESTILACION']
        cu_api = C(cu, 'CALIDAD API'); cu_visc = C(cu, 'Viscosidad Cst')
        
        spg_col = C(cu, 'SPG')
        dens_col = C(cu, 'DENSIDAD (BBL/TONMETRICA)')
        iv50_col = C(cu, 'Iv50')
        
        if cu_api and spg_col:
            cu[spg_col] = [spg_from_api(x) for x in cu[cu_api]]
        if cu_api and dens_col:
            cu[dens_col] = [density_from_spg(spg_from_api(x)) for x in cu[cu_api]]
        if cu_visc and iv50_col:
            cu[iv50_col] = [iv50_from_visc(x) for x in cu[cu_visc]]

        # Columna de verificacion: suma de FRACCION % por crudo. El Excel la
        # escribe solo en la primera fila de cada grupo de crudo.
        cu_crudo = C(cu, 'CRUDO ORIGEN')
        cu_frac = C(cu, 'FRACCION %')
        total_col = C(cu, 'TOTAL (DEBE SER IGUAL A 1)')
        if cu_crudo and cu_frac and total_col:
            sumas = {}
            for crudo, frac in zip(cu[cu_crudo], cu[cu_frac]):
                key = _norm(crudo)
                sumas[key] = sumas.get(key, 0.0) + (_to_float(frac) or 0.0)
            vistos = set()
            totales = []
            for crudo in cu[cu_crudo]:
                key = _norm(crudo)
                if key and key not in vistos:
                    vistos.add(key)
                    totales.append(sumas[key])
                else:
                    totales.append(None)
            cu[total_col] = totales

    # --- 13.PRECIOS_VENTA (antes de 6.VENTAS, que depende de ella) ---
    if HOJA_PRECIOS in sheets:
        pv = sheets[HOJA_PRECIOS]
        c_corr = C(pv, 'CORRIENTE COMERCIAL')
        c_aj = C(pv, 'AJUSTE VS BRENT (USD/BBL)')
        c_pre = C(pv, 'PRECIO CALCULADO (USD/BBL)')
        if c_corr and c_aj and c_pre:
            pv[c_pre] = [
                (BRENT + (_to_float(a) or 0.0)) if not _is_blank(a) else None
                for a in pv[c_aj]
            ]

    # --- 6.VENTAS ---
    if '6.VENTAS' in sheets:
        v = sheets['6.VENTAS']
        # El precio sale de la hoja de precios cuando esa hoja tiene datos; si
        # no, se respeta el que ya trae la fila (compatibilidad hacia atras).
        tabla_precios = sheets.get(HOJA_PRECIOS)
        if tabla_precios is not None and len(tabla_precios):
            c_corr = C(tabla_precios, 'CORRIENTE COMERCIAL')
            c_pre = C(tabla_precios, 'PRECIO CALCULADO (USD/BBL)')
            v_flujo0 = C(v, 'FLUJO')
            v_precio0 = C(v, 'Precio de venta (USD/BPD)')
            if c_corr and c_pre and v_flujo0 and v_precio0:
                mapa_precio = {
                    _norm(k): _to_float(p)
                    for k, p in zip(tabla_precios[c_corr], tabla_precios[c_pre])
                    if not _is_blank(k)
                }
                v[v_precio0] = [
                    mapa_precio.get(_norm(corriente_comercial(f)), actual)
                    for f, actual in zip(v[v_flujo0], v[v_precio0])
                ]
        v_flujo = C(v, 'FLUJO'); v_punto = C(v, 'PUNTO DE VENTA'); v_den = C(v, 'VENTA DENSIDAD')
        
        orig_col = C(v, 'CRUDO ORIGEN')
        tipo_col = C(v, 'TIPO FLUJO')
        nodo_venta_col = C(v, 'VALIDACION TIPO DE NODO VENTA')
        origen_blend_col = C(v, 'ORIGEN BLEND')
        dc_col = C(v, 'Densidad (BBL/TONMetrica) Contractual')
        crudo_d_col = C(v, 'CRUDO: DENSIDAD  (BBL/TONMETRICA)')
        refi_d_col = C(v, 'REFINADO: DENSIDAD  (BBL/TONMETRICA)')
        relacion_col = C(v, 'RELACION  DCONTRAC/DENTREGA')
        
        if v_flujo and orig_col:
            v[orig_col] = [cat['refinados'].get(str(x), '') for x in v[v_flujo]]
        if v_flujo and tipo_col:
            v[tipo_col] = [cat['flujos'].get(str(x), '') for x in v[v_flujo]]
        if v_punto and nodo_venta_col:
            v[nodo_venta_col] = [cat['nodos'].get(str(x), '') for x in v[v_punto]]
            
        if v_flujo and v_den and origen_blend_col:
            v[origen_blend_col] = [cat['mezcla_centro'].get(str(f), '') if _is_den(d) else ''
                                   for f, d in zip(v[v_flujo], v[v_den])]
            
        if v_den and dc_col:
            dc_new = []
            for d, existing in zip(v[v_den], v[dc_col]):
                if _is_den(d):
                    dc_new.append(existing if not _is_blank(existing) else DENSIDAD_CONTRACTUAL_DEN)
                else:
                    dc_new.append('')
            v[dc_col] = dc_new
            
            if v_flujo and crudo_d_col:
                v_crudo_d = [cat['crudo_dens'].get(str(f), 1) if _is_den(d) else ''
                             for f, d in zip(v[v_flujo], v[v_den])]
                v[crudo_d_col] = v_crudo_d
            else:
                v_crudo_d = [1] * len(v)
                
            if v_flujo and refi_d_col:
                v_refi_d = [cat['refi_dens'].get(str(f), 1) if _is_den(d) else ''
                            for f, d in zip(v[v_flujo], v[v_den])]
                v[refi_d_col] = v_refi_d
            else:
                v_refi_d = [1] * len(v)
                
            if relacion_col:
                relacion = []
                for d, dc_val, crudo_dd, refi_dd in zip(v[v_den], v[dc_col], v_crudo_d, v_refi_d):
                    if _is_den(d):
                        J = _to_float(dc_val)
                        M = _to_float(crudo_dd); N = _to_float(refi_dd)
                        relacion.append(J / (N * M) if (J and M and N) else 1)
                    else:
                        relacion.append(1)
                v[relacion_col] = relacion

    # --- 7.COSTOS_REFINACION ---
    if '7.COSTOS_REFINACION' in sheets:
        cr = sheets['7.COSTOS_REFINACION']
        cr_planta = C(cr, 'PLANTA REFINACION')
        tipo_nodo_col = C(cr, 'TIPO DE NODO')
        if cr_planta and tipo_nodo_col:
            cr[tipo_nodo_col] = [cat['nodos'].get(str(x), '') for x in cr[cr_planta]]

    # --- 8.LIMITES_CALIDAD ---
    if '8.LIMITES_CALIDAD' in sheets:
        lim = sheets['8.LIMITES_CALIDAD']
        lim_minapi = C(lim, 'MINIMO CALIDAD API'); lim_maxapi = C(lim, 'MAXIMO CALIDAD API')
        lim_minv = C(lim, 'MINIMO VISCOSIDAD'); lim_maxv = C(lim, 'MAXIMO VISCOSIDAD')
        
        min_spg_col = C(lim, 'MINIMO CALIDAD SPG')
        max_spg_col = C(lim, 'MAXIMO CALIDAD SPG')
        min_v50_col = C(lim, 'MINIMO V50')
        max_v50_col = C(lim, 'MAXIMO V50')
        
        if lim_maxapi and min_spg_col:
            lim[min_spg_col] = [spg_from_api(x) for x in lim[lim_maxapi]]
        if lim_minapi and max_spg_col:
            lim[max_spg_col] = [spg_from_api(x) for x in lim[lim_minapi]]
        if lim_minv and min_v50_col:
            lim[min_v50_col] = [iv50_from_visc(x) for x in lim[lim_minv]]
        if lim_maxv and max_v50_col:
            lim[max_v50_col] = [iv50_from_visc(x) for x in lim[lim_maxv]]

    # --- 9.FLASH_RESTR ---
    if '9.FLASH_RESTR' in sheets:
        fl = sheets['9.FLASH_RESTR']
        fl_flujo = C(fl, 'FLUJO'); fl_relax = C(fl, 'RELAX RESTRICION?')
        fl_min = C(fl, 'MINIMO REQUERIDO EN DESTINO')
        
        tipo_col = C(fl, 'TIPO FLUJO')
        fp_col = C(fl, 'FLASH POINT C')
        binary_col = C(fl, 'BINARY COND')
        
        if fl_flujo and tipo_col:
            fl[tipo_col] = [cat['flujos'].get(str(x), '') for x in fl[fl_flujo]]
        if fl_flujo and fp_col:
            fl_fp = [cat['flash'].get(str(x), '') for x in fl[fl_flujo]]
            fl[fp_col] = fl_fp
        else:
            fl_fp = [''] * len(fl)
            
        if fl_relax and fl_min and binary_col:
            binary = []
            for relax, fp_val, mn in zip(fl[fl_relax], fl_fp, fl[fl_min]):
                if str(relax).strip().upper() == 'NO':
                    f = _to_float(fp_val); m = _to_float(mn)
                    binary.append(1 if (f is not None and m is not None and f >= m) else 0)
                else:
                    binary.append(1)
            fl[binary_col] = binary

    # --- 10.CVC y 11.REL_CRUDO_MEZCLA (mismos lookups) ---
    for sheet_name in ('10.CVC', '11.REL_CRUDO_MEZCLA'):
        if sheet_name in sheets:
            s = sheets[sheet_name]
            s_flujo = C(s, 'FLUJO')
            orig_col = C(s, 'CRUDO ORIGEN')
            tipo_col = C(s, 'TIPO FLUJO')
            if s_flujo and orig_col:
                s[orig_col] = [cat['refinados'].get(str(x), '') for x in s[s_flujo]]
            if s_flujo and tipo_col:
                s[tipo_col] = [cat['flujos'].get(str(x), '') for x in s[s_flujo]]

    # --- 11.REL_CRUDO_MEZCLA deduplicacion y limpieza ---
    if '11.REL_CRUDO_MEZCLA' in sheets:
        s = sheets['11.REL_CRUDO_MEZCLA']
        s_flujo = C(s, 'FLUJO')
        orig_col = C(s, 'CRUDO ORIGEN')
        tipo_col = C(s, 'TIPO FLUJO')
        dest_col = C(s, 'DESTINO')
        mezcla_col = C(s, 'MEZCLA A PERTENECER')
        
        if s_flujo and dest_col and mezcla_col and tipo_col:
            def norm(val):
                if pd.isna(val) or val is None:
                    return ''
                return str(val).strip().upper()
                
            keep_indices = []
            seen_cp = set()
            seen_ref = set()
            
            for idx, row in s.iterrows():
                if s.loc[idx].isna().all():
                    continue
                    
                flujo = norm(row[s_flujo])
                tipo = norm(row[tipo_col])
                destino = norm(row[dest_col])
                mezcla = norm(row[mezcla_col])
                orig = norm(row[orig_col]) if orig_col else ''
                
                if tipo not in ('CRUDO', 'PRODUCTO COMPRADO', 'PRODUCTO REFINADO'):
                    continue
                    
                if tipo in ('CRUDO', 'PRODUCTO COMPRADO'):
                    if not flujo or not destino or not mezcla:
                        continue
                    key = (flujo, destino, mezcla)
                    if key in seen_cp:
                        continue
                    seen_cp.add(key)
                elif tipo == 'PRODUCTO REFINADO':
                    if not flujo or not orig or not destino or not mezcla:
                        continue
                    key = (flujo, orig, destino, mezcla)
                    if key in seen_ref:
                        continue
                    seen_ref.add(key)
                
                keep_indices.append(idx)
                
            sheets['11.REL_CRUDO_MEZCLA'] = s.loc[keep_indices].copy()

    # --- 12.COSTOS_OPERACIONALES ---
    if '12.COSTOS_OPERACIONALES' in sheets:
        co = sheets['12.COSTOS_OPERACIONALES']
        comp_cols = [C(co, name) for name in COSTO_OPER_COMPONENTES if C(co, name)]
        total_col = C(co, 'Costo Operacional (USD/BBL)')
        if total_col and comp_cols:
            co[total_col] = co[comp_cols].apply(
                lambda row: sum((_to_float(x) or 0.0) for x in row), axis=1
            )

    return sheets


# ---------------------------------------------------------------------------
# Extraccion (para sembrar la BD desde un Excel base - flujo hibrido)
# ---------------------------------------------------------------------------

def extraer_escenario(excel_path):
    """Lee el Excel base y devuelve {'economics': {...}, 'sheets': {hoja: DataFrame}}."""
    xls = pd.ExcelFile(excel_path)
    sheets = {s: pd.read_excel(xls, sheet_name=s) for s in MODEL_SHEETS}

    # Un Excel viejo no trae las columnas que el sistema haya agregado despues
    # (por ejemplo el minimo de venta). Se agregan vacias para que aparezcan en
    # el editor y el usuario pueda usarlas sin rehacer el archivo.
    # Hoja auxiliar de precios: si el Excel no la trae, se deriva de 6.VENTAS
    if HOJA_PRECIOS in xls.sheet_names:
        sheets[HOJA_PRECIOS] = pd.read_excel(xls, sheet_name=HOJA_PRECIOS)

    try:
        from modelo_esquema import COLUMNAS
        for hoja, df in sheets.items():
            faltantes = [c for c in COLUMNAS.get(hoja, []) if c not in df.columns]
            for c in faltantes:
                df[c] = None
    except ImportError:
        pass
    economics = {}
    if '0.ECONOMICS' in xls.sheet_names:
        econ = pd.read_excel(xls, sheet_name='0.ECONOMICS')
        idx_col = _find_col(econ, 'INDEX'); val_col = _find_col(econ, 'VALUE')
        if idx_col and val_col:
            for _, row in econ.iterrows():
                economics[str(row[idx_col]).strip()] = row[val_col]
    if HOJA_PRECIOS not in sheets:
        sheets[HOJA_PRECIOS] = construir_precios_desde_ventas(
            sheets['6.VENTAS'], _to_float(economics.get('BRENT')) or 0.0)

    return {'economics': economics, 'sheets': sheets}


# ---------------------------------------------------------------------------
# Construccion del Excel temporal que consume el solver
# ---------------------------------------------------------------------------

def construir_excel(base_template_path, sheets, economics, out_path):
    """
    Reconstruye un Parametros.xlsx valido para el solver.

    - Parte de `base_template_path` (preserva hojas auxiliares y rangos con nombre).
    - Recalcula las columnas derivadas.
    - Escribe en cada hoja del modelo los datos de `sheets` (dict hoja -> DataFrame).
    - Actualiza 0.ECONOMICS con los valores de `economics`.
    """
    sheets = {k: v.copy() for k, v in sheets.items()}
    recompute_derived(sheets, economics)

    wb = load_workbook(base_template_path)

    # Actualizar 0.ECONOMICS (INDEX/VALUE)
    if '0.ECONOMICS' in wb.sheetnames and economics:
        ws = wb['0.ECONOMICS']
        for row in range(2, ws.max_row + 1):
            key = ws.cell(row=row, column=1).value
            if key is not None and str(key).strip() in economics:
                ws.cell(row=row, column=2).value = economics[str(key).strip()]

    # Reescribir cada hoja del modelo
    for sheet_name in MODEL_SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        df = sheets[sheet_name]

        # Mapear columnas del DataFrame a columnas de la hoja por nombre de encabezado
        headers = {}
        for col_idx in range(1, ws.max_column + 1):
            h = ws.cell(row=1, column=col_idx).value
            if h is not None:
                headers[_norm(h)] = col_idx

        # Limpiar filas de datos previas
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)

        # Escribir filas nuevas
        for i, (_, row) in enumerate(df.iterrows()):
            excel_row = i + 2
            for col_name in df.columns:
                col_idx = headers.get(_norm(col_name))
                if col_idx is None:
                    continue
                value = row[col_name]
                if isinstance(value, float) and math.isnan(value):
                    value = None
                elif value == '':
                    value = None
                else:
                    value = _coerce_num(value)
                cell = ws.cell(row=excel_row, column=col_idx)
                cell.value = value
                # Un texto que empieza por '=' (p. ej. la columna documental
                # 'Formula de Precio de venta' = '=+BRENT+15') seria interpretado
                # por openpyxl como formula y se guardaria sin valor cacheado, con
                # lo que pandas lo leeria vacio. Lo forzamos a texto.
                if isinstance(value, str) and value.startswith('='):
                    cell.data_type = 's'

    wb.save(out_path)
    return out_path


# ---------------------------------------------------------------------------
# Serializacion a JSON (para persistir en BD y editar en la pagina)
# ---------------------------------------------------------------------------

def _py(value):
    """Convierte un valor de pandas/numpy a un tipo JSON-serializable nativo."""
    if value is None:
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass
    if hasattr(value, 'item'):  # escalar de numpy
        try:
            return value.item()
        except Exception:
            pass
    return value


def columnas_editables(sheet_name, columns):
    """Columnas que el usuario edita en la pagina (todas menos las derivadas)."""
    derived = {_norm(c) for c in DERIVED_COLUMNS.get(sheet_name, [])}
    return [c for c in columns if _norm(c) not in derived]


def _sheets_a_payload(sheets, economics):
    """Serializa {hoja: DataFrame} + economics a la estructura JSON de la pagina."""
    sheets_payload = {}
    for name, df in sheets.items():
        cols = [str(c) for c in df.columns]
        rows = [{str(c): _py(r[c]) for c in df.columns} for _, r in df.iterrows()]
        sheets_payload[name] = {
            'columns': cols,
            'editable': columnas_editables(name, cols),
            'rows': rows,
        }
    econ = {str(k): _py(v) for k, v in (economics or {}).items()}
    return {'economics': econ, 'sheets': sheets_payload}


def escenario_a_payload(excel_path):
    """
    Lee un Excel base y devuelve una estructura JSON-serializable lista para la
    pagina / BD:
        {
          'economics': {'BRENT': 107, 'TRM': 3800, ...},
          'sheets': {
             '3.COMPRAS': {'columns': [...], 'editable': [...], 'rows': [{col: val}, ...]},
             ...
          }
        }
    """
    data = extraer_escenario(excel_path)
    recompute_derived(data['sheets'], data['economics'])
    return _sheets_a_payload(data['sheets'], data['economics'])


def escenario_vacio(base_template_path=None):
    """Payload con la MISMA estructura (columnas) pero SIN filas, para empezar de cero.

    Sin `base_template_path` la estructura sale de `modelo_esquema`, asi que se
    puede crear un escenario en blanco aunque no exista ningun Excel. Si se pasa
    una ruta, se toma la estructura y los valores economicos de ese archivo.
    """
    if base_template_path:
        data = extraer_escenario(base_template_path)
        vacios = {name: df.iloc[0:0].copy() for name, df in data['sheets'].items()}
        return _sheets_a_payload(vacios, data['economics'])

    from modelo_esquema import payload_vacio
    return payload_vacio()


def recalcular_payload(payload):
    """Recibe un payload editado y devuelve el mismo con las columnas derivadas
    recalculadas (autoritativo: misma logica que usa el solver)."""
    economics, sheets = payload_a_sheets(payload)
    recompute_derived(sheets, economics)
    out = _sheets_a_payload(sheets, economics)
    out['economics'] = {str(k): _py(v) for k, v in (payload.get('economics') or {}).items()}
    return out


def payload_a_sheets(payload):
    """Convierte el payload JSON de la pagina/BD a (economics, {hoja: DataFrame}).

    Devuelve tambien las hojas auxiliares (precios); el solver ignora las que
    no necesita.
    """
    sheets = {}
    for name in list(MODEL_SHEETS) + [h for h in payload.get('sheets', {})
                                      if h not in MODEL_SHEETS]:
        sp = payload['sheets'][name]
        cols = sp['columns']
        rows = [{c: row.get(c) for c in cols} for row in sp['rows']]
        sheets[name] = pd.DataFrame(rows, columns=cols)
    return payload.get('economics', {}), sheets


def construir_excel_desde_payload(base_template_path, payload, out_path):
    """Atajo: payload JSON -> Excel temporal valido para el solver."""
    economics, sheets = payload_a_sheets(payload)
    return construir_excel(base_template_path, sheets, economics, out_path)


# ---------------------------------------------------------------------------
# Auto-validacion: ¿el motor de derivacion reproduce el archivo real?
# ---------------------------------------------------------------------------

def _self_test(excel_path):
    import warnings
    warnings.filterwarnings('ignore')
    data = extraer_escenario(excel_path)
    economics, original = data['economics'], data['sheets']
    print(f"Economics: { {k: economics[k] for k in list(economics)[:5]} }")

    # Copia con derivadas borradas, luego recalculadas
    sheets = {k: v.copy() for k, v in original.items()}
    for sn, cols in DERIVED_COLUMNS.items():
        for col in cols:
            real = _find_col(sheets[sn], col)
            if real is not None:
                sheets[sn][real] = ''
    recompute_derived(sheets, economics)

    total_ok = total = 0
    for sn, cols in DERIVED_COLUMNS.items():
        for col in cols:
            real = _find_col(original[sn], col)
            if real is None:
                continue
            # Alinear por el indice original: recompute_derived puede eliminar
            # filas (p. ej. duplicados exactos en 11.REL_CRUDO_MEZCLA), y comparar
            # por posicion desalinearia todo lo que viene despues.
            b = sheets[sn][_find_col(sheets[sn], col)]
            a = original[sn][real].reindex(b.index)
            n = len(b); ok = bad = 0; ex = None
            for k in range(n):
                x, y = a.iloc[k], b.iloc[k]
                if _is_blank(x) and _is_blank(y):
                    ok += 1; continue
                fx, fy = _to_float(x), _to_float(y)
                if fx is not None and fy is not None:
                    if abs(fx - fy) <= 1e-2: ok += 1
                    else: bad += 1; ex = ex or (k, x, y)
                elif str(x).strip() == str(y).strip():
                    ok += 1
                else:
                    bad += 1; ex = ex or (k, x, y)
            total_ok += ok; total += n
            flag = 'OK ' if bad == 0 else 'XX '
            extra = f' | fila{ex[0]} cache={ex[1]!r} calc={ex[2]!r}' if ex else ''
            print(f"  {flag}{sn} :: {col}: {ok}/{n}{extra}")
    print(f"\nTOTAL: {total_ok}/{total} celdas derivadas reproducidas "
          f"({100.0 * total_ok / total:.2f}%)")


if __name__ == '__main__':
    import sys
    path = sys.argv[1] if len(sys.argv) > 1 else (
        r'C:/Users/Juan Diego Ayala/OneDrive - conquerstrading/'
        r'Documentos/ACTIVIDADES FINANCIERAS/MODELO MATEMATICO/Parametros.xlsx'
    )
    _self_test(path)

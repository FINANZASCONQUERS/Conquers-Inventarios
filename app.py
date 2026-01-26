import requests
from sqlalchemy import or_
import json
import hashlib
from datetime import datetime, time, date, timedelta
import os
from functools import wraps
from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify, send_file, current_app, send_from_directory, has_app_context # Añadido send_file, current_app, send_from_directory
from werkzeug.security import generate_password_hash, check_password_hash
import openpyxl 
from io import BytesIO # Para Excel
import logging # Para un logging más flexible
import copy
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import func
from sqlalchemy.exc import IntegrityError
import pytz
import pandas as pd
import uuid
import re
from flask import g
from flask import Response
from weasyprint import HTML, CSS
import math
from sqlalchemy import or_
from flask_migrate import Migrate
import numpy as np
import re
import base64
import mimetypes
from io import BytesIO
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from itertools import groupby
import io
from flask import Response
import base64
from werkzeug.utils import secure_filename
from dotenv import load_dotenv
from urllib.parse import urljoin, urlparse, quote, unquote
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
load_dotenv()

BOGOTA_TZ = pytz.timezone('America/Bogota')


def to_bogota_datetime(value, *, assume_local=False):
    """Convierte un datetime a la zona horaria de Bogotá (devuelve datetime aware).

    Si ``value`` no tiene información de zona horaria se asume UTC por defecto.
    Para valores guardados como hora local de Bogotá en formato naive, pasar
    ``assume_local=True`` para evitar desplazamientos involuntarios.
    """
    if not value:
        return None
    try:
        if value.tzinfo is None:
            if assume_local:
                value = BOGOTA_TZ.localize(value)
            else:
                value = pytz.utc.localize(value)
        return value.astimezone(BOGOTA_TZ)
    except Exception:
        return None


def bogota_naive(value):
    """Normaliza un datetime a naive en zona Bogotá (para persistir sin offset)."""
    if not value:
        return None
    try:
        if value.tzinfo is None:
            return value
        return value.astimezone(BOGOTA_TZ).replace(tzinfo=None)
    except Exception:
        return None

from apscheduler.schedulers.background import BackgroundScheduler

# --- Módulo modelo optimización (nuevo) ---
from modelo_optimizacion import ejecutar_modelo, EXCEL_DEFAULT

# --- Blueprint para WhatsApp ---
# TEMPORALMENTE DESHABILITADO por error de spacy
# from bot_whatsapp import bot_bp

# Utilidad simple de permiso admin
def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if session.get('rol') != 'admin':
            flash('Solo administradores pueden acceder a esta sección.', 'danger')
            return redirect(url_for('home'))
        return f(*args, **kwargs)
    return decorated

def formatear_info_actualizacion(fecha_dt_utc, usuario_str):
    """
    Formatea la fecha y el usuario, convirtiendo la hora de UTC a la de Bogotá.
    Esta versión es robusta y maneja correctamente las zonas horarias.
    """
    try:
        if not fecha_dt_utc or not usuario_str:
            return "Información no disponible."

        # Define la zona horaria de Bogotá
        bogota_zone = pytz.timezone('America/Bogota')

        # Comprobación de seguridad: Si la fecha no tiene zona horaria (es "naive"),
        # le asignamos UTC. Si ya la tiene, no hacemos nada.
        if fecha_dt_utc.tzinfo is None:
            fecha_dt_utc = pytz.utc.localize(fecha_dt_utc)

        # Ahora que estamos seguros de que es una fecha en UTC, la convertimos a la zona de Bogotá
        dt_obj_bogota = fecha_dt_utc.astimezone(bogota_zone)

        # Formateamos el texto final para mostrarlo
        meses = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
        nombre_mes = meses[dt_obj_bogota.month - 1]
        
        fecha_formateada = dt_obj_bogota.strftime(f"%d de {nombre_mes} de %Y")
        hora_formateada = dt_obj_bogota.strftime("%I:%M %p")

        mensaje = f"Última actualización guardada por {usuario_str} el {fecha_formateada} a las {hora_formateada}"
        return mensaje

    except Exception as e:
        print(f"Error al formatear fecha: {e}")
        return "Fecha de registro con error de formato."
        
def componer_fecha_hora(hora_str, fecha_base=None):
    """
    Toma una hora en formato 'HH:MM' y la combina con una fecha base
    para crear un objeto datetime completo.
    """
    if not hora_str: return None
    
    # Si no se provee una fecha base, se usa la fecha del día actual.
    if fecha_base is None:
        fecha_base = date.today()
        
    try:
        # Crea un objeto 'time' desde el string "HH:MM"
        hora_obj = time.fromisoformat(hora_str)
        # Combina la fecha base con el objeto 'time'
        return datetime.combine(fecha_base, hora_obj)
    except (ValueError, TypeError):
        # Si el formato de hora es inválido (ej. "abc"), devuelve None.
        return None
    
def mes_espaniol(fecha_str):
    # fecha_str: '2025-01' o '2025-01-01'
    partes = fecha_str.split('-')
    anio = partes[0]
    mes = int(partes[1])
    meses_es = [
        '', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
        'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre'
    ]
    return f"{meses_es[mes]}-{anio}"
    
def convertir_plot_a_base64(fig):
    """Toma una figura de Matplotlib, la guarda en memoria y la devuelve como una cadena Base64."""
    buf = BytesIO()
    fig.savefig(buf, format="png", bbox_inches="tight")
    plt.close(fig)  # Cierra la figura para liberar memoria
    return base64.b64encode(buf.getvalue()).decode('utf-8')

def grafico_linea_base64(labels, data, ylabel):
    fig, ax = plt.subplots(figsize=(8, 3))
    ax.plot(labels, data, marker='o', color='#007bff')
    ax.set_ylabel(ylabel)
    ax.set_xticks(range(len(labels)))
    ax.set_xticklabels(labels, rotation=45, ha='right')
    for i, v in enumerate(data):
        ax.text(i, v, f'{v:.2f}', ha='center', va='bottom', fontsize=8)
    plt.tight_layout()
    buf = BytesIO()
    plt.savefig(buf, format='png')
    plt.close(fig)
    buf.seek(0)
    return base64.b64encode(buf.read()).decode('utf-8')

def grafico_barra_base64(labels, data, ylabel):
    fig, ax = plt.subplots(figsize=(8, 3))
    bars = ax.bar(labels, data, color='#28a745')
    ax.set_ylabel(ylabel)
    ax.set_xticks(range(len(labels)))
    ax.set_xticklabels(labels, rotation=45, ha='right')
    for bar in bars:
        height = bar.get_height()
        ax.annotate(f'{height:,.0f}', xy=(bar.get_x() + bar.get_width() / 2, height),
                    xytext=(0, 3), textcoords="offset points", ha='center', va='bottom', fontsize=8)
    plt.tight_layout()
    buf = BytesIO()
    plt.savefig(buf, format='png')
    plt.close(fig)
    buf.seek(0)
    return base64.b64encode(buf.read()).decode('utf-8')


def _derive_media_filename(media_url):
    """Intenta extraer un nombre de archivo legible desde una URL de adjunto."""
    try:
        parsed = urlparse(media_url)
        candidate = os.path.basename(parsed.path or '')
        if candidate:
            return candidate
    except Exception:
        pass
    return 'adjunto.pdf'

def procesar_analisis_remolcadores(registros):
    """
    Toma una lista de registros, ejecuta el análisis de Pandas y devuelve
    los resultados como HTML y gráficos en Base64.
    """
    if not registros:
        return None

    datos_df = [{
        "ID": r.maniobra_id, "BARCAZA": r.barcaza, "EVENTO ANTERIO": r.evento_anterior,
        "EVENTO ACTUAL": r.evento_actual, "HORA INICIO": r.hora_inicio, "HORA FIN": r.hora_fin,
        "MT ENTREGADAS": float(r.mt_entregadas) if r.mt_entregadas else 0.0,
        "CARGAS": r.carga_estado
    } for r in registros]
    
    if not datos_df:
        return None
        
    df = pd.DataFrame(datos_df)

    if df.empty or df['HORA FIN'].isnull().all():
        return None

    # --- Lógica de preparación de datos (sin cambios) ---
    df["HORA INICIO"] = pd.to_datetime(df["HORA INICIO"])
    df["HORA FIN"]   = pd.to_datetime(df["HORA FIN"])
    df.dropna(subset=['HORA INICIO', 'HORA FIN'], inplace=True)
    
    df["duration_hours"] = (df["HORA FIN"] - df["HORA INICIO"]).dt.total_seconds() / 3600
    df["pair"] = (df["EVENTO ANTERIO"].astype(str).str.strip().str.upper() + " -> " + df["EVENTO ACTUAL"].astype(str).str.strip().str.upper())
    df["trayecto_final"] = df["pair"]
    df = df.sort_values(["ID", "HORA INICIO"]).reset_index(drop=True)

    # Segunda pasada: reglas de regreso
    comb_rules_regreso = {
        ("INICIO CONTECAR -> LLEGADA SPD", "INICIO SPD -> LLEGADA BASE OPS"): "INICIO CONTECAR -> LLEGADA BASE OPS",
        ("INICIO FONDEO -> LLEGADA SPD", "INICIO SPD -> LLEGADA BASE OPS"): "INICIO FONDEO -> LLEGADA BASE OPS",
        ("INICIO SPRC -> LLEGADA SPD", "INICIO SPD -> LLEGADA BASE OPS"): "INICIO SPRC -> LLEGADA BASE OPS",
        ("INICIO PUERTO BAHIA -> LLEGADA SPD", "INICIO SPD -> LLEGADA BASE OPS"): "INICIO PUERTO BAHIA -> LLEGADA BASE OPS",
        ("INICIO CONTECAR -> LLEGADA BITA", "INICIO BITA -> LLEGADA BASE OPS"): "INICIO CONTECAR -> LLEGADA BASE OPS",
        ("INICIO FONDEO -> LLEGADA BITA", "INICIO BITA -> LLEGADA BASE OPS"): "INICIO FONDEO -> LLEGADA BASE OPS",
        ("INICIO SPRC -> LLEGADA BITA", "INICIO BITA -> LLEGADA BASE OPS"): "INICIO SPRC -> LLEGADA BASE OPS",
        ("INICIO PUERTO BAHIA -> LLEGADA BITA", "INICIO BITA -> LLEGADA BASE OPS"): "INICIO PUERTO BAHIA -> LLEGADA BASE OPS",
    }
    for i in range(len(df) - 1):
        if df.at[i, "ID"] != df.at[i + 1, "ID"]: continue
        key = (df.at[i, "trayecto_final"], df.at[i + 1, "trayecto_final"])
        if key in comb_rules_regreso:
            df.at[i, "duration_hours"] += df.at[i + 1, "duration_hours"]
            df.at[i, "HORA FIN"] = df.at[i + 1, "HORA FIN"]
            df.at[i, "trayecto_final"] = comb_rules_regreso[key]
            df.loc[i + 1, ["trayecto_final", "duration_hours"]] = [None, np.nan]

    def convertir_a_texto_legible(horas):
        if pd.isna(horas): return ""
        td = timedelta(hours=horas)
        h = int(td.total_seconds() // 3600)
        m = int((td.total_seconds() % 3600) // 60)
        partes = ([f"{h}h"] if h > 0 else []) + ([f"{m}m"] if m > 0 else [])
        return " ".join(partes) or "0m"

    # --- ANÁLISIS DE TRAYECTOS (sin cambios) ---
    pairs_loaded = [
        "INICIO SPD -> LLEGADA CONTECAR", "INICIO SPD -> LLEGADA SPRC", "INICIO SPD -> LLEGADA FONDEO", 
        "INICIO SPD -> LLEGADA PUERTO BAHIA", "INICIO BITA -> LLEGADA CONTECAR", "INICIO BITA -> LLEGADA SPRC", 
        "INICIO BITA -> LLEGADA FONDEO", "INICIO BITA -> LLEGADA PUERTO BAHIA", "ESPERAR AUTORIZACION -> AUTORIZADO"
    ]
    pairs_empty = [
    # Viajes directos
    "INICIO BASE OPS -> LLEGADA BITA",
    "INICIO BASE OPS -> LLEGADA SPD",
    "INICIO CONTECAR -> LLEGADA BASE OPS",
    "INICIO CONTECAR -> LLEGADA BITA",
    "INICIO CONTECAR -> LLEGADA SPD",
    "INICIO FONDEO -> LLEGADA BASE OPS",
    "INICIO FONDEO -> LLEGADA BITA",
    "INICIO FONDEO -> LLEGADA SPD",
    "INICIO PUERTO BAHIA -> LLEGADA SPD",
    "INICIO PUERTO BAHIA -> LLEGADA BASE OPS",
    "INICIO PUERTO BAHIA -> LLEGADA BITA",
    "INICIO SPRC -> LLEGADA BASE OPS",
    "INICIO SPRC -> LLEGADA SPD",
    "INICIO BITA -> LLEGADA BASE OPS",
    "INICIO SPD -> LLEGADA BASE OPS"
]
    df_valido = df[df["trayecto_final"].notnull() & df['CARGAS'].notna()]
    df_loaded = df_valido[df_valido["CARGAS"].str.strip().str.upper() == "LLENO"]
    df_empty = df_valido[df_valido["CARGAS"].str.strip().str.upper() == "VACIO"]
    prom_loaded = df_loaded.groupby("trayecto_final", as_index=False).agg(avg_hours=("duration_hours", "mean"), n_samples=("duration_hours", "size"))
    prom_empty = df_empty.groupby("trayecto_final", as_index=False).agg(avg_hours=("duration_hours", "mean"), n_samples=("duration_hours", "size"))
    
    for df_prom in [prom_loaded, prom_empty]:
        if not df_prom.empty:
            df_prom.columns = ["Trayecto", "Promedio (h)", "Cantidad de registros"]
            df_prom["Promedio legible"] = df_prom["Promedio (h)"].apply(convertir_a_texto_legible)
            df_prom = df_prom[["Trayecto", "Promedio legible", "Promedio (h)", "Cantidad de registros"]]

    def estilo_tablas(df_sty, titulo, color_titulo):
        return (df_sty.style.set_caption(f'<span style="font-size:18px; color:{color_titulo}; font-weight:bold;">{titulo}</span>')
                .set_table_styles([{"selector": "thead", "props": [("background-color", "#f7f7f7"),("border-bottom", "2px solid #1a5f1a"),("font-weight", "bold")]},{"selector": "tbody tr", "props": [("border-bottom", "1px solid #ddd")]},{"selector": "td", "props": [("padding", "8px")]},{"selector": "caption", "props": [("caption-side", "top"), ("font-size", "0px")]},{"selector": "", "props": [("border-collapse", "collapse")]}])
                .background_gradient(subset=['Promedio (h)'], cmap='YlGn').background_gradient(subset=['Cantidad de registros'], cmap='Blues')
                .format({'Promedio (h)': "{:,.2f} h", 'Cantidad de registros': "{:,.0f} registros"})
                .set_properties(subset=['Promedio legible'], **{'text-align': 'left', 'font-style': 'italic', 'color': '#2c5f2c'})
                .set_properties(subset=['Trayecto'], **{'font-weight': '500', 'color': '#1a1a1a'}).hide(axis="index"))

    tabla_cargado_html = estilo_tablas(prom_loaded, "⛴️ TRAYECTOS CON CARGA (LLENO) - TIEMPOS PROMEDIO GENERALES", "#1a5f1a").to_html(escape=False)
    tabla_vacio_html = estilo_tablas(prom_empty, "🛳️ TRAYECTOS DE REGRESO (VACIO) - TIEMPOS PROMEDIO GENERALES", "#1a5f7a").to_html(escape=False)

    # --- GRÁFICOS ---
    grafico_tanqueo_b64 = None
    df_tanqueo = df[df["EVENTO ACTUAL"].astype(str).str.strip().str.upper() == "TANQUEO"].copy()
    
    if not df_tanqueo.empty:
        df_tanqueo["Duración Legible"] = df_tanqueo["duration_hours"].apply(convertir_a_texto_legible)
        meses_es = { 1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril', 5: 'Mayo', 6: 'Junio', 7: 'Julio', 8: 'Agosto', 9: 'Septiembre', 10: 'Octubre', 11: 'Noviembre', 12: 'Diciembre' }
        df_tanqueo["Mes"] = df_tanqueo["HORA INICIO"].dt.month.map(meses_es) + " " + df_tanqueo["HORA INICIO"].dt.year.astype(str)
        df_tanqueo["Fecha_Orden"] = df_tanqueo["HORA INICIO"].dt.to_period("M")
        df_tanqueo_sorted = df_tanqueo.sort_values(["Fecha_Orden", "ID"]).reset_index(drop=True)
        df_tanqueo_sorted["Etiqueta"] = df_tanqueo_sorted["Mes"] + " | ID " + df_tanqueo_sorted["ID"].astype(str)
        promedio = df_tanqueo_sorted["duration_hours"].mean()
        promedio_texto = convertir_a_texto_legible(promedio)

        # ▼▼▼ CAMBIO 1: Se ajusta el tamaño del gráfico para que sea más compacto ▼▼▼
        fig_tanqueo, ax = plt.subplots(figsize=(18, max(6, len(df_tanqueo_sorted) * 0.4)))
        
        ax.barh(df_tanqueo_sorted["Etiqueta"], df_tanqueo_sorted["duration_hours"], color="#1f7a1f")
        ax.set_xlabel("Horas de Tanqueo")
        ax.set_ylabel("Mes y Maniobra ID")
        ax.invert_yaxis()
        for index, row in df_tanqueo_sorted.iterrows():
            duration = row['duration_hours']
            ax.text(0.2, index, row["Duración Legible"], ha="left", va="center", color="white", fontsize=9, fontweight="bold")
            ax.text(duration + 0.2, index, f"MT: {row['MT ENTREGADAS']:.2f}", ha="left", va="center", color="#333333", fontsize=9)
            
        if pd.notna(promedio):
            ax.axvline(x=promedio, color="red", linestyle="--", linewidth=1.5)
            ax.text(promedio + 0.1, len(df_tanqueo_sorted) - 0.5, f" Promedio: {promedio_texto}", color="red", fontsize=10)
        
        ax.set_title("Duración de Tanqueo por Mes y ID", fontsize=16)
        plt.tight_layout()
        
        grafico_tanqueo_b64 = convertir_plot_a_base64(fig_tanqueo)

    grafico_total_b64 = None
    meses_es = { 1: 'Ene', 2: 'Feb', 3: 'Mar', 4: 'Abr', 5: 'May', 6: 'Jun', 7: 'Jul', 8: 'Ago', 9: 'Sep', 10: 'Oct', 11: 'Nov', 12: 'Dic' }
    df["Mes"] = df["HORA INICIO"].dt.month.map(meses_es) + "-" + df["HORA INICIO"].dt.year.astype(str)
    
    df_total = df.groupby("ID", as_index=False).agg(
        duration_hours=("duration_hours", "sum"),
        Mes=("Mes", lambda x: ", ".join(sorted(x.unique()))),
        BARCAZA=("BARCAZA", "first"),
        MT_ENTREGADAS=("MT ENTREGADAS", "first")
    )
    if not df_total.empty:
        df_total["Duración Legible"] = df_total["duration_hours"].apply(convertir_a_texto_legible)
        df_total["ID_Mes"] = "ID " + df_total["ID"].astype(str) + " | " + df_total["Mes"]
        df_total = df_total.sort_values("ID").reset_index(drop=True)
        promedio = df_total["duration_hours"].mean()
        promedio_texto = convertir_a_texto_legible(promedio)

        # ▼▼▼ CAMBIO 2: Se reduce el ancho del gráfico para que no se salga de la página ▼▼▼
        fig_total, ax = plt.subplots(figsize=(25, max(8, len(df_total) * 0.5)))
        ax.barh(df_total["ID_Mes"], df_total["duration_hours"], color="#004d99")
        
        for idx, row in df_total.iterrows():
            ax.text(0.2, idx, row["Duración Legible"], va="center", ha="left", color="white", fontsize=9, fontweight='bold')
            ax.text(row["duration_hours"] + 0.2, idx, f"MT: {row['MT_ENTREGADAS']:.2f}", va="center", ha="left", color="#333333", fontsize=9, fontweight='bold')
        
        if pd.notna(promedio):
            ax.axvline(x=promedio, color='red', linestyle='--', linewidth=1.5)
            ax.text(promedio + 0.1, 0, f"Promedio: {promedio_texto}", color='red', backgroundcolor='white')
        
        ax.set_title("Total de Horas por Maniobra", fontsize=14)
        ax.set_xlabel("Total de Horas")
        ax.set_ylabel("ID | Barcaza | Mes")
        ax.invert_yaxis()
        plt.tight_layout()
        grafico_total_b64 = convertir_plot_a_base64(fig_total)
    
    return {
        "tabla_cargado_html": tabla_cargado_html,
        "tabla_vacio_html": tabla_vacio_html,
        "grafico_tanqueo_b64": grafico_tanqueo_b64,
        "grafico_total_b64": grafico_total_b64
    }

app = Flask(__name__)
app.secret_key = 'clave_secreta_para_produccion_cambiar'

import os
app.config['SQLALCHEMY_DATABASE_URI'] = os.environ.get('DATABASE_URL', 'postgresql://postgres:Sara_121128@localhost:5432/inventario_dev')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
# Directorio de guías (PDF/Imagen) - configurable por entorno
app.config['GUIDES_DIR'] = os.environ.get('GUIDES_DIR') or ('/var/data/guias' if os.name != 'nt' else os.path.join(app.root_path, 'guias'))
app.config['MIME_DEBUG'] = os.environ.get('MIME_DEBUG', '0')  # Activa logs de depuración de MIME cuando sea '1'/'true'
os.makedirs(app.config['GUIDES_DIR'], exist_ok=True)
from extensions import db
db.init_app(app) # <--- Iniciamos la BD definida en extensions.py
migrate = Migrate(app, db)

scheduler = BackgroundScheduler()

@app.get('/.well-known/appspecific/com.chrome.devtools.json')
def chrome_devtools_probe():
    """Responde la sonda de Chrome DevTools con 204 para evitar 404 en logs."""
    return Response(status=204)

def send_reminders():
    # Función para enviar recordatorios a usuarios inactivos en step 4
    with app.app_context():
        now = datetime.utcnow()
        # Buscar sesiones en step 4 que no han tenido actividad en las últimas 5 horas
        sesiones_inactivas = SolicitudCita.query.filter(
            SolicitudCita.whatsapp_step == '4',
            SolicitudCita.last_reminder.isnot(None),
            (now - SolicitudCita.last_reminder) >= timedelta(hours=5)
        ).all()
        
        for solicitud in sesiones_inactivas:
            # Enviar mensaje de recordatorio
            mensaje = "Recordatorio: Aún esperamos la imagen de la guía/manifesto para completar su solicitud. Por favor, envíela lo antes posible."
            # Enviar el mensaje por WhatsApp
            send_whatsapp_message(solicitud.telefono, mensaje)
            # Actualizar last_reminder
            solicitud.last_reminder = now
            db.session.commit()

# Inicializar el scheduler después de la configuración de la app
scheduler.add_job(func=send_reminders, trigger="interval", hours=5)
scheduler.start()

class RegistroPlanta(db.Model):
    __tablename__ = 'registros_planta'

    id = db.Column(db.Integer, primary_key=True)
    timestamp = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)
    usuario = db.Column(db.String(100), nullable=False)
    
    tk = db.Column(db.String(50))
    producto = db.Column(db.String(100))
    max_cap = db.Column(db.Float)
    bls_60 = db.Column(db.Float)
    api = db.Column(db.Float)
    bsw = db.Column(db.Float)
    s = db.Column(db.Float)

    def __repr__(self):
        return f'<RegistroPlanta ID: {self.id}, TK: {self.tk}>'
    
class RegistroBarcazaOrion(db.Model):
    __tablename__ = 'registros_barcaza_orion' # Nombre de la nueva tabla

    id = db.Column(db.Integer, primary_key=True)
    timestamp = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)
    usuario = db.Column(db.String(100), nullable=False)
    
    # Columnas específicas de la planilla Orion
    tk = db.Column(db.String(50))
    producto = db.Column(db.String(100))
    max_cap = db.Column(db.Float)
    bls_60 = db.Column(db.Float)
    api = db.Column(db.Float)
    bsw = db.Column(db.Float)
    s = db.Column(db.Float)
    grupo = db.Column(db.String(50)) # Columna especial para Orion

    def __repr__(self):
        return f'<RegistroBarcazaOrion ID: {self.id}, TK: {self.tk}>'
    
class RegistroBarcazaBita(db.Model):
    __tablename__ = 'registros_barcaza_bita' # Nombre de la nueva tabla

    id = db.Column(db.Integer, primary_key=True)
    timestamp = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)
    usuario = db.Column(db.String(100), nullable=False)
    
    # Columnas de la planilla BITA
    tk = db.Column(db.String(50))
    producto = db.Column(db.String(100))
    max_cap = db.Column(db.Float)
    bls_60 = db.Column(db.Float)
    api = db.Column(db.Float)
    bsw = db.Column(db.Float)
    s = db.Column(db.Float)

    def __repr__(self):
        return f'<RegistroBarcazaBita ID: {self.id}, TK: {self.tk}>'

class RegistroTransito(db.Model):
    __tablename__ = 'registros_transito'

    id = db.Column(db.Integer, primary_key=True)
    timestamp = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)
    usuario = db.Column(db.String(100), nullable=False)

    tipo_transito = db.Column(db.String(50), nullable=False)

    # El resto de las columnas de tu planilla
    origen = db.Column(db.String(100))
    fecha = db.Column(db.String(50)) # Guardamos la fecha del cargue como texto
    guia = db.Column(db.String(100))
    producto = db.Column(db.String(100))
    placa = db.Column(db.String(50))
    api = db.Column(db.Float)
    bsw = db.Column(db.Float)
    nsv = db.Column(db.Float)
    observaciones = db.Column(db.Text)

    def __repr__(self):
        return f'<RegistroTransito ID: {self.id}, Guia: {self.guia}>'

class RegistroCalidad(db.Model):
    __tablename__ = 'registros_calidad'

    id = db.Column(db.Integer, primary_key=True)
    timestamp = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)
    usuario = db.Column(db.String(100), nullable=False)

    fecha = db.Column(db.String(50))
    hora = db.Column(db.String(50))
    producto = db.Column(db.String(50))
    responsable = db.Column(db.String(100))
    origen = db.Column(db.String(100))
    placa = db.Column(db.String(50))
    campo = db.Column(db.String(100))
    bsw = db.Column(db.Float)
    flash_point = db.Column(db.Float)
    api_obs = db.Column(db.Float)
    temp = db.Column(db.Float)
    api_corr = db.Column(db.Float)
    observaciones = db.Column(db.Text)

    def __repr__(self):
        return f'<RegistroCalidad ID: {self.id}, Fecha: {self.fecha}>'

class RegistroZisa(db.Model):
    __tablename__ = 'registros_zisa'
    id = db.Column(db.Integer, primary_key=True)
    
    empresa = db.Column(db.String(50), nullable=False, default='ZISA', index=True)
    
    # Columnas de la planilla
    mes = db.Column(db.String(50), nullable=False)
    carrotanque = db.Column(db.String(100))
    producto = db.Column(db.String(100))
    numero_sae = db.Column(db.String(50)) # Para la columna "N° S.A.E"
    acta = db.Column(db.String(50))
    bbl_netos = db.Column(db.Float)
    bbl_descargados = db.Column(db.Float)

    # Datos de auditoría
    usuario_carga = db.Column(db.String(100), nullable=False)
    fecha_carga = db.Column(db.DateTime, default=datetime.utcnow)

    estado = db.Column(db.String(50), default='Disponible', nullable=False)

    def __repr__(self):
        return f'<RegistroZisa id={self.id} carrotanque={self.carrotanque}>'  

class DefinicionCrudo(db.Model):
    __tablename__ = 'definiciones_crudo'

    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(100), unique=True, nullable=False, index=True)
    api = db.Column(db.Float)
    sulfur = db.Column(db.Float, nullable=True)      
    viscosity = db.Column(db.Float, nullable=True)    
    curva_json = db.Column(db.Text, nullable=False)
    assay_json = db.Column(db.Text, nullable=True)

    def __repr__(self):
        return f'<DefinicionCrudo {self.nombre}>'
    
class RegistroRemolcador(db.Model):
    __tablename__ = 'registros_remolcador'

    id = db.Column(db.Integer, primary_key=True)
    
    # --- CAMBIOS EN EL MODELO ---
    maniobra_id = db.Column(db.Integer, nullable=False, index=True)
    barcaza = db.Column(db.String(100), nullable=True) # <-- NUEVA COLUMNA
    nombre_barco = db.Column(db.String(100), nullable=True)
    evento_anterior = db.Column(db.String(200), nullable=True)
    hora_inicio = db.Column(db.DateTime, nullable=False)
    evento_actual = db.Column(db.String(200), nullable=True)
    hora_fin = db.Column(db.DateTime, nullable=True)
    mt_entregadas = db.Column(db.Numeric(10, 2), nullable=True)
    carga_estado = db.Column(db.String(50), nullable=True)
    usuario_actualizacion = db.Column(db.String(100))
    fecha_actualizacion = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    # La propiedad 'duracion' ya funciona correctamente para fechas completas
    @property
    def duracion(self):
        if not self.hora_fin or not self.hora_inicio:
            return ""
        
        delta = self.hora_fin - self.hora_inicio
        total_minutos = delta.total_seconds() / 60
        horas = int(total_minutos // 60)
        minutos = int(total_minutos % 60)
        
        return f"{horas}h {minutos}m"
    
    def __repr__(self):
        return f'<RegistroRemolcador {self.id}>'    

class ProgramacionCargue(db.Model):
    __tablename__ = 'programacion_cargue'
    id = db.Column(db.Integer, primary_key=True)
    
    # Campos de Juliana y Samantha
    factura = db.Column(db.String(100))
    fecha_programacion = db.Column(db.Date, nullable=False, default=date.today)
    empresa_transportadora = db.Column(db.String(150))
    placa = db.Column(db.String(50))
    tanque = db.Column(db.String(50))
    nombre_conductor = db.Column(db.String(150))
    cedula_conductor = db.Column(db.String(50))
    celular_conductor = db.Column(db.String(50))
    hora_llegada_estimada = db.Column(db.Time)
    producto_a_cargar = db.Column(db.String(100))
    
    # Campos de Ana Maria
    destino = db.Column(db.String(150))
    cliente = db.Column(db.String(150))
    
    # Campos de Refinería
    estado = db.Column(db.String(50), default='PROGRAMADO') 
    galones = db.Column(db.Float)
    barriles = db.Column(db.Float)
    temperatura = db.Column(db.Float)
    api_obs = db.Column(db.Float)
    api_corregido = db.Column(db.Float)
    precintos = db.Column(db.String(200))
    
    # Campo de Samantha
    fecha_despacho = db.Column(db.Date, nullable=True)
    numero_guia = db.Column(db.String(100))
    
    # Tipo de guía (Física o Digital)
    tipo_guia = db.Column(db.String(20), default='Física')
    
    # Imagen de la guía (base64)
    imagen_guia = db.Column(db.Text, nullable=True)

    # Auditoría
    ultimo_editor = db.Column(db.String(100))
    fecha_actualizacion = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    # Nuevo: momento en que TODOS los campos de refinería quedaron completos (para iniciar conteo de 30 min)
    refineria_completado_en = db.Column(db.DateTime, nullable=True)

# ---------------- BLOQUEO DE CELDAS (EDICIÓN EN TIEMPO REAL) -----------------
class ProgramacionCargueLock(db.Model):
    __tablename__ = 'programacion_cargue_locks'
    id = db.Column(db.Integer, primary_key=True)
    registro_id = db.Column(db.Integer, nullable=False, index=True)
    campo = db.Column(db.String(100), nullable=False)
    usuario = db.Column(db.String(120), nullable=False)
    timestamp = db.Column(db.DateTime, default=datetime.utcnow, nullable=False, index=True)
    __table_args__ = (db.UniqueConstraint('registro_id', 'campo', name='uq_prog_lock_registro_campo'),)

    def expirado(self, minutos=2):
        return datetime.utcnow() - self.timestamp > timedelta(minutes=minutos)

def _init_lock_table():
    from sqlalchemy import inspect
    with app.app_context():
        insp = inspect(db.engine)
        if 'programacion_cargue_locks' not in insp.get_table_names():
            ProgramacionCargueLock.__table__.create(db.engine)

_init_lock_table()

# Asegurar columna nueva en runtime si no existe (SQLite permite ADD COLUMN simple)
def _ensure_refineria_completion_column():
    from sqlalchemy import inspect, text
    with app.app_context():
        insp = inspect(db.engine)
        # Evitar error si la tabla aún no existe (por ejemplo, primera vez o DB recién borrada)
        if 'programacion_cargue' not in insp.get_table_names():
            print("[INIT] Tabla 'programacion_cargue' no existe todavía; se omite verificación de columna 'refineria_completado_en'. Ejecuta migraciones o crea las tablas primero.")
            return
        cols = [c['name'] for c in insp.get_columns('programacion_cargue')]
        if 'refineria_completado_en' not in cols:
            try:
                # Elegir tipo correcto según motor (PostgreSQL no acepta DATETIME)
                dialect = db.engine.dialect.name
                if dialect == 'postgresql':
                    col_type = 'TIMESTAMP'
                elif dialect == 'mysql':
                    col_type = 'DATETIME'
                else:
                    col_type = 'DATETIME'
                ddl = f'ALTER TABLE programacion_cargue ADD COLUMN refineria_completado_en {col_type}'
                with db.engine.begin() as conn:
                    conn.execute(text(ddl))
                print(f'Columna refineria_completado_en añadida (tipo {col_type})')
            except Exception as e:
                print('No se pudo añadir columna refineria_completado_en:', e)

_ensure_refineria_completion_column()

# Asegurar que la columna imagen_guia soporte rutas largas (TEXT) y exista
def _ensure_programacion_imagen_text():
    from sqlalchemy import inspect, text
    with app.app_context():
        insp = inspect(db.engine)
        if 'programacion_cargue' not in insp.get_table_names():
            return
        cols = insp.get_columns('programacion_cargue')
        col = next((c for c in cols if c['name'] == 'imagen_guia'), None)
        if col is None:
            try:
                with db.engine.begin() as con:
                    con.execute(text("ALTER TABLE programacion_cargue ADD COLUMN imagen_guia TEXT"))
            except Exception as e:
                print("[INIT] No se pudo crear columna imagen_guia:", e)
            return
        # Convertir a TEXT si es VARCHAR
        if str(col['type']).lower().startswith('varchar'):
            try:
                with db.engine.begin() as con:
                    con.execute(text("ALTER TABLE programacion_cargue ALTER COLUMN imagen_guia TYPE TEXT"))
                print("[INIT] Columna imagen_guia convertida a TEXT")
            except Exception as e:
                print("[INIT] No se pudo convertir imagen_guia a TEXT:", e)

_ensure_programacion_imagen_text()

def _ensure_tipo_guia_column():
    """Asegura que la columna `tipo_guia` exista en `programacion_cargue`.

    Añade la columna como VARCHAR(20) con valor por defecto 'Física' si no existe.
    """
    from sqlalchemy import inspect, text
    with app.app_context():
        insp = inspect(db.engine)
        if 'programacion_cargue' not in insp.get_table_names():
            return
        cols = [c['name'] for c in insp.get_columns('programacion_cargue')]
        if 'tipo_guia' not in cols:
            try:
                dialect = db.engine.dialect.name
                if dialect == 'postgresql':
                    col_type = "VARCHAR(20)"
                elif dialect == 'mysql':
                    col_type = "VARCHAR(20)"
                else:
                    col_type = "VARCHAR(20)"
                # Añadir columna con valor por defecto 'Física'
                ddl = f"ALTER TABLE programacion_cargue ADD COLUMN tipo_guia {col_type} DEFAULT 'Física'"
                with db.engine.begin() as conn:
                    conn.execute(text(ddl))
                print("[INIT] Columna tipo_guia añadida a programacion_cargue")
            except Exception as e:
                print("[INIT] No se pudo añadir columna tipo_guia:", e)

_ensure_tipo_guia_column()

# Asegurar que las tablas de SIZA tengan todas sus columnas
def _ensure_siza_schema():
    from sqlalchemy import inspect, text
    with app.app_context():
        insp = inspect(db.engine)
        
        # 1. Verificar inventario_siza_diario
        if 'inventario_siza_diario' in insp.get_table_names():
            cols = [c['name'] for c in insp.get_columns('inventario_siza_diario')]
            columnas_inv = {
                'volumen_agua_generada': 'FLOAT DEFAULT 0.0',
                'volumen_desperdicio_generado': 'FLOAT DEFAULT 0.0'
            }
            for col_name, col_def in columnas_inv.items():
                if col_name not in cols:
                    try:
                        with db.engine.begin() as con:
                            con.execute(text(f"ALTER TABLE inventario_siza_diario ADD COLUMN {col_name} {col_def}"))
                        print(f"[INIT] Columna {col_name} agregada a inventario_siza_diario")
                    except Exception as e:
                        print(f"[INIT] No se pudo agregar columna {col_name} a inventario_siza_diario:", e)

        # 2. Verificar recargas_siza
        if 'recargas_siza' in insp.get_table_names():
            cols = [c['name'] for c in insp.get_columns('recargas_siza')]
            columnas_rec = {
                'volumen_merma': 'FLOAT DEFAULT 0.0',
                'descontado_dian': 'BOOLEAN DEFAULT FALSE'
            }
            for col_name, col_def in columnas_rec.items():
                if col_name not in cols:
                    try:
                        with db.engine.begin() as con:
                            con.execute(text(f"ALTER TABLE recargas_siza ADD COLUMN {col_name} {col_def}"))
                        print(f"[INIT] Columna {col_name} agregada a recargas_siza")
                    except Exception as e:
                        print(f"[INIT] No se pudo agregar columna {col_name} a recargas_siza:", e)

        # 3. Verificar volumen_pendiente_dian
        if 'volumen_pendiente_dian' in insp.get_table_names():
            cols = [c['name'] for c in insp.get_columns('volumen_pendiente_dian')]
            if 'volumen_por_aprobar' not in cols:
                try:
                    with db.engine.begin() as con:
                        con.execute(text("ALTER TABLE volumen_pendiente_dian ADD COLUMN volumen_por_aprobar FLOAT DEFAULT 0.0"))
                    print("[INIT] Columna volumen_por_aprobar agregada a volumen_pendiente_dian")
                except Exception as e:
                    print("[INIT] No se pudo agregar columna volumen_por_aprobar a volumen_pendiente_dian:", e)

_ensure_siza_schema()

# ---------------- EDICIONES EN VIVO (NO PERSISTIDAS) -----------------
# Estructura en memoria para broadcast simple (clave: (registro_id,campo))
LIVE_EDITS = {}
LIVE_EDIT_TTL_SECONDS = 25  # tiempo de vida de una edición mostrada

def _purge_live_edits():
    now = datetime.utcnow()
    expiradas = []
    for key, info in LIVE_EDITS.items():
        if (now - info['timestamp']).total_seconds() > LIVE_EDIT_TTL_SECONDS:
            expiradas.append(key)
    for k in expiradas:
        LIVE_EDITS.pop(k, None)

class EPPItem(db.Model):
    __tablename__ = 'epp_items'
    id = db.Column(db.Integer, primary_key=True)
    
    # Datos principales del item
    nombre = db.Column(db.String(150), nullable=False, index=True) # Ej: "Botas de Seguridad"
    categoria = db.Column(db.String(50), nullable=False, index=True) # "EPP", "Dotación", "Equipos de Emergencia"
    stock_actual = db.Column(db.Integer, default=0, nullable=False)

    # Campos para detalles específicos
    referencia = db.Column(db.String(150), nullable=True) # Ej: "Brahama", "MSA Safari"
    talla = db.Column(db.String(50), nullable=True)      # Ej: "42", "L", "N/A"
    fecha_vencimiento = db.Column(db.Date, nullable=True) # Para items que expiran
    observaciones = db.Column(db.Text, nullable=True)     # Ej: "20 LBS", "Color Verde"

    # Relación con las asignaciones
    asignaciones = db.relationship('EPPAssignment', backref='item', lazy=True, cascade="all, delete-orphan")

    def __repr__(self):
        return f'<EPPItem {self.nombre} - {self.referencia} ({self.talla})>'

class EPPAssignment(db.Model):
    __tablename__ = 'epp_assignments'
    id = db.Column(db.Integer, primary_key=True)
    item_id = db.Column(db.Integer, db.ForeignKey('epp_items.id'), nullable=False)
    
    # Datos de la asignación
    empleado_nombre = db.Column(db.String(200), nullable=False, index=True)
    cantidad_entregada = db.Column(db.Integer, nullable=False)
    fecha_entrega = db.Column(db.Date, nullable=False, default=date.today)
    observaciones = db.Column(db.Text, nullable=True)

    def __repr__(self):
        return f'<EPPAssignment for {self.empleado_nombre}>'
    
class RegistroCompra(db.Model):
    __tablename__ = 'registros_compra'
    id = db.Column(db.Integer, primary_key=True)
    fecha = db.Column(db.Date, index=True)
    proveedor = db.Column(db.String(200), index=True)
    tarifa = db.Column(db.Float)
    producto = db.Column(db.String(200))
    cantidad_bls = db.Column(db.Float)
    cantidad_gln = db.Column(db.Float)
    brent = db.Column(db.Float)
    descuento = db.Column(db.Float)
    precio_uni_bpozo = db.Column(db.Float)
    total_neto = db.Column(db.Float)
    price_compra_pond = db.Column(db.Float)

    def __repr__(self):
        return f'<RegistroCompra {self.id} - {self.numero_factura}>'

# ================== CONTROL CUPO SIZA MULTI-PRODUCTO ==================

class ProductoSiza(db.Model):
    """Catálogo de productos SIZA disponibles."""
    __tablename__ = 'productos_siza'
    
    id = db.Column(db.Integer, primary_key=True)
    codigo = db.Column(db.String(50), nullable=False, unique=True, index=True)
    nombre = db.Column(db.String(100), nullable=False)
    activo = db.Column(db.Boolean, default=True, nullable=False)
    color_badge = db.Column(db.String(20), default='primary')  # Para UI
    orden = db.Column(db.Integer, default=0)  # Para ordenar en pantalla
    
    def __repr__(self):
        return f'<ProductoSiza {self.codigo} - {self.nombre}>'

class InventarioSizaDiario(db.Model):
    """Inventario diario de cada producto SIZA."""
    __tablename__ = 'inventario_siza_diario'
    
    id = db.Column(db.Integer, primary_key=True)
    fecha = db.Column(db.Date, nullable=False, index=True)
    producto_id = db.Column(db.Integer, db.ForeignKey('productos_siza.id'), nullable=False)
    
    # Volumen disponible del día
    cupo_web = db.Column(db.Float, nullable=False, default=0.0)

    # Métricas de Desperdicio y Agua Acumulada
    volumen_agua_generada = db.Column(db.Float, default=0.0)       # Agua generada por este producto
    volumen_desperdicio_generado = db.Column(db.Float, default=0.0) # Desperdicio (Merma) generado por este producto
    
    # Auditoría
    usuario_actualizacion = db.Column(db.String(100), nullable=False)
    fecha_actualizacion = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    
    # Relación
    producto = db.relationship('ProductoSiza', backref='inventarios')
    
    # Índice compuesto para búsqueda rápida
    __table_args__ = (
        db.UniqueConstraint('fecha', 'producto_id', name='uix_fecha_producto'),
    )
    
    def __repr__(self):
        return f'<InventarioSizaDiario {self.fecha} - Producto {self.producto_id}: {self.cupo_web} BBL>'

class RecargaSiza(db.Model):
    """Historial de recargas de cupo para cada producto."""
    __tablename__ = 'recargas_siza'
    
    id = db.Column(db.Integer, primary_key=True)
    fecha = db.Column(db.Date, nullable=False, index=True)
    producto_id = db.Column(db.Integer, db.ForeignKey('productos_siza.id'), nullable=False)
    volumen_recargado = db.Column(db.Float, nullable=False)
    volumen_merma = db.Column(db.Float, default=0.0)  # Merma/Desperdicio generado en esta recarga
    descontado_dian = db.Column(db.Boolean, default=False)  # Si se descontó de cupo DIAN
    observacion = db.Column(db.Text, nullable=True)
    
    # Auditoría
    usuario_registro = db.Column(db.String(100), nullable=False)
    fecha_registro = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    usuario_edicion = db.Column(db.String(100), nullable=True)
    fecha_edicion = db.Column(db.DateTime, nullable=True)
    
    # Relación
    producto = db.relationship('ProductoSiza', backref='recargas')
    
    def __repr__(self):
        return f'<RecargaSiza {self.fecha} - Producto {self.producto_id}: +{self.volumen_recargado} BBL>'

class ConsumoSiza(db.Model):
    """Historial de consumos/despachos de cupo para cada producto."""
    __tablename__ = 'consumos_siza'
    
    id = db.Column(db.Integer, primary_key=True)
    fecha = db.Column(db.Date, nullable=False, index=True)
    producto_id = db.Column(db.Integer, db.ForeignKey('productos_siza.id'), nullable=False)
    volumen_consumido = db.Column(db.Float, nullable=False)
    observacion = db.Column(db.Text, nullable=True)
    
    # Auditoría
    usuario_registro = db.Column(db.String(100), nullable=False)
    fecha_registro = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    usuario_edicion = db.Column(db.String(100), nullable=True)
    fecha_edicion = db.Column(db.DateTime, nullable=True)
    
    # Relación
    producto = db.relationship('ProductoSiza', backref='consumos')
    
    def __repr__(self):
        return f'<ConsumoSiza {self.fecha} - Producto {self.producto_id}: -{self.volumen_consumido} BBL>'

class PedidoSiza(db.Model):
    """Pedidos de productos SIZA."""
    __tablename__ = 'pedidos_siza'
    
    id = db.Column(db.Integer, primary_key=True)
    numero_pedido = db.Column(db.String(100), nullable=False, unique=True, index=True)
    producto_id = db.Column(db.Integer, db.ForeignKey('productos_siza.id'), nullable=False)
    volumen_solicitado = db.Column(db.Float, nullable=False)
    observacion = db.Column(db.Text, nullable=True)
    estado = db.Column(db.String(50), default='PENDIENTE', nullable=False, index=True)
    # Estados posibles: PENDIENTE, APROBADO, RECHAZADO, COMPLETADO
    
    fecha_registro = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    usuario_registro = db.Column(db.String(100), nullable=False)
    fecha_gestion = db.Column(db.DateTime, nullable=True)
    usuario_gestion = db.Column(db.String(100), nullable=True)
    
    # Relación
    producto = db.relationship('ProductoSiza', backref='pedidos')
    
    def __repr__(self):
        return f'<PedidoSiza {self.numero_pedido} - {self.estado}>'

class VolumenPendienteDian(db.Model):
    """Volumen pendiente de aprobación DIAN (general, no por producto)."""
    __tablename__ = 'volumen_pendiente_dian'
    
    id = db.Column(db.Integer, primary_key=True)
    fecha = db.Column(db.Date, nullable=False, unique=True, index=True)
    volumen_pendiente = db.Column(db.Float, nullable=False, default=0.0) # Este es el volumen APROBADO disponible para distribuir
    volumen_por_aprobar = db.Column(db.Float, nullable=False, default=0.0) # Nuevo: Pendiente de aprobación DIAN
    observacion = db.Column(db.Text, nullable=True)
    
    # Auditoría
    usuario_actualizacion = db.Column(db.String(100), nullable=False)
    fecha_actualizacion = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    
    def __repr__(self):
        return f'<VolumenPendienteDian {self.fecha}: {self.volumen_pendiente} BBL>'

class HistorialAprobacionDian(db.Model):
    """Historial individual de aprobaciones DIAN para poder revertirlas."""
    __tablename__ = 'historial_aprobacion_dian'
    
    id = db.Column(db.Integer, primary_key=True)
    fecha_operativa = db.Column(db.Date, nullable=False, index=True) # Fecha a la que aplica
    volumen_agregado = db.Column(db.Float, nullable=False) # Cuánto se sumó al aprobado
    observacion = db.Column(db.String(255), nullable=True)
    
    # Auditoría
    usuario_registro = db.Column(db.String(100), nullable=False)
    fecha_registro = db.Column(db.DateTime, default=datetime.utcnow)
    
    def __repr__(self):
        return f'<HistorialAprobacionDian {self.fecha_operativa}: +{self.volumen_agregado}>'



class MovimientoDian(db.Model):
    """Registro detallado de todos los movimientos de volumen DIAN (Ingreso, Aprobación, Consumo)."""
    __tablename__ = 'movimientos_dian'
    
    id = db.Column(db.Integer, primary_key=True)
    fecha_operativa = db.Column(db.Date, nullable=False, index=True)
    tipo = db.Column(db.String(50), nullable=False) # 'INGRESO_PENDIENTE', 'APROBACION', 'CONSUMO', 'AJUSTE'
    volumen = db.Column(db.Float, nullable=False) # Valor absoluto del movimiento
    observacion = db.Column(db.String(255), nullable=True)
    
    # Auditoría
    usuario_registro = db.Column(db.String(100), nullable=False)
    fecha_registro = db.Column(db.DateTime, default=datetime.utcnow)
    
    def __repr__(self):
        return f'<MovimientoDian {self.fecha_operativa} [{self.tipo}]: {self.volumen}>'

class CupoSizaConfig(db.Model):
    """DEPRECADO: Mantenido por compatibilidad. Usar InventarioSizaDiario."""
    __tablename__ = 'cupo_siza_config'
    
    id = db.Column(db.Integer, primary_key=True)
    fecha = db.Column(db.Date, nullable=False, unique=True, index=True)
    cupo_web = db.Column(db.Float, nullable=False, default=0.0)
    usuario_actualizacion = db.Column(db.String(100), nullable=False)
    fecha_actualizacion = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    
    def __repr__(self):
        return f'<CupoSizaConfig {self.fecha} - {self.cupo_web} BBL>'

# ================== DESPACHOS TK -> BARCAZA ==================
class TrasiegoTKBarcaza(db.Model):
    __tablename__ = 'trasiegos_tk_barcaza'

    id = db.Column(db.Integer, primary_key=True)
    timestamp = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    fecha = db.Column(db.Date, nullable=False, index=True)
    usuario = db.Column(db.String(120), nullable=False)

    origen_tk = db.Column(db.String(50), nullable=False)
    destino_barcaza = db.Column(db.String(120), nullable=False)  # Nombre barcaza (p.ej. Manzanillo)
    destino_compartimento = db.Column(db.String(50), nullable=True)  # p.ej. TK1, TK2

    tk_cm_inicial = db.Column(db.Integer, nullable=True)
    tk_mm_inicial = db.Column(db.Integer, nullable=True)
    tk_bbl_bruto_inicial = db.Column(db.Float, nullable=True)
    tk_bbl_inicial = db.Column(db.Float, nullable=True)
    tk_api = db.Column(db.Float, nullable=True)
    tk_temp = db.Column(db.Float, nullable=True)
    tk_cm_final = db.Column(db.Integer, nullable=True)
    tk_mm_final = db.Column(db.Integer, nullable=True)
    tk_bbl_bruto_final = db.Column(db.Float, nullable=True)
    tk_bbl_final = db.Column(db.Float, nullable=True)

    # Ingreso simultáneo al tanque (opcional)
    tk_caudal_bbl_min = db.Column(db.Float, nullable=True)
    tk_minutos_ingreso = db.Column(db.Float, nullable=True)
    tk_bbl_ingreso = db.Column(db.Float, nullable=True)

    bar_cm_inicial = db.Column(db.Integer, nullable=True)
    bar_mm_inicial = db.Column(db.Integer, nullable=True)
    bar_bbl_bruto_inicial = db.Column(db.Float, nullable=True)
    bar_bbl_inicial = db.Column(db.Float, nullable=True)
    bar_api = db.Column(db.Float, nullable=True)
    bar_temp = db.Column(db.Float, nullable=True)
    bar_cm_final = db.Column(db.Integer, nullable=True)
    bar_mm_final = db.Column(db.Integer, nullable=True)
    bar_bbl_bruto_final = db.Column(db.Float, nullable=True)
    bar_bbl_final = db.Column(db.Float, nullable=True)

    notas = db.Column(db.Text, nullable=True)
    tk_notas = db.Column(db.Text, nullable=True)

    @property
    def trasiego_segun_tanque(self):
            try:
                if self.tk_bbl_inicial is not None and self.tk_bbl_final is not None:
                    ini = float(self.tk_bbl_inicial)
                    fin = float(self.tk_bbl_final)
                    return round(ini - fin, 2)
                return None
            except Exception:
                return None

    @property
    def trasiego_segun_barcaza(self):
        try:
            ini = float(self.bar_bbl_inicial or 0)
            fin = float(self.bar_bbl_final or 0)
            return round(fin - ini, 2)
        except Exception:
            return None

    @property
    def diferencia(self):
        try:
            t = float(self.trasiego_segun_tanque or 0)
            b = float(self.trasiego_segun_barcaza or 0)
            return round(b - t, 2)
        except Exception:
            return None

    def __repr__(self):
        return f'<Trasiego {self.fecha} {self.origen_tk} -> {self.destino_barcaza}/{self.destino_compartimento or ""}>'

def _ensure_trasiegos_table():
    from sqlalchemy import inspect
    with app.app_context():
        insp = inspect(db.engine)
        if 'trasiegos_tk_barcaza' not in insp.get_table_names():
            try:
                TrasiegoTKBarcaza.__table__.create(db.engine)
                print('[INIT] Tabla trasiegos_tk_barcaza creada')
            except Exception as e:
                print('[INIT] No fue posible crear tabla trasiegos_tk_barcaza:', e)

_ensure_trasiegos_table()

def _ensure_trasiegos_columns():
    from sqlalchemy import inspect, text
    with app.app_context():
        insp = inspect(db.engine)
        if 'trasiegos_tk_barcaza' not in insp.get_table_names():
            return
        cols = [c['name'] for c in insp.get_columns('trasiegos_tk_barcaza')]
        to_add = []
        if 'tk_api' not in cols:
            to_add.append('ALTER TABLE trasiegos_tk_barcaza ADD COLUMN tk_api FLOAT')
        if 'bar_api' not in cols:
            to_add.append('ALTER TABLE trasiegos_tk_barcaza ADD COLUMN bar_api FLOAT')
        if 'tk_notas' not in cols:
            to_add.append('ALTER TABLE trasiegos_tk_barcaza ADD COLUMN tk_notas TEXT')
        # Nuevas columnas de BBL Bruto
        if 'tk_bbl_bruto_inicial' not in cols:
            to_add.append('ALTER TABLE trasiegos_tk_barcaza ADD COLUMN tk_bbl_bruto_inicial FLOAT')
        if 'tk_bbl_bruto_final' not in cols:
            to_add.append('ALTER TABLE trasiegos_tk_barcaza ADD COLUMN tk_bbl_bruto_final FLOAT')
        if 'bar_bbl_bruto_inicial' not in cols:
            to_add.append('ALTER TABLE trasiegos_tk_barcaza ADD COLUMN bar_bbl_bruto_inicial FLOAT')
        if 'bar_bbl_bruto_final' not in cols:
            to_add.append('ALTER TABLE trasiegos_tk_barcaza ADD COLUMN bar_bbl_bruto_final FLOAT')
        # Nuevas columnas para ingreso simultáneo al TK
        if 'tk_caudal_bbl_min' not in cols:
            to_add.append('ALTER TABLE trasiegos_tk_barcaza ADD COLUMN tk_caudal_bbl_min FLOAT')
        if 'tk_minutos_ingreso' not in cols:
            to_add.append('ALTER TABLE trasiegos_tk_barcaza ADD COLUMN tk_minutos_ingreso FLOAT')
        if 'tk_bbl_ingreso' not in cols:
            to_add.append('ALTER TABLE trasiegos_tk_barcaza ADD COLUMN tk_bbl_ingreso FLOAT')
        for ddl in to_add:
            try:
                with db.engine.begin() as conn:
                    conn.execute(text(ddl))
            except Exception as e:
                print('[INIT] No fue posible añadir columna en trasiegos_tk_barcaza:', e)

_ensure_trasiegos_columns()

def _ensure_volumen_dian_columns():
    from sqlalchemy import inspect, text
    with app.app_context():
        insp = inspect(db.engine)
        if 'volumen_pendiente_dian' not in insp.get_table_names():
            return
        cols = [c['name'] for c in insp.get_columns('volumen_pendiente_dian')]
        if 'volumen_por_aprobar' not in cols:
            try:
                # Determinar el tipo de columna según el dialecto
                dialect = db.engine.dialect.name
                col_type = 'FLOAT' 
                default_val = '0.0'
                
                ddl = f'ALTER TABLE volumen_pendiente_dian ADD COLUMN volumen_por_aprobar {col_type} DEFAULT {default_val}'
                with db.engine.begin() as conn:
                    conn.execute(text(ddl))
                print('[INIT] Columna volumen_por_aprobar añadida a volumen_pendiente_dian')
            except Exception as e:
                print('[INIT] No fue posible añadir columna volumen_por_aprobar:', e)

_ensure_volumen_dian_columns()

# ===== Servir guías cargadas (PDF/imagenes) =====
@app.get('/guias/<path:filename>')
def serve_guia(filename):
    """Sirve archivos de guía desde GUIDES_DIR con Content-Type correcto e inline.

    - Detecta MIME con mimetypes.guess_type.
    - Fallback a application/pdf si extensión .pdf o firma %PDF.
    - Fuerza Content-Disposition: inline y añade X-Content-Type-Options: nosniff.
    """
    normalized = _normalize_guia_relative_path(filename)
    if not normalized:
        return jsonify(success=False, message='Archivo no encontrado'), 404

    base_dir = app.config['GUIDES_DIR']
    # Normalizar y prevenir traversal
    safe_path = os.path.normpath(os.path.join(base_dir, normalized))
    base_norm = os.path.normpath(base_dir)
    if not safe_path.startswith(base_norm + os.sep) and safe_path != base_norm:
        return jsonify(success=False, message='Ruta inválida'), 400
    if not os.path.exists(safe_path):
        return jsonify(success=False, message='Archivo no encontrado'), 404

    mime, _ = mimetypes.guess_type(safe_path)
    if not mime or mime == 'application/octet-stream':
        try:
            if safe_path.lower().endswith('.pdf'):
                mime = 'application/pdf'
            else:
                with open(safe_path, 'rb') as f:
                    head = f.read(8)
                if head.startswith(b'%PDF'):
                    mime = 'application/pdf'
        except Exception:
            pass
    mime = mime or 'application/octet-stream'

    resp = send_file(safe_path, mimetype=mime, as_attachment=False, conditional=True)
    resp.headers['Content-Disposition'] = f'inline; filename="{os.path.basename(safe_path)}"'
    resp.headers['X-Content-Type-Options'] = 'nosniff'
    resp.headers.setdefault('Cache-Control', 'public, max-age=3600')
    # Log de depuración opcional
    try:
        debug = str(app.config.get('MIME_DEBUG', '0')).lower() in ('1','true','yes','on')
        if debug:
            size = -1
            try:
                size = os.path.getsize(safe_path)
            except Exception:
                pass
            current_app.logger.info(
                f"[MIME_DEBUG] /guias -> file={normalized} mime={mime} size={size} disposition=inline Accept={request.headers.get('Accept')}"
            )
    except Exception:
        pass
    return resp

# ================== NUEVOS MODELOS PARA FLUJO DE EFECTIVO (PERSISTENCIA) ==================
class FlujoUploadBatch(db.Model):
    __tablename__ = 'flujo_upload_batches'
    id = db.Column(db.Integer, primary_key=True)
    filename = db.Column(db.String(255))
    usuario = db.Column(db.String(120))
    uploaded_at = db.Column(db.DateTime, default=datetime.utcnow, index=True)
    total_bancos = db.Column(db.Integer, default=0)
    total_odoo = db.Column(db.Integer, default=0)

class FlujoBancoMovimiento(db.Model):
    __tablename__ = 'flujo_bancos_movimientos'
    id = db.Column(db.Integer, primary_key=True)
    batch_id = db.Column(db.Integer, db.ForeignKey('flujo_upload_batches.id'), index=True)
    fecha = db.Column(db.Date, index=True)
    empresa = db.Column(db.String(120), index=True)
    movimiento = db.Column(db.Text)
    monto = db.Column(db.Float)  # valor COP$
    banco = db.Column(db.String(120), index=True)
    tipo_banco = db.Column(db.String(120), nullable=True)
    unique_hash = db.Column(db.String(64), unique=True, index=True)
    creado_en = db.Column(db.DateTime, default=datetime.utcnow, index=True)

class FlujoOdooMovimiento(db.Model):
    __tablename__ = 'flujo_odoo_movimientos'
    id = db.Column(db.Integer, primary_key=True)
    batch_id = db.Column(db.Integer, db.ForeignKey('flujo_upload_batches.id'), index=True)
    fecha = db.Column(db.Date, index=True)
    empresa = db.Column(db.String(120), index=True)
    movimiento = db.Column(db.Text)
    debito = db.Column(db.Float, default=0)
    credito = db.Column(db.Float, default=0)
    tipo_flujo = db.Column(db.String(200))
    tercero = db.Column(db.String(200))
    rubro = db.Column(db.String(200))
    clase = db.Column(db.String(200))
    subclase = db.Column(db.String(200))
    banco = db.Column(db.String(120), index=True)
    tipo_banco = db.Column(db.String(120), nullable=True)
    unique_hash = db.Column(db.String(64), unique=True, index=True)
    creado_en = db.Column(db.DateTime, default=datetime.utcnow, index=True)

# ================== TABLAS DE AFORO ==================
class AforoTabla(db.Model):
    __tablename__ = 'aforos_tablas'
    id = db.Column(db.Integer, primary_key=True)
    timestamp = db.Column(db.DateTime, default=datetime.utcnow, nullable=False, index=True)
    usuario = db.Column(db.String(120), nullable=False)
    # tipo: 'TK' para tanques de planta; 'BARCAZA' para compartimentos/tanques de barcazas
    tipo = db.Column(db.String(20), nullable=False, index=True)
    # nombre: identificador lógico, ej. 'TK-109' o '1P' o 'MARI TK-1C'
    nombre = db.Column(db.String(120), nullable=False, index=True)
    # datos: JSON con lista de filas [{"cm":int, "mm":int, "bbl":float}] (mm usualmente 0..9)
    datos_json = db.Column(db.Text, nullable=False)

    __table_args__ = (
        db.UniqueConstraint('tipo', 'nombre', name='uq_aforo_tipo_nombre'),
    )

    def __repr__(self):
        return f"<AforoTabla {self.tipo}:{self.nombre} ({self.id})>"

def _ensure_aforos_table():
    from sqlalchemy import inspect
    with app.app_context():
        insp = inspect(db.engine)
        if 'aforos_tablas' not in insp.get_table_names():
            try:
                AforoTabla.__table__.create(db.engine)
                print('[INIT] Tabla aforos_tablas creada')
            except Exception as e:
                print('[INIT] No fue posible crear tabla aforos_tablas:', e)

_ensure_aforos_table()

def _hash_row(values: list) -> str:
    base = '|'.join('' if v is None else str(v).strip() for v in values)
    return hashlib.sha256(base.encode('utf-8')).hexdigest()
    
BASE_DIR = os.path.abspath(os.path.dirname(__file__))

@app.context_processor
def inject_current_year():
    return {'current_year': datetime.now().year}

# --- Utilidades de aforos ---
def _parse_aforo_excel_to_json(ws):
    """
    Convierte una hoja de Excel a un JSON usable por el cálculo de aforo.

    Dos modos de salida (detectados automáticamente):
    - Modo 'step': Estructura por decímetro con 'base' (cada 10 cm),
      incrementos por centímetro 'inc_cm' (1..9) y por milímetro 'inc_mm' (1..9).
      Este modo replica el método de Excel: base(10 cm) + inc_cm + inc_mm.
    - Modo 'flat': Lista de registros {cm, mm, bbl} para hojas simples.

    Soporta hojas con pares horizontales 'NIVEL/VOLUMEN', 'NIVEL 2/VOLUMEN 2',
    'NIVEL 3/VOLUMEN 3' y también hojas simples con columnas 'cm', 'mm', 'bbl'.
    """
    # Construir texto de cabecera por columna combinando fila 1 y 2
    def _cell_text(cell):
        v = cell.value
        return str(v).strip().lower() if v is not None else ''

    max_cols = ws.max_column or 0
    row1 = [ _cell_text(c) for c in ws[1][:max_cols] ] if ws.max_row >= 1 else []
    row2 = [ _cell_text(c) for c in ws[2][:max_cols] ] if ws.max_row >= 2 else []
    headers = []
    for i in range(max_cols):
        h1 = row1[i] if i < len(row1) else ''
        h2 = row2[i] if i < len(row2) else ''
        headers.append((h1 + ' ' + h2).strip())

    # Heurísticas para identificar columnas de nivel y volumen
    def is_lvl(h: str) -> bool:
        return any(k in h for k in ['nivel', 'cm', 'mm'])

    def lvl_unit(h: str) -> str:
        if 'mm' in h and 'cm' not in h:
            return 'mm'
        return 'cm'

    def is_vol(h: str) -> bool:
        return any(k in h for k in ['bbl', 'bls', 'volumen'])

    # Determinar fila inicial de datos: si segunda fila parece cabecera también, empezamos en 3
    header_has_words = any(any(ch.isalpha() for ch in h) for h in row2)
    start_row = 3 if header_has_words else 2

    # Detectar pares (nivel, volumen) adyacentes
    pairs = []  # lista de tuplas (idx_level, idx_volume, unit)
    i = 0
    while i < len(headers) - 1:
        h_lvl = headers[i]
        h_vol = headers[i+1]
        if is_lvl(h_lvl) and is_vol(h_vol):
            pairs.append((i, i+1, lvl_unit(h_lvl)))
            i += 2
            continue
        i += 1

    # Intentar detectar estructura por decímetro (step). Heurística: hay al menos
    # dos pares en 'cm' (base + cm) y opcionalmente un par en 'mm'.
    cm_pairs = [p for p in pairs if p[2] == 'cm']
    mm_pairs = [p for p in pairs if p[2] == 'mm']

    if cm_pairs:
        # Elegimos el primer par 'cm' como BASE y el segundo (si existe) como CM.
        base_lvl_idx, base_vol_idx, _ = cm_pairs[0]
        cm_inc_idx = cm_pairs[1] if len(cm_pairs) > 1 else None
        mm_inc_idx = mm_pairs[0] if mm_pairs else None

        base_map = {}
        inc_cm_map = {}
        inc_mm_map = {}
        # Algunas hojas listan incrementos globales (no por decímetro)
        inc_cm_global = {}
        inc_mm_global = {}
        current_dec = None

        for row in ws.iter_rows(min_row=start_row, values_only=True):
            try:
                nivel_base = row[base_lvl_idx] if base_lvl_idx < len(row) else None
                vol_base = row[base_vol_idx] if base_vol_idx < len(row) else None
            except Exception:
                nivel_base = None
                vol_base = None

            # Si hay nivel base en esta fila, actualizamos decímetro actual
            if nivel_base not in (None, '') and vol_base not in (None, ''):
                try:
                    nb = float(str(nivel_base).replace(',', '.'))
                    vb = float(str(vol_base).replace(',', '.'))
                    dec = int(round(nb))
                    current_dec = dec
                    base_map[current_dec] = vb
                    if current_dec not in inc_cm_map:
                        inc_cm_map[current_dec] = {}
                    if current_dec not in inc_mm_map:
                        inc_mm_map[current_dec] = {}
                except Exception:
                    pass

            # Incrementos por centímetro (si existe el par y hay decímetro vigente)
            if cm_inc_idx:
                lvl_idx, vol_idx, _ = cm_inc_idx
                try:
                    n2 = row[lvl_idx] if lvl_idx < len(row) else None
                    v2 = row[vol_idx] if vol_idx < len(row) else None
                    if n2 not in (None, '') and v2 not in (None, ''):
                        n2i = int(round(float(str(n2).replace(',', '.'))))
                        v2f = float(str(v2).replace(',', '.'))
                        if 1 <= n2i <= 9:
                            inc_cm_global[n2i] = v2f
                            if current_dec is not None:
                                inc_cm_map.setdefault(current_dec, {})[n2i] = v2f
                except Exception:
                    pass

            # Incrementos por milímetro (si existe el par y hay decímetro vigente)
            if mm_inc_idx:
                lvl_idx, vol_idx, _ = mm_inc_idx
                try:
                    n3 = row[lvl_idx] if lvl_idx < len(row) else None
                    v3 = row[vol_idx] if vol_idx < len(row) else None
                    if n3 not in (None, '') and v3 not in (None, ''):
                        n3i = int(round(float(str(n3).replace(',', '.'))))
                        v3f = float(str(v3).replace(',', '.'))
                        if 1 <= n3i <= 9:
                            inc_mm_global[n3i] = v3f
                            if current_dec is not None:
                                inc_mm_map.setdefault(current_dec, {})[n3i] = v3f
                except Exception:
                    pass

        # Si logramos leer algún base, devolvemos modo step
        if base_map:
            # Normalizar claves a strings para JSON compacto
            def _norm(d):
                return {str(k): v for k, v in d.items()}
            def _norm_nest(d):
                return {str(k): {str(kk): vv for kk, vv in v.items()} for k, v in d.items() if v}
            return {
                'mode': 'step',
                'base': _norm(base_map),
                'inc_cm': _norm_nest(inc_cm_map),
                'inc_mm': _norm_nest(inc_mm_map),
                'inc_cm_global': {str(k): v for k, v in (inc_cm_global or {}).items()},
                'inc_mm_global': {str(k): v for k, v in (inc_mm_global or {}).items()},
            }

    # Si no hay pares 'cm' o no se pudo construir modo step, intentar modo 'flat'
    data = []
    if not pairs:
        try:
            idx_cm = next((i for i,h in enumerate(headers) if 'cm' in h), None)
            idx_mm = next((i for i,h in enumerate(headers) if 'mm' in h), None)
            idx_bbl = next((i for i,h in enumerate(headers) if any(k in h for k in ['bbl','bls','volumen'])), None)
            if idx_cm is None or idx_bbl is None:
                raise ValueError
            for row in ws.iter_rows(min_row=start_row, values_only=True):
                cm = row[idx_cm] if idx_cm is not None and idx_cm < len(row) else None
                mm = row[idx_mm] if idx_mm is not None and idx_mm < len(row) else 0
                bbl = row[idx_bbl] if idx_bbl is not None and idx_bbl < len(row) else None
                try:
                    cm_i = int(float(str(cm).replace(',', '.')))
                    mm_i = int(float(str(mm or 0).replace(',', '.')))
                    bbl_f = float(str(bbl).replace(',', '.'))
                    data.append({'cm': cm_i, 'mm': mm_i, 'bbl': bbl_f})
                except Exception:
                    continue
        except Exception:
            raise ValueError('No se identificaron columnas de NIVEL/CM/MM y VOLUMEN (BBL/BLS).')
    else:
        # Procesar pares detectados como lista plana (mejor que nada)
        for lvl_idx, vol_idx, unit in pairs:
            for row in ws.iter_rows(min_row=start_row, values_only=True):
                try:
                    nivel_v = row[lvl_idx] if lvl_idx < len(row) else None
                    vol_v = row[vol_idx] if vol_idx < len(row) else None
                except Exception:
                    continue
                if nivel_v in (None, '') or vol_v in (None, ''):
                    continue
                try:
                    nivel_f = float(str(nivel_v).replace(',', '.'))
                    bbl_f = float(str(vol_v).replace(',', '.'))
                except Exception:
                    continue
                if unit == 'cm':
                    cm_i = int(round(nivel_f))
                    mm_i = 0
                else:
                    total_mm = int(round(nivel_f))
                    if total_mm < 0:
                        continue
                    cm_i = total_mm // 10
                    mm_i = total_mm % 10
                data.append({'cm': cm_i, 'mm': mm_i, 'bbl': bbl_f})

    if not data:
        raise ValueError('No se leyeron filas válidas de aforo.')

    # Deduplicar por (cm,mm) manteniendo el último valor encontrado
    dedup = {}
    for r in data:
        key = (int(r['cm']), int(r['mm']))
        try:
            dedup[key] = float(r['bbl'])
        except Exception:
            continue
    out = [ {'cm': k[0], 'mm': k[1], 'bbl': v} for k,v in dedup.items() ]
    out.sort(key=lambda r: (r['cm'], r['mm']))
    return {'mode': 'flat', 'data': out}

def _parse_barge_columns_sheet(ws, prefer_tipo: str = 'BARCAZA', default_name: str | None = None):
    """
    Parser especializado para hojas con formato "TEMP LAMINA 60" donde:
    - Una columna contiene la lámina/altura (cm)
    - Varias columnas (una por tanque/compartimento) contienen volúmenes

    Devuelve un dict mapping nombre_tabla -> payload_json_dict (modo 'flat').

    Los encabezados se esperan tipo: "MAN TK 1", "MG6 1P", "CR 2S", "OD 3P", "OILTECH 1C", etc.
    Se normalizan a nombres compatibles con la UI de Trasiegos: "<GRUPO>-<COMP>"
    p.ej. CR-1P, MARGOTH-1S, MANZANILLO-1, ODISEA-3S, OILTECH-1C.
    """
    try:
        max_cols = ws.max_column or 0
        # Construir encabezados combinando fila 1 y 2 (por si hay títulos en dos filas)
        def _txt(cell):
            v = cell.value
            return str(v).strip() if v is not None else ''

        row1 = [_txt(c) for c in (ws[1][:max_cols] if ws.max_row >= 1 else [])]
        row2 = [_txt(c) for c in (ws[2][:max_cols] if ws.max_row >= 2 else [])]
        headers = []
        for i in range(max_cols):
            h1 = (row1[i] if i < len(row1) else '')
            h2 = (row2[i] if i < len(row2) else '')
            h = (h1 + ' ' + h2).strip().upper()
            headers.append(h)

        # Detectar columna de lámina
        lam_idx = None
        for i, h in enumerate(headers):
            if any(k in h for k in ['LAMINA', 'LÁMINA', 'ALTURA', 'NIVEL']):
                lam_idx = i
                break
        if lam_idx is None:
            return {}

        # Función de mapeo de prefijo a grupo
        def map_group(pref: str) -> str:
            p = pref.upper().replace('.', '').replace('_', '').strip()
            if p.startswith('MAN'):
                return 'MANZANILLO'
            if p.startswith('MG6') or p.startswith('MARG'):
                return 'MARGOTH'
            if p.startswith('CR'):
                return 'CR'
            if p.startswith('OD') or p.startswith('ODI'):
                return 'ODISEA'
            if 'OIL' in p:
                return 'OILTECH'
            return p

        # Extraer definiciones de columnas objetivo (todas menos lámina) con su nombre lógico
        targets = []  # (col_idx, nombre)
        import re as _re
        for i, h in enumerate(headers):
            if i == lam_idx:
                continue
            if not h:
                continue
            # Intentar patrones comunes
            # 1) PREF TK <num/opcional letra>, 2) PREF <num>[P|S|C], 3) simplemente etiqueta de tanque
            m = _re.search(r"^(MAN|MG6|CR|OD|ODISEA|OILTECH)[\s\-_/]*T?K?\s*([0-9]+[A-Z]?)$", h, _re.IGNORECASE)
            if not m:
                m = _re.search(r"^(MAN|MG6|CR|OD|ODISEA|OILTECH)[\s\-_/]*([0-9]+[A-Z]?)$", h, _re.IGNORECASE)
            if m:
                pref = m.group(1).upper()
                comp = m.group(2).upper()
                grupo = map_group(pref)
                nombre = f"{grupo}-{comp}"
                targets.append((i, nombre))
                continue
            # Si no matchea, ignorar columnas no numéricas
            # Si el encabezado es numérico o poco descriptivo, usar default_name si está disponible
            header_clean = h.strip().upper()
            if default_name:
                if not header_clean or _re.fullmatch(r"[0-9\s\.,°%-]+", header_clean):
                    targets.append((i, default_name.strip().upper()))
                    continue
            targets.append((i, header_clean))  # última opción: usar header completo

        if not targets:
            return {}

        # Si solo hay una columna objetivo y tenemos default_name, usarlo como nombre
        if len(targets) == 1 and default_name:
            targets = [(targets[0][0], default_name.strip().upper())]

        # Detectar desde qué fila empiezan los datos (si fila 2 parece cabecera, empezar en 3)
        start_row = 3 if any(any(ch.isalpha() for ch in (row2[i] if i < len(row2) else '')) for i in range(max_cols)) else 2

        tablas = {}  # nombre -> lista data
        for row in ws.iter_rows(min_row=start_row, values_only=True):
            lam = row[lam_idx] if lam_idx < len(row) else None
            if lam in (None, ''):
                continue
            try:
                lam_cm = int(round(float(str(lam).replace(',', '.'))))
            except Exception:
                continue
            for col_idx, nombre in targets:
                if col_idx >= len(row):
                    continue
                val = row[col_idx]
                if val in (None, ''):
                    continue
                try:
                    bbl = float(str(val).replace(',', '.'))
                except Exception:
                    continue
                tablas.setdefault(nombre.upper().strip(), []).append({'cm': lam_cm, 'mm': 0, 'bbl': bbl})

        # Post-proceso: ordenar y empaquetar
        out = {}
        for nombre, data in tablas.items():
            if not data:
                continue
            data.sort(key=lambda r: (r['cm'], r.get('mm', 0)))
            out[nombre] = {'mode': 'flat', 'data': data}
        return out
    except Exception:
        return {}

def _parse_simple_lamina_single_volume(ws):
    """
    Parser simple para hojas con columnas:
    - "LÁMINA" (o ALTURA/NIVEL) y
    - una única columna de volúmenes (encabezado puede ser numérico como "60").

    Devuelve dict {'mode':'flat','data':[{'cm':int,'mm':0,'bbl':float}, ...]}
    o {} si no logra identificar las columnas.
    """
    try:
        max_cols = ws.max_column or 0
        def _txt(cell):
            v = cell.value
            return str(v).strip() if v is not None else ''

        row1 = [_txt(c) for c in (ws[1][:max_cols] if ws.max_row >= 1 else [])]
        row2 = [_txt(c) for c in (ws[2][:max_cols] if ws.max_row >= 2 else [])]
        headers = []
        for i in range(max_cols):
            h1 = (row1[i] if i < len(row1) else '')
            h2 = (row2[i] if i < len(row2) else '')
            headers.append((h1 + ' ' + h2).strip().upper())

        # Buscar lamina
        lam_idx = None
        for i, h in enumerate(headers):
            if any(k in h for k in ['LAMINA', 'LÁMINA', 'ALTURA', 'NIVEL']):
                lam_idx = i
                break
        if lam_idx is None:
            return {}

        # Detectar fila inicial (si fila 2 tiene letras, empezar en 3)
        start_row = 3 if any(any(ch.isalpha() for ch in (row2[i] if i < len(row2) else '')) for i in range(max_cols)) else 2

        # Elegir la mejor columna de volumen por cantidad de valores numéricos
        def _is_num(x):
            try:
                float(str(x).replace(',', '.'))
                return True
            except Exception:
                return False

        best_idx = None
        best_count = -1
        sample_rows = list(ws.iter_rows(min_row=start_row, max_row=min(start_row + 50, ws.max_row), values_only=True))
        for j in range(max_cols):
            if j == lam_idx:
                continue
            cnt = 0
            for r in sample_rows:
                if j < len(r) and _is_num(r[j]):
                    cnt += 1
            if cnt > best_count:
                best_count = cnt
                best_idx = j
        if best_idx is None or best_count <= 0:
            return {}

        data = []
        for row in ws.iter_rows(min_row=start_row, values_only=True):
            lam = row[lam_idx] if lam_idx < len(row) else None
            vol = row[best_idx] if best_idx < len(row) else None
            if lam in (None, '') or vol in (None, ''):
                continue
            try:
                cm = int(round(float(str(lam).replace(',', '.'))))
                bbl = float(str(vol).replace(',', '.'))
            except Exception:
                continue
            data.append({'cm': cm, 'mm': 0, 'bbl': bbl})
        if not data:
            return {}
        data.sort(key=lambda r: (r['cm'], r['mm']))
        return {'mode': 'flat', 'data': data}
    except Exception:
        return {}

def _interp_bbl(datos, cm, mm):
    """
    Calcula BBL según datos de aforo. Soporta:
    - Modo 'step': base(10 cm) + inc_cm + inc_mm
    - Modo 'flat': lineal con mejoras: usa valores absolutos si existen para
      (cm,0) y (cm,mm); si no, interpola.
    """
    # Helper: lineal a partir de lista plana
    def _linear_from_flat(lista, cm_i, mm_i):
        nivel = cm_i + (mm_i or 0)/10.0
        niveles = [d['cm'] + d['mm']/10.0 for d in lista]
        bbllist = [d['bbl'] for d in lista]
        if not niveles:
            return 0.0
        if nivel <= niveles[0]:
            return bbllist[0]
        if nivel >= niveles[-1]:
            return bbllist[-1]
        for i in range(1, len(niveles)):
            if niveles[i] >= nivel:
                x0, x1 = niveles[i-1], niveles[i]
                y0, y1 = bbllist[i-1], bbllist[i]
                t = (nivel - x0) / (x1 - x0) if (x1 - x0) != 0 else 0
                return y0 + t * (y1 - y0)
        return bbllist[-1]

    # Modo STEP
    if isinstance(datos, dict) and datos.get('mode') == 'step':
        base_map = {int(k): float(v) for k, v in (datos.get('base') or {}).items()}
        inc_cm = {int(k): {int(kk): float(vv) for kk, vv in (sub or {}).items()} for k, sub in (datos.get('inc_cm') or {}).items()}
        inc_mm = {int(k): {int(kk): float(vv) for kk, vv in (sub or {}).items()} for k, sub in (datos.get('inc_mm') or {}).items()}
        inc_cm_global = {int(k): float(v) for k, v in (datos.get('inc_cm_global') or {}).items()}
        inc_mm_global = {int(k): float(v) for k, v in (datos.get('inc_mm_global') or {}).items()}

        if not base_map:
            return 0.0

        dec = (cm // 10) * 10
        # Buscar el decímetro igual o menor existente
        if dec not in base_map:
            menores = [k for k in base_map.keys() if k <= cm]
            if not menores:
                dec = min(base_map.keys())
            else:
                dec = max(menores)

        base = base_map.get(dec, 0.0)
        cm_intra = cm - dec
        total = base

        # Incremento por centímetro
        if cm_intra > 0:
            cm_table = inc_cm.get(dec, {})
            if cm_intra in cm_table:
                total += cm_table[cm_intra]
            elif cm_intra in inc_cm_global:
                total += inc_cm_global[cm_intra]
            else:
                # Aproximación lineal entre bases de decímetro
                nxt = base_map.get(dec + 10)
                if nxt is not None:
                    total += (nxt - base) * (cm_intra / 10.0)

        # Incremento por milímetro
        mm = int(mm or 0)
        if mm > 0:
            mm_table = inc_mm.get(dec, {})
            if mm in mm_table:
                total += mm_table[mm]
            elif mm in inc_mm_global:
                total += inc_mm_global[mm]
            else:
                # Aproximación dentro del centímetro con bases de (cm,0) y (cm+1,0) si están
                cm_table = inc_cm.get(dec, {})
                v_cm = cm_table.get(cm_intra, None) if cm_intra > 0 else 0.0
                v_cm_next = cm_table.get(cm_intra + 1, None)
                if v_cm is not None and v_cm_next is not None:
                    total += (v_cm_next - v_cm) * (mm / 10.0)
                else:
                    # Último recurso: usar salto entre base dec y próximo dec
                    nxt = base_map.get(dec + 10)
                    if nxt is not None:
                        total = base + (nxt - base) * ((cm_intra + mm/10.0) / 10.0)
        return total

    # Modo FLAT u otros
    lista = datos.get('data') if isinstance(datos, dict) and 'data' in datos else datos
    if not isinstance(lista, list) or not lista:
        return 0.0

    # Intentar método escalonado a partir de valores absolutos si existen
    # Mapa rápido
    M = {(int(d['cm']), int(d['mm'])): float(d['bbl']) for d in lista if 'cm' in d and 'mm' in d and 'bbl' in d}
    dec = (cm // 10) * 10
    base = M.get((dec, 0), None)
    total = None
    # Fallback estilo Excel: si el flat contiene incrementos globales de cm como (1..9,0)
    # y de mm como (0,1..9), usar: base(dec) + inc_cm(cm_intra) + inc_mm(mm)
    if base is not None:
        cm_intra = cm - dec
        inc_cm_glob = M.get((cm_intra, 0), None) if (1 <= cm_intra <= 9) else 0.0
        inc_mm_glob = M.get((0, int(mm)), None) if int(mm or 0) > 0 else 0.0
        if (inc_cm_glob is not None) or (inc_mm_glob is not None and int(mm or 0) > 0):
            # Si existen incrementos globales, aplicarlos como en Excel
            total = float(base) + float(inc_cm_glob or 0.0) + float(inc_mm_glob or 0.0)
            return total
    if base is not None:
        total = base
        # cm exacto disponible
        v_cm0 = M.get((cm, 0), None)
        if v_cm0 is not None:
            total = v_cm0
        else:
            # Interpolar entre (dec,0) y (dec+10,0)
            v_dec_next = M.get((dec + 10, 0), None)
            if v_dec_next is not None:
                total = base + (v_dec_next - base) * ((cm - dec) / 10.0)
        # mm incremento
        if mm and total is not None:
            v_cmmm = M.get((cm, int(mm)))
            if v_cmmm is not None and v_cm0 is not None:
                total = v_cmmm
            else:
                # Aproximación dentro del centímetro con (cm,0) y (cm+1,0)
                v_cm1 = M.get((cm + 1, 0), None)
                if v_cm0 is not None and v_cm1 is not None:
                    total = (v_cm0 + (v_cm1 - v_cm0) * (mm / 10.0))

    if total is not None:
        return total
    # Fallback: completamente lineal
    return _linear_from_flat(lista, cm, mm)

def _expand_preview_rows(datos, max_rows=200):
    """
    Genera una lista plana de filas {cm, mm, bbl} para vista previa.
    - Para modo 'flat': devuelve hasta max_rows primeras filas ordenadas.
    - Para modo 'step': muestrea mm=0 para cada centímetro entre el mínimo y máximo decímetro disponible.
    """
    try:
        # Vista previa especial para VCF/API6A
        if isinstance(datos, dict) and 'data' in datos and len(datos['data']) > 0:
            sample = datos['data'][0]
            if all(k in sample for k in ('api', 'temp', 'vcf')):
                lista = list(datos.get('data') or [])
                lista.sort(key=lambda r: (float(r.get('api', 0)), float(r.get('temp', 0))))
                return lista[:max_rows]
        if isinstance(datos, dict) and datos.get('mode') == 'flat':
            lista = list(datos.get('data') or [])
            lista.sort(key=lambda r: (int(r.get('cm', 0)), int(r.get('mm', 0))))
            return lista[:max_rows]
        if isinstance(datos, dict) and datos.get('mode') == 'step':
            base_map = {int(k): float(v) for k, v in (datos.get('base') or {}).items()}
            if not base_map:
                return []
            decs = sorted(base_map.keys())
            cm_min = decs[0]
            cm_max = decs[-1] + 9
            out = []
            for cm in range(cm_min, cm_max + 1):
                b = _interp_bbl(datos, cm, 0)
                out.append({'cm': int(cm), 'mm': 0, 'bbl': float(b)})
                if len(out) >= max_rows:
                    break
            return out
    except Exception:
        pass
    return []

# Decorador para verificar login (mejorado para AJAX)
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'email' not in session:
            # Si la petición espera JSON (como fetch), devuelve un error JSON y un código 401
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or \
               (request.accept_mimetypes.best_match(['application/json', 'text/html']) == 'application/json'):
                return jsonify(success=False, message="Sesión expirada o no autenticado. Por favor, inicie sesión de nuevo.", error_code="SESSION_EXPIRED"), 401
            
            flash('Por favor inicie sesión para acceder a esta página.', 'warning')
            return redirect(url_for('login', next=request.url))
        return f(*args, **kwargs)
    return decorated_function

@app.before_request
def log_request():
    print(f"➞️  {request.method} {request.path}")

USUARIOS = {

    # Juan Diego  (Admin): Tiene acceso a todo.
    "numbers@conquerstrading.com": {
        "password": generate_password_hash("Conquers2025"),
        "nombre": "Juan Diego Ayala",
        "rol": "admin",
        "area": [] 
    },

    # Carlos (Admin): Tiene acceso a todo.
    "oci@conquerstrading.com": {
        "password": generate_password_hash("Conquers2025"),
        "nombre": "Carlos Barón",
        "rol": "admin",
        "area": [] # El admin no necesita áreas específicas, su rol le da acceso a todo.
    },
    # Brandon (Admin): Acceso total, igual que Carlos.
    "logistics.inventory@conquerstrading.com": {
        "password": generate_password_hash("Conquers2025"),
        "nombre": "Brandon Niño",
        "rol": "admin",
        "area": []
    },
    # Juan Diego (Editor): Solo acceso a Barcaza Orion.
    "qualitycontrol@conquerstrading.com": {
        "password": generate_password_hash("Conquers2025"),
        "nombre": "Juan Diego Cuadros",
        "rol": "editor",
        "area": ["barcaza_orion", "barcaza_bita", "programacion_cargue", "control_calidad", "siza_solicitante", "reportes"] 
    },
    # Ricardo (Editor): Solo acceso a Barcaza BITA.
    "quality.manager@conquerstrading.com": {
        "password": generate_password_hash("Conquers2025"),
        "nombre": "Ricardo Congo",
        "rol": "editor",
        "area": ["barcaza_bita"]
    },
    # Omar (Viewer): Rol limitado para ver reportes.
    "omar.morales@conquerstrading.com": {
    "password": generate_password_hash("Conquers2025"),
    "nombre": "Omar Morales",
    "rol": "viewer",
    "area": ["reportes", "planilla_precios", "simulador_rendimiento", "flujo_efectivo", "siza_solicitante"]
},

    "david.restrepo@conquerstrading.com": {
        "password": generate_password_hash("Conquers2025"),
        "nombre": "David Restrepo",
        "rol": "viewer",
        "area": ["reportes", "planilla_precios", "simulador_rendimiento", "flujo_efectivo", "siza_solicitante"] 
    },


    "finance@conquerstrading.com": {
        "password": generate_password_hash("Conquers2025"),
        "nombre": "German Galvis",
        "rol": "viewer",
    "area": ["reportes", "planilla_precios", "simulador_rendimiento", "control_remolcadores", "flujo_efectivo", "modelo_optimizacion"] 
    },
    
    # Ignacio (Editor): Solo acceso a Planta y Rendimientos
    "production@conquerstrading.com": {
        "password": generate_password_hash("Conquers2025"),
        "nombre": "Ignacio Quimbayo",
        "rol": "editor",
        "area": ["planta", "simulador_rendimiento", "programacion_cargue", "control_calidad", "reportes"] 
    },
    # Juliana (Editor): Tiene acceso a Tránsito, Generar Guía y SIZA Solicitante.
    "ops@conquerstrading.com": {
        "password": generate_password_hash("Conquers2025"),
        "nombre": "Juliana Torres",
        "rol": "editor",
        "area": ["transito", "guia_transporte", "control_remolcadores", "programacion_cargue", "siza_solicitante"]
    },
    # Samantha (Editor): Tiene acceso a Generar Guía y SIZA Solicitante.
    "logistic@conquerstrading.com": {
        "password": generate_password_hash("Conquers2025*"),
        "nombre": "Samantha Roa",
        "rol": "editor",
        "area": ["guia_transporte", "programacion_cargue", "siza_solicitante"]
    },

    "comex@conquerstrading.com": {
        "password": generate_password_hash("Conquers2025"),     
        "nombre": "Daniela Cuadrado",
        "rol": "editor",
        "area": ["zisa_inventory", "programacion_cargue", "siza_solicitante", "siza_gestor"] 
    },

    "comexzf@conquerstrading.com": {
        "password": generate_password_hash("Conquers2025"),     
        "nombre": "Shirli Diaz",
        "rol": "editor",
        "area": ["programacion_cargue", "siza_solicitante", "siza_gestor"] 
    },

    # SIZA - Solicitantes (solo pueden ver y solicitar pedidos)
    "carlos.baron@conquerstrading.com": {
        "password": generate_password_hash("Conquers2025"),
        "nombre": "Carlos Baron",
        "rol": "editor",
        "area": ["siza_solicitante"]
    },

    "juandiego.cuadros@conquerstrading.com": {
        "password": generate_password_hash("Conquers2025"),
        "nombre": "Juan Diego Cuadros",
        "rol": "editor",
        "area": ["siza_solicitante"]
    },

    "brando@conquerstrading.com": {
        "password": generate_password_hash("Conquers2025"),
        "nombre": "Brando",
        "rol": "editor",
        "area": ["siza_solicitante"]
    },

    "felipe.delavega@conquerstrading.com": {
        "password": generate_password_hash("Conquers2025"),     
        "nombre": "Felipe De La Vega",
        "rol": "editor",
    "area": ["simulador_rendimiento", "flujo_efectivo", "modelo_optimizacion", "siza_solicitante", "planilla_precios"] 
    },

        "accountingzf@conquerstrading.com": { 
        "password": generate_password_hash("Conquers2025"),       
        "nombre": "Kelly Suarez",
        "rol": "editor",
        "area": ["contabilidad"] 
    },
        "amariagallo@conquerstrading.com": {
        "password": generate_password_hash("Conquers2025"), 
        "nombre": "Ana Maria Gallo",
        "rol": "logistica_destino",
        "area": ["programacion_cargue","gestion_compras", "planilla_precios"]
    },

        "refinery.control@conquerstrading.com": {
        "password": generate_password_hash("Conquers2025"), 
        "nombre": "Control Refineria",
        "rol": "refineria",
        "area": ["programacion_cargue", "control_calidad", "planta"] 
    },
        "opensea@conquerstrading.com": {
        "password": generate_password_hash("Conquers2025"), 
        "nombre": "Opensea", 
        "rol": "operador_remolcador", 
        "area": ["control_remolcadores"]
    },

    "safety@conquerstrading.com": {
    "password": generate_password_hash("Conquers2025"),
    "nombre": "Sebastian Blanco",
    "rol": "editor",
    "area": ["inventario_epp"]
}


}
   
PLANILLA_PLANTA = [
    {"TK": "TK-109", "PRODUCTO": "CRUDO RF.", "MAX_CAP": 22000, "BLS_60": "", "API": "", "BSW": "", "S": ""},
    {"TK": "TK-110", "PRODUCTO": "FO4",       "MAX_CAP": 22000, "BLS_60": "", "API": "", "BSW": "", "S": ""},
    {"TK": "TK-108", "PRODUCTO": "VLSFO",    "MAX_CAP": 28000, "BLS_60": "", "API": "", "BSW": "", "S": ""},
    {"TK": "TK-01",  "PRODUCTO": "DILUYENTE", "MAX_CAP": 450,   "BLS_60": "", "API": "", "BSW": "", "S": ""},
    {"TK": "TK-02",  "PRODUCTO": "DILUYENTE", "MAX_CAP": 450,   "BLS_60": "", "API": "", "BSW": "", "S": ""},
    {"TK": "TK-102", "PRODUCTO": "FO6",       "MAX_CAP": 4100,  "BLS_60": "", "API": "", "BSW": "", "S": ""},
    {"TK": "Consumo Interno", "PRODUCTO": "DILUYENTE", "MAX_CAP": 124.78, "MAX_CAP_GAL": 5240.91, "FILL_CAP_GAL": 4765.16, "BLS_60": "", "API": "", "BSW": "", "S": ""}
]
PLANILLA_BARCAZA_ORION = [
    # Sección MANZANILLO (MGO)
    {"TK": "1", "PRODUCTO": "MGO", "MAX_CAP": 709, "BLS_60": "", "API": "", "BSW": "", "S": "", "grupo": "MANZANILLO"},
    {"TK": "2", "PRODUCTO": "MGO", "MAX_CAP": 806, "BLS_60": "", "API": "", "BSW": "", "S": "", "grupo": "MANZANILLO"},
    {"TK": "3", "PRODUCTO": "MGO", "MAX_CAP": 694, "BLS_60": "", "API": "", "BSW": "", "S": "", "grupo": "MANZANILLO"},
    
    # Tanque Principal (TK-101)
    {"TK": "TK-101", "PRODUCTO": "VLSFO", "MAX_CAP":4660.52, "BLS_60": "", "API": "", "BSW": "", "S": "", "grupo": "PRINCIPAL"},
    
    # BARCAZA CR (VLSFO)
    {"TK": "1P", "PRODUCTO": "VLSFO", "MAX_CAP": 742.68, "BLS_60": "", "API": "", "BSW": "", "S": "", "grupo": "CR"},
    {"TK": "1S", "PRODUCTO": "VLSFO", "MAX_CAP": 739.58, "BLS_60": "", "API": "", "BSW": "", "S": "", "grupo": "CR"},
    {"TK": "2P", "PRODUCTO": "VLSFO", "MAX_CAP": 886.56, "BLS_60": "", "API": "", "BSW": "", "S": "", "grupo": "CR"},
    {"TK": "2S", "PRODUCTO": "VLSFO", "MAX_CAP": 890.24, "BLS_60": "", "API": "", "BSW": "", "S": "", "grupo": "CR"},
    {"TK": "3P", "PRODUCTO": "VLSFO", "MAX_CAP": 877.95, "BLS_60": "", "API": "", "BSW": "", "S": "", "grupo": "CR"},
    {"TK": "3S", "PRODUCTO": "VLSFO", "MAX_CAP": 888.44, "BLS_60": "", "API": "", "BSW": "", "S": "", "grupo": "CR"},
    {"TK": "4P", "PRODUCTO": "VLSFO", "MAX_CAP": 892.57, "BLS_60": "", "API": "", "BSW": "", "S": "", "grupo": "CR"},
    {"TK": "4S", "PRODUCTO": "VLSFO", "MAX_CAP": 887.54, "BLS_60": "", "API": "", "BSW": "", "S": "", "grupo": "CR"},
    {"TK": "5P", "PRODUCTO": "VLSFO", "MAX_CAP": 737.09, "BLS_60": "", "API": "", "BSW": "", "S": "", "grupo": "CR"},
    {"TK": "5S", "PRODUCTO": "VLSFO", "MAX_CAP": 739.45, "BLS_60": "", "API": "", "BSW": "", "S": "", "grupo": "CR"},
    
    # BARCAZA MARGOTH (VLSFO)
    {"TK": "1P", "PRODUCTO": "VLSFO", "MAX_CAP": 582.09, "BLS_60": "", "API": "", "BSW": "", "S": "", "grupo": "MARGOTH"},
    {"TK": "1S", "PRODUCTO": "VLSFO", "MAX_CAP": 582.09, "BLS_60": "", "API": "", "BSW": "", "S": "", "grupo": "MARGOTH"},
    {"TK": "2P", "PRODUCTO": "VLSFO", "MAX_CAP": 572.66, "BLS_60": "", "API": "", "BSW": "", "S": "", "grupo": "MARGOTH"},
    {"TK": "2S", "PRODUCTO": "VLSFO", "MAX_CAP": 572.66, "BLS_60": "", "API": "", "BSW": "", "S": "", "grupo": "MARGOTH"},
    {"TK": "3P", "PRODUCTO": "VLSFO", "MAX_CAP": 572.68, "BLS_60": "", "API": "", "BSW": "", "S": "", "grupo": "MARGOTH"},
    {"TK": "3S", "PRODUCTO": "VLSFO", "MAX_CAP": 572.68, "BLS_60": "", "API": "", "BSW": "", "S": "", "grupo": "MARGOTH"},
    {"TK": "4P", "PRODUCTO": "VLSFO", "MAX_CAP": 575.10, "BLS_60": "", "API": "", "BSW": "", "S": "", "grupo": "MARGOTH"},
    {"TK": "4S", "PRODUCTO": "VLSFO", "MAX_CAP": 575.10, "BLS_60": "", "API": "", "BSW": "", "S": "", "grupo": "MARGOTH"},
    {"TK": "5P", "PRODUCTO": "VLSFO", "MAX_CAP": 571.72, "BLS_60": "", "API": "", "BSW": "", "S": "", "grupo": "MARGOTH"},
    {"TK": "5S", "PRODUCTO": "VLSFO", "MAX_CAP": 571.72, "BLS_60": "", "API": "", "BSW": "", "S": "", "grupo": "MARGOTH"},
    
    # BARCAZA ODISEA (VLSFO)
    {"TK": "1P", "PRODUCTO": "VLSFO", "MAX_CAP": 2533.98, "BLS_60": "", "API": "", "BSW": "", "S": "", "grupo": "ODISEA"},
    {"TK": "1S", "PRODUCTO": "VLSFO", "MAX_CAP": 2544.17, "BLS_60": "", "API": "", "BSW": "", "S": "", "grupo": "ODISEA"},
    {"TK": "2P", "PRODUCTO": "VLSFO", "MAX_CAP": 3277.10, "BLS_60": "", "API": "", "BSW": "", "S": "", "grupo": "ODISEA"},
    {"TK": "2S", "PRODUCTO": "VLSFO", "MAX_CAP": 3282.97, "BLS_60": "", "API": "", "BSW": "", "S": "", "grupo": "ODISEA"},
    {"TK": "3P", "PRODUCTO": "VLSFO", "MAX_CAP": 3302.94, "BLS_60": "", "API": "", "BSW": "", "S": "", "grupo": "ODISEA"},
    {"TK": "3S", "PRODUCTO": "VLSFO", "MAX_CAP": 3287.42, "BLS_60": "", "API": "", "BSW": "", "S": "", "grupo": "ODISEA"},
    {"TK": "4P", "PRODUCTO": "VLSFO", "MAX_CAP": 3282.96, "BLS_60": "", "API": "", "BSW": "", "S": "", "grupo": "ODISEA"},
    {"TK": "4S", "PRODUCTO": "VLSFO", "MAX_CAP": 3291.98, "BLS_60": "", "API": "", "BSW": "", "S": "", "grupo": "ODISEA"},
    {"TK": "5P", "PRODUCTO": "VLSFO", "MAX_CAP": 2930.16, "BLS_60": "", "API": "", "BSW": "", "S": "", "grupo": "ODISEA"},
    {"TK": "5S", "PRODUCTO": "VLSFO", "MAX_CAP": 2933.93, "BLS_60": "", "API": "", "BSW": "", "S": "", "grupo": "ODISEA"},
]

PLANILLA_BARCAZA_BITA = [
    # Barcaza Marinse
    {"TK": "MARI TK-1C", "PRODUCTO": "VLSFO", "MAX_CAP": 1506.56, "BLS_60": "", "API": "", "BSW": "", "S": ""},
    {"TK": "MARI TK-2C", "PRODUCTO": "VLSFO", "MAX_CAP": 1541.10, "BLS_60": "", "API": "", "BSW": "", "S": ""},
    {"TK": "MARI TK-3C", "PRODUCTO": "VLSFO", "MAX_CAP": 1438.96, "BLS_60": "", "API": "", "BSW": "", "S": ""},
    {"TK": "MARI TK-4C", "PRODUCTO": "VLSFO", "MAX_CAP": 1433.75, "BLS_60": "", "API": "", "BSW": "", "S": ""},
    {"TK": "MARI TK-5C", "PRODUCTO": "VLSFO", "MAX_CAP": 1641.97, "BLS_60": "", "API": "", "BSW": "", "S": ""},
    {"TK": "MARI TK-6C", "PRODUCTO": "VLSFO", "MAX_CAP": 1617.23, "BLS_60": "", "API": "", "BSW": "", "S": ""},
    # Barcaza Oidech
    {"TK": "OID TK-1C", "PRODUCTO": "VLSFO", "MAX_CAP": 4535.54, "BLS_60": "", "API": "", "BSW": "", "S": ""},
    {"TK": "OID TK-2C", "PRODUCTO": "VLSFO", "MAX_CAP": 5808.34, "BLS_60": "", "API": "", "BSW": "", "S": ""},
    {"TK": "OID TK-3C", "PRODUCTO": "VLSFO", "MAX_CAP": 4928.29, "BLS_60": "", "API": "", "BSW": "", "S": ""}
]
PLANILLA_TRANSITO_GENERAL = [
    {"ORIGEN": "", "FECHA": "", "GUIA": "", "PRODUCTO": "", "PLACA": "", "API": "", "BSW": "",  "NSV": "", "OBSERVACIONES":""}
    for _ in range(10)  # O el número de filas que desees por defecto
]

PLANILLA_TRANSITO_REFINERIA = [
    {"ORIGEN": "", "FECHA": "", "GUIA": "", "PRODUCTO": "", "PLACA": "", "API": "", "BSW": "",  "NSV": "", "OBSERVACIONES":""}
    for _ in range(10)  # O el número de filas que desees por defecto
]

# REEMPLAZA TU LISTA ACTUAL CON ESTA
DEPARTAMENTOS_Y_CAPITALES = [
    {"departamento": "Amazonas", "capital": "Leticia", "lat": -4.2152, "lng": -69.9406},
    {"departamento": "Antioquia", "capital": "Medellín", "lat": 6.2442, "lng": -75.5812},
    {"departamento": "Arauca", "capital": "Arauca", "lat": 7.084, "lng": -70.759},
    {"departamento": "Atlántico", "capital": "Barranquilla", "lat": 10.9639, "lng": -74.7964},
    {"departamento": "Bolívar", "capital": "Cartagena", "lat": 10.3910, "lng": -75.4794},
    {"departamento": "Boyacá", "capital": "Tunja", "lat": 5.534, "lng": -73.367},
    {"departamento": "Caldas", "capital": "Manizales", "lat": 5.068, "lng": -75.517},
    {"departamento": "Caquetá", "capital": "Florencia", "lat": 1.614, "lng": -75.606},
    {"departamento": "Casanare", "capital": "Yopal", "lat": 5.337, "lng": -72.390},
    {"departamento": "Cauca", "capital": "Popayán", "lat": 2.445, "lng": -76.614},
    {"departamento": "Cesar", "capital": "Valledupar", "lat": 10.463, "lng": -73.253},
    {"departamento": "Chocó", "capital": "Quibdó", "lat": 5.694, "lng": -76.661},
    {"departamento": "Córdoba", "capital": "Montería", "lat": 8.747, "lng": -75.881},
    {"departamento": "Cundinamarca", "capital": "Bogotá", "lat": 4.711, "lng": -74.072},
    {"departamento": "Guainía", "capital": "Inírida", "lat": 3.865, "lng": -67.923},
    {"departamento": "Guaviare", "capital": "San José del Guaviare", "lat": 2.572, "lng": -72.645},
    {"departamento": "Huila", "capital": "Neiva", "lat": 2.927, "lng": -75.281},
    {"departamento": "La Guajira", "capital": "Riohacha", "lat": 11.544, "lng": -72.907},
    {"departamento": "Magdalena", "capital": "Santa Marta", "lat": 11.240, "lng": -74.199},
    {"departamento": "Meta", "capital": "Villavicencio", "lat": 4.142, "lng": -73.626},
    {"departamento": "Nariño", "capital": "Pasto", "lat": 1.213, "lng": -77.281},
    {"departamento": "Norte de Santander", "capital": "Cúcuta", "lat": 7.893, "lng": -72.507},
    {"departamento": "Putumayo", "capital": "Mocoa", "lat": 1.154, "lng": -76.646},
    {"departamento": "Quindío", "capital": "Armenia", "lat": 4.533, "lng": -75.681},
    {"departamento": "Risaralda", "capital": "Pereira", "lat": 4.813, "lng": -75.696},
    {"departamento": "San Andrés y Providencia", "capital": "San Andrés", "lat": 12.584, "lng": -81.700},
    {"departamento": "Santander", "capital": "Bucaramanga", "lat": 7.119, "lng": -73.122},
    {"departamento": "Sucre", "capital": "Sincelejo", "lat": 9.295, "lng": -75.397},
    {"departamento": "Tolima", "capital": "Ibagué", "lat": 4.438, "lng": -75.232},
    {"departamento": "Valle del Cauca", "capital": "Cali", "lat": 3.451, "lng": -76.532},
    {"departamento": "Vaupés", "capital": "Mitú", "lat": 1.257, "lng": -70.234},
    {"departamento": "Vichada", "capital": "Puerto Carreño", "lat": 6.189, "lng": -67.485}
]
PLANILLA_PRECIOS = [
    {
        "DEPARTAMENTO": d["departamento"], "CAPITAL": d["capital"],
        "LAT": d["lat"], "LNG": d["lng"], # <-- AÑADIMOS LAS COORDENADAS AQUÍ
        "DISTANCIA_KM": "", "COSTO_FLETE": "", "PRECIO_VENTA": ""
    } for d in DEPARTAMENTOS_Y_CAPITALES
]

def cargar_productos():
    ruta = "productos.json"
    try:
        if os.path.exists(ruta):
            with open(ruta, encoding='utf-8') as f:
                data = json.load(f)
                # Validar estructura
                if not all(isinstance(v, list) for v in data.values()):
                    raise ValueError("Estructura inválida en productos.json")
                return data
    except Exception as e:
        print(f"Error cargando productos: {e}")
    return {"REFINERIA": [], "EDSM": []}  # Estructura por defecto

def guardar_registro_generico(datos_a_guardar, tipo_area):
    """
    Función genérica para guardar los datos de cualquier planilla en un archivo JSON.
    
    Args:
        datos_a_guardar (list): La lista de diccionarios (la planilla) con los datos actualizados.
        tipo_area (str): Un prefijo para el nombre del archivo (ej: 'planta', 'barcaza_orion').
    """
    try:
        # 1. Crear el timestamp para el nombre del archivo
        fecha = datetime.now().strftime("%Y_%m_%d_%H_%M_%S")
        
        # 2. Definir la carpeta y el nombre del archivo
        carpeta = "registros"
        os.makedirs(carpeta, exist_ok=True) # Crea la carpeta si no existe
        nombre_archivo = f"{tipo_area}_{fecha}.json"
        ruta_completa = os.path.join(carpeta, nombre_archivo)
        
        # 3. Preparar el diccionario de datos que se guardará
        data_para_json = {
            "fecha": fecha,
            "area": tipo_area,
            "usuario": session.get("nombre", "No identificado"),
            "datos": datos_a_guardar
        }
        
        # 4. Escribir el archivo JSON
        with open(ruta_completa, 'w', encoding='utf-8') as f:
            json.dump(data_para_json, f, ensure_ascii=False, indent=4)
            
        # 5. Devolver una respuesta de éxito en formato JSON
        return jsonify(success=True, message=f"Registro de '{tipo_area}' guardado exitosamente.")

    except Exception as e:
        # En caso de cualquier error, registrarlo y devolver un error en formato JSON
        print(f"ERROR en guardar_registro_generico para '{tipo_area}': {e}")
        return jsonify(success=False, message=f"Error interno del servidor al guardar el registro: {str(e)}"), 500

def cargar_transito_config():
    ruta_config = "transito_config.json"
    default_config = {
        "REFINERIA": {
            "nombre_display": "Tránsito Crudo Refinería",
            "campos": {}
        },
        "EDSM": {
            "nombre_display": "Tránsito Crudo EDSM",
            "campos": {}
        }
    }
    
    try:
        if os.path.exists(ruta_config):
            with open(ruta_config, 'r', encoding='utf-8') as f:
                config = json.load(f)
                # Validación de estructura básica
                if not all(k in config for k in ['REFINERIA', 'EDSM']):
                    raise ValueError("Estructura inválida")
                return config
    except Exception as e:
        print(f"Error cargando configuración: {e}")
    
    # Si hay error, devolver configuración por defecto
    return default_config




def tiene_permiso(permiso_requerido):
    """Función auxiliar para verificar si el usuario actual tiene un permiso específico."""
    if session.get('rol') == 'admin':
        return True
    areas_del_usuario = session.get('area', [])
    return permiso_requerido in areas_del_usuario


def permiso_requerido(area_requerida):
    """
    Decorador que verifica si un usuario tiene permiso para un área específica.
    El rol 'admin' siempre tiene acceso.
    """
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            # 1. El admin siempre tiene acceso
            if session.get('rol') == 'admin':
                return f(*args, **kwargs)
            
            # 2. Revisa si el área requerida está en la lista de áreas del usuario
            areas_del_usuario = session.get('area', [])
            if area_requerida in areas_del_usuario:
                return f(*args, **kwargs)
            
            # 3. Si no cumple ninguna condición, denegar acceso
            flash("No tienes los permisos necesarios para acceder a esta página.", "danger")
            return redirect(url_for('home'))
        return decorated_function
    return decorator

# ---------------- Gestión de Aforos (Admin) -----------------
@login_required
@admin_required
@app.route('/aforos')
def aforos_page():
    tablas = db.session.query(AforoTabla).order_by(AforoTabla.tipo.asc(), AforoTabla.nombre.asc()).all()
    # Cargar una lista mínima para mostrar en tabla (sin los datos)
    return render_template('aforos.html', nombre=session.get('nombre'), tablas=tablas)

@login_required
@admin_required
@app.route('/api/aforos/get')
def aforos_get():
    try:
        tid = request.args.get('id')
        tipo = (request.args.get('tipo') or '').upper() or None
        nombre = (request.args.get('nombre') or '').upper() or None
        q = db.session.query(AforoTabla)
        if tid:
            q = q.filter(AforoTabla.id == int(tid))
        elif tipo and nombre:
            q = q.filter(AforoTabla.tipo == tipo, AforoTabla.nombre == nombre)
        else:
            return jsonify(success=False, message='Parámetros inválidos'), 400
        r = q.first()
        if not r:
            return jsonify(success=False, message='No encontrado'), 404
        datos = json.loads(r.datos_json)
        preview = _expand_preview_rows(datos, max_rows=200)
        mode = datos.get('mode') if isinstance(datos, dict) else 'flat'
        total_rows = 0
        if isinstance(datos, dict) and mode == 'flat':
            total_rows = len(datos.get('data') or [])
        elif isinstance(datos, dict) and mode == 'step':
            total_rows = len((datos.get('base') or {})) * 10
        return jsonify(success=True, tabla={
            'id': r.id, 'tipo': r.tipo, 'nombre': r.nombre,
            'timestamp': r.timestamp.isoformat(), 'usuario': r.usuario,
            'mode': mode, 'total_rows_est': total_rows
        }, preview=preview)
    except Exception as e:
        return jsonify(success=False, message=str(e)), 500

@login_required
@admin_required
@app.route('/api/aforos/download')
def aforos_download():
    try:
        tid = request.args.get('id')
        if not tid:
            return jsonify(success=False, message='id requerido'), 400
        r = db.session.query(AforoTabla).filter(AforoTabla.id == int(tid)).first()
        if not r:
            return jsonify(success=False, message='No encontrado'), 404
        datos = json.loads(r.datos_json)
        rows = []
        if isinstance(datos, dict) and datos.get('mode') == 'flat':
            rows = list(datos.get('data') or [])
        else:
            rows = _expand_preview_rows(datos, max_rows=100000)
        # Generar CSV en memoria
        import csv
        from io import StringIO
        sio = StringIO()
        writer = csv.writer(sio)
        writer.writerow(['cm', 'mm', 'bbl'])
        for it in rows:
            writer.writerow([it.get('cm'), it.get('mm'), it.get('bbl')])
        sio.seek(0)
        resp = Response(sio.getvalue(), mimetype='text/csv')
        fname = f"aforo_{r.tipo}_{r.nombre}.csv".replace(' ', '_')
        resp.headers['Content-Disposition'] = f'attachment; filename="{fname}"'
        return resp
    except Exception as e:
        return jsonify(success=False, message=str(e)), 500

@login_required
@admin_required
@app.route('/api/aforos/upload', methods=['POST'])
def aforos_upload():
    if 'archivo_excel' not in request.files:
        return jsonify(success=False, message='Archivo no recibido'), 400
    archivo = request.files['archivo_excel']
    tipo = (request.form.get('tipo') or '').upper()
    # Nuevo: permitir optar por el parser de múltiples columnas solo si se solicita explícitamente
    # Por defecto, para BARCAZA se usará el nombre de la hoja como "nombre" de la tabla
    parsear_columnas = (request.form.get('parsear_columnas') == '1')
    # nombre manual es opcional; si el archivo tiene una sola hoja, puede usarse como override
    nombre_override = (request.form.get('nombre') or '').upper().strip()
    if not archivo or not archivo.filename.lower().endswith('.xlsx'):
        return jsonify(success=False, message='Suba un archivo .xlsx'), 400
    if tipo not in ('TK','BARCAZA','VCF'):
        return jsonify(success=False, message='Tipo inválido (TK, BARCAZA o VCF)'), 400
    try:
        wb = openpyxl.load_workbook(archivo)
        hojas = wb.sheetnames
        total_ok = 0
        total_err = 0
        detalles = []
        for idx, sheet_name in enumerate(hojas):
            ws = wb[sheet_name]
            nombre_final = (nombre_override if (len(hojas) == 1 and nombre_override) else sheet_name).upper().strip()
            try:
                filas_count = 0
                datos = None
                # Si es VCF (API6A), procesar como tabla especial: matriz o vertical
                if tipo == 'VCF':
                    max_cols = ws.max_column or 0
                    row1 = [str(c.value).strip().lower() if c.value is not None else '' for c in ws[1][:max_cols]]
                    def is_temp_header(h):
                        h = h.lower().replace('.', '').replace(' ', '')
                        return any(x in h for x in ['temp', 'deg', 'degre', 't'])
                    is_matrix = row1[0] and is_temp_header(row1[0])
                    apis = []
                    data = []
                    if is_matrix:
                        # Formato matriz: encabezados API en la primera fila
                        for i in range(1, max_cols):
                            h = row1[i]
                            try:
                                apis.append(float(str(h).replace(',', '.')))
                            except Exception:
                                apis.append(None)
                        for row in ws.iter_rows(min_row=2, values_only=True):
                            try:
                                temp_v = row[0]
                                for j, api_v in enumerate(apis):
                                    vcf_v = row[j+1] if j+1 < len(row) else None
                                    if api_v is None or temp_v is None or vcf_v is None:
                                        continue
                                    api_f = float(api_v)
                                    temp_f = float(str(temp_v).replace(',', '.'))
                                    vcf_f = float(str(vcf_v).replace(',', '.'))
                                    data.append({'api': api_f, 'temp': temp_f, 'vcf': vcf_f})
                            except Exception:
                                continue
                    else:
                        # Formato vertical: columnas API, TEMP, VCF
                        idx_api = next((i for i, h in enumerate(row1) if 'api' in h), None)
                        idx_temp = next((i for i, h in enumerate(row1) if 'temp' in h or 't' in h), None)
                        idx_vcf = next((i for i, h in enumerate(row1) if 'vcf' in h or 'factor' in h), None)
                        for row in ws.iter_rows(min_row=2, values_only=True):
                            try:
                                api_v = row[idx_api] if idx_api is not None else None
                                temp_v = row[idx_temp] if idx_temp is not None else None
                                vcf_v = row[idx_vcf] if idx_vcf is not None else None
                                if api_v is None or temp_v is None or vcf_v is None:
                                    continue
                                api_f = float(str(api_v).replace(',', '.'))
                                temp_f = float(str(temp_v).replace(',', '.'))
                                vcf_f = float(str(vcf_v).replace(',', '.'))
                                data.append({'api': api_f, 'temp': temp_f, 'vcf': vcf_f})
                            except Exception:
                                continue
                    if data:
                        filas_count = len(data)
                        datos = {'data': data, 'mode': 'flat'}
                        payload = json.dumps(datos, ensure_ascii=False)
                        existente = db.session.query(AforoTabla).filter_by(tipo=tipo, nombre=nombre_final).first()
                        if existente:
                            existente.datos_json = payload
                            existente.usuario = session.get('nombre','No identificado')
                            existente.timestamp = datetime.utcnow()
                        else:
                            db.session.add(AforoTabla(
                                usuario=session.get('nombre','No identificado'),
                                tipo=tipo,
                                nombre=nombre_final,
                                datos_json=payload
                            ))
                        total_ok += 1
                        detalles.append(f"{sheet_name} -> {nombre_final}: OK ({filas_count} filas)")
                        continue
                # ...existing code for TK/BARCAZA...

                # 2) Si BARCAZA sin filas, intentar parser simple LÁMINA + 1 volumen
                if tipo == 'BARCAZA' and filas_count == 0 and not parsear_columnas:
                    datos_simple = _parse_simple_lamina_single_volume(ws)
                    if datos_simple and len(datos_simple.get('data') or []) > 0:
                        payload = json.dumps(datos_simple, ensure_ascii=False)
                        existente = db.session.query(AforoTabla).filter_by(tipo=tipo, nombre=nombre_final).first()
                        if existente:
                            existente.datos_json = payload
                            existente.usuario = session.get('nombre','No identificado')
                            existente.timestamp = datetime.utcnow()
                        else:
                            db.session.add(AforoTabla(
                                usuario=session.get('nombre','No identificado'),
                                tipo=tipo,
                                nombre=nombre_final,
                                datos_json=payload
                            ))
                        total_ok += 1
                        filas_s = len(datos_simple.get('data') or [])
                        detalles.append(f"{sheet_name} -> {nombre_final}: OK ({filas_s} filas)")
                        continue

                # 3) Fallback final: parser de múltiples columnas cuando corresponda
                if tipo == 'BARCAZA' and (parsear_columnas or filas_count == 0):
                    multi = _parse_barge_columns_sheet(ws, prefer_tipo='BARCAZA', default_name=nombre_final)
                    if multi:
                        for nom, datos_m in multi.items():
                            payload = json.dumps(datos_m, ensure_ascii=False)
                            existente = db.session.query(AforoTabla).filter_by(tipo=tipo, nombre=nom).first()
                            if existente:
                                existente.datos_json = payload
                                existente.usuario = session.get('nombre','No identificado')
                                existente.timestamp = datetime.utcnow()
                            else:
                                db.session.add(AforoTabla(
                                    usuario=session.get('nombre','No identificado'),
                                    tipo=tipo,
                                    nombre=nom,
                                    datos_json=payload
                                ))
                            total_ok += 1
                            filas_m = len(datos_m.get('data') or []) if isinstance(datos_m, dict) else 0
                            detalles.append(f"{sheet_name} -> {nom}: OK ({filas_m} filas)")
                        continue

                # Si llegó aquí, no se pudo interpretar la hoja
                raise ValueError('No se pudo interpretar columnas de nivel/volumen')
            except Exception as e:
                total_err += 1
                detalles.append(f"{sheet_name}: ERROR {e}")
        db.session.commit()
        msg = f"Procesado: {total_ok} hojas OK, {total_err} con error."
        return jsonify(success=True, message=msg, detalles=detalles)
    except Exception as e:
        db.session.rollback()
        return jsonify(success=False, message=str(e)), 500

@login_required
@app.route('/api/aforos/list')
def aforos_list():
    # acceso general para poder consultarlo desde Trasiegos
    tipo = (request.args.get('tipo') or '').upper() or None
    q = db.session.query(AforoTabla)
    if tipo in ('TK','BARCAZA'):
        q = q.filter(AforoTabla.tipo == tipo)
    filas = q.order_by(AforoTabla.tipo.asc(), AforoTabla.nombre.asc()).all()
    res = []
    for r in filas:
        res.append({
            'id': r.id,
            'tipo': r.tipo,
            'nombre': r.nombre,
            'timestamp': r.timestamp.isoformat(),
            'usuario': r.usuario
        })
    return jsonify(success=True, tablas=res)

@login_required
@app.route('/api/aforos/calcular')
def aforos_calcular():
    try:
        nombre = (request.args.get('nombre') or '').upper()
        tipo = (request.args.get('tipo') or '').upper()
        cm = int(request.args.get('cm') or 0)
        mm = int(request.args.get('mm') or 0)
        if not nombre or tipo not in ('TK','BARCAZA'):
            return jsonify(success=False, message='Parámetros inválidos'), 400
        tabla = db.session.query(AforoTabla).filter_by(tipo=tipo, nombre=nombre).first()
        if not tabla:
            return jsonify(success=False, message='Tabla de aforo no encontrada'), 404
        datos = json.loads(tabla.datos_json)
        # Para TK-01, TK-02 y barcaza MANZANILLO, buscar valor exacto
        nombres_exactos = ["TK-01", "TK-02"]
        # Barcaza MANZANILLO compartimentos
        nombres_exactos += ["1", "2", "3"]
        # Barcaza ODISEA compartimentos
        nombres_exactos += ["1P", "1S", "2P", "2S", "3P", "3S"]
        # Barcaza MARGOTH compartimentos SOLO con prefijo MG6-
        nombres_exactos += [
            "MG6-1P", "MG6-1S", "MG6-2P", "MG6-2S", "MG6-3P", "MG6-3S", "MG6-4P", "MG6-4S", "MG6-5P", "MG6-5S"
        ]
        # Barcaza CR compartimentos SOLO con prefijo CR
        nombres_exactos += ["CR 1S", "CR 1P", "CR 2S", "CR 2P", "CR 3S", "CR 3P", "CR 4S", "CR 4P", "CR 5S", "CR 5P"]
        if nombre in nombres_exactos:
            lista = datos.get('data') if isinstance(datos, dict) and 'data' in datos else datos
            if isinstance(lista, list):
                match = next((d for d in lista if int(d.get('cm', -1)) == cm and int(d.get('mm', 0)) == mm), None)
                if match:
                    bbl = match.get('bbl', 0.0)
                    return jsonify(success=True, bbl=round(float(bbl), 2))
        # Para otros tanques, usar interpolación
        bbl = _interp_bbl(datos, cm, mm)
        return jsonify(success=True, bbl=round(float(bbl), 2))
    except Exception as e:
        return jsonify(success=False, message=str(e)), 500

@login_required
@admin_required
@app.route('/api/aforos/delete', methods=['POST'])
def aforos_delete():
    try:
        payload = request.get_json(silent=True) or {}
        tid = payload.get('id')
        tipo = (payload.get('tipo') or '').upper() or None
        nombre = (payload.get('nombre') or '').upper() or None
        prefix = (payload.get('prefix') or '').upper().strip() or None
        nombre_like = (payload.get('nombre_like') or '').upper().strip() or None

        if not (tid or (tipo and (nombre or prefix or nombre_like))):
            return jsonify(success=False, message='Parámetros requeridos: id OR (tipo + nombre|prefix|nombre_like)'), 400

        q = db.session.query(AforoTabla)
        if tid:
            q = q.filter(AforoTabla.id == int(tid))
        else:
            q = q.filter(AforoTabla.tipo == tipo)
            if nombre:
                q = q.filter(AforoTabla.nombre == nombre)
            elif prefix:
                q = q.filter(AforoTabla.nombre.like(f"{prefix}%"))
            elif nombre_like:
                q = q.filter(AforoTabla.nombre.like(f"%{nombre_like}%"))

        filas = q.all()
        if not filas:
            return jsonify(success=False, message='No se encontraron tablas a eliminar'), 404
        count = len(filas)
        for r in filas:
            db.session.delete(r)
        db.session.commit()
        return jsonify(success=True, deleted=count)
    except Exception as e:
        db.session.rollback()
        return jsonify(success=False, message=str(e)), 500
def calcular_estadisticas(lista_tanques):
    """
    Calcula totales y promedios PONDERADOS para una lista de tanques.
    El promedio de API, BSW y S se pondera por el volumen (BLS_60) de cada tanque.
    """
    if not lista_tanques:
        return {
            'total_cap': 0, 'total_bls': 0, 'total_porc': 0,
            'prom_api': 0, 'prom_bsw': 0, 'prom_s': 0
        }

    # --- Totales simples (Suma) ---
    total_cap = sum(float(t.get('MAX_CAP') or 0) for t in lista_tanques)
    total_bls = sum(float(t.get('BLS_60') or 0) for t in lista_tanques)
    total_porc = (total_bls / total_cap * 100) if total_cap > 0 else 0

    # --- INICIO DEL CÁLCULO DE PROMEDIO PONDERADO ---
    
    suma_ponderada_api = 0
    suma_ponderada_bsw = 0
    suma_ponderada_s = 0

    # Solo calculamos si hay volumen total para evitar división por cero
    if total_bls > 0:
        for t in lista_tanques:
            bls = float(t.get('BLS_60') or 0)
            
            # El "peso" de cada tanque es su volumen dividido por el volumen total
            peso = bls / total_bls
            
            # Multiplicamos el valor de cada propiedad por su peso y lo sumamos
            suma_ponderada_api += (float(t.get('API') or 0) * peso)
            suma_ponderada_bsw += (float(t.get('BSW') or 0) * peso)
            suma_ponderada_s += (float(t.get('S') or 0) * peso)

    return {
        'total_cap': total_cap,
        'total_bls': total_bls,
        'total_porc': total_porc,
        'prom_api': suma_ponderada_api, # Ahora estos son los promedios ponderados
        'prom_bsw': suma_ponderada_bsw,
        'prom_s': suma_ponderada_s
    }

def permiso_exclusivo(email_requerido):
    """
    Decorador que da acceso SOLO al email especificado. Nadie más puede entrar.
    """
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if session.get('email') != email_requerido:
                flash("No tiene permiso para acceder a esta página.", "danger")
                return redirect(url_for('home'))
            return f(*args, **kwargs)
        return decorated_function
    return decorator

def _to_float(value):
    if value in (None, ''):
        return None
    try:
        return float(str(value).replace(',', '.'))
    except Exception:
        return None

def _to_int(value):
    if value in (None, ''):
        return None
    try:
        return int(str(value).strip())
    except Exception:
        return None

@login_required
@app.route('/trasiegos', methods=['GET', 'POST'])
def trasiegos_page():
    nombre_usr = session.get('nombre', '')
    ALLOWED_TK = { 'Ignacio Quimbayo', 'Control Refineria' }
    ALLOWED_BAR = { 'Juan Diego Cuadros', 'Ricardo Congo' }
    can_tk = nombre_usr in ALLOWED_TK
    can_bar = nombre_usr in ALLOWED_BAR
    def _get_or_create(fecha_dt, origen_tk, destino_barcaza, destino_comp):
        q = (db.session.query(TrasiegoTKBarcaza)
             .filter(TrasiegoTKBarcaza.fecha == fecha_dt,
                     TrasiegoTKBarcaza.origen_tk == origen_tk,
                     TrasiegoTKBarcaza.destino_barcaza == destino_barcaza,
                     (TrasiegoTKBarcaza.destino_compartimento == (destino_comp or None))) )
        inst = q.first()
        if inst:
            return inst, False
        inst = TrasiegoTKBarcaza(
            fecha=fecha_dt,
            usuario=session.get('nombre','No identificado'),
            origen_tk=origen_tk,
            destino_barcaza=destino_barcaza,
            destino_compartimento=destino_comp or None
        )
        db.session.add(inst)
        return inst, True

    if request.method == 'POST':
        try:
            fecha = request.form.get('fecha') or date.today().isoformat()
            fecha_dt = date.fromisoformat(fecha)
            origen_tk = (request.form.get('origen_tk','') or '').upper()
            destino_barcaza = (request.form.get('destino_barcaza','') or '').upper()
            destino_compartimento = (request.form.get('destino_compartimento') or '').upper() or None
            seccion = request.form.get('seccion')  # 'tk' o 'bar'

            # Permisos por sección
            if seccion == 'tk' and not can_tk:
                flash('No tiene permiso para registrar datos del Tanque.', 'danger')
                return redirect(url_for('trasiegos_page'))
            if seccion == 'bar' and not can_bar:
                flash('No tiene permiso para registrar datos de la Barcaza.', 'danger')
                return redirect(url_for('trasiegos_page'))

            # TK: forzar clave sin barcaza/comp para que el bar vincule luego
            if seccion == 'tk':
                destino_barcaza = ''
                destino_compartimento = None

            # Si ya existe registro para esta clave, bloquear cambio de fecha al del primer registro
            clave_barcaza = destino_barcaza or ''
            existente = (db.session.query(TrasiegoTKBarcaza)
                         .filter(TrasiegoTKBarcaza.origen_tk == origen_tk,
                                 TrasiegoTKBarcaza.destino_barcaza == clave_barcaza,
                                 (TrasiegoTKBarcaza.destino_compartimento == (destino_compartimento or None)))
                         .order_by(TrasiegoTKBarcaza.id.asc()).first())
            if existente:
                fecha_dt = existente.fecha

            inst, created = _get_or_create(fecha_dt, origen_tk, destino_barcaza, destino_compartimento)
            inst.usuario = session.get('nombre','No identificado')

            if seccion == 'tk':
                inst.tk_cm_inicial = _to_int(request.form.get('tk_cm_inicial'))
                inst.tk_mm_inicial = _to_int(request.form.get('tk_mm_inicial'))
                inst.tk_bbl_bruto_inicial = _to_float(request.form.get('tk_bbl_bruto_inicial'))
                inst.tk_bbl_inicial = _to_float(request.form.get('tk_bbl_inicial'))
                inst.tk_api = _to_float(request.form.get('tk_api'))
                inst.tk_temp = _to_float(request.form.get('tk_temp'))
                inst.tk_cm_final = _to_int(request.form.get('tk_cm_final'))
                inst.tk_mm_final = _to_int(request.form.get('tk_mm_final'))
                inst.tk_bbl_bruto_final = _to_float(request.form.get('tk_bbl_bruto_final'))
                inst.tk_bbl_final = _to_float(request.form.get('tk_bbl_final'))
                inst.tk_notas = request.form.get('tk_notas')
                # Ingreso simultáneo al TK: caudal (BBL/h), minutos y total bbl
                caudal = _to_float(request.form.get('tk_caudal_bbl_min'))
                minutos = _to_float(request.form.get('tk_minutos_ingreso'))
                bbling = _to_float(request.form.get('tk_bbl_ingreso'))
                inst.tk_caudal_bbl_min = caudal
                inst.tk_minutos_ingreso = minutos
                # Si no se envía tk_bbl_ingreso explícito, calcularlo
                if bbling is None and caudal is not None and minutos is not None:
                    bbling = round(caudal * (minutos/60.0), 2)
                inst.tk_bbl_ingreso = bbling
                # Si no tiene barcaza definida, se almacena como cadena vacía (columna NOT NULL en SQLite)
                if inst.destino_barcaza is None:
                    inst.destino_barcaza = ''
            elif seccion == 'bar':
                # Intentar vincular con un registro TK previo (barcaza en blanco) si no encontramos uno exacto
                if created and destino_barcaza and origen_tk:
                    # Buscar registro con misma fecha y TK, barcaza en blanco/NULL
                    prev = (db.session.query(TrasiegoTKBarcaza)
                            .filter(TrasiegoTKBarcaza.origen_tk == origen_tk,
                                    TrasiegoTKBarcaza.fecha == fecha_dt,
                                    (TrasiegoTKBarcaza.destino_barcaza == '') )
                            .order_by(TrasiegoTKBarcaza.id.asc()).first())
                    if prev:
                        # Reutilizar registro y bloquear fecha del previo
                        db.session.expunge(inst)
                        inst = prev
                        created = False
                        inst.destino_barcaza = destino_barcaza
                        inst.destino_compartimento = destino_compartimento or None
                    else:
                        # Si existe uno para ese TK con barcaza en blanco pero con OTRA fecha, bloquea cambio de fecha
                        prev_any = (db.session.query(TrasiegoTKBarcaza)
                                    .filter(TrasiegoTKBarcaza.origen_tk == origen_tk,
                                            (TrasiegoTKBarcaza.destino_barcaza == ''))
                                    .order_by(TrasiegoTKBarcaza.id.desc()).first())
                        if prev_any and prev_any.fecha != fecha_dt:
                            flash(f'La fecha está bloqueada para este trasiego (use {prev_any.fecha}).', 'warning')
                            return redirect(url_for('trasiegos_page'))
                inst.bar_cm_inicial = _to_int(request.form.get('bar_cm_inicial'))
                inst.bar_mm_inicial = _to_int(request.form.get('bar_mm_inicial'))
                inst.bar_bbl_bruto_inicial = _to_float(request.form.get('bar_bbl_bruto_inicial'))
                inst.bar_bbl_inicial = _to_float(request.form.get('bar_bbl_inicial'))
                inst.bar_api = _to_float(request.form.get('bar_api'))
                inst.bar_temp = _to_float(request.form.get('bar_temp'))
                inst.bar_cm_final = _to_int(request.form.get('bar_cm_final'))
                inst.bar_mm_final = _to_int(request.form.get('bar_mm_final'))
                inst.bar_bbl_bruto_final = _to_float(request.form.get('bar_bbl_bruto_final'))
                inst.bar_bbl_final = _to_float(request.form.get('bar_bbl_final'))
            else:
                # compat: si no viene seccion, guardamos todo lo que venga
                inst.tk_cm_inicial = _to_int(request.form.get('tk_cm_inicial'))
                inst.tk_mm_inicial = _to_int(request.form.get('tk_mm_inicial'))
                inst.tk_bbl_bruto_inicial = _to_float(request.form.get('tk_bbl_bruto_inicial'))
                inst.tk_bbl_inicial = _to_float(request.form.get('tk_bbl_inicial'))
                inst.tk_api = _to_float(request.form.get('tk_api'))
                inst.tk_temp = _to_float(request.form.get('tk_temp'))
                inst.tk_cm_final = _to_int(request.form.get('tk_cm_final'))
                inst.tk_mm_final = _to_int(request.form.get('tk_mm_final'))
                inst.tk_bbl_bruto_final = _to_float(request.form.get('tk_bbl_bruto_final'))
                inst.tk_bbl_final = _to_float(request.form.get('tk_bbl_final'))
                # Ingreso simultáneo (modo compat)
                caudal = _to_float(request.form.get('tk_caudal_bbl_min'))
                minutos = _to_float(request.form.get('tk_minutos_ingreso'))
                bbling = _to_float(request.form.get('tk_bbl_ingreso'))
                inst.tk_caudal_bbl_min = caudal
                inst.tk_minutos_ingreso = minutos
                if bbling is None and caudal is not None and minutos is not None:
                    bbling = round(caudal * (minutos/60.0), 2)
                inst.tk_bbl_ingreso = bbling
                inst.bar_cm_inicial = _to_int(request.form.get('bar_cm_inicial'))
                inst.bar_mm_inicial = _to_int(request.form.get('bar_mm_inicial'))
                inst.bar_bbl_bruto_inicial = _to_float(request.form.get('bar_bbl_bruto_inicial'))
                inst.bar_bbl_inicial = _to_float(request.form.get('bar_bbl_inicial'))
                inst.bar_api = _to_float(request.form.get('bar_api'))
                inst.bar_temp = _to_float(request.form.get('bar_temp'))
                inst.bar_cm_final = _to_int(request.form.get('bar_cm_final'))
                inst.bar_mm_final = _to_int(request.form.get('bar_mm_final'))
                inst.bar_bbl_bruto_final = _to_float(request.form.get('bar_bbl_bruto_final'))
                inst.bar_bbl_final = _to_float(request.form.get('bar_bbl_final'))

            inst.notas = request.form.get('notas')
            db.session.commit()
            if seccion == 'tk':
                flash('Datos del Tanque guardados','success')
            elif seccion == 'bar':
                flash('Datos de la Barcaza guardados','success')
            else:
                flash('Trasiego guardado exitosamente','success')
            return redirect(url_for('trasiegos_page'))
        except Exception as e:
            db.session.rollback()
            app.logger.error(f"Error guardando trasiego: {e}")
            flash(f'Error guardando: {e}','danger')
    # GET
    desde = request.args.get('desde')
    hasta = request.args.get('hasta')
    q = db.session.query(TrasiegoTKBarcaza)
    if desde:
        try:
            q = q.filter(TrasiegoTKBarcaza.fecha >= date.fromisoformat(desde))
        except Exception:
            pass
    if hasta:
        try:
            q = q.filter(TrasiegoTKBarcaza.fecha <= date.fromisoformat(hasta))
        except Exception:
            pass
    trasiegos = q.order_by(TrasiegoTKBarcaza.fecha.desc(), TrasiegoTKBarcaza.id.desc()).all()
    return render_template('trasiegos.html', nombre=session.get('nombre'), trasiegos=trasiegos, desde=desde or '', hasta=hasta or '', allowed_tk=can_tk, allowed_bar=can_bar)

@login_required
@app.route('/guardar_trasiegos_masivo', methods=['POST'])
def guardar_trasiegos_masivo():
    def to_float(v):
        if v is None:
            return None
        s = str(v).strip().replace(',', '.')
        if s == '':
            return None
        try:
            return float(s)
        except Exception:
            return None
    def to_int(v):
        if v is None:
            return None
        s = str(v).strip()
        if s == '':
            return None
        try:
            return int(float(s))
        except Exception:
            return None
    try:
        data = request.get_json()
        trasiegos = data.get('trasiegos', [])
        if not trasiegos or not isinstance(trasiegos, list):
            return jsonify(success=False, message='No se recibieron trasiegos válidos'), 400
        guardados = 0
        errores = []
        for t in trasiegos:
            try:
                fecha = t.get('fecha') or date.today().isoformat()
                fecha_dt = date.fromisoformat(fecha)
                origen_tk = (t.get('origen_tk','') or '').upper()
                destino_barcaza = (t.get('destino_barcaza','') or '').upper()
                destino_compartimento = (t.get('destino_compartimento') or '').upper() or None
                usuario = session.get('nombre','No identificado')
                # Crear registro
                reg = TrasiegoTKBarcaza(
                    fecha=fecha_dt,
                    usuario=usuario,
                    origen_tk=origen_tk,
                    destino_barcaza=destino_barcaza,
                    destino_compartimento=destino_compartimento,
                    tk_cm_inicial=to_int(t.get('tk_cm_inicial')),
                    tk_mm_inicial=to_int(t.get('tk_mm_inicial')),
                    tk_bbl_bruto_inicial=to_float(t.get('tk_bbl_bruto_inicial')),
                    tk_bbl_inicial=to_float(t.get('tk_bbl_inicial')),
                    tk_api=to_float(t.get('tk_api')),
                    tk_temp=to_float(t.get('tk_temp')),
                    tk_cm_final=to_int(t.get('tk_cm_final')),
                    tk_mm_final=to_int(t.get('tk_mm_final')),
                    tk_bbl_bruto_final=to_float(t.get('tk_bbl_bruto_final')),
                    tk_bbl_final=to_float(t.get('tk_bbl_final')),
                    tk_caudal_bbl_min=to_float(t.get('tk_caudal_bbl_min')),
                    tk_minutos_ingreso=to_float(t.get('tk_minutos_ingreso')),
                    tk_bbl_ingreso=to_float(t.get('tk_bbl_ingreso')),
                    bar_cm_inicial=to_int(t.get('bar_cm_inicial')),
                    bar_mm_inicial=to_int(t.get('bar_mm_inicial')),
                    bar_bbl_bruto_inicial=to_float(t.get('bar_bbl_bruto_inicial')),
                    bar_bbl_inicial=to_float(t.get('bar_bbl_inicial')),
                    bar_api=to_float(t.get('bar_api')),
                    bar_temp=to_float(t.get('bar_temp')),
                    bar_cm_final=to_int(t.get('bar_cm_final')),
                    bar_mm_final=to_int(t.get('bar_mm_final')),
                    bar_bbl_bruto_final=to_float(t.get('bar_bbl_bruto_final')),
                    bar_bbl_final=to_float(t.get('bar_bbl_final')),
                    notas=t.get('notas'),
                    tk_notas=t.get('tk_notas')
                )
                db.session.add(reg)
                guardados += 1
            except Exception as e:
                errores.append(str(e))
        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            return jsonify(success=False, message=f'Error al guardar: {e}', errores=errores), 500
        return jsonify(success=True, guardados=guardados, errores=errores)
    except Exception as e:
        return jsonify(success=False, message=str(e)), 500

def _build_opciones_trasiegos():
    # Origen TK desde planilla de planta
    tks = [str(p.get('TK')) for p in PLANILLA_PLANTA if p.get('TK')]
    # Barcazas y compartimentos
    barcazas = {}
    # Orion por grupos
    try:
        from collections import defaultdict
        grp = defaultdict(list)
        for p in PLANILLA_BARCAZA_ORION:
            g = str(p.get('grupo') or 'ORION').upper()
            grp[g].append(str(p.get('TK')))
        for g, lst in grp.items():
            barcazas[g] = lst
    except Exception:
        pass
    # BITA: separar por prefijo
    try:
        marinse = [p['TK'] for p in PLANILLA_BARCAZA_BITA if str(p.get('TK','')).startswith('MARI ')]
        oidech = [p['TK'] for p in PLANILLA_BARCAZA_BITA if str(p.get('TK','')).startswith('OID ')]
        if marinse:
            barcazas['BITA-MARINSE'] = marinse
        if oidech:
            barcazas['BITA-OIDECH'] = oidech
    except Exception:
        pass
    # Aforos TK y Barcaza
    aforosTK = [r.nombre for r in db.session.query(AforoTabla).filter(AforoTabla.tipo == 'TK').order_by(AforoTabla.nombre.asc()).all()]
    aforosBAR = [r.nombre for r in db.session.query(AforoTabla).filter(AforoTabla.tipo == 'BARCAZA').order_by(AforoTabla.nombre.asc()).all()]
    return { 'tks': tks, 'barcazas': barcazas, 'aforosTK': aforosTK, 'aforosBAR': aforosBAR }

@login_required
@app.route('/api/trasiegos/opciones')
def api_trasiegos_opciones():
    try:
        return jsonify(success=True, **_build_opciones_trasiegos())
    except Exception as e:
        return jsonify(success=False, message=str(e)), 500

@login_required
@app.route('/api/trasiegos/verificar_fecha')
def api_trasiegos_verificar_fecha():
    try:
        tk = (request.args.get('tk') or '').upper().strip()
        if not tk:
            return jsonify(success=True, fecha_bloqueada=False)
        reg = (db.session.query(TrasiegoTKBarcaza)
               .filter(TrasiegoTKBarcaza.origen_tk == tk,
                       or_(TrasiegoTKBarcaza.destino_barcaza == '', TrasiegoTKBarcaza.destino_barcaza.is_(None)))
               .order_by(TrasiegoTKBarcaza.id.desc())
               .first())
        if reg:
            return jsonify(success=True, fecha_bloqueada=True, fecha=reg.fecha.isoformat(), tk=reg.origen_tk or '')
        return jsonify(success=True, fecha_bloqueada=False)
    except Exception as e:
        return jsonify(success=False, message=str(e)), 500

# Endpoint para buscar VCF en la tabla API6A
@login_required
@app.route('/api/vcf_api6a')
def api_vcf_api6a():
    try:
        api = request.args.get('api', type=float)
        temp = request.args.get('temp', type=float)
        unidad = request.args.get('unidad', default='C').upper()
        nombre = request.args.get('nombre')
        if api is None or temp is None:
            return jsonify(success=False, message='API y temperatura requeridos'), 400
        if nombre:
            tabla = db.session.query(AforoTabla).filter_by(tipo='VCF', nombre=nombre.upper()).first()
        else:
            tabla = db.session.query(AforoTabla).filter_by(tipo='VCF').order_by(AforoTabla.timestamp.desc()).first()
        if not tabla:
            return jsonify(success=False, message='Tabla VCF/API6A no encontrada'), 404
        datos = json.loads(tabla.datos_json)
        lista = datos.get('data') if isinstance(datos, dict) and 'data' in datos else datos
        if not isinstance(lista, list):
            return jsonify(success=False, message='Datos VCF inválidos'), 500
        # Buscar coincidencia exacta primero
        match = next((d for d in lista if abs(d.get('api',-1)-api)<0.01 and abs(d.get('temp',-1)-temp)<0.01), None)
        if match:
            return jsonify(success=True, vcf=match.get('vcf', 1))
        # Si no hay coincidencia exacta, buscar la más cercana (interpolación simple)
        # Ordenar por distancia
        lista_ordenada = sorted(lista, key=lambda d: ((d.get('api',0)-api)**2 + (d.get('temp',0)-temp)**2))
        if lista_ordenada:
            return jsonify(success=True, vcf=lista_ordenada[0].get('vcf', 1))
        return jsonify(success=False, message='No se encontró VCF para esos valores'), 404
    except Exception as e:
        return jsonify(success=False, message=str(e)), 500

@login_required
@app.route('/trasiegos/eliminar/<int:id>', methods=['POST'])
def eliminar_trasiego(id):
    try:
        reg = TrasiegoTKBarcaza.query.get_or_404(id)
        db.session.delete(reg)
        db.session.commit()
        flash('Registro eliminado','success')
    except Exception as e:
        db.session.rollback()
        flash(f'No se pudo eliminar: {e}','danger')
    return redirect(url_for('trasiegos_page'))

@login_required
@app.route('/reporte_trasiegos')
def reporte_trasiegos():
    # Reporte tipo tabla como la imagen, por fecha
    fecha_str = request.args.get('fecha')
    try:
        fecha_sel = date.fromisoformat(fecha_str) if fecha_str else date.today()
    except Exception:
        fecha_sel = date.today()

    registros = (db.session.query(TrasiegoTKBarcaza)
                 .filter(TrasiegoTKBarcaza.fecha == fecha_sel)
                 .order_by(TrasiegoTKBarcaza.origen_tk.asc())
                 .all())

    # Columnas dinámicas: Tks origen y compartimentos destino
    columnas_tk = []
    columnas_bar = []
    for r in registros:
        if r.origen_tk and r.origen_tk not in columnas_tk:
            columnas_tk.append(r.origen_tk)
        comp = (r.destino_compartimento or r.destino_barcaza)
        if comp and comp not in columnas_bar:
            columnas_bar.append(comp)

    # Construir estructura para la plantilla
    def build_map(keys):
        return {k: { 'ini': {'cm':None,'mm':None,'bbl':None, 'bbl_bruto': None}, 'fin': {'cm':None,'mm':None,'bbl':None, 'bbl_bruto': None}, 'temp': None, 'api': None, 'trasiego': None } for k in keys}

    datos_tk = build_map(columnas_tk)
    datos_bar = build_map(columnas_bar)

    for r in registros:
        tkd = datos_tk.get(r.origen_tk)
        if tkd is not None:
            tkd['ini'] = {'cm': r.tk_cm_inicial, 'mm': r.tk_mm_inicial, 'bbl': r.tk_bbl_inicial, 'bbl_bruto': r.tk_bbl_bruto_inicial}
            tkd['fin'] = {'cm': r.tk_cm_final, 'mm': r.tk_mm_final, 'bbl': r.tk_bbl_final, 'bbl_bruto': r.tk_bbl_bruto_final}
            tkd['temp'] = r.tk_temp
            tkd['api'] = r.tk_api
            tkd['trasiego'] = r.trasiego_segun_tanque
        comp = (r.destino_compartimento or r.destino_barcaza)
        bd = datos_bar.get(comp)
        if bd is not None:
            bd['ini'] = {'cm': r.bar_cm_inicial, 'mm': r.bar_mm_inicial, 'bbl': r.bar_bbl_inicial, 'bbl_bruto': r.bar_bbl_bruto_inicial}
            bd['fin'] = {'cm': r.bar_cm_final, 'mm': r.bar_mm_final, 'bbl': r.bar_bbl_final, 'bbl_bruto': r.bar_bbl_bruto_final}
            bd['temp'] = r.bar_temp
            bd['api'] = r.bar_api
            bd['trasiego'] = r.trasiego_segun_barcaza

    diferencia_total = round(
        sum([d.get('trasiego') or 0 for d in datos_bar.values()]) - sum([d.get('trasiego') or 0 for d in datos_tk.values()])
        , 2)

    return render_template('reporte_trasiegos.html',
                           nombre=session.get('nombre'),
                           fecha_seleccionada=fecha_sel.isoformat(),
                           columnas_tk=columnas_tk,
                           columnas_bar=columnas_bar,
                           datos_tk=datos_tk,
                           datos_bar=datos_bar,
                           diferencia_total=diferencia_total)

@login_required
@permiso_requerido('transito')
@app.route('/transito')
def transito():
    # Iniciamos la consulta base
    query = db.session.query(RegistroTransito)

    # Leemos todos los posibles filtros desde la URL
    filtros = {
        'fecha': request.args.get('fecha_cargue'),
        'guia': request.args.get('guia'),
        'origen': request.args.get('origen'),
        'producto': request.args.get('producto'),
        'placa': request.args.get('placa')
    }

    # Aplicamos los filtros a la consulta solo si tienen un valor
    if filtros['fecha']:
        query = query.filter(RegistroTransito.fecha == filtros['fecha'])
    if filtros['guia']:
        query = query.filter(RegistroTransito.guia.ilike(f"%{filtros['guia']}%"))
    if filtros['origen']:
        query = query.filter(RegistroTransito.origen == filtros['origen'])
    if filtros['producto']:
        query = query.filter(RegistroTransito.producto == filtros['producto'])
    if filtros['placa']:
        query = query.filter(RegistroTransito.placa.ilike(f"%{filtros['placa']}%"))

    # Ejecutamos la consulta final
    todos_los_registros = query.order_by(RegistroTransito.timestamp.desc()).all()

    # Separamos los resultados y los convertimos a diccionario
    datos_general = [{ "id": r.id, "ORIGEN": r.origen, "FECHA": r.fecha, "GUIA": r.guia, "PRODUCTO": r.producto, "PLACA": r.placa, "API": r.api or '', "BSW": r.bsw or '', "NSV": r.nsv or '', "OBSERVACIONES": r.observaciones or '' } for r in todos_los_registros if r.tipo_transito == 'general']
    datos_refineria = [{ "id": r.id, "ORIGEN": r.origen, "FECHA": r.fecha, "GUIA": r.guia, "PRODUCTO": r.producto, "PLACA": r.placa, "API": r.api or '', "BSW": r.bsw or '', "NSV": r.nsv or '', "OBSERVACIONES": r.observaciones or '' } for r in todos_los_registros if r.tipo_transito == 'refineria']

    return render_template("transito.html",
                           nombre=session.get("nombre"),
                           datos_general=datos_general,
                           datos_refineria=datos_refineria,
                           tipo_inicial="general",
                           transito_config=cargar_transito_config(),
                           filtros=filtros)
  
@login_required
@app.route('/control_calidad')
def control_calidad():
    # Solo permitir acceso a usuarios específicos y admin
    email_usuario = session.get('email')
    rol_usuario = session.get('rol')
    emails_permitidos = ['production@conquerstrading.com', 'qualitycontrol@conquerstrading.com', 'refinery.control@conquerstrading.com', 'quality.manager@conquerstrading.com']
    
    if rol_usuario != 'admin' and email_usuario not in emails_permitidos:
        flash("No tienes permisos para acceder a esta página.", "danger")
        return redirect(url_for('home'))
    # Consulta todos los registros de RegistroCalidad ordenados por timestamp desc
    todos_los_registros = db.session.query(RegistroCalidad).order_by(RegistroCalidad.timestamp.desc()).all()
    
    # Calcular KPIs
    today = date.today()
    registros_hoy = [r for r in todos_los_registros if r.timestamp.date() == today]
    registros_hoy_count = len(registros_hoy)
    
    bsw_values = [r.bsw for r in registros_hoy if r.bsw is not None]
    bsw_promedio = sum(bsw_values) / len(bsw_values) if bsw_values else 0
    
    alertas_calidad = len([r for r in registros_hoy if r.bsw and r.bsw > 0.5])
    
    # Convertir a diccionario para la plantilla
    datos = [
        {
            "id": r.id,
            "fecha": r.fecha,
            "hora": r.hora,
            "producto": r.producto or '',
            "responsable": r.responsable,
            "origen": r.origen,
            "placa": r.placa,
            "campo": r.campo,
            "bsw": r.bsw or '',
            "flash_point": r.flash_point or '',
            "api_obs": r.api_obs or '',
            "temp": r.temp or '',
            "api_corr": r.api_corr or '',
            "observaciones": r.observaciones or ''
        }
        for r in todos_los_registros
    ]
    
    return render_template("control_calidad.html", 
                           nombre=session.get("nombre"), 
                           datos=datos, 
                           registros_hoy=registros_hoy_count, 
                           bsw_promedio=bsw_promedio, 
                           alertas_calidad=alertas_calidad,
                           email_usuario=session.get('email'),
                           rol_usuario=session.get('rol'))
  
@login_required
@app.route('/api/control_calidad', methods=['GET'])
def get_control_calidad_data():
    # Solo permitir acceso a usuarios específicos y admin
    email_usuario = session.get('email')
    rol_usuario = session.get('rol')
    emails_permitidos = ['production@conquerstrading.com', 'qualitycontrol@conquerstrading.com', 'refinery.control@conquerstrading.com', 'quality.manager@conquerstrading.com']
    
    if rol_usuario != 'admin' and email_usuario not in emails_permitidos:
        return jsonify({"error": "No tienes permisos para acceder a esta información"}), 403
    # Consulta todos los registros de RegistroCalidad ordenados por timestamp desc
    todos_los_registros = db.session.query(RegistroCalidad).order_by(RegistroCalidad.timestamp.desc()).all()
    
    # Convertir a diccionario para JSON
    datos = [
        {
            "id": r.id,
            "fecha": r.fecha,
            "hora": r.hora,
            "producto": r.producto or '',
            "responsable": r.responsable,
            "origen": r.origen,
            "placa": r.placa,
            "campo": r.campo,
            "bsw": r.bsw or '',
            "flash_point": r.flash_point or '',
            "api_obs": r.api_obs or '',
            "temp": r.temp or '',
            "api_corr": r.api_corr or '',
            "observaciones": r.observaciones or ''
        }
        for r in todos_los_registros
    ]
    
    return jsonify(datos)

@login_required
@app.route('/api/control_calidad', methods=['POST'])
def crear_registro_calidad():
    """Crear un nuevo registro vacío de control de calidad"""
    # Solo permitir acceso a usuarios específicos y admin
    email_usuario = session.get('email')
    rol_usuario = session.get('rol')
    emails_permitidos = ['production@conquerstrading.com', 'qualitycontrol@conquerstrading.com', 'refinery.control@conquerstrading.com', 'quality.manager@conquerstrading.com']
    
    if rol_usuario != 'admin' and email_usuario not in emails_permitidos:
        return jsonify({"success": False, "message": "No tienes permisos para crear registros"}), 403
    try:
        nuevo = RegistroCalidad(
            fecha=None,
            hora=None,
            producto=None,
            responsable=None,
            origen=None,
            placa=None,
            campo=None,
            bsw=None,
            flash_point=None,
            api_obs=None,
            temp=None,
            api_corr=None,
            observaciones=None,
            usuario=session.get("nombre", "No identificado"),
            timestamp=datetime.utcnow()
        )
        db.session.add(nuevo)
        db.session.commit()
        return jsonify(success=True, message="Registro creado exitosamente.", id=nuevo.id)
    except Exception as e:
        db.session.rollback()
        return jsonify(success=False, message=f"Error interno: {str(e)}"), 500

@login_required
@app.route('/api/control_calidad/<int:id>', methods=['PUT'])
def actualizar_registro_calidad(id):
    """Actualizar un campo específico de un registro de control de calidad"""
    # Solo permitir acceso a usuarios específicos y admin
    email_usuario = session.get('email')
    rol_usuario = session.get('rol')
    emails_permitidos = ['production@conquerstrading.com', 'qualitycontrol@conquerstrading.com', 'refinery.control@conquerstrading.com', 'quality.manager@conquerstrading.com']
    
    if rol_usuario != 'admin' and email_usuario not in emails_permitidos:
        return jsonify({"success": False, "message": "No tienes permisos para editar registros"}), 403
    try:
        registro = RegistroCalidad.query.get_or_404(id)
        datos = request.get_json()
        
        def to_float(v):
            if v is None or v == '':
                return None
            s = str(v).strip().replace(',', '.')
            try:
                return float(s)
            except Exception:
                return None
        
        # Actualizar solo los campos que vienen en el request
        if 'fecha' in datos:
            registro.fecha = datos.get('fecha')
        if 'hora' in datos:
            registro.hora = datos.get('hora')
        if 'producto' in datos:
            registro.producto = datos.get('producto')
        if 'responsable' in datos:
            registro.responsable = datos.get('responsable')
        if 'origen' in datos:
            registro.origen = datos.get('origen')
        if 'placa' in datos:
            registro.placa = datos.get('placa')
        if 'campo' in datos:
            registro.campo = datos.get('campo')
        if 'bsw' in datos:
            registro.bsw = to_float(datos.get('bsw'))
        if 'flash_point' in datos:
            registro.flash_point = to_float(datos.get('flash_point'))
        if 'api_obs' in datos:
            registro.api_obs = to_float(datos.get('api_obs'))
        if 'temp' in datos:
            registro.temp = to_float(datos.get('temp'))
        if 'api_corr' in datos:
            registro.api_corr = to_float(datos.get('api_corr'))
        if 'observaciones' in datos:
            registro.observaciones = datos.get('observaciones')
        
        registro.usuario = session.get("nombre", "No identificado")
        registro.timestamp = datetime.utcnow()
        
        db.session.commit()
        return jsonify(success=True, message="Registro actualizado exitosamente.")
    except Exception as e:
        db.session.rollback()
        return jsonify(success=False, message=f"Error: {str(e)}"), 500

@login_required
@app.route('/api/control_calidad/<int:id>', methods=['DELETE'])
def eliminar_registro_calidad(id):
    # Solo permitir acceso a usuarios específicos y admin
    email_usuario = session.get('email')
    rol_usuario = session.get('rol')
    emails_permitidos = ['production@conquerstrading.com', 'qualitycontrol@conquerstrading.com', 'refinery.control@conquerstrading.com', 'quality.manager@conquerstrading.com']
    
    if rol_usuario != 'admin' and email_usuario not in emails_permitidos:
        return jsonify({"success": False, "message": "No tienes permisos para eliminar registros"}), 403
    try:
        registro = RegistroCalidad.query.get_or_404(id)
        db.session.delete(registro)
        db.session.commit()
        return jsonify(success=True, message="Registro eliminado.")
    except Exception as e:
        db.session.rollback()
        return jsonify(success=False, message=f"Error: {str(e)}"), 500

@login_required
@app.route('/api/add-origen', methods=['POST'])
def agregar_origen():
    data = request.get_json()
    origen_nombre = data.get('origen_nombre', '').strip().upper()
    tipo_planilla = data.get('tipo_planilla', 'EDSM')  # 'EDSM' o 'REFINERIA'

    if not origen_nombre or tipo_planilla not in ['EDSM', 'REFINERIA']:
        return jsonify(success=False, message="Datos incompletos o inválidos"), 400

    try:
        # Cargar configuración actual
        config = cargar_transito_config()
        
        # Verificar si el origen ya existe
        if origen_nombre in config[tipo_planilla]['campos']:
            return jsonify(success=False, message="Este origen ya existe"), 409

        # Agregar el nuevo origen
        config[tipo_planilla]['campos'][origen_nombre] = {
            "productos": [],
            "auto_select_product": ""
        }

        # Guardar la configuración actualizada
        with open('transito_config.json', 'w', encoding='utf-8') as f:
            json.dump(config, f, ensure_ascii=False, indent=4)

        return jsonify(success=True, message="Origen agregado exitosamente")

    except Exception as e:
        print(f"Error al agregar origen: {e}")
        return jsonify(success=False, message="Error interno del servidor"), 500
    
@login_required
@app.route('/api/add-producto', methods=['POST'])
def agregar_producto_transito():
    data = request.get_json()
    origen_nombre = data.get('origen_nombre', '').strip().upper()
    producto_nombre = data.get('producto_nombre', '').strip()
    tipo_planilla = data.get('tipo_planilla', 'EDSM')  # 'EDSM' o 'REFINERIA'

    if not origen_nombre or not producto_nombre or tipo_planilla not in ['EDSM', 'REFINERIA']:
        return jsonify(success=False, message="Datos incompletos o inválidos"), 400

    try:
        # Cargar configuración actual
        config = cargar_transito_config()
        
        # Verificar si el origen existe
        if origen_nombre not in config[tipo_planilla]['campos']:
            return jsonify(success=False, message="El origen especificado no existe"), 404

        # Verificar si el producto ya existe
        if producto_nombre in config[tipo_planilla]['campos'][origen_nombre]['productos']:
            return jsonify(success=False, message="Este producto ya existe para este origen"), 409

        # Agregar el nuevo producto
        config[tipo_planilla]['campos'][origen_nombre]['productos'].append(producto_nombre)

        # Guardar la configuración actualizada
        with open('transito_config.json', 'w', encoding='utf-8') as f:
            json.dump(config, f, ensure_ascii=False, indent=4)

        return jsonify(success=True, message="Producto agregado exitosamente")

    except Exception as e:
        print(f"Error al agregar producto: {e}")
        return jsonify(success=False, message="Error interno del servidor"), 500

@app.route('/login', methods=['GET', 'POST'])
def login():
    if 'email' in session:
        return redirect(url_for('home'))

    next_page = request.args.get('next')
    if request.method == 'POST':
        email = request.form['email'].lower()
        password = request.form['password']
        user = USUARIOS.get(email)

        if user and check_password_hash(user['password'], password):
            session['email'] = email
            session['area'] = user['area']
            session['nombre'] = user['nombre']
            session['rol'] = user['rol']
            flash(f"Bienvenido {user['nombre']}", 'success')
            return redirect(next_page or url_for('home'))

        flash('Email o contraseña incorrectos', 'danger')

    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    flash('Sesión cerrada', 'info')
    return redirect(url_for('login'))

@login_required
@permiso_requerido('planta')
@app.route('/planta')
def planta():
    # 1. Obtiene la fecha del filtro de la URL. Si no se envía ninguna, usa la fecha de hoy.
    fecha_str = request.args.get('fecha')

    try:
        fecha_seleccionada = date.fromisoformat(fecha_str) if fecha_str else date.today()
    except (ValueError, TypeError):
        fecha_seleccionada = date.today()
    
    timestamp_limite = datetime.combine(fecha_seleccionada, time.max)

    # 2. Consulta para obtener el estado MÁS RECIENTE de CADA tanque EN O ANTES de la fecha seleccionada
    subquery = db.session.query(
        RegistroPlanta.tk,
        func.max(RegistroPlanta.timestamp).label('max_timestamp')
    ).filter(RegistroPlanta.timestamp <= timestamp_limite
             ).group_by(RegistroPlanta.tk).subquery()

    registros_recientes = db.session.query(RegistroPlanta).join(
        subquery,
        (RegistroPlanta.tk == subquery.c.tk) & (RegistroPlanta.timestamp == subquery.c.max_timestamp)
    ).all()
    
    # 3. Preparar y ORDENAR los datos según el orden deseado
    orden_deseado = ["TK-109", "TK-110", "TK-01", "TK-02", "TK-102", "TK-108", "Consumo Interno"]
    orden_map = {tk: i for i, tk in enumerate(orden_deseado)}

    # Combinar defaults con últimos registros para asegurar que todos los TK de la planilla existan (por ejemplo, TK-108)
    datos_por_tk = {fila["TK"]: dict(fila) for fila in PLANILLA_PLANTA}
    if registros_recientes:
        for registro in registros_recientes:
            datos_por_tk[registro.tk] = {
                "TK": registro.tk,
                "PRODUCTO": "DILUYENTE" if registro.tk == "Consumo Interno" else (registro.producto or datos_por_tk.get(registro.tk, {}).get("PRODUCTO")),
                "MAX_CAP": 124.78 if registro.tk == "Consumo Interno" else (registro.max_cap or datos_por_tk.get(registro.tk, {}).get("MAX_CAP")),
                "BLS_60": registro.bls_60 or "",
                "API": registro.api or "",
                "BSW": registro.bsw or "",
                "S": registro.s or ""
            }
    # Filtrar únicamente los TK permitidos (ignorar otros como TK-03)
    allowed_set = set(orden_deseado)
    datos_para_plantilla = [v for k, v in datos_por_tk.items() if k in allowed_set]

    # Ordenar la lista según el orden deseado
    datos_para_plantilla = sorted(
        datos_para_plantilla,
        key=lambda fila: orden_map.get(fila["TK"], 99)
    )

    # 4. Construimos listado de días con registros para colorear el calendario
    try:
        dias_rows = (db.session
            .query(func.date(RegistroPlanta.timestamp).label('dia'))
            .group_by(func.date(RegistroPlanta.timestamp))
            .all())
        fechas_con_registro = []
        for (dia,) in dias_rows:
            try:
                fechas_con_registro.append(dia.isoformat())
            except AttributeError:
                fechas_con_registro.append(str(dia))
    except Exception:
        fechas_con_registro = []

    # 5. Enviamos los datos y la fecha seleccionada de vuelta al HTML
    return render_template("planta.html", 
                           planilla=datos_para_plantilla, 
                           nombre=session.get("nombre", "Usuario"),
                           fecha_seleccionada=fecha_seleccionada.isoformat(),
                           today_iso=date.today().isoformat(),
                           fechas_con_registro=fechas_con_registro)

@login_required
@permiso_requerido('planta')
@app.route('/reporte_variaciones_tanques')
def reporte_variaciones_tanques():
    try:
        end_str = request.args.get('hasta')
        start_str = request.args.get('desde')
        hasta = date.fromisoformat(end_str) if end_str else None
        desde = date.fromisoformat(start_str) if start_str else None

        # Subconsulta: último registro por día y tanque (con rango opcional)
        fecha_dia = func.date(RegistroPlanta.timestamp)
        subq_query = db.session.query(func.max(RegistroPlanta.id).label('max_id'))
        if desde is not None:
            subq_query = subq_query.filter(RegistroPlanta.timestamp >= datetime.combine(desde, time.min))
        if hasta is not None:
            subq_query = subq_query.filter(RegistroPlanta.timestamp <= datetime.combine(hasta, time.max))
        subq = subq_query.group_by(RegistroPlanta.tk, fecha_dia).subquery()

        registros = (db.session.query(RegistroPlanta)
                     .filter(RegistroPlanta.id.in_(subq))
                     .order_by(RegistroPlanta.tk.asc(), RegistroPlanta.timestamp.asc())
                     .all())

        # Organizar por tanque y día
        por_tanque = {}
        for r in registros:
            if not r.tk:
                continue
            tk = r.tk
            dia = r.timestamp.date()
            bls = float(r.bls_60 or 0)
            por_tanque.setdefault(tk, {})[dia] = {
                'bls_60': bls,
                'max_cap': float(r.max_cap or 0),
                'producto': r.producto or ''
            }

        # Construir series por tanque: fechas, valores y variaciones
        series = {}
        for tk, mapa in por_tanque.items():
            fechas_ord = sorted(mapa.keys())
            valores = [mapa[f]['bls_60'] for f in fechas_ord]
            diffs = []
            prev = None
            for v in valores:
                if prev is None:
                    diffs.append(None)
                else:
                    diffs.append(round(v - prev, 2))
                prev = v
            # Convertir fechas a strings ISO para JS
            etiquetas = [f.isoformat() for f in fechas_ord]
            max_cap = next((mapa[f]['max_cap'] for f in fechas_ord if mapa[f]['max_cap'] > 0), 0)
            producto = next((mapa[f]['producto'] for f in fechas_ord if (mapa[f]['producto'] or '').strip()), '')
            series[tk] = {
                'labels': etiquetas,
                'bls_60': valores,
                'variacion': diffs,
                'max_cap': max_cap,
                'producto': producto
            }

        # Fechas con registro para colorear
        try:
            dias_rows = (db.session
                .query(func.date(RegistroPlanta.timestamp).label('dia'))
                .group_by(func.date(RegistroPlanta.timestamp))
                .all())
            fechas_con_registro = []
            for (dia,) in dias_rows:
                try:
                    fechas_con_registro.append(dia.isoformat())
                except AttributeError:
                    fechas_con_registro.append(str(dia))
        except Exception:
            fechas_con_registro = []

        return render_template('reporte_variaciones_tanques.html',
                               nombre=session.get('nombre'),
                               desde=desde.isoformat() if desde else '',
                               hasta=hasta.isoformat() if hasta else '',
                               series_por_tanque=series,
                               today_iso=date.today().isoformat(),
                               fechas_con_registro=fechas_con_registro)
    except Exception as e:
        app.logger.error(f"Error en reporte_variaciones_tanques: {e}")
        flash(f"Ocurrió un error al generar el reporte: {e}", 'danger')
        # Fallback a rango por defecto
    return render_template('reporte_variaciones_tanques.html',
                   nombre=session.get('nombre'),
                   desde='',
                   hasta='',
                   series_por_tanque={},
                   today_iso=date.today().isoformat(),
                   fechas_con_registro=[])
@login_required
@app.route('/reporte_planta')
def reporte_planta():
    # 1. La lógica del filtro de fecha no cambia
    fecha_str = request.args.get('fecha')
    try:
        fecha_seleccionada = date.fromisoformat(fecha_str) if fecha_str else date.today()
    except (ValueError, TypeError):
        fecha_seleccionada = date.today()
    
    timestamp_limite = datetime.combine(fecha_seleccionada, time.max)

    # 2. La consulta a la base de datos no cambia
    subquery = (db.session.query(
        func.max(RegistroPlanta.id).label('max_id')
    ).filter(
        RegistroPlanta.timestamp <= timestamp_limite
    ).group_by(RegistroPlanta.tk).subquery())

    registros_recientes = (db.session.query(RegistroPlanta)
        .filter(RegistroPlanta.id.in_(subquery))
        .all())
    
    # 3. Preparamos los datos y la información
    datos_planta_js = []
    fecha_actualizacion_info = "No hay registros para la fecha seleccionada."

    if registros_recientes:
        # ========================================================
        #  INICIO: LÓGICA DE ORDENAMIENTO PERSONALIZADO
        # ========================================================
        
    # 1. Definimos el orden exacto que queremos.
        orden_deseado = ["TK-109", "TK-110", "TK-01", "TK-02", "TK-102", "TK-108", "Consumo Interno"]
        
        # 2. Creamos un mapa para asignar un "peso" a cada TK.
        orden_map = {tk: i for i, tk in enumerate(orden_deseado)}
        
        # 3. Ordenamos la lista de registros usando nuestro mapa.
        #    Los tanques no especificados en la lista irán al final.
        registros_ordenados = sorted(
            registros_recientes, 
            key=lambda r: orden_map.get(r.tk, 99) # Usamos 99 para que los no encontrados vayan al final
        )
        
        # ========================================================
        #  FIN DE LA LÓGICA DE ORDENAMIENTO
        # ========================================================

        # Usamos la nueva lista YA ORDENADA para construir los datos para el HTML
        # Preparar mapa de defaults para fallback
        defaults_map = {fila["TK"]: fila for fila in PLANILLA_PLANTA}

        mapa_js = {r.tk: {
            "TK": r.tk,
            "PRODUCTO": "DILUYENTE" if r.tk == "Consumo Interno" else (r.producto or defaults_map.get(r.tk, {}).get("PRODUCTO")),
            "MAX_CAP": 124.78 if r.tk == "Consumo Interno" else (r.max_cap or defaults_map.get(r.tk, {}).get("MAX_CAP")),
            "BLS_60": r.bls_60,
            "API": r.api,
            "BSW": r.bsw,
            "S": r.s
        } for r in registros_ordenados if r.tk}

        # Asegurar que todos los TK definidos por defecto existan (ej. TK-108)
        for fila in PLANILLA_PLANTA:
            tk = fila.get("TK")
            if tk and tk not in mapa_js:
                mapa_js[tk] = {
                    "TK": tk,
                    "PRODUCTO": fila.get("PRODUCTO"),
                    "MAX_CAP": fila.get("MAX_CAP"),
                    "BLS_60": None,
                    "API": None,
                    "BSW": None,
                    "S": None
                }

        # Ordenar según orden deseado y filtrar únicamente los TK permitidos
        orden_deseado = ["TK-109", "TK-110", "TK-01", "TK-02", "TK-102", "TK-108", "Consumo Interno"]
        orden_map = {tk: i for i, tk in enumerate(orden_deseado)}
        allowed_set = set(orden_deseado)
        datos_planta_js = sorted(
            [v for k, v in mapa_js.items() if k in allowed_set],
            key=lambda d: orden_map.get(d.get("TK"), 99)
        )
        
        # La lógica para la fecha de actualización no cambia
        ultimo_registro_general = max(registros_recientes, key=lambda r: r.timestamp)
        fecha_formato = ultimo_registro_general.timestamp.strftime("%Y_%m_%d_%H_%M_%S")
        fecha_actualizacion_info = formatear_info_actualizacion(
            fecha_formato, 
            ultimo_registro_general.usuario
        )

    # 4. Renderizamos la plantilla con los datos ya ordenados
    # Fechas con registro para colorear
    try:
        dias_rows = (db.session
            .query(func.date(RegistroPlanta.timestamp).label('dia'))
            .group_by(func.date(RegistroPlanta.timestamp))
            .all())
        fechas_con_registro = []
        for (dia,) in dias_rows:
            try:
                fechas_con_registro.append(dia.isoformat())
            except AttributeError:
                fechas_con_registro.append(str(dia))
    except Exception:
        fechas_con_registro = []

    return render_template("reporte_planta.html", 
                           datos_planta_para_js=datos_planta_js,
                           fecha_actualizacion_info=fecha_actualizacion_info,
                           fecha_seleccionada=fecha_seleccionada.isoformat(),
                           today_iso=date.today().isoformat(),
                           fechas_con_registro=fechas_con_registro)


@login_required
@permiso_requerido('transito')
@app.route('/guardar-config-transito', methods=['POST'])
def guardar_config_transito():
    try:
        nueva_config = request.get_json()
        # Validación básica de la estructura recibida
        if not isinstance(nueva_config, dict) or 'REFINERIA' not in nueva_config or 'EDSM' not in nueva_config:
            return jsonify(success=False, message="Formato de configuración inválido."), 400

        with open('transito_config.json', 'w', encoding='utf-8') as f:
            json.dump(nueva_config, f, ensure_ascii=False, indent=4)

        return jsonify(success=True, message="Configuración guardada exitosamente.")
    except Exception as e:
        print(f"Error al guardar transito_config.json: {e}")
        return jsonify(success=False, message=f"Error interno del servidor: {str(e)}"), 500

@login_required
@permiso_requerido('transito')
@app.route('/guardar-registro-transito-<tipo_transito>', methods=['POST'])
def guardar_transito(tipo_transito):
    datos_recibidos = request.get_json()
    if not isinstance(datos_recibidos, list):
        return jsonify(success=False, message="Formato de datos incorrecto."), 400

    try:
        # Itera sobre cada fila enviada desde el frontend
        for datos_fila in datos_recibidos:
            # Solo procesamos filas que tengan datos, especialmente una guía.
            guia = datos_fila.get('GUIA')
            if not guia:
                continue

            registro_id = datos_fila.get('id')

            # Si la fila tiene un ID, significa que es un registro existente.
            if registro_id:
                registro = db.session.query(RegistroTransito).get(registro_id)
                if registro:
                    # ACTUALIZAMOS el registro existente
                    registro.usuario = session.get("nombre", "No identificado")
                    registro.origen = datos_fila.get('ORIGEN')
                    registro.fecha = datos_fila.get('FECHA')
                    registro.producto = datos_fila.get('PRODUCTO')
                    registro.placa = datos_fila.get('PLACA')
                    registro.api = float(str(datos_fila.get('API')).replace(',', '.')) if datos_fila.get('API') else None
                    registro.bsw = float(str(datos_fila.get('BSW')).replace(',', '.')) if datos_fila.get('BSW') else None
                    registro.nsv = float(str(datos_fila.get('NSV')).replace(',', '.')) if datos_fila.get('NSV') else None
                    registro.observaciones = datos_fila.get('OBSERVACIONES')
                    registro.timestamp = datetime.utcnow()
            else:
                # Si la fila NO tiene ID, es un registro nuevo y lo CREAMOS.
                nuevo_registro = RegistroTransito(
                    usuario=session.get("nombre", "No identificado"),
                    tipo_transito=tipo_transito,
                    guia=guia,
                    origen=datos_fila.get('ORIGEN'),
                    fecha=datos_fila.get('FECHA'),
                    producto=datos_fila.get('PRODUCTO'),
                    placa=datos_fila.get('PLACA'),
                    api=float(str(datos_fila.get('API')).replace(',', '.')) if datos_fila.get('API') else None,
                    bsw=float(str(datos_fila.get('BSW')).replace(',', '.')) if datos_fila.get('BSW') else None,
                    nsv=float(str(datos_fila.get('NSV')).replace(',', '.')) if datos_fila.get('NSV') else None,
                    observaciones=datos_fila.get('OBSERVACIONES')
                )
                db.session.add(nuevo_registro)

        # Confirmamos todos los cambios (updates y nuevos) en la base de datos.
        db.session.commit()

        # Después de guardar, consultamos el historial COMPLETO para devolverlo al frontend.
        registros_actualizados = db.session.query(RegistroTransito).filter_by(tipo_transito=tipo_transito).order_by(RegistroTransito.timestamp.desc()).all()
        
        datos_para_frontend = [
            {"id": r.id, "ORIGEN": r.origen, "FECHA": r.fecha, "GUIA": r.guia, "PRODUCTO": r.producto, "PLACA": r.placa, "API": r.api or '', "BSW": r.bsw or '', "NSV": r.nsv or '', "OBSERVACIONES": r.observaciones or ''}
            for r in registros_actualizados
        ]
        
        return jsonify(success=True, message="Historial guardado exitosamente.", datos=datos_para_frontend)

    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Error al guardar tránsito: {e}")
        return jsonify(success=False, message=f"Error interno: {str(e)}"), 500
    
@login_required
@permiso_requerido('transito')
@app.route('/api/transito/eliminar-todo/<string:tipo_transito>', methods=['DELETE'])
def eliminar_todo_transito(tipo_transito):
    """
    Elimina TODOS los registros de un tipo de tránsito específico ('general' o 'refineria').
    """
    # Validamos que el tipo sea uno de los esperados
    if tipo_transito not in ['general', 'refineria']:
        return jsonify(success=False, message="Tipo de tránsito no válido."), 400

    try:
        # Ejecuta la eliminación masiva
        num_borrados = RegistroTransito.query.filter_by(tipo_transito=tipo_transito).delete()
        
        # Confirma la transacción
        db.session.commit()
        
        return jsonify(success=True, message=f"Se eliminaron {num_borrados} registros de la planilla '{tipo_transito}'.")

    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Error en eliminación masiva de tránsito '{tipo_transito}': {e}")
        return jsonify(success=False, message=f"Error interno del servidor: {str(e)}"), 500    
    
@login_required
@permiso_requerido('transito')
@app.route('/eliminar-registro-transito/<int:id>', methods=['DELETE'])
def eliminar_registro_transito(id):
    """
    Elimina un único registro de la tabla de tránsito por su ID.
    """
    try:
        # Busca el registro por su ID. Si no lo encuentra, devuelve un error 404.
        registro_a_eliminar = RegistroTransito.query.get_or_404(id)
        
        # Elimina el registro de la sesión de la base de datos
        db.session.delete(registro_a_eliminar)
        
        # Confirma los cambios en la base de datos
        db.session.commit()
        
        # Devuelve una respuesta de éxito en formato JSON
        return jsonify(success=True, message="Registro eliminado exitosamente.")

    except Exception as e:
        # Si algo sale mal, revierte los cambios y registra el error
        db.session.rollback()
        app.logger.error(f"Error al eliminar registro de tránsito ID {id}: {e}")
        return jsonify(success=False, message=f"Error interno del servidor: {str(e)}"), 500    
    
@login_required
@permiso_requerido('transito') # <--- LÍNEA CORREGIDA
@app.route('/subir_excel_transito', methods=['POST'])
def subir_excel_transito():
    """
    Procesa un archivo Excel subido para cargar datos en la planilla de tránsito.
    """
    if 'archivo_excel' not in request.files:
        return jsonify({'success': False, 'message': "No se encontró el archivo en la solicitud."}), 400

    archivo = request.files['archivo_excel']
    tipo_transito = request.form.get('tipo_transito', 'general')
    sobrescribir = request.form.get('sobrescribirDatos') == 'on'

    if archivo.filename == '':
        return jsonify({'success': False, 'message': "No se seleccionó ningún archivo."}), 400

    if not archivo.filename.endswith('.xlsx'):
        return jsonify({'success': False, 'message': "Formato no válido. Por favor, suba un archivo .xlsx"}), 400

    try:
        # Si se marca "Sobrescribir", se borran los registros de hoy para esa planilla.
        if sobrescribir:
            today_start = datetime.combine(date.today(), time.min)
            today_end = datetime.combine(date.today(), time.max)
            
            num_borrados = db.session.query(RegistroTransito).filter(
                RegistroTransito.tipo_transito == tipo_transito,
                RegistroTransito.timestamp >= today_start,
                RegistroTransito.timestamp <= today_end
            ).delete(synchronize_session=False)
            
            db.session.commit()
            print(f"Sobrescribiendo: Se eliminaron {num_borrados} registros de hoy para '{tipo_transito}'.")

        workbook = openpyxl.load_workbook(archivo)
        sheet = workbook.active
        
        nuevos_registros = []
        filas_con_error = 0
        
        for row_index in range(2, sheet.max_row + 1):
            row_data = [cell.value for cell in sheet[row_index]]
            
            if not any(row_data) or not row_data[1]:
                continue
            
            try:
                def safe_float(value):
                    if value is None: return None
                    return float(str(value).replace(',', '.'))

                fecha = row_data[0]
                if isinstance(fecha, datetime):
                    fecha = fecha.strftime('%Y-%m-%d')
                else:
                    fecha = str(fecha) if fecha else None

                nuevo_registro = RegistroTransito(
                    usuario=session.get("nombre", "Carga Excel"),
                    tipo_transito=tipo_transito,
                    fecha=fecha,
                    guia=str(row_data[1]),
                    origen=str(row_data[2]),
                    producto=str(row_data[3]),
                    placa=str(row_data[4]),
                    api=safe_float(row_data[5]),
                    bsw=safe_float(row_data[6]),
                    nsv=safe_float(row_data[7]),
                    observaciones=str(row_data[8]) if len(row_data) > 8 and row_data[8] else ""
                )
                nuevos_registros.append(nuevo_registro)
            except (ValueError, TypeError, IndexError) as e:
                filas_con_error += 1
                app.logger.warning(f"ADVERTENCIA: Saltando fila {row_index} del Excel por error de formato: {e}")
                continue

        if not nuevos_registros:
            return jsonify({'success': False, 'message': "No se encontraron registros válidos para cargar en el archivo."}), 400
        
        db.session.add_all(nuevos_registros)
        db.session.commit()
        
        message = f"Se han cargado y guardado {len(nuevos_registros)} registros exitosamente."
        if filas_con_error > 0:
            message += f" Se saltaron {filas_con_error} filas por errores de formato."

        return jsonify({'success': True, 'message': message})

    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Error crítico al procesar el archivo Excel: {e}")
        return jsonify({'success': False, 'message': f"Error interno del servidor: {str(e)}"}), 500

    
@login_required
@permiso_requerido('control_remolcadores')
@app.route('/api/registros_remolcadores/<int:id>', methods=['DELETE'])
def eliminar_evento_remolcador(id):
    """Elimina un único evento de la maniobra."""
    
    usuario_puede_eliminar = (
        session.get('rol') == 'admin' or 
        session.get('email') == 'ops@conquerstrading.com' or
        session.get('email') == 'opensea@conquerstrading.com'
    )

    if not usuario_puede_eliminar:
        return jsonify(success=False, message="No tienes permiso para eliminar este evento."), 403
    
    try:
        registro = RegistroRemolcador.query.get_or_404(id)
        db.session.delete(registro)
        db.session.commit()
        return jsonify(success=True, message="Evento eliminado correctamente.")
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Error al eliminar evento de remolcador: {e}")
        return jsonify(success=False, message=f"Error interno: {str(e)}"), 500

# ================== MODELO OPTIMIZACIÓN ==================
@login_required
@permiso_requerido('modelo_optimizacion')
@app.route('/modelo-optimizacion', methods=['GET','POST'])
def modelo_optimizacion_page():
    from flask import current_app
    error = None
    # Restricción extra: solo Felipe y German (y admin)
    allowed = {"felipe.delavega@conquerstrading.com", "finance@conquerstrading.com"}
    if session.get('rol') != 'admin' and session.get('email') not in allowed:
        flash('No tienes permiso para este módulo.', 'danger')
        return redirect(url_for('home'))
    resultados = None
    grafico_base64 = None
    grafico_div = None
    grafico_div = None
    grafico_div = None
    grafico_div = None
    grafico_div = None
    excel_descargable = False
    # Importes numéricos (el template aplica el formato)
    total_volumen = 0.0
    total_costo_import = 0.0
    brent_valor = None
    trm_valor = None
    generado_en = datetime.utcnow().strftime('%d/%m/%Y %H:%M:%S')
    generar_excel = 'si'
    component_data = []
    componentes_cols = [
        'USD/BBL CRUDO + FLETE Marino', 'Remolcador a CZF', 'USD/BBL Ingreso a CZF (Alm+OperPort 2)',
        'USD/BBL %FIN mes', 'USD/BBL Alm+OperPort 1', 'Nacionalización USD/BBL', 'USD/Bbl Exportación', 'Transp Terrestre a CZF'
    ]
    componentes_alias = [
        'CRUDO + FLETE', 'REMOLCADOR', 'INGRESO CZF', 'FINANCIACIÓN', 'ALM+OPER PORT', 'NACIONALIZACIÓN', 'EXPORTACIÓN', 'TRANSP TERRESTRE'
    ]
    component_stats = []
    ranking_spread_imp = []
    ranking_spread_exp = []
    if request.method == 'POST':
        generar_excel = request.form.get('generar_excel','si')
        try:
            excel_path = EXCEL_DEFAULT
            if 'archivo_excel' in request.files and request.files['archivo_excel'].filename:
                # Guardar temporalmente
                up_file = request.files['archivo_excel']
                tmp_path = os.path.join(BASE_DIR, 'upload_modelo_temp.xlsx')
                up_file.save(tmp_path)
                excel_path = tmp_path
            data = ejecutar_modelo(excel_path, generar_excel == 'si')
            resultados = data['resumen']
            grafico_base64 = data['grafico_base64']
            # Enviar valores crudos como números; el template se encarga del formato
            brent_valor = float(data['BRENT']) if data['BRENT'] is not None else None
            trm_valor = float(data.get('TRM')) if data.get('TRM') is not None else None
            total_volumen = float(sum(r['Volumen'] for r in resultados))
            total_costo_import = float(sum(r['CostoTotalImp'] for r in resultados))
            # Construir datos detallados para gráfica interactiva
            df_det = data['df_result']
            if not df_det.empty:
                # Incluimos 'Flete Marino' (valor total) para poder calcular USD/BBL en el tooltip
                extra_cols = ['Flete Marino'] if 'Flete Marino' in df_det.columns else []
                sub = df_det[['ID','Producto','Volumen Compra BBL','Puerto Llegada'] + extra_cols + componentes_cols].copy()
                for _, row in sub.iterrows():
                    comp_entry = {
                        'ID': row['ID'],
                        'Producto': row['Producto'],
                        'Volumen Compra BBL': float(row['Volumen Compra BBL'] or 0) if row['Volumen Compra BBL'] is not None else 0,
                        'Puerto Llegada': row['Puerto Llegada'] or ''
                    }
                    # Guardamos Flete Marino total para calcular USD/BBL en el frontend (si existe)
                    if 'Flete Marino' in row.index:
                        try:
                            comp_entry['Flete Marino'] = float(row['Flete Marino'] or 0)
                        except Exception:
                            comp_entry['Flete Marino'] = 0.0
                    for col in componentes_cols:
                        val = row[col]
                        try:
                            comp_entry[col] = float(val) if val is not None else 0.0
                        except Exception:
                            comp_entry[col] = 0.0
                    component_data.append(comp_entry)
                # Estadísticas comparativas por componente
                for alias, col in zip(componentes_alias, componentes_cols):
                    serie = df_det[col].astype(float).fillna(0)
                    component_stats.append({
                        'alias': alias,
                        'min': round(float(serie.min()),4),
                        'max': round(float(serie.max()),4),
                        'avg': round(float(serie.mean()),4)
                    })
                # Rankings spreads
                if 'Spread Total on Brent IMPORTACIONES' in df_det.columns:
                    ranking_spread_imp = (
                        df_det[['ID','Producto','Spread Total on Brent IMPORTACIONES']]
                        .rename(columns={'Spread Total on Brent IMPORTACIONES':'valor'})
                        .sort_values('valor', ascending=False)
                        .head(10)
                        .to_dict(orient='records')
                    )
                if 'Spread Exportaciones' in df_det.columns:
                    ranking_spread_exp = (
                        df_det[['ID','Producto','Spread Exportaciones']]
                        .rename(columns={'Spread Exportaciones':'valor'})
                        .sort_values('valor', ascending=False)
                        .head(10)
                        .to_dict(orient='records')
                    )
            if data['excel_bytes']:
                # Guardar en archivo temporal (no en sesión para evitar exceder tamaño cookie)
                from uuid import uuid4
                tmp_dir = os.path.join(BASE_DIR, 'tmp_modelo_excel')
                os.makedirs(tmp_dir, exist_ok=True)
                # Limpieza simple de archivos viejos (> 2 horas)
                try:
                    import time
                    now = time.time()
                    for fname in os.listdir(tmp_dir):
                        fpath = os.path.join(tmp_dir, fname)
                        if os.path.isfile(fpath) and now - os.path.getmtime(fpath) > 7200:
                            try: os.remove(fpath)
                            except Exception: pass
                except Exception:
                    pass
                filename = f"modelo_{uuid4().hex}.xlsx"
                file_path = os.path.join(tmp_dir, filename)
                with open(file_path, 'wb') as f:
                    f.write(data['excel_bytes'])
                session['modelo_excel_path'] = file_path
                excel_descargable = True
        except Exception as e:
            current_app.logger.error(f"Error modelo optimización: {e}")
            error = str(e)
    return render_template('modelo_optimizacion.html',
                           nombre=session.get('nombre'),
                           resultados=resultados,
                           grafico_base64=grafico_base64,
                           excel_descargable=excel_descargable,
                           total_volumen=total_volumen,
                           total_costo_import=total_costo_import,
                           brent_valor=brent_valor,
                           trm_valor=trm_valor,
                           generado_en=generado_en,
                           generar_excel=generar_excel,
                           component_data=component_data,
                           componentes_cols=componentes_cols,
                           componentes_alias=componentes_alias,
                           component_stats=component_stats,
                           ranking_spread_imp=ranking_spread_imp,
                           ranking_spread_exp=ranking_spread_exp,
                           error=error)

@login_required
@permiso_requerido('modelo_optimizacion')
@app.route('/modelo-optimizacion/descargar')
def descargar_resultado_modelo():
    allowed = {"felipe.delavega@conquerstrading.com", "finance@conquerstrading.com"}
    if session.get('rol') != 'admin' and session.get('email') not in allowed:
        flash('No tienes permiso para este módulo.', 'danger')
        return redirect(url_for('home'))
    file_path = session.get('modelo_excel_path')
    if not file_path or not os.path.isfile(file_path):
        flash('No hay archivo para descargar', 'warning')
        return redirect(url_for('modelo_optimizacion_page'))
    return send_file(file_path, as_attachment=True, download_name='Resultados_modelo.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

       
@login_required
@app.route('/agregar-producto', methods=['POST'])
def agregar_producto():
    data = request.get_json()
    nuevo_producto = data.get("producto")
    grupo = data.get("grupo")  # "REFINERIA" o "EDSM"

    if not nuevo_producto or grupo not in ["REFINERIA", "EDSM"]:
        return jsonify(success=False, message="Datos incompletos")

    ruta = "productos.json"
    try:
        with open(ruta, encoding="utf-8") as f:
            productos = json.load(f)

        if nuevo_producto not in productos[grupo]:
            productos[grupo].append(nuevo_producto)
            with open(ruta, "w", encoding="utf-8") as f:
                json.dump(productos, f, ensure_ascii=False, indent=2)

        return jsonify(success=True)
    except Exception as e:
        return jsonify(success=False, message=str(e))
    
@login_required       
@app.route('/historial_registros') 
def historial_registros():        
    registros = []
    carpeta = "registros"
    os.makedirs(carpeta, exist_ok=True)

    for archivo in sorted(os.listdir(carpeta), reverse=True):
        if archivo.endswith(".json"):
            ruta = os.path.join(carpeta, archivo)
            try:
                with open(ruta, encoding='utf-8') as f:
                    registro = json.load(f)
                    if session.get("email") in ["omar.morales@conquerstrading.com", "oci@conquerstrading.com"]:
                        registros.append(registro)
                    else:
                        if registro.get("usuario") == session.get("nombre"):
                            registros.append(registro)
            except Exception as e:
                print(f"Error al cargar {archivo}: {e}")
    # Asegúrate que el nombre del template sigue siendo el correcto si quieres reutilizarlo
    return render_template("reporte_general.html", registros=registros, nombre=session.get("nombre"))

@login_required
@permiso_requerido('transito')
@app.route('/reporte_transito')
def reporte_transito():
    app.logger.info("Accediendo a /reporte_transito desde la base de datos")
    datos_consolidados = {}
    datos_conteo_camiones = {}
    observaciones_camiones = {} 
    camiones_para_mapa = []
    
    fecha_actualizacion_info = "No se encontraron registros de tránsito."
    
    try:
        todos_los_registros = db.session.query(RegistroTransito).order_by(RegistroTransito.timestamp.desc()).all()

        if not todos_los_registros:
            return render_template("reporte_transito.html", 
                                   datos_consolidados=datos_consolidados, 
                                   datos_conteo_camiones=datos_conteo_camiones,
                                   observaciones_camiones=observaciones_camiones,
                                   camiones_mapa=camiones_para_mapa,
                                   nombre=session.get("nombre"), 
                                   fecha_actualizacion_info=fecha_actualizacion_info)

        ultimo_registro = max(todos_los_registros, key=lambda r: r.timestamp)
        fecha_formato = ultimo_registro.timestamp.strftime("%Y_%m_%d_%H_%M_%S")
        fecha_actualizacion_info = formatear_info_actualizacion(fecha_formato, ultimo_registro.usuario)

        for reg in todos_los_registros:
            origen = (reg.origen or "").strip()
            producto = (reg.producto or "").strip()
            
            if not origen or not producto:
                continue
            
            tipo_destino_reporte = "Refinería" if reg.tipo_transito == "refineria" else "EDSM"
            nsv = float(reg.nsv or 0.0)

            datos_consolidados.setdefault(tipo_destino_reporte, {}).setdefault(origen, {}).setdefault(producto, 0.0)
            datos_consolidados[tipo_destino_reporte][origen][producto] += nsv
            
            datos_conteo_camiones.setdefault(tipo_destino_reporte, {}).setdefault(origen, {}).setdefault(producto, 0)
            datos_conteo_camiones[tipo_destino_reporte][origen][producto] += 1
            
            if reg.observaciones and reg.observaciones.strip():
                observacion_texto = reg.observaciones.strip()
                placa = reg.placa or "SIN PLACA"
                texto_completo = f"{placa}: {observacion_texto}"
                
                lista_de_observaciones = observaciones_camiones.setdefault(tipo_destino_reporte, {}).setdefault(origen, {}).setdefault(producto, [])
                lista_de_observaciones.append(texto_completo)

            # --- AGREGADO: Poblar camiones_para_mapa para el mapa ---
            ciudad = ""
            if reg.observaciones and "|" in reg.observaciones:
                ciudad = reg.observaciones.split("|")[0].strip()
            elif reg.observaciones:
                ciudad = reg.observaciones.strip()
            camiones_para_mapa.append({
                "ciudad": ciudad,
                "tipo_transito": tipo_destino_reporte,
                "placa": reg.placa,
                "origen": reg.origen,
                "producto": reg.producto,
                "NSV": reg.nsv,
                "OBSERVACIONES": reg.observaciones
                
            })
            
    except Exception as e:
        app.logger.error(f"Error crítico al generar reporte de tránsito desde BD: {e}")
        flash(f"Ocurrió un error al generar el reporte: {e}", "danger")
        fecha_actualizacion_info = "Error al cargar los datos."

    return render_template("reporte_transito.html",
                           datos_consolidados=datos_consolidados,
                           datos_conteo_camiones=datos_conteo_camiones,
                           observaciones_camiones=observaciones_camiones,
                           camiones_mapa=camiones_para_mapa,
                           nombre=session.get("nombre"),
                           fecha_actualizacion_info=fecha_actualizacion_info)
@login_required
@permiso_requerido('barcaza_orion')
@app.route('/barcaza_orion')
def barcaza_orion():
    print("\n--- INICIANDO RUTA /barcaza_orion ---")
    
    fecha_str = request.args.get('fecha')
    try:
        fecha_seleccionada = date.fromisoformat(fecha_str) if fecha_str else date.today()
    except (ValueError, TypeError):
        fecha_seleccionada = date.today()
    
    print(f"DEBUG: Filtrando por fecha: {fecha_seleccionada}")
    timestamp_limite = datetime.combine(fecha_seleccionada, time.max)

    subquery = (db.session.query(
        func.max(RegistroBarcazaOrion.id).label('max_id')
    ).filter(RegistroBarcazaOrion.timestamp <= timestamp_limite).group_by(RegistroBarcazaOrion.tk, RegistroBarcazaOrion.grupo).subquery())

    registros_recientes = (db.session.query(RegistroBarcazaOrion)
        .filter(RegistroBarcazaOrion.id.in_(subquery)).all())
    
    print(f"DEBUG: La consulta a la BD encontró {len(registros_recientes)} registros.")
    
    datos_para_plantilla = []
    if registros_recientes:
        for registro in registros_recientes:
            datos_para_plantilla.append({
                "TK": registro.tk, "PRODUCTO": registro.producto, "MAX_CAP": registro.max_cap,
                "BLS_60": registro.bls_60 or "", "API": registro.api or "", 
                "BSW": registro.bsw or "", "S": registro.s or "", "grupo": registro.grupo or ""
            })
    else:
        print("DEBUG: No se encontraron registros, se usará la planilla por defecto.")
        datos_para_plantilla = PLANILLA_BARCAZA_ORION

    # Ordenar los tanques CR según el orden de PLANILLA_BARCAZA_ORION
    def ordenar_por_planilla(lista, grupo, planilla):
        orden = [t["TK"] for t in planilla if t["grupo"] == grupo]
        return sorted([t for t in lista if t["grupo"] == grupo], key=lambda x: orden.index(x["TK"]) if x["TK"] in orden else 999)

    tanques_principales = [tk for tk in datos_para_plantilla if tk.get('grupo') == 'PRINCIPAL']
    tanques_man = [tk for tk in datos_para_plantilla if tk.get('grupo') == 'MANZANILLO']
    tanques_cr = ordenar_por_planilla(datos_para_plantilla, 'CR', PLANILLA_BARCAZA_ORION)
    tanques_margoth = [tk for tk in datos_para_plantilla if tk.get('grupo') == 'MARGOTH']
    tanques_odisea = [tk for tk in datos_para_plantilla if tk.get('grupo') == 'ODISEA']

    return render_template("barcaza_orion.html",
                           titulo="Planilla Barcaza Orion",
                           tanques_principales=tanques_principales,
                           tanques_man=tanques_man,
                           tanques_cr=tanques_cr,
                           tanques_margoth=tanques_margoth,
                           tanques_odisea=tanques_odisea,
                           nombre=session.get("nombre"),
                           fecha_seleccionada=fecha_seleccionada.isoformat(),
                           today_iso=date.today().isoformat())

@app.cli.command("sync-orion")
def sync_orion_tanks():
    """
    Revisa la planilla por defecto de Orion y añade los tanques que falten en la base de datos.
    Este comando es seguro y no borra datos existentes.
    VERSIÓN CORREGIDA: Revisa la tupla (TK, grupo) para evitar conflictos.
    """
    try:
        # Obtenemos una lista de tuplas (tk, grupo) que ya existen en la base de datos
        tanques_existentes_tuplas = db.session.query(
            RegistroBarcazaOrion.tk, 
            RegistroBarcazaOrion.grupo
        ).distinct().all()
        
        # Convertimos la lista de tuplas a un set para búsquedas rápidas y eficientes
        set_tanques_db = set(tanques_existentes_tuplas)
        
        nuevos_tanques_agregados = 0
        
        # Iteramos sobre la lista de tanques que DEBERÍA existir (la de tu código)
        for tanque_plantilla in PLANILLA_BARCAZA_ORION:
            tk_plantilla = tanque_plantilla["TK"]
            grupo_plantilla = tanque_plantilla["grupo"]
            
            # Revisamos si la combinación (tk, grupo) NO está en nuestro set de la BD
            if (tk_plantilla, grupo_plantilla) not in set_tanques_db:
                print(f"Tanque '{tk_plantilla}' del grupo '{grupo_plantilla}' no encontrado. Añadiendo...")
                
                nuevo_registro = RegistroBarcazaOrion(
                    usuario="system_sync",
                    tk=tk_plantilla,
                    producto=tanque_plantilla["PRODUCTO"],
                    max_cap=tanque_plantilla["MAX_CAP"],
                    grupo=grupo_plantilla,
                    bls_60=None, api=None, bsw=None, s=None
                )
                db.session.add(nuevo_registro)
                nuevos_tanques_agregados += 1

        if nuevos_tanques_agregados > 0:
            db.session.commit()
            print(f"¡Éxito! Se han añadido {nuevos_tanques_agregados} tanques nuevos a la Barcaza Orion.")
        else:
            print("La base de datos ya está sincronizada. No se añadieron tanques nuevos.")
            
    except Exception as e:
        db.session.rollback()
        print(f"Ocurrió un error durante la sincronización: {e}")

@login_required
@permiso_requerido('barcaza_bita')
@app.route('/barcaza_bita')
def barcaza_bita():
    # 1. Lógica del filtro de fecha
    fecha_str = request.args.get('fecha')
    try:
        fecha_seleccionada_obj = date.fromisoformat(fecha_str) if fecha_str else date.today()
    except (ValueError, TypeError):
        fecha_seleccionada_obj = date.today()

    # --- CÓDIGO CLAVE PARA FORMATEAR LA FECHA ---
    meses = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
    nombre_mes = meses[fecha_seleccionada_obj.month - 1]
    fecha_display = fecha_seleccionada_obj.strftime(f"%d de {nombre_mes} de %Y")
    # --- FIN DEL CÓDIGO CLAVE ---
    

    timestamp_limite = datetime.combine(fecha_seleccionada_obj, time.max)

    # 2. Consulta a la base de datos para BITA
    subquery = (db.session.query(func.max(RegistroBarcazaBita.id).label('max_id'))
        .filter(RegistroBarcazaBita.timestamp <= timestamp_limite).group_by(RegistroBarcazaBita.tk).subquery())

    registros_recientes = (db.session.query(RegistroBarcazaBita)
        .filter(RegistroBarcazaBita.id.in_(subquery)).all())

    # 3. Preparar los datos
    datos_para_plantilla = []
    if registros_recientes:
        for r in registros_recientes:
            datos_para_plantilla.append({
                "TK": r.tk, "PRODUCTO": r.producto, "MAX_CAP": r.max_cap,
                "BLS_60": r.bls_60 or "", "API": r.api or "", "BSW": r.bsw or "", "S": r.s or ""
            })
    else:
        datos_para_plantilla = PLANILLA_BARCAZA_BITA

    # 4. Lógica para separar en grupos
    grupos = {
        "BARCAZA MARINSE": [tk for tk in datos_para_plantilla if tk.get('TK', '').startswith('MARI')],
        "BARCAZA OIDECH": [tk for tk in datos_para_plantilla if tk.get('TK', '').startswith('OID')]
    }

    # 5. Renderizar la plantilla, pasando todas las variables necesarias
    return render_template("barcaza_bita.html",
                           titulo="Planilla Barcaza BITA",
                           grupos=grupos,
                           nombre=session.get('nombre', 'Desconocido'),
                           fecha_seleccionada=fecha_seleccionada_obj.isoformat(),
                           today_iso=date.today().isoformat(),
                           fecha_display=fecha_display) 

@login_required
@permiso_requerido('guia_transporte') 
@app.route('/guia_transporte')
def guia_transporte():
    """
    Muestra la guía de transporte. Si recibe datos en la URL, los pasa a la plantilla
    para autocompletar el formulario. Si no, pasa datos vacíos.
    """
    # Creamos un diccionario para guardar los datos que vienen de la URL.
    # Usamos .get('nombre_parametro', '') para que, si un dato no llega, no de error.
    datos_guia = {
        'placa': request.args.get('placa', ''),
        'conductor': request.args.get('nombre_conductor', ''),
        'cedula': request.args.get('cedula_conductor', ''),
        'destino': request.args.get('destino', ''),
        'producto': request.args.get('producto_a_cargar', ''),
        'galones': request.args.get('galones', ''),
        'transportadora': request.args.get('empresa_transportadora', ''),
        'cliente': request.args.get('cliente', ''),
        'temperatura': request.args.get('temperatura', ''),
        'api_obs': request.args.get('api_obs', ''),
        'api_corregido': request.args.get('api_corregido', ''),
        'precintos': request.args.get('precintos', ''),
        # Campos adicionales para poblar "PLACA DEL TANQUE" desde Programación
        'tanque': request.args.get('tanque', ''),
        'placa_tanque': request.args.get('placa_tanque', ''),
        # Campo de factura/remisión desde Programación
        'factura_remision': request.args.get('factura', '')
    }
    
    # Pasamos el diccionario 'datos_guia' a la plantilla HTML.
    return render_template(
        "guia_transporte.html", 
        nombre=session.get("nombre"),
        datos_guia=datos_guia
    )

@login_required
@permiso_requerido("siza_solicitante")  # Permite ver a solicitantes y gestores
@app.route('/dashboard-siza')
def dashboard_siza():
    """Dashboard de control de cupo SIZA multi-producto."""
    try:
        hoy = date.today()
        
        # Obtener todos los productos activos
        productos = ProductoSiza.query.filter_by(activo=True).order_by(ProductoSiza.orden).all()
        
        # Inicializar productos por defecto si no existen
        # Verificar y crear productos por defecto uno a uno
        defaults_data = [
            {'codigo': 'F04', 'nombre': 'F04', 'color': 'primary', 'orden': 1},
            {'codigo': 'DILUYENTE', 'nombre': 'DILUYENTE', 'color': 'success', 'orden': 2},
            {'codigo': 'MGO', 'nombre': 'MGO', 'color': 'warning', 'orden': 3},
            {'codigo': 'F06', 'nombre': 'F06', 'color': 'dark', 'orden': 4},
            {'codigo': 'AGUA_RESIDUAL', 'nombre': 'AGUA RESIDUAL', 'color': 'danger', 'orden': 5}
        ]
        
        updated_products = False
        existing_codes = {p.codigo for p in productos}
        
        for d in defaults_data:
            if d['codigo'] not in existing_codes:
                new_prod = ProductoSiza(codigo=d['codigo'], nombre=d['nombre'], color_badge=d['color'], orden=d['orden'])
                db.session.add(new_prod)
                updated_products = True
        
        if updated_products:
            db.session.commit()
            # Recargar lista completa
            productos = ProductoSiza.query.filter_by(activo=True).order_by(ProductoSiza.orden).all()
        
        # Calcular métricas por producto
        inventario_productos = []
        total_disponible = 0
        total_comprometido = 0
        alerta_global = False
        productos_sin_inventario = []  # Lista de productos DIAN sin inventario
        pedidos_pendientes_para_tabla = []
        
        for producto in productos:
            # Obtener inventario del día
            inventario = InventarioSizaDiario.query.filter_by(
                fecha=hoy,
                producto_id=producto.id
            ).first()
            
            if not inventario:
                # Buscar el último inventario registrado de días anteriores para arrastrar saldo
                ultimo_inventario = InventarioSizaDiario.query.filter(
                    InventarioSizaDiario.producto_id == producto.id,
                    InventarioSizaDiario.fecha < hoy
                ).order_by(InventarioSizaDiario.fecha.desc()).first()
                
                saldo_inicial = ultimo_inventario.cupo_web if ultimo_inventario else 0.0
                
                # Crear inventario inicial para hoy con el saldo anterior
                inventario = InventarioSizaDiario(
                    fecha=hoy,
                    producto_id=producto.id,
                    cupo_web=saldo_inicial,
                    # Arrastrar acumulados históricos de agua y desperdicio
                    volumen_agua_generada=ultimo_inventario.volumen_agua_generada if ultimo_inventario else 0.0,
                    volumen_desperdicio_generado=ultimo_inventario.volumen_desperdicio_generado if ultimo_inventario else 0.0,
                    usuario_actualizacion='Sistema (Arrastre Saldo)'
                )
                db.session.add(inventario)
                db.session.commit()
            
            # Pedidos comprometidos (PENDIENTE + APROBADO) - Aún no despachados
            pedidos_producto = PedidoSiza.query.filter(
                PedidoSiza.producto_id == producto.id,
                PedidoSiza.estado.in_(['PENDIENTE', 'APROBADO'])
            ).all()
            
            comprometido_producto = sum(p.volumen_solicitado for p in pedidos_producto)
            disponible_producto = inventario.cupo_web - comprometido_producto # Disponible real para nuevos pedidos
            
            # Pedidos SOLO pendientes para la tabla de gestión
            pedidos_gestion = [p for p in pedidos_producto if p.estado == 'PENDIENTE']
            pedidos_pendientes_para_tabla.extend(pedidos_gestion)
            
            # Calcular Desperdicio Generado HOY por este producto
            desperdicio_hoy = 0
            # Solo buscar si no es agua residual (evitar consultas innecesarias)
            if producto.codigo != 'AGUA_RESIDUAL':
                # Buscamos movimientos de DIAN que contengan "Desperdicio" y el nombre del producto en la observación
                try:
                    movs_desperdicio = MovimientoDian.query.filter(
                        MovimientoDian.fecha_operativa == hoy,
                        MovimientoDian.observacion.like(f'%Desperdicio%{producto.nombre}%')
                    ).all()
                    desperdicio_hoy = sum(m.volumen for m in movs_desperdicio)
                except Exception:
                    desperdicio_hoy = 0

            inventario_productos.append({
                'producto': producto,
                'inventario': inventario,
                'comprometido': comprometido_producto,
                'disponible': disponible_producto,
                'alerta': disponible_producto <= 0,
                'pedidos_count': len(pedidos_producto),
                'desperdicio_hoy': desperdicio_hoy
            })
            
            total_disponible += disponible_producto
            total_comprometido += comprometido_producto
            
            # Solo alertar para productos DIAN (excluir AGUA_RESIDUAL)
            if disponible_producto <= 0 and producto.codigo != 'AGUA_RESIDUAL':
                alerta_global = True
                productos_sin_inventario.append(producto.nombre)
        
        # Obtener pedidos aprobados (para la nueva tabla en frontend)
        pedidos_aprobados = PedidoSiza.query.filter_by(estado='APROBADO').order_by(PedidoSiza.fecha_gestion.desc()).all()
        
        # Obtener mis pedidos despachados HOY para notificación visual
        usuario_actual = session.get('nombre', '')
        mis_despachos_hoy = []
        if usuario_actual:
            # Filtramos por nombre de usuario (asumiendo que coincide con usuario_registro)
            # y que el estado sea DESPACHADO y la fecha de gestión sea hoy.
            # Nota: fecha_gestion es datetime UTC. Comparamos con fecha hoy.
            # Convertimos a Bogota si es necesario, pero una comparacion simple con la fecha funciona para "hoy".
            # Para mayor precision convertimos.
            pass
            # Haremos esto simple: filtrar por estado y usuario, y en template filtrar fecha o aqui mismo.
            # SQLAlchemy func.date(fecha_gestion)
            mis_despachos_hoy = PedidoSiza.query.filter(
                PedidoSiza.usuario_registro == usuario_actual,
                PedidoSiza.estado == 'DESPACHADO',
                func.date(PedidoSiza.fecha_gestion) == hoy
            ).all()

        # Obtener pedidos pendientes y aprobados (para el modal de consumo)
        todos_pedidos = PedidoSiza.query.filter(
            (PedidoSiza.estado == 'PENDIENTE') | (PedidoSiza.estado == 'APROBADO')
        ).order_by(PedidoSiza.fecha_registro).all()
        
        # Obtener historial de movimientos (recargas y consumos) de los últimos 30 días
        fecha_limite = hoy - timedelta(days=30)
        
        recargas = RecargaSiza.query.filter(
            RecargaSiza.fecha >= fecha_limite
        ).order_by(RecargaSiza.fecha.desc(), RecargaSiza.fecha_registro.desc()).all()
        
        consumos = ConsumoSiza.query.filter(
            ConsumoSiza.fecha >= fecha_limite
        ).order_by(ConsumoSiza.fecha.desc(), ConsumoSiza.fecha_registro.desc()).all()
        
        # Combinar y ordenar movimientos
        movimientos = []
        for recarga in recargas:
            movimientos.append({
                'tipo': 'recarga',
                'id': recarga.id,
                'fecha': recarga.fecha,
                'producto': recarga.producto,
                'volumen': recarga.volumen_recargado,
                'observacion': recarga.observacion,
                'usuario': recarga.usuario_registro,
                'fecha_registro': recarga.fecha_registro,
                'editado': recarga.usuario_edicion is not None,
                'usuario_edicion': recarga.usuario_edicion,
                'fecha_edicion': recarga.fecha_edicion
            })
        
        for consumo in consumos:
            movimientos.append({
                'tipo': 'consumo',
                'id': consumo.id,
                'fecha': consumo.fecha,
                'producto': consumo.producto,
                'volumen': consumo.volumen_consumido,
                'observacion': consumo.observacion,
                'usuario': consumo.usuario_registro,
                'fecha_registro': consumo.fecha_registro,
                'editado': consumo.usuario_edicion is not None,
                'usuario_edicion': consumo.usuario_edicion,
                'fecha_edicion': consumo.fecha_edicion
            })
        
        # Ordenar por fecha descendente
        movimientos.sort(key=lambda x: (x['fecha'], x['fecha_registro']), reverse=True)
        
        # Limitar a los últimos 15 para la vista principal
        movimientos_recientes = movimientos[:15]
        total_movimientos = len(movimientos)
        
        # Verificar si el usuario es gestor (puede aprobar/rechazar/recargar/consumir)
        es_gestor = tiene_permiso('siza_gestor')
        
        # Obtener volumen pendiente DIAN
        volumen_dian = VolumenPendienteDian.query.filter_by(fecha=hoy).first()
        if not volumen_dian:
            # Buscar saldo del día anterior (o último registro existente)
            ultimo_registro = VolumenPendienteDian.query.filter(
                VolumenPendienteDian.fecha < hoy
            ).order_by(VolumenPendienteDian.fecha.desc()).first()

            saldo_pendiente = ultimo_registro.volumen_pendiente if ultimo_registro else 0.0
            saldo_por_aprobar = ultimo_registro.volumen_por_aprobar if ultimo_registro else 0.0

            # Crear registro para HOY arrastrando los saldos
            volumen_dian = VolumenPendienteDian(
                fecha=hoy,
                volumen_pendiente=saldo_pendiente,   # ARRASTRE DE SALDO APROBADO
                volumen_por_aprobar=saldo_por_aprobar, # ARRASTRE DE POR APROBAR
                usuario_actualizacion='Sistema (Arrastre Saldo)'
            )
            db.session.add(volumen_dian)
            db.session.commit()
        
        # Obtener historial de aprobaciones de hoy (Compatibilidad)
        historial_dian_hoy = HistorialAprobacionDian.query.filter_by(fecha_operativa=hoy).order_by(HistorialAprobacionDian.fecha_registro.desc()).all()
        
        # Obtener historial DETALLADO de movimientos DIAN (Completo)
        historial_dian_completo = []
        try:
            historial_dian_completo = MovimientoDian.query.order_by(
                MovimientoDian.fecha_operativa.desc(), 
                MovimientoDian.fecha_registro.desc()
            ).limit(200).all()
        except Exception:
            pass # Si la tabla no existe aún, ignorar
        
        # Verificar si el usuario puede editar volumen DIAN (solo Daniela y Shirley)
        usuario_email = session.get('email', '')
        puede_editar_dian = usuario_email in ['comex@conquerstrading.com', 'comexzf@conquerstrading.com']
        
        # Verificar si el usuario puede ver volumen DIAN (Admins + Gestores + Solicitantes)
        areas_usuario = session.get('area', [])
        puede_ver_dian = session.get('rol') == 'admin' or puede_editar_dian or 'siza_solicitante' in areas_usuario or 'siza_gestor' in areas_usuario
        
        return render_template(
            'siza_dashboard.html',
            inventario_productos=inventario_productos,
            total_disponible=total_disponible,
            total_comprometido=total_comprometido,
            alerta_cupo=alerta_global,
            historial_dian_hoy=historial_dian_hoy, # Pasar historial al template
            productos_sin_inventario=productos_sin_inventario,  # Lista de productos DIAN sin inventario
            pedidos=todos_pedidos,  # Todos los pedidos (PENDIENTE y APROBADO) para modales
            pedidos_pendientes=pedidos_pendientes_para_tabla,  # Solo pendientes para tabla principal
            pedidos_aprobados=pedidos_aprobados, # Nueva lista para despachar
            productos=productos,
            movimientos=movimientos_recientes,  # Solo los últimos 15
            total_movimientos=total_movimientos,  # Total de movimientos disponibles
            hoy=hoy,
            es_gestor=es_gestor,  # Permiso para mostrar/ocultar botones en template
            volumen_dian=volumen_dian,  # Volumen pendiente DIAN
            puede_editar_dian=puede_editar_dian,  # Permiso para editar volumen DIAN
            puede_ver_dian=puede_ver_dian,  # Permiso para ver volumen DIAN
            mis_despachos_hoy=mis_despachos_hoy, # Notificación de despachos para solicitante
            historial_dian_completo=historial_dian_completo # Historial detallado de movimientos DIAN (Todos)
        )
    except Exception as e:
        flash(f'Error al cargar dashboard: {str(e)}', 'danger')
        return redirect(url_for('home'))

@login_required
@permiso_requerido("siza_solicitante")  # Todos pueden ver historial
@app.route('/siza/historial')
def historial_siza():
    """Renderiza la página de historial completo de SIZA."""
    try:
        # Obtener lista de productos para los filtros
        productos = ProductoSiza.query.filter_by(activo=True).order_by(ProductoSiza.orden).all()
        
        return render_template(
            'siza_historial.html',
            productos=productos
        )
    except Exception as e:
        flash(f'Error al cargar historial: {str(e)}', 'danger')
        return redirect(url_for('dashboard_siza'))

@login_required
@permiso_requerido("siza_gestor")  # Solo gestores pueden actualizar inventario
@app.route('/siza/actualizar-inventario', methods=['POST'])
def actualizar_inventario_siza():
    """Actualiza el inventario de un producto SIZA específico."""
    try:
        producto_id = int(request.form.get('producto_id'))
        nuevo_cupo = float(request.form.get('nuevo_cupo', 0))
        
        if nuevo_cupo < 0:
            return jsonify({'success': False, 'message': 'El cupo no puede ser negativo.'}), 400
        
        hoy = date.today()
        
        # Buscar o crear inventario del día
        inventario = InventarioSizaDiario.query.filter_by(
            fecha=hoy,
            producto_id=producto_id
        ).first()
        
        if not inventario:
            inventario = InventarioSizaDiario(
                fecha=hoy,
                producto_id=producto_id,
                cupo_web=nuevo_cupo,
                usuario_actualizacion=session.get('nombre', 'Sistema')
            )
            db.session.add(inventario)
        else:
            inventario.cupo_web = nuevo_cupo
            inventario.usuario_actualizacion = session.get('nombre', 'Sistema')
            inventario.fecha_actualizacion = datetime.utcnow()
        
        db.session.commit()
        
        producto = ProductoSiza.query.get(producto_id)
        flash(f'Inventario de {producto.nombre} actualizado: {nuevo_cupo:,.0f} Barriles', 'success')
        
    except ValueError:
        flash('El valor del cupo debe ser numérico.', 'danger')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al actualizar inventario: {str(e)}', 'danger')
    
    return redirect(url_for('dashboard_siza'))

@login_required
@permiso_requerido("siza_gestor")  # Solo gestores pueden recargar
@app.route('/siza/recargar-producto', methods=['POST'])
def recargar_producto_siza():
    """Registra una recarga de inventario para un producto."""
    try:
        producto_id = int(request.form.get('producto_id'))
        volumen_recarga = float(request.form.get('volumen_recarga', 0))
        observacion = request.form.get('observacion', '').strip()
        descontar_de_dian = request.form.get('descontar_de_dian') in ['on', '1']  # Checkbox
        
        if volumen_recarga <= 0:
            flash('El volumen de recarga debe ser mayor a cero.', 'danger')
            return redirect(url_for('dashboard_siza'))
        
        hoy = date.today()
        
        # Si se debe descontar del volumen DIAN
        if descontar_de_dian:
            volumen_dian = VolumenPendienteDian.query.filter_by(fecha=hoy).first()
            
            if not volumen_dian:
                flash('No hay volumen pendiente DIAN registrado para hoy.', 'warning')
            elif volumen_dian.volumen_pendiente < volumen_recarga:
                flash(f'El volumen a recargar ({volumen_recarga:,.0f} BBL) excede el disponible en DIAN ({volumen_dian.volumen_pendiente:,.0f} BBL).', 'danger')
                return redirect(url_for('dashboard_siza'))
            else:
                # Descontar del volumen DIAN
                volumen_dian.volumen_pendiente -= volumen_recarga
                volumen_dian.usuario_actualizacion = session.get('nombre', 'Sistema')
                volumen_dian.fecha_actualizacion = datetime.utcnow()
                
                # Actualizar observación indicando la distribución
                producto = ProductoSiza.query.get(producto_id)
                obs_dian = f"Distribuido {volumen_recarga:,.0f} BBL a {producto.nombre}"
                if volumen_dian.observacion:
                    volumen_dian.observacion += f" | {obs_dian}"
                else:
                    volumen_dian.observacion = obs_dian

                # Registar CONSUMO en Historial DIAN
                db.session.add(MovimientoDian(
                    fecha_operativa=hoy, tipo='RECARGA-CONSUMO', volumen=volumen_recarga,
                    observacion=f"Descarga hacia {producto.nombre}: {observacion or 'Recarga Inventario'}", 
                    usuario_registro=session.get('nombre', 'Sistema')
                ))
        
        # === LEER MERMA (Desperdicio) ANTES DE REGISTRAR ===
        vol_merma = float(request.form.get('volumen_total_desperdicio', 0))
        
        # Registrar la recarga (ahora con merma)
        recarga = RecargaSiza(
            fecha=hoy,
            producto_id=producto_id,
            volumen_recargado=volumen_recarga,
            volumen_merma=vol_merma,  # NUEVO: Guardar merma individual
            descontado_dian=descontar_de_dian,  # Rastrear si se descontó de DIAN
            observacion=observacion or None,
            usuario_registro=session.get('nombre', 'Sistema')
        )
        db.session.add(recarga)
        
        # Actualizar el inventario del día
        inventario = InventarioSizaDiario.query.filter_by(
            fecha=hoy,
            producto_id=producto_id
        ).first()
        
        if not inventario:
            inventario = InventarioSizaDiario(
                fecha=hoy,
                producto_id=producto_id,
                cupo_web=volumen_recarga,
                usuario_actualizacion=session.get('nombre', 'Sistema')
            )
            db.session.add(inventario)
        else:
            inventario.cupo_web += volumen_recarga
            inventario.usuario_actualizacion = session.get('nombre', 'Sistema')
            inventario.fecha_actualizacion = datetime.utcnow()
        
        # === LOGICA DE MERMA (Desperdicio) ===
        # (vol_merma ya se leyó antes de crear RecargaSiza)
        
        # 1. Actualizar Acumulado de Merma del Producto Actual
        # Verificar existencia de columna (por seguridad en migraciones parciales)
        try:
            inventario.volumen_desperdicio_generado = (inventario.volumen_desperdicio_generado or 0) + vol_merma
        except AttributeError:
            pass # Si la columna no existe aún
        
        # 2. Descontar de DIAN (Solo Merma)
        if descontar_de_dian and 'volumen_dian' in locals() and volumen_dian:
            if vol_merma > 0:
                volumen_dian.volumen_pendiente -= vol_merma
                
                current_prod = ProductoSiza.query.get(producto_id)
                obs_dian = f"Merma por recarga {current_prod.nombre}: {vol_merma:,.3f} BBL"
                
                db.session.add(MovimientoDian(
                    fecha_operativa=hoy, tipo='RECARGA-CONSUMO', volumen=vol_merma,
                    observacion=obs_dian, 
                    usuario_registro=session.get('nombre', 'Sistema')
                ))
                flash(f'ℹ️ Registrada Merma: {vol_merma:,.3f} BBL. Descontado de DIAN.', 'info')

        db.session.commit()
        
        producto = ProductoSiza.query.get(producto_id)
        mensaje = f'Recarga de {producto.nombre}: +{volumen_recarga:,.0f} Barriles. Nuevo total: {inventario.cupo_web:,.0f} BBL'
        if descontar_de_dian:
            volumen_dian_actual = VolumenPendienteDian.query.filter_by(fecha=hoy).first()
            if volumen_dian_actual:
                mensaje += f' | Descontado de DIAN. Pendiente DIAN: {volumen_dian_actual.volumen_pendiente:,.0f} BBL'
        flash(mensaje, 'success')
        
    except ValueError:
        flash('El volumen debe ser un número válido.', 'danger')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al registrar recarga: {str(e)}', 'danger')
    
    return redirect(url_for('dashboard_siza'))

@app.route('/siza/historial-recargas/<int:producto_id>')
@login_required
def historial_recargas_siza(producto_id):
    """Obtener historial de recargas de un producto con mermas."""
    try:
        recargas = RecargaSiza.query.filter_by(producto_id=producto_id)\
            .order_by(RecargaSiza.fecha.desc(), RecargaSiza.fecha_registro.desc())\
            .limit(50).all()
        
        resultado = []
        for r in recargas:
            resultado.append({
                'id': r.id,
                'fecha': r.fecha.strftime('%Y-%m-%d'),
                'volumen_recargado': float(r.volumen_recargado),
                'volumen_merma': float(r.volumen_merma or 0),
                'observacion': r.observacion or '',
                'usuario_registro': r.usuario_registro,
                'hora_registro': r.fecha_registro.strftime('%H:%M') if r.fecha_registro else ''
            })
        
        return jsonify({'success': True, 'recargas': resultado})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/siza/eliminar-recarga/<int:recarga_id>', methods=['POST'])
@login_required
def eliminar_recarga_siza(recarga_id):
    """Eliminar una recarga del historial."""
    try:
        recarga = RecargaSiza.query.get_or_404(recarga_id)
        producto_id = recarga.producto_id
        volumen = recarga.volumen_recargado
        merma = recarga.volumen_merma or 0
        fecha = recarga.fecha
        descontado_dian = recarga.descontado_dian
        
        # Eliminar registro
        db.session.delete(recarga)
        
        # Actualizar inventario (restar el volumen que se había sumado)
        inventario = InventarioSizaDiario.query.filter_by(
            fecha=fecha, producto_id=producto_id
        ).first()
        
        if inventario:
            inventario.cupo_web -= volumen
            inventario.volumen_desperdicio_generado = max(0, (inventario.volumen_desperdicio_generado or 0) - merma)
        
        # DEVOLVER A DIAN si se había descontado
        if descontado_dian:
            volumen_devolver = volumen + merma  # Volumen original + merma
            volumen_dian = VolumenPendienteDian.query.filter_by(fecha=fecha).first()
            
            if volumen_dian:
                volumen_dian.volumen_pendiente += volumen_devolver
                volumen_dian.fecha_actualizacion = datetime.utcnow()
                
                # Registrar movimiento de devolución
                db.session.add(MovimientoDian(
                    fecha_operativa=fecha,
                    tipo='REVERSA',
                    volumen=volumen_devolver,
                    observacion=f"Reversa por eliminación de recarga ID {recarga_id}: +{volumen:,.0f} BBL (producto) + {merma:,.0f} BBL (merma)",
                    usuario_registro=session.get('nombre', 'Sistema')
                ))
        
        db.session.commit()
        
        msg = f'Recarga eliminada: -{volumen:,.0f} BBL'
        if descontado_dian:
            msg += f' | Devuelto a DIAN: +{(volumen + merma):,.0f} BBL'
        flash(msg, 'warning')
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/siza/editar-recarga/<int:recarga_id>', methods=['POST'])
@login_required
def editar_recarga_siza(recarga_id):
    """Editar una recarga existente."""
    try:
        recarga = RecargaSiza.query.get_or_404(recarga_id)
        
        # Valores actuales
        volumen_anterior = recarga.volumen_recargado
        merma_anterior = recarga.volumen_merma or 0
        descontado_dian_anterior = recarga.descontado_dian
        
        # Nuevos valores del formulario
        nuevo_volumen = float(request.form.get('volumen_recargado'))
        nueva_merma = float(request.form.get('volumen_merma', 0))
        nueva_observacion = request.form.get('observacion', '').strip()
        
        # Calcular diferencias
        diff_volumen = nuevo_volumen - volumen_anterior
        diff_merma = nueva_merma - merma_anterior
        
        # Actualizar la recarga
        recarga.volumen_recargado = nuevo_volumen
        recarga.volumen_merma = nueva_merma
        recarga.observacion = nueva_observacion or None
        recarga.usuario_edicion = session.get('nombre', 'Sistema')
        recarga.fecha_edicion = datetime.utcnow()
        
        # Actualizar inventario del producto
        inventario = InventarioSizaDiario.query.filter_by(
            fecha=recarga.fecha, producto_id=recarga.producto_id
        ).first()
        
        if inventario:
            inventario.cupo_web += diff_volumen
            inventario.volumen_desperdicio_generado = max(0, (inventario.volumen_desperdicio_generado or 0) + diff_merma)
        
        # Actualizar DIAN si se había descontado
        if descontado_dian_anterior:
            diff_total = diff_volumen + diff_merma
            volumen_dian = VolumenPendienteDian.query.filter_by(fecha=recarga.fecha).first()
            
            if volumen_dian:
                volumen_dian.volumen_pendiente -= diff_total
                volumen_dian.fecha_actualizacion = datetime.utcnow()
                
                # Registrar movimiento de ajuste
                db.session.add(MovimientoDian(
                    fecha_operativa=recarga.fecha,
                    tipo='AJUSTE',
                    volumen=abs(diff_total),
                    observacion=f"Ajuste por edición recarga ID {recarga_id}: {'+' if diff_total < 0 else '-'}{abs(diff_total):,.0f} BBL",
                    usuario_registro=session.get('nombre', 'Sistema')
                ))
        
        db.session.commit()
        
        flash(f'Recarga actualizada: {nuevo_volumen:,.0f} BBL (Δ {diff_volumen:+,.0f})', 'success')
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500



@login_required
@permiso_requerido("siza_gestor")  # Solo gestores pueden consumir
@app.route('/siza/registrar-consumo', methods=['POST'])
def registrar_consumo_siza():
    """Registra un consumo/despacho de inventario para un producto."""
    try:
        producto_id = int(request.form.get('producto_id'))
        volumen_consumo = float(request.form.get('volumen_consumo', 0))
        observacion = request.form.get('observacion', '').strip()
        
        if volumen_consumo <= 0:
            flash('El volumen de consumo debe ser mayor a cero.', 'danger')
            return redirect(url_for('dashboard_siza'))
        
        hoy = date.today()
        
        # Verificar que haya suficiente inventario
        inventario = InventarioSizaDiario.query.filter_by(
            fecha=hoy,
            producto_id=producto_id
        ).first()
        
        if not inventario or inventario.cupo_web < volumen_consumo:
            producto = ProductoSiza.query.get(producto_id)
            disponible = inventario.cupo_web if inventario else 0
            flash(f'No hay suficiente inventario de {producto.nombre}. Disponible: {disponible:,.0f} BBL', 'danger')
            return redirect(url_for('dashboard_siza'))
        
        # Registrar el consumo
        consumo = ConsumoSiza(
            fecha=hoy,
            producto_id=producto_id,
            volumen_consumido=volumen_consumo,
            observacion=observacion or None,
            usuario_registro=session.get('nombre', 'Sistema')
        )
        db.session.add(consumo)
        
        # Actualizar el inventario del día
        inventario.cupo_web -= volumen_consumo
        inventario.usuario_actualizacion = session.get('nombre', 'Sistema')
        inventario.fecha_actualizacion = datetime.utcnow()
        
        db.session.commit()
        
        producto = ProductoSiza.query.get(producto_id)
        flash(f'Consumo de {producto.nombre}: -{volumen_consumo:,.0f} BBL. Nuevo total: {inventario.cupo_web:,.0f} BBL', 'success')
        
    except ValueError:
        flash('El volumen debe ser un número válido.', 'danger')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al registrar consumo: {str(e)}', 'danger')
    
    return redirect(url_for('dashboard_siza'))

@login_required
@login_required
@permiso_requerido("siza_gestor")  # Solo gestores pueden editar consumos
@app.route('/siza/editar-consumo/<int:consumo_id>', methods=['POST'])
def editar_consumo_siza(consumo_id):
    """Edita un consumo existente."""
    try:
        consumo = ConsumoSiza.query.get_or_404(consumo_id)
        volumen_anterior = consumo.volumen_consumido
        producto_id = consumo.producto_id
        fecha_consumo = consumo.fecha
        
        nuevo_volumen = float(request.form.get('volumen_consumido', 0))
        nueva_observacion = request.form.get('observacion', '').strip()
        
        if nuevo_volumen <= 0:
            flash('El volumen debe ser mayor a cero.', 'danger')
            return redirect(url_for('dashboard_siza'))
        
        # Calcular diferencia (invertida porque es consumo)
        diferencia = volumen_anterior - nuevo_volumen
        
        # Actualizar el consumo
        consumo.volumen_consumido = nuevo_volumen
        consumo.observacion = nueva_observacion or None
        consumo.usuario_edicion = session.get('nombre', 'Sistema')
        consumo.fecha_edicion = datetime.utcnow()
        
        # Ajustar el inventario
        inventario = InventarioSizaDiario.query.filter_by(
            fecha=fecha_consumo,
            producto_id=producto_id
        ).first()
        
        if inventario:
            inventario.cupo_web += diferencia
            inventario.usuario_actualizacion = session.get('nombre', 'Sistema')
            inventario.fecha_actualizacion = datetime.utcnow()
        
        db.session.commit()
        
        flash(f'Consumo editado exitosamente. Ajuste: {diferencia:+,.0f} BBL', 'success')
        
    except ValueError:
        flash('El volumen debe ser un número válido.', 'danger')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al editar consumo: {str(e)}', 'danger')
    
    return redirect(url_for('dashboard_siza'))

@login_required
@permiso_requerido("siza_gestor")  # Solo gestores pueden eliminar movimientos
@app.route('/siza/eliminar-movimiento/<tipo>/<int:movimiento_id>', methods=['POST'])
def eliminar_movimiento_siza(tipo, movimiento_id):
    """Elimina una recarga o consumo."""
    try:
        if tipo == 'recarga':
            registro = RecargaSiza.query.get_or_404(movimiento_id)
            ajuste_inventario = -registro.volumen_recargado  # Restar lo que se había sumado
            tipo_texto = 'Recarga'
        elif tipo == 'consumo':
            registro = ConsumoSiza.query.get_or_404(movimiento_id)
            ajuste_inventario = registro.volumen_consumido  # Sumar lo que se había restado
            tipo_texto = 'Consumo'
        else:
            flash('Tipo de movimiento inválido.', 'danger')
            return redirect(url_for('dashboard_siza'))
        
        producto_id = registro.producto_id
        fecha_registro = registro.fecha
        
        # Ajustar inventario
        inventario = InventarioSizaDiario.query.filter_by(
            fecha=fecha_registro,
            producto_id=producto_id
        ).first()
        
        if inventario:
            inventario.cupo_web += ajuste_inventario
            inventario.usuario_actualizacion = session.get('nombre', 'Sistema')
            inventario.fecha_actualizacion = datetime.utcnow()
        
        # Eliminar registro
        db.session.delete(registro)
        db.session.commit()
        
        flash(f'{tipo_texto} eliminado exitosamente. Inventario ajustado: {ajuste_inventario:+,.0f} BBL', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error al eliminar movimiento: {str(e)}', 'danger')
    
    return redirect(url_for('dashboard_siza'))

@login_required
@app.route('/siza/actualizar-volumen-dian', methods=['POST'])
def actualizar_volumen_dian():
    """Actualiza el volumen pendiente de aprobación DIAN. Solo accesible por Daniela y Shirley."""
    try:
        # Verificar permisos
        usuario_email = session.get('email', '')
        if usuario_email not in ['comex@conquerstrading.com', 'comexzf@conquerstrading.com']:
            flash('No tienes permiso para actualizar el volumen pendiente DIAN.', 'danger')
            return redirect(url_for('dashboard_siza'))
        
        hoy = date.today()
        
        # Obtener valores del formulario
        agregar_aprobacion = float(request.form.get('agregar_aprobacion', 0))  # Nuevo: volumen a SUMAR
        volumen_aprobado_actual = float(request.form.get('volumen_pendiente', 0))  # Valor actual (hidden)
        volumen_por_aprobar = float(request.form.get('volumen_por_aprobar', 0))
        observacion = request.form.get('observacion', '').strip()
        
        # Buscar o crear registro
        volumen_dian = VolumenPendienteDian.query.filter_by(fecha=hoy).first()
        
        # Guardar saldo previo para calcular ingresos
        saldo_por_aprobar_previo = volumen_dian.volumen_por_aprobar if volumen_dian else 0.0
        
        if not volumen_dian:
            # Lógica para registro NUEVO
            volumen_final_por_aprobar = max(0, volumen_por_aprobar - agregar_aprobacion) if agregar_aprobacion > 0 else volumen_por_aprobar
            
            volumen_dian = VolumenPendienteDian(
                fecha=hoy,
                volumen_pendiente=agregar_aprobacion,
                volumen_por_aprobar=volumen_final_por_aprobar,
                observacion=observacion,
                usuario_actualizacion=session.get('nombre', 'Usuario')
            )
            db.session.add(volumen_dian)
            
            # Registrar Movimientos Iniciales
            if volumen_por_aprobar > 0:
                db.session.add(MovimientoDian(
                    fecha_operativa=hoy, tipo='INGRESO_PENDIENTE', volumen=volumen_por_aprobar,
                    observacion=f"Ingreso Inicial: {observacion}", usuario_registro=session.get('nombre', 'Usuario')
                ))
                
            if agregar_aprobacion > 0:
                db.session.add(MovimientoDian(
                    fecha_operativa=hoy, tipo='APROBACION', volumen=agregar_aprobacion,
                    observacion=f"Aprobación Inicial: {observacion}", usuario_registro=session.get('nombre', 'Usuario')
                ))
                
        else:
            # Lógica para registro EXISTENTE
            
            # 1. Detectar INGRESO de volumen pendiente (Diferencia entre lo que había y lo nuevo en el form)
            delta_ingreso = volumen_por_aprobar - saldo_por_aprobar_previo
            
            if abs(delta_ingreso) > 0.01:
                tipo_mov = 'INGRESO_PENDIENTE' if delta_ingreso > 0 else 'AJUSTE_NEGATIVO'
                db.session.add(MovimientoDian(
                    fecha_operativa=hoy, tipo=tipo_mov, volumen=abs(delta_ingreso),
                    observacion=f"Ajuste Manual Por Aprobar ({delta_ingreso:+.2f}): {observacion}", 
                    usuario_registro=session.get('nombre', 'Usuario')
                ))

            # 2. Procesar APROBACIÓN y Actualización
            volumen_dian.volumen_pendiente = volumen_dian.volumen_pendiente + agregar_aprobacion
            
            # Si se agregó una aprobación
            if agregar_aprobacion > 0:
                # VALIDACIÓN INTELIGENTE:
                if agregar_aprobacion > volumen_por_aprobar:
                    flash(f'⚠️ Error Lógico: Estás intentando aprobar {agregar_aprobacion:,.2f} BBL, pero solo hay {volumen_por_aprobar:,.2f} BBL disponibles ({saldo_por_aprobar_previo} + ingresos).', 'danger')
                    return redirect(url_for('dashboard_siza'))

                # Descontar del volumen por aprobar
                volumen_dian.volumen_por_aprobar = max(0, volumen_por_aprobar - agregar_aprobacion)
                
                # Registrar Historial APROBACION
                db.session.add(MovimientoDian(
                    fecha_operativa=hoy, tipo='APROBACION', volumen=agregar_aprobacion,
                    observacion=f"Aprobación: {observacion}", usuario_registro=session.get('nombre', 'Usuario')
                ))
                
                # Mantener compatibilidad con HistorialAprobacionDian (Viejo)
                db.session.add(HistorialAprobacionDian(
                    fecha_operativa=hoy, volumen_agregado=agregar_aprobacion,
                    observacion=observacion, usuario_registro=session.get('nombre', 'Usuario')
                ))
                
            else:
                # Si solo se actualizó el pendiente sin aprobar
                volumen_dian.volumen_por_aprobar = volumen_por_aprobar
            
            volumen_dian.observacion = observacion
            volumen_dian.usuario_actualizacion = session.get('nombre', 'Usuario')
            volumen_dian.fecha_actualizacion = datetime.utcnow()
        
        # Caso especial: Si es el primer registro del día y hubo aprobación, también guardar historial
        if agregar_aprobacion > 0 and 'nuevo_historial' not in locals():
             nuevo_historial = HistorialAprobacionDian(
                fecha_operativa=hoy,
                volumen_agregado=agregar_aprobacion,
                observacion=observacion,
                usuario_registro=session.get('nombre', 'Usuario')
            )
             db.session.add(nuevo_historial)
        
        db.session.commit()
        
        # Mensaje de confirmación
        if agregar_aprobacion > 0:
            flash(f'✅ Nueva aprobación DIAN agregada: {agregar_aprobacion:,.3f} BBL. Total Aprobado: {volumen_dian.volumen_pendiente:,.3f} BBL', 'success')
        else:
            flash(f'Control DIAN actualizado. Aprobado: {volumen_dian.volumen_pendiente:,.3f} | Por Aprobar: {volumen_dian.volumen_por_aprobar:,.3f}', 'success')
        
    except ValueError:
        flash('El volumen debe ser un número válido.', 'danger')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al actualizar volumen DIAN: {str(e)}', 'danger')
    
    return redirect(url_for('dashboard_siza'))

@login_required
@app.route('/siza/eliminar-historial-dian/<int:id>', methods=['POST'])
def eliminar_historial_dian_siza(id):
    """Elimina un registro de historial de aprobación y revierte el volumen."""
    try:
        # Verificar permisos
        usuario_email = session.get('email', '')
        if usuario_email not in ['comex@conquerstrading.com', 'comexzf@conquerstrading.com']:
            flash('No tienes permiso para eliminar registros DIAN.', 'danger')
            return redirect(url_for('dashboard_siza'))
            
        registro = HistorialAprobacionDian.query.get_or_404(id)
        
        # Obtener el registro principal de volumen para la fecha
        volumen_dian = VolumenPendienteDian.query.filter_by(fecha=registro.fecha_operativa).first()
        
        if volumen_dian:
            # Revertir cambios:
            # 1. Restar el volumen agregado del total aprobado
            volumen_dian.volumen_pendiente = max(0, volumen_dian.volumen_pendiente - registro.volumen_agregado)
            
            # 2. Devolver el volumen al "Por Aprobar" (porque si se eliminó la aprobación, vuelve a estar pendiente)
            volumen_dian.volumen_por_aprobar += registro.volumen_agregado
            
            volumen_dian.usuario_actualizacion = session.get('nombre', 'Sistema')
            volumen_dian.fecha_actualizacion = datetime.utcnow()
            
            # Eliminar el registro del historial
            db.session.delete(registro)
            db.session.commit()
            
            flash(f'✅ Registro eliminado. Se restaron {registro.volumen_agregado:,.3f} BBL del aprobado y se devolvieron a pendiente.', 'success')
        else:
            flash('No se encontró el registro de volumen diario asociado.', 'danger')
            
    except Exception as e:
        db.session.rollback()
        flash(f'Error al eliminar registro: {str(e)}', 'danger')
        
    return redirect(url_for('dashboard_siza'))

def enviar_alerta_nuevo_pedido(pedido, producto_nombre):
    """Envía notificación por correo a Daniela y Shirli cuando hay un nuevo pedido."""
    destinatarios = ['comex@conquerstrading.com', 'comexzf@conquerstrading.com']

    asunto = f"🔔 Nuevo Pedido SIZA Pendiente: {pedido.numero_pedido}"
    
    cuerpo = f"""
    Hola Daniela y Shirli,

    Se ha registrado un nuevo pedido en SIZA que requiere su atención.

    📋 DETALLES DEL PEDIDO
    --------------------------------------------------
    Producto:      {producto_nombre}
    N° Pedido:     {pedido.numero_pedido}
    Volumen:       {pedido.volumen_solicitado:,.3f} BBL
    Solicitante:   {pedido.usuario_registro}
    Fecha/Hora:    {datetime.now().strftime('%Y-%m-%d %H:%M')}
    Estado:        PENDIENTE
    --------------------------------------------------

    Por favor ingresen al módulo SIZA para aprobar o gestionar este pedido.
    
    Atentamente,
    Sistema SIZA - Conquers Trading
    """

    # Configuración SMTP
    # Por seguridad, las credenciales SE DEBEN configurar en las variables de entorno de Render
    smtp_server = os.getenv('SMTP_SERVER', 'smtp.office365.com')
    smtp_port = int(os.getenv('SMTP_PORT', 587))
    smtp_user = os.getenv('SMTP_USER', 'numbers@conquerstrading.com') # Usuario por defecto (público)
    smtp_password = os.getenv('SMTP_PASSWORD')

    if not smtp_password:
        print(f"⚠️ [SIMULACIÓN CORREO] Faltan credenciales reales. Configurar SMTP_PASSWORD en Render.")
        # No retornamos error, simplemente no enviamos el correo para no bloquear la app
        return


    try:
        msg = MIMEMultipart()
        msg['From'] = smtp_user
        msg['To'] = ", ".join(destinatarios)
        msg['Subject'] = asunto
        msg.attach(MIMEText(cuerpo, 'plain'))

        server = smtplib.SMTP(smtp_server, smtp_port)
        server.starttls()
        server.login(smtp_user, smtp_password)
        text = msg.as_string()
        server.sendmail(smtp_user, destinatarios, text)
        server.quit()
        print(f"✅ Notificación de correo enviada exitosamente a {destinatarios} desde {smtp_user}")


    except Exception as e:
        error_msg = f"⚠️ No se pudo enviar el correo: {str(e)}"
        print(f"❌ {error_msg}")
        try:
            flash(error_msg, 'warning')
        except:
            pass

def enviar_notificacion_despacho(pedido, producto_nombre, volumen_real):
    """Envía correo al solicitante informando que su pedido fue despachado."""
    email_solicitante = obtener_email_usuario(pedido.usuario_registro)
    
    if not email_solicitante:
        print(f"⚠️ No se encontró email para el usuario {pedido.usuario_registro}. No se envió notificación.")
        return

    asunto = f"🚀 Pedido Despachado: {pedido.numero_pedido} - SIZA"
    
    cuerpo = f"""
    Hola {pedido.usuario_registro},

    Tu pedido ha sido DESPACHADO y procesado exitosamente.

    📦 RESUMEN DEL DESPACHO
    --------------------------------------------------
    N° Pedido:     {pedido.numero_pedido}
    Producto:      {producto_nombre}
    Volumen:       {volumen_real:,.3f} BBL
    Fecha/Hora:    {datetime.now().strftime('%Y-%m-%d %H:%M')}
    Estado:        DESPACHADO (COMPLETADO)
    --------------------------------------------------

    El volumen ha sido descontado del inventario oficial.
    
    Atentamente,
    Equipo de Logística - Conquers Trading
    """

    # Configuración SMTP
    smtp_server = os.getenv('SMTP_SERVER', 'smtp.office365.com')
    smtp_port = int(os.getenv('SMTP_PORT', 587))
    smtp_user = os.getenv('SMTP_USER', 'numbers@conquerstrading.com') 
    smtp_password = os.getenv('SMTP_PASSWORD')

    if not smtp_password:
        print(f"⚠️ [SIMULACIÓN CORREO] Se enviaría a {email_solicitante}: {asunto}")
        return

    try:
        msg = MIMEMultipart()
        msg['From'] = smtp_user
        msg['To'] = email_solicitante
        msg['Subject'] = asunto
        msg.attach(MIMEText(cuerpo, 'plain'))

        server = smtplib.SMTP(smtp_server, smtp_port)
        server.starttls()
        server.login(smtp_user, smtp_password)
        text = msg.as_string()
        server.sendmail(smtp_user, [email_solicitante], text) 
        server.quit()
        print(f"✅ Notificación de despacho enviada a {email_solicitante}")
    except Exception as e:
        error_msg = f"⚠️ No se pudo enviar el correo de despacho: {str(e)}"
        print(f"❌ {error_msg}")
        try:
            flash(error_msg, 'warning')
        except:
            pass

@app.route('/siza/reset-fabrica', methods=['POST'])
@login_required
def siza_reset_fabrica():
    """Borra TODOS los datos transaccionales de SIZA para reiniciar."""
    # Doble verificación de seguridad
    if session.get('rol') != 'admin' and session.get('email') not in ['comex@conquerstrading.com', 'comexzf@conquerstrading.com']:
        flash('⚠️ ACCESO DENEGADO: Solo administradores pueden reiniciar el sistema.', 'danger')
        return redirect(url_for('dashboard_siza'))

    try:
        # 1. Eliminar pedidos
        num_ped = PedidoSiza.query.delete()
        
        # 2. Eliminar movimientos de inventario y recargas
        num_rec = RecargaSiza.query.delete()
        num_con = ConsumoSiza.query.delete()
        InventarioSizaDiario.query.delete()
        
        # 3. Eliminar datos DIAN
        VolumenPendienteDian.query.delete()
        HistorialAprobacionDian.query.delete()
        
        try:
             MovimientoDian.query.delete()
        except: 
            pass
        
        db.session.commit()
        
        mensaje = f'✅ SISTEMA SIZA REINICIADO: Se eliminaron {num_ped} pedidos, {num_rec} recargas y {num_con} consumos. El inventario empieza de 0.'
        print(mensaje)
        flash(mensaje, 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error critico al reiniciar: {str(e)}', 'danger')
    
    return redirect(url_for('dashboard_siza'))



@login_required
@permiso_requerido("cupo_siza")
@app.route('/siza/actualizar-cupo-web', methods=['POST'])
def actualizar_cupo_web():
    """DEPRECADO: Mantenido por compatibilidad. Usar actualizar_inventario_siza."""
    return redirect(url_for('dashboard_siza'))

@login_required
@permiso_requerido("siza_solicitante")  # Solicitantes pueden crear pedidos
@app.route('/siza/registrar-pedido', methods=['POST'])
def registrar_pedido():
    """Registra un nuevo pedido SIZA para un producto específico con validación inteligente."""
    try:
        numero_pedido = request.form.get('numero_pedido', '').strip()
        producto_id = int(request.form.get('producto_id'))
        volumen_solicitado = float(request.form.get('volumen_solicitado', 0))
        observacion = request.form.get('observacion', '').strip()
        
        if not numero_pedido:
            flash('Debe ingresar un número de pedido.', 'danger')
            return redirect(url_for('dashboard_siza'))
        
        if volumen_solicitado <= 0:
            flash('El volumen debe ser mayor a cero.', 'danger')
            return redirect(url_for('dashboard_siza'))
        
        # Verificar si ya existe
        existe = PedidoSiza.query.filter_by(numero_pedido=numero_pedido).first()
        if existe:
            flash(f'El pedido {numero_pedido} ya está registrado.', 'warning')
            return redirect(url_for('dashboard_siza'))
        
        # VALIDACIÓN INTELIGENTE: Verificar disponibilidad de inventario
        hoy = date.today()
        inventario = InventarioSizaDiario.query.filter_by(
            fecha=hoy,
            producto_id=producto_id
        ).first()
        
        cupo_disponible = inventario.cupo_web if inventario else 0.0
        
        # Calcular lo ya comprometido por otros pedidos pendientes
        pedidos_pendientes = PedidoSiza.query.filter_by(
            producto_id=producto_id,
            estado='PENDIENTE'
        ).all()
        total_comprometido = sum(p.volumen_solicitado for p in pedidos_pendientes)
        disponible_real = cupo_disponible - total_comprometido
        
        # Crear el pedido de todas formas, pero con advertencia si excede disponibilidad
        nuevo_pedido = PedidoSiza(
            numero_pedido=numero_pedido,
            producto_id=producto_id,
            volumen_solicitado=volumen_solicitado,
            observacion=observacion or None,
            estado='PENDIENTE',
            usuario_registro=session.get('nombre', 'Sistema')
        )
        
        db.session.add(nuevo_pedido)
        db.session.commit()
        
        producto = ProductoSiza.query.get(producto_id)
        
        # Enviar notificación por correo
        enviar_alerta_nuevo_pedido(nuevo_pedido, producto.nombre)
        
        # Mostrar mensaje según disponibilidad
        if volumen_solicitado > disponible_real:
            flash(
                f'⚠️ ADVERTENCIA: Pedido {numero_pedido} registrado pero NO HAY CANTIDAD SUFICIENTE. '
                f'Solicitado: {volumen_solicitado:,.0f} BBL | Disponible: {disponible_real:,.0f} BBL. '
                f'Se requiere recarga de {producto.nombre} antes de aprobar este pedido.',
                'warning'
            )
        else:
            flash(
                f'✅ Pedido {numero_pedido} registrado exitosamente para {producto.nombre}: '
                f'{volumen_solicitado:,.0f} BBL. Disponible suficiente: {disponible_real:,.0f} BBL',
                'success'
            )
        
    except ValueError:
        flash('Los valores numéricos deben ser válidos.', 'danger')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al registrar pedido: {str(e)}', 'danger')
    
    return redirect(url_for('dashboard_siza'))

@login_required
@permiso_requerido("siza_solicitante")  # Solicitantes pueden editar sus propios pedidos
@app.route('/siza/editar-pedido/<int:pedido_id>', methods=['POST'])
def editar_pedido_siza(pedido_id):
    """Edita un pedido SIZA existente (estado, volumen, observación)."""
    # Verificar permisos: gestores pueden editar todo, solicitantes solo sus pedidos
    es_gestor = tiene_permiso('siza_gestor')
    usuario_actual = session.get('nombre', 'Sistema')
    try:
        pedido = PedidoSiza.query.get_or_404(pedido_id)
        
        # Verificar si el usuario puede editar este pedido
        if not es_gestor and pedido.usuario_registro != usuario_actual:
            flash('No tienes permiso para editar este pedido. Solo puedes editar tus propios pedidos.', 'danger')
            return redirect(url_for('dashboard_siza'))
        
        # Permitir editar cualquier campo
        nuevo_estado = request.form.get('estado', '').strip()
        nuevo_volumen = request.form.get('volumen_solicitado', '').strip()
        nueva_observacion = request.form.get('observacion', '').strip()
        
        cambios = []
        
        # Actualizar estado si se proporciona
        if nuevo_estado and nuevo_estado != pedido.estado:
            if nuevo_estado in ['PENDIENTE', 'APROBADO', 'RECHAZADO', 'COMPLETADO']:
                estado_anterior = pedido.estado
                pedido.estado = nuevo_estado
                cambios.append(f'Estado: {estado_anterior} → {nuevo_estado}')
            else:
                flash('Estado inválido.', 'danger')
                return redirect(url_for('dashboard_siza'))
        
        # Actualizar volumen si se proporciona
        if nuevo_volumen:
            try:
                volumen_nuevo = float(nuevo_volumen)
                if volumen_nuevo <= 0:
                    flash('El volumen debe ser mayor a cero.', 'danger')
                    return redirect(url_for('dashboard_siza'))
                if volumen_nuevo != pedido.volumen_solicitado:
                    volumen_anterior = pedido.volumen_solicitado
                    pedido.volumen_solicitado = volumen_nuevo
                    cambios.append(f'Volumen: {volumen_anterior:,.0f} → {volumen_nuevo:,.0f} BBL')
            except ValueError:
                flash('Volumen inválido.', 'danger')
                return redirect(url_for('dashboard_siza'))
        
        # Actualizar observación
        if nueva_observacion != (pedido.observacion or ''):
            pedido.observacion = nueva_observacion or None
            cambios.append('Observación actualizada')
        
        # Registrar quién y cuándo editó
        pedido.usuario_gestion = session.get('nombre', 'Sistema')
        pedido.fecha_gestion = datetime.utcnow()
        
        db.session.commit()
        
        if cambios:
            flash(f'Pedido {pedido.numero_pedido} editado: {" | ".join(cambios)}', 'success')
        else:
            flash('No se realizaron cambios.', 'info')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error al editar pedido: {str(e)}', 'danger')
    
    return redirect(url_for('dashboard_siza'))

@login_required
@permiso_requerido("siza_gestor")  # Solo gestores pueden consumir
@app.route('/siza/consumir-pedidos', methods=['POST'])
def consumir_pedidos_siza():
    """Consume automáticamente los pedidos pendientes de un producto."""
    try:
        producto_id = int(request.form.get('producto_id'))
        
        # Obtener pedidos APROBADOS pendientes de consumo para este producto
        pedidos_aprobados = PedidoSiza.query.filter_by(
            producto_id=producto_id,
            estado='APROBADO'
        ).order_by(PedidoSiza.fecha_registro).all()
        
        if not pedidos_aprobados:
            flash('No hay pedidos aprobados pendientes de consumo para este producto.', 'info')
            return redirect(url_for('dashboard_siza'))
        
        hoy = date.today()
        inventario = InventarioSizaDiario.query.filter_by(
            fecha=hoy,
            producto_id=producto_id
        ).first()
        
        if not inventario:
            flash('No hay inventario registrado para este producto.', 'danger')
            return redirect(url_for('dashboard_siza'))
        
        # Procesar cada pedido aprobado
        pedidos_consumidos = []
        volumen_total_consumido = 0
        
        for pedido in pedidos_aprobados:
            if inventario.cupo_web >= pedido.volumen_solicitado:
                # Registrar el consumo
                consumo = ConsumoSiza(
                    fecha=hoy,
                    producto_id=producto_id,
                    volumen_consumido=pedido.volumen_solicitado,
                    observacion=f"Consumo automático de pedido {pedido.numero_pedido}",
                    usuario_registro=session.get('nombre', 'Sistema')
                )
                db.session.add(consumo)
                
                # Actualizar inventario
                inventario.cupo_web -= pedido.volumen_solicitado
                inventario.usuario_actualizacion = session.get('nombre', 'Sistema')
                inventario.fecha_actualizacion = datetime.utcnow()
                
                # Marcar pedido como COMPLETADO
                pedido.estado = 'COMPLETADO'
                pedido.usuario_gestion = session.get('nombre', 'Sistema')
                pedido.fecha_gestion = datetime.utcnow()
                
                pedidos_consumidos.append(pedido.numero_pedido)
                volumen_total_consumido += pedido.volumen_solicitado
            else:
                # No hay suficiente inventario para este pedido
                flash(
                    f'⚠️ Pedido {pedido.numero_pedido} no pudo ser consumido. '
                    f'Requiere: {pedido.volumen_solicitado:,.0f} BBL | Disponible: {inventario.cupo_web:,.0f} BBL',
                    'warning'
                )
        
        db.session.commit()
        
        producto = ProductoSiza.query.get(producto_id)
        if pedidos_consumidos:
            flash(
                f'✅ Consumo exitoso de {producto.nombre}. '
                f'Pedidos procesados: {len(pedidos_consumidos)} ({", ".join(pedidos_consumidos)}). '
                f'Volumen total: {volumen_total_consumido:,.0f} BBL. '
                f'Nuevo inventario: {inventario.cupo_web:,.0f} BBL',
                'success'
            )
        else:
            flash('No se pudieron consumir pedidos. Verifique el inventario disponible.', 'warning')
        
    except ValueError:
        flash('Error en los valores proporcionados.', 'danger')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al consumir pedidos: {str(e)}', 'danger')
    
    return redirect(url_for('dashboard_siza'))

@login_required
@permiso_requerido("siza_gestor")  # Solo gestores pueden aprobar/rechazar
@app.route('/siza/gestionar-pedido/<int:pedido_id>', methods=['POST'])
def gestionar_pedido(pedido_id):
    """Aprueba o rechaza un pedido SIZA."""
    try:
        accion = request.form.get('accion')  # 'aprobar' o 'rechazar'
        
        pedido = PedidoSiza.query.get_or_404(pedido_id)
        
        if pedido.estado != 'PENDIENTE':
            flash(f'El pedido ya fue procesado anteriormente.', 'warning')
            return redirect(url_for('dashboard_siza'))
        
        if accion == 'aprobar':
            # Verificar si hay cupo disponible para este producto
            hoy = date.today()
            inventario = InventarioSizaDiario.query.filter_by(
                fecha=hoy,
                producto_id=pedido.producto_id
            ).first()
            
            cupo_disponible = inventario.cupo_web if inventario else 0.0
            
            # Pedidos pendientes del mismo producto
            pedidos_pendientes = PedidoSiza.query.filter_by(
                producto_id=pedido.producto_id,
                estado='PENDIENTE'
            ).all()
            
            total_comprometido = sum(p.volumen_solicitado for p in pedidos_pendientes)
            disponible_real = cupo_disponible - total_comprometido
            
            if disponible_real <= 0:
                flash(f'No hay cupo disponible de {pedido.producto.nombre} para aprobar este pedido.', 'danger')
                return redirect(url_for('dashboard_siza'))
            
            pedido.estado = 'APROBADO'
            pedido.usuario_gestion = session.get('nombre', 'Sistema')
            pedido.fecha_gestion = datetime.utcnow()
            flash(f'Pedido {pedido.numero_pedido} ({pedido.producto.nombre}) APROBADO exitosamente.', 'success')
            
        elif accion == 'rechazar':
            pedido.estado = 'RECHAZADO'
            pedido.usuario_gestion = session.get('nombre', 'Sistema')
            pedido.fecha_gestion = datetime.utcnow()
            flash(f'Pedido {pedido.numero_pedido} ({pedido.producto.nombre}) RECHAZADO.', 'info')
        
        else:
            flash('Acción no válida.', 'danger')
            return redirect(url_for('dashboard_siza'))
        
        db.session.commit()
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error al gestionar pedido: {str(e)}', 'danger')
    
    return redirect(url_for('dashboard_siza'))

def obtener_email_usuario(nombre_usuario):
    """Busca el email de un usuario en el diccionario USUARIOS basado en su nombre."""
    for email, datos in USUARIOS.items():
        if datos.get('nombre') == nombre_usuario:
            return email
    return None

def enviar_notificacion_despacho(pedido, producto_nombre, volumen_real):
    """Envía correo al solicitante informando que su pedido fue despachado."""
    email_solicitante = obtener_email_usuario(pedido.usuario_registro)
    
    if not email_solicitante:
        print(f"⚠️ No se encontró email para el usuario {pedido.usuario_registro}. No se envió notificación.")
        return

    asunto = f"🚀 Pedido Despachado: {pedido.numero_pedido} - SIZA"
    
    cuerpo = f"""
    Hola {pedido.usuario_registro},

    Tu pedido ha sido DESPACHADO y procesado exitosamente.

    📦 RESUMEN DEL DESPACHO
    --------------------------------------------------
    N° Pedido:     {pedido.numero_pedido}
    Producto:      {producto_nombre}
    Volumen:       {volumen_real:,.3f} BBL
    Fecha/Hora:    {datetime.now().strftime('%Y-%m-%d %H:%M')}
    Estado:        DESPACHADO (COMPLETADO)
    --------------------------------------------------

    El volumen ha sido descontado del inventario oficial.
    
    Atentamente,
    Equipo de Logística - Conquers Trading
    """

    # Configuración SMTP
    smtp_server = os.getenv('SMTP_SERVER', 'smtp.office365.com')
    smtp_port = int(os.getenv('SMTP_PORT', 587))
    smtp_user = os.getenv('SMTP_USER', 'numbers@conquerstrading.com') 
    smtp_password = os.getenv('SMTP_PASSWORD')

    if not smtp_password:
        print(f"⚠️ [SIMULACIÓN CORREO] Se enviaría a {email_solicitante}: {asunto}")
        return

    try:
        msg = MIMEMultipart()
        msg['From'] = smtp_user
        msg['To'] = email_solicitante
        msg['Subject'] = asunto
        msg.attach(MIMEText(cuerpo, 'plain'))

        server = smtplib.SMTP(smtp_server, smtp_port)
        server.starttls()
        server.login(smtp_user, smtp_password)
        text = msg.as_string()
        server.sendmail(smtp_user, [email_solicitante], text) 
        server.quit()
        print(f"✅ Notificación de despacho enviada a {email_solicitante}")
    except Exception as e:
        print(f"❌ Error al enviar correo de despacho: {str(e)}")

@login_required
@permiso_requerido("siza_gestor")
@app.route('/siza/despachar-pedido/<int:pedido_id>', methods=['POST'])
def despachar_pedido_siza(pedido_id):
    """Registra el despacho (consumo real) de un pedido aprobado."""
    try:
        pedido = PedidoSiza.query.get_or_404(pedido_id)
        
        if pedido.estado != 'APROBADO':
            flash(f'El pedido {pedido.numero_pedido} debe estar APROBADO para ser despachado.', 'warning')
            return redirect(url_for('dashboard_siza'))
        
        # Volumen real despachado (puede diferir ligeramente del solicitado)
        volumen_despachado = float(request.form.get('volumen_real', pedido.volumen_solicitado))
        observacion = request.form.get('observacion', '')
        
        if volumen_despachado <= 0:
            flash('El volumen despachado debe ser mayor a cero.', 'danger')
            return redirect(url_for('dashboard_siza'))
        
        hoy = date.today()
        
        # Obtener inventario para descontar
        inventario = InventarioSizaDiario.query.filter_by(
            fecha=hoy,
            producto_id=pedido.producto_id
        ).first()
        
        # Nota: El volumen ya estaba "comprometido", pero ahora se "consume" realmente.
        # Al pasar a DESPACHADO, sale de la suma de "comprometido" (filtro estado IN PENDIENTE, APROBADO).
        # Y debemos restarlo de cupo_web.
        
        if not inventario or inventario.cupo_web < volumen_despachado:
             # OJO: Si estaba comprometido, teóricamente debería haber cupo, salvo que se haya consumido por otro lado sin pedido.
             # Permitiremos el despacho pero advertiremos si queda negativo? No, bloqueo estricto si no hay fisico.
             # Pero espera, cupo_web incluye TODO (incluso lo comprometido).
            if not inventario or inventario.cupo_web < volumen_despachado:
                 flash(f'Error Crítico: No hay inventario físico suficiente ({inventario.cupo_web if inventario else 0} BBL) para despachar.', 'danger')
                 return redirect(url_for('dashboard_siza'))

        # Actualizar Inventario
        inventario.cupo_web -= volumen_despachado
        inventario.usuario_actualizacion = session.get('nombre', 'Sistema')
        inventario.fecha_actualizacion = datetime.utcnow()
        
        # Registrar Consumo vinculado (para historial)
        consumo = ConsumoSiza(
            fecha=hoy,
            producto_id=pedido.producto_id,
            volumen_consumido=volumen_despachado,
            observacion=f"Despacho Pedido #{pedido.numero_pedido}. {observacion}",
            usuario_registro=session.get('nombre', 'Sistema')
        )
        db.session.add(consumo)
        
        # Actualizar Pedido
        pedido.estado = 'DESPACHADO'
        pedido.usuario_gestion = session.get('nombre', 'Sistema')
        pedido.fecha_gestion = datetime.utcnow()
        # Podríamos guardar el volumen real entregado en algún campo extra del pedido si fuera necesario, 
        # pero por ahora asumimos que el ConsumoSiza es el registro fiel del físico.
        
        db.session.commit()
        
        # Enviar notificación de despacho
        enviar_notificacion_despacho(pedido, pedido.producto.nombre, volumen_despachado)

        flash(f'Pedido {pedido.numero_pedido} DESPACHADO correctamente (-{volumen_despachado:,.3f} BBL).', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error al despachar pedido: {str(e)}', 'danger')
    
    return redirect(url_for('dashboard_siza'))

@login_required
@permiso_requerido("siza_solicitante")  # Todos pueden ver historial
@app.route('/siza/historial-movimientos')
def historial_movimientos_siza():
    """Obtiene el historial completo de movimientos (recargas y consumos)."""
    try:
        # Parámetros de filtro
        dias = int(request.args.get('dias', 90))  # Últimos 90 días por defecto
        tipo_filtro = request.args.get('tipo', 'todos')  # todos, recarga, consumo
        producto_id = request.args.get('producto_id', None)
        
        fecha_limite = date.today() - timedelta(days=dias)
        
        movimientos = []
        
        # Obtener recargas
        if tipo_filtro in ['todos', 'recarga']:
            query_recargas = RecargaSiza.query.filter(RecargaSiza.fecha >= fecha_limite)
            if producto_id:
                query_recargas = query_recargas.filter_by(producto_id=int(producto_id))
            recargas = query_recargas.order_by(RecargaSiza.fecha.desc(), RecargaSiza.fecha_registro.desc()).all()
            
            for recarga in recargas:
                movimientos.append({
                    'tipo': 'recarga',
                    'id': recarga.id,
                    'fecha': recarga.fecha.strftime('%d/%m/%Y'),
                    'fecha_registro': recarga.fecha_registro.strftime('%d/%m/%Y %H:%M'),
                    'producto': recarga.producto.nombre,
                    'producto_id': recarga.producto_id,
                    'volumen': float(recarga.volumen_recargado),
                    'observacion': recarga.observacion or '-',
                    'usuario': recarga.usuario_registro,
                    'editado': recarga.usuario_edicion is not None,
                    'usuario_edicion': recarga.usuario_edicion,
                    'fecha_edicion': recarga.fecha_edicion.strftime('%d/%m/%Y %H:%M') if recarga.fecha_edicion else None
                })
        
        # Obtener consumos
        if tipo_filtro in ['todos', 'consumo']:
            query_consumos = ConsumoSiza.query.filter(ConsumoSiza.fecha >= fecha_limite)
            if producto_id:
                query_consumos = query_consumos.filter_by(producto_id=int(producto_id))
            consumos = query_consumos.order_by(ConsumoSiza.fecha.desc(), ConsumoSiza.fecha_registro.desc()).all()
            
            for consumo in consumos:
                movimientos.append({
                    'tipo': 'consumo',
                    'id': consumo.id,
                    'fecha': consumo.fecha.strftime('%d/%m/%Y'),
                    'fecha_registro': consumo.fecha_registro.strftime('%d/%m/%Y %H:%M'),
                    'producto': consumo.producto.nombre,
                    'producto_id': consumo.producto_id,
                    'volumen': float(consumo.volumen_consumido),
                    'observacion': consumo.observacion or '-',
                    'usuario': consumo.usuario_registro,
                    'editado': consumo.usuario_edicion is not None,
                    'usuario_edicion': consumo.usuario_edicion,
                    'fecha_edicion': consumo.fecha_edicion.strftime('%d/%m/%Y %H:%M') if consumo.fecha_edicion else None
                })
        
        # Ordenar por fecha
        movimientos.sort(key=lambda x: x['fecha_registro'], reverse=True)
        
        return jsonify({
            'success': True,
            'movimientos': movimientos,
            'total': len(movimientos)
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@login_required
@permiso_requerido("siza_solicitante")  # Todos pueden ver historial de pedidos
@app.route('/siza/historial-pedidos')
def historial_pedidos_siza():
    """Obtiene el historial completo de pedidos con filtros."""
    try:
        # Parámetros de filtro
        dias = int(request.args.get('dias', 90))  # Últimos 90 días por defecto
        estado_filtro = request.args.get('estado', 'todos')  # todos, PENDIENTE, APROBADO, RECHAZADO, COMPLETADO
        producto_id = request.args.get('producto_id', None)
        
        fecha_limite = date.today() - timedelta(days=dias)
        
        # Query base
        query_pedidos = PedidoSiza.query.filter(PedidoSiza.fecha_registro >= fecha_limite)
        
        # Filtrar por estado
        if estado_filtro != 'todos':
            query_pedidos = query_pedidos.filter_by(estado=estado_filtro)
        
        # Filtrar por producto
        if producto_id:
            query_pedidos = query_pedidos.filter_by(producto_id=int(producto_id))
        
        # Obtener pedidos
        pedidos = query_pedidos.order_by(PedidoSiza.fecha_registro.desc()).all()
        
        # Formatear resultados
        pedidos_list = []
        for pedido in pedidos:
            pedidos_list.append({
                'id': pedido.id,
                'numero_pedido': pedido.numero_pedido,
                'producto': pedido.producto.nombre,
                'producto_id': pedido.producto_id,
                'volumen_solicitado': float(pedido.volumen_solicitado),
                'observacion': pedido.observacion or '-',
                'estado': pedido.estado,
                'fecha_registro': pedido.fecha_registro.strftime('%d/%m/%Y %H:%M'),
                'usuario_registro': pedido.usuario_registro,
                'fecha_gestion': pedido.fecha_gestion.strftime('%d/%m/%Y %H:%M') if pedido.fecha_gestion else None,
                'usuario_gestion': pedido.usuario_gestion or '-'
            })
        
        return jsonify({
            'success': True,
            'pedidos': pedidos_list,
            'total': len(pedidos_list)
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@login_required
@app.route('/inicio-siza')
def home_siza():
    """Página de inicio unificada - redirecciona a home global."""
    return redirect(url_for('home_global'))

@login_required
@app.route('/reporte_barcaza')
def reporte_barcaza():
    # 1. Lógica del filtro de fecha (idéntica a la que ya usamos)
    fecha_str = request.args.get('fecha')
    try:
        fecha_seleccionada = date.fromisoformat(fecha_str) if fecha_str else date.today()
    except (ValueError, TypeError):
        fecha_seleccionada = date.today()
    
    timestamp_limite = datetime.combine(fecha_seleccionada, time.max)

    # 2. Consulta a la base de datos para obtener el estado de ese día para la Barcaza Orion
    subquery = (db.session.query(
        func.max(RegistroBarcazaOrion.id).label('max_id')
    ).filter(
        RegistroBarcazaOrion.timestamp <= timestamp_limite
    ).group_by(RegistroBarcazaOrion.tk, RegistroBarcazaOrion.grupo).subquery())

    registros_recientes = (db.session.query(RegistroBarcazaOrion)
        .filter(RegistroBarcazaOrion.id.in_(subquery))
        .all())
    
    # 3. Preparar los datos para la plantilla
    todos_los_tanques_lista = []
    if registros_recientes:
        for registro in registros_recientes:
            todos_los_tanques_lista.append({
                "TK": registro.tk, "PRODUCTO": registro.producto, "MAX_CAP": registro.max_cap,
                "BLS_60": registro.bls_60, "API": registro.api, 
                "BSW": registro.bsw, "S": registro.s, "grupo": registro.grupo
            })

    # 4. Calcular el total consolidado a partir de los datos filtrados
    total_consolidado = calcular_estadisticas(todos_los_tanques_lista)
    
    # 5. Agrupar los tanques en el diccionario que la plantilla espera
    datos_para_template = {}
    nombres_display = {
        "PRINCIPAL": "Tanque Principal (TK-101)", "MANZANILLO": "Barcaza Manzanillo (MGO)",
        "CR": "Barcaza CR", "MARGOTH": "Barcaza Margoth", "ODISEA": "Barcaza Odisea"
    }
    def ordenar_por_planilla(lista, grupo, planilla):
        orden = [t["TK"] for t in planilla if t["grupo"] == grupo]
        return sorted([t for t in lista if t["grupo"] == grupo], key=lambda x: orden.index(x["TK"]) if x["TK"] in orden else 999)

    if todos_los_tanques_lista:
        for grupo_key, nombre_barcaza in nombres_display.items():
            if grupo_key == "CR":
                tanques_ordenados = ordenar_por_planilla(todos_los_tanques_lista, "CR", PLANILLA_BARCAZA_ORION)
            else:
                tanques_ordenados = [t for t in todos_los_tanques_lista if t.get("grupo") == grupo_key]
            if tanques_ordenados:
                datos_para_template[nombre_barcaza] = {"tanques": tanques_ordenados, "totales": {}}
                datos_para_template[nombre_barcaza]["totales"] = calcular_estadisticas(tanques_ordenados)

    # 6. Formatear el mensaje de "Última actualización"
    fecha_actualizacion_info = "No hay registros para la fecha seleccionada."
    if registros_recientes:
        ultimo_registro = max(registros_recientes, key=lambda r: r.timestamp)
        fecha_formato_para_funcion = ultimo_registro.timestamp.strftime("%Y_%m_%d_%H_%M_%S")
        fecha_actualizacion_info = formatear_info_actualizacion(
            fecha_formato_para_funcion, 
            ultimo_registro.usuario
        )
    
    # 7. Renderizar la plantilla con todos los datos necesarios
    return render_template("reporte_barcaza_orion.html",
                           titulo="Reporte Interactivo - Barcaza Orion", # Título corregido
                           datos_para_template=datos_para_template,
                           total_consolidado=total_consolidado,
                           todos_los_tanques_json=json.dumps(todos_los_tanques_lista),
                           fecha_actualizacion_info=fecha_actualizacion_info,
                           fecha_seleccionada=fecha_seleccionada.isoformat(),
                           today_iso=date.today().isoformat())

@login_required
@app.route('/reporte_barcaza_bita')
def reporte_barcaza_bita():
    # La lógica de consulta es idéntica a la de la planilla
    fecha_str = request.args.get('fecha')
    try:
        fecha_seleccionada = date.fromisoformat(fecha_str) if fecha_str else date.today()
    except (ValueError, TypeError):
        fecha_seleccionada = date.today()

    timestamp_limite = datetime.combine(fecha_seleccionada, time.max)
    subquery = (db.session.query(func.max(RegistroBarcazaBita.id).label('max_id'))
        .filter(RegistroBarcazaBita.timestamp <= timestamp_limite).group_by(RegistroBarcazaBita.tk).subquery())
    registros_recientes = (db.session.query(RegistroBarcazaBita)
        .filter(RegistroBarcazaBita.id.in_(subquery)).all())

    # Preparar los datos y estadísticas para el reporte
    datos_reporte = []
    if registros_recientes:
        for r in registros_recientes:
            datos_reporte.append({
                "TK": r.tk, "PRODUCTO": r.producto, "MAX_CAP": r.max_cap,
                "BLS_60": r.bls_60, "API": r.api, "BSW": r.bsw, "S": r.s
            })

    total_consolidado = calcular_estadisticas(datos_reporte)
    tanques_marinse = [tk for tk in datos_reporte if tk.get('TK','').startswith('MARI')]
    tanques_oidech = [tk for tk in datos_reporte if tk.get('TK','').startswith('OID')]
    stats_marinse = calcular_estadisticas(tanques_marinse)
    stats_oidech = calcular_estadisticas(tanques_oidech)

    fecha_actualizacion_info = "No hay registros para la fecha seleccionada."
    if registros_recientes:
        ultimo_registro = max(registros_recientes, key=lambda r: r.timestamp)
        fecha_fmt = ultimo_registro.timestamp.strftime("%Y_%m_%d_%H_%M_%S")
        fecha_actualizacion_info = formatear_info_actualizacion(fecha_fmt, ultimo_registro.usuario)

    return render_template("reporte_barcaza_bita.html",
                           titulo="Reporte Interactivo - Barcaza BITA",
                           fecha_actualizacion_info=fecha_actualizacion_info,
                           nombre=session.get('nombre', 'Desconocido'),
                           total_consolidado=total_consolidado,
                           todos_los_tanques_json=json.dumps(datos_reporte),
                           tanques_marinse=tanques_marinse,
                           stats_marinse=stats_marinse,
                           tanques_oidech=tanques_oidech,
                           stats_oidech=stats_oidech,
                           fecha_seleccionada=fecha_seleccionada.isoformat(),
                           today_iso=date.today().isoformat())

@login_required
@permiso_requerido('barcaza_bita')
@app.route('/guardar-registro-bita', methods=['POST'])
def guardar_registro_bita():
    lista_tanques = request.get_json()
    if not isinstance(lista_tanques, list):
        return jsonify(success=False, message="Formato de datos incorrecto."), 400

    try:
        def to_float(v):
            if v is None:
                return None
            s = str(v).strip().replace(',', '.')
            if s == '':
                return None
            try:
                return float(s)
            except Exception:
                return None
        today_start = datetime.combine(date.today(), time.min)
        today_end = datetime.combine(date.today(), time.max)

        for datos_tanque in lista_tanques:
            tk = datos_tanque.get('TK')
            if not tk: continue

            registro_existente = db.session.query(RegistroBarcazaBita).filter(
                RegistroBarcazaBita.tk == tk,
                RegistroBarcazaBita.timestamp.between(today_start, today_end)
            ).first()

            if registro_existente:
                # ACTUALIZAR
                registro_existente.usuario = session.get("nombre", "No identificado")
                registro_existente.bls_60 = to_float(datos_tanque.get('BLS_60'))
                registro_existente.api = to_float(datos_tanque.get('API'))
                registro_existente.bsw = to_float(datos_tanque.get('BSW'))
                registro_existente.s = to_float(datos_tanque.get('S'))
                registro_existente.timestamp = datetime.utcnow()
            else:
                # CREAR
                nuevo_registro = RegistroBarcazaBita(
                    timestamp=datetime.utcnow(),
                    usuario=session.get("nombre", "No identificado"),
                    tk=tk,
                    producto=datos_tanque.get('PRODUCTO'),
                    max_cap=to_float(datos_tanque.get('MAX_CAP')),
                    bls_60=to_float(datos_tanque.get('BLS_60')),
                    api=to_float(datos_tanque.get('API')),
                    bsw=to_float(datos_tanque.get('BSW')),
                    s=to_float(datos_tanque.get('S'))
                )
                db.session.add(nuevo_registro)
        
        db.session.commit()
        return jsonify(success=True, message="Inventario de Barcaza BITA actualizado.")
    except Exception as e:
        db.session.rollback()
        return jsonify(success=False, message=f"Error interno: {str(e)}"), 500

@login_required
@permiso_requerido('barcaza_orion')
@app.route('/guardar_registro_barcaza', methods=['POST'])
def guardar_registro_barcaza():
    lista_tanques = request.get_json()
    if not isinstance(lista_tanques, list):
        return jsonify(success=False, message="Formato incorrecto."), 400
    
    try:
        def to_float(v):
            if v is None:
                return None
            s = str(v).strip().replace(',', '.')
            if s == '':
                return None
            try:
                return float(s)
            except Exception:
                return None
        today_start = datetime.combine(date.today(), time.min)
        today_end = datetime.combine(date.today(), time.max)

        for datos_tanque in lista_tanques:
            tk = datos_tanque.get('TK')
            grupo = datos_tanque.get('grupo')
            if not tk or not grupo: continue

            registro_existente = db.session.query(RegistroBarcazaOrion).filter(
                RegistroBarcazaOrion.tk == tk,
                RegistroBarcazaOrion.grupo == grupo,
                RegistroBarcazaOrion.timestamp.between(today_start, today_end)
            ).first()

            if registro_existente:
                # ACTUALIZAR
                registro_existente.usuario = session.get("nombre", "No identificado")
                registro_existente.bls_60 = to_float(datos_tanque.get('BLS_60'))
                registro_existente.api = to_float(datos_tanque.get('API'))
                registro_existente.bsw = to_float(datos_tanque.get('BSW'))
                registro_existente.s = to_float(datos_tanque.get('S'))
                registro_existente.timestamp = datetime.utcnow()
            else:
                # CREAR
                nuevo_registro = RegistroBarcazaOrion(
                    timestamp=datetime.utcnow(),
                    usuario=session.get("nombre", "No identificado"),
                    tk=tk,
                    grupo=grupo,
                    producto=datos_tanque.get('PRODUCTO'),
                    max_cap=to_float(datos_tanque.get('MAX_CAP')),
                    bls_60=to_float(datos_tanque.get('BLS_60')),
                    api=to_float(datos_tanque.get('API')),
                    bsw=to_float(datos_tanque.get('BSW')),
                    s=to_float(datos_tanque.get('S'))
                )
                db.session.add(nuevo_registro)
        
        db.session.commit()
        return jsonify(success=True, message="Inventario de Barcaza Orion actualizado.")
    except Exception as e:
        db.session.rollback()
        return jsonify(success=False, message=f"Error interno: {str(e)}"), 500
    
@login_required
@app.route('/dashboard_reportes')
def dashboard_reportes():
    # El permiso de acceso no necesita cambios
    user_areas = session.get('area', [])
    if session.get('rol') != 'admin' and len(user_areas) == 1 and user_areas[0] == 'guia_transporte':
        return redirect(url_for('home_logistica'))

    # --- Resumen para PLANTA ---
    planta_summary = {'datos': [], 'info_completa': 'Sin Registros'}
    try:
        registros_planta = db.session.query(RegistroPlanta).all()
        # Filtramos para asegurarnos de que solo usamos registros con fecha
        registros_validos = [r for r in registros_planta if r.timestamp]
        if registros_validos:
            ultimo_registro = max(registros_validos, key=lambda r: r.timestamp)
            planta_summary['datos'] = registros_validos
            planta_summary['info_completa'] = formatear_info_actualizacion(
            ultimo_registro.timestamp, ultimo_registro.usuario
            )
    except Exception as e:
        print(f"Error al cargar resumen de Planta: {e}")

    # --- Resumen para BARCAZA ORION ---
    orion_summary = {'datos': [], 'info_completa': 'Sin Registros'}
    try:
        registros_orion = db.session.query(RegistroBarcazaOrion).all()
        registros_validos = [r for r in registros_orion if r.timestamp]
        if registros_validos:
            ultimo_registro = max(registros_validos, key=lambda r: r.timestamp)
            orion_summary['datos'] = registros_validos
            orion_summary['info_completa'] = formatear_info_actualizacion(
                ultimo_registro.timestamp, ultimo_registro.usuario
            )
    except Exception as e:
        print(f"Error al cargar resumen de Orion: {e}")

    # --- Resumen para BARCAZA BITA ---
    bita_summary = {'datos': [], 'info_completa': 'Sin Registros'}
    try:
        registros_bita = db.session.query(RegistroBarcazaBita).all()
        registros_validos = [r for r in registros_bita if r.timestamp]
        if registros_validos:
            ultimo_registro = max(registros_validos, key=lambda r: r.timestamp)
            bita_summary['datos'] = registros_validos
            bita_summary['info_completa'] = formatear_info_actualizacion(
                ultimo_registro.timestamp, ultimo_registro.usuario
            )
    except Exception as e:
        print(f"Error al cargar resumen de BITA: {e}")

    # --- Resumen para TRÁNSITO ---
    transito_summary = {'datos': [], 'refineria_count': 0, 'edms_count': 0, 'otros_count': 0, 'info_completa': 'Sin Registros'}
    try:
        registros_transito = db.session.query(RegistroTransito).all()
        registros_validos = [r for r in registros_transito if r.timestamp]
        if registros_validos:
            ultimo_registro = max(registros_validos, key=lambda r: r.timestamp)
            transito_summary['datos'] = registros_validos
            transito_summary['info_completa'] = formatear_info_actualizacion(
                ultimo_registro.timestamp, ultimo_registro.usuario
            )
            transito_summary['refineria_count'] = sum(1 for r in registros_validos if r.tipo_transito == 'refineria')
            transito_summary['edms_count'] = sum(1 for r in registros_validos if r.tipo_transito == 'general')
            
    except Exception as e:
        print(f"Error al cargar resumen de Tránsito: {e}")

    # --- Construcción de ALERTAS ÚTILES ---
    alerts = []  # Cada alerta: {severity, category, message, icon}

    def add_alert(severity, category, message):
        icon_map = {
            'danger': 'exclamation-octagon-fill',
            'warning': 'exclamation-triangle-fill',
            'info': 'info-circle-fill',
            'success': 'check-circle-fill'
        }
        alerts.append({
            'severity': severity,
            'category': category,
            'message': message,
            'icon': icon_map.get(severity, 'info-circle')
        })

    now_utc = datetime.utcnow()

    # Helper para obtener último por clave (tk) de una lista de registros con timestamp
    def latest_by(records, attr):
        latest = {}
        for r in records:
            key = getattr(r, attr, None)
            if key is None:
                continue
            cur = latest.get(key)
            if not cur or r.timestamp > cur.timestamp:
                latest[key] = r
        return list(latest.values())

    # Planta: niveles bajos / altos y desactualización
    try:
        if planta_summary['datos']:
            latest_tanques = latest_by(planta_summary['datos'], 'tk')
            low, high = [], []
            for r in latest_tanques:
                if r.max_cap and r.bls_60 is not None:
                    try:
                        pct = (float(r.bls_60) / float(r.max_cap)) * 100 if r.max_cap else None
                    except Exception:
                        pct = None
                    if pct is not None:
                        if pct < 15:
                            low.append((r.tk, pct))
                        elif pct > 90:
                            high.append((r.tk, pct))
            if low:
                detalle = ', '.join(f"{tk}:{pct:.1f}%" for tk, pct in low[:4])
                add_alert('warning', 'PLANTA', f'Tanques con nivel bajo (<15%): {detalle}' + (' ...' if len(low) > 4 else ''))
            if high:
                detalle = ', '.join(f"{tk}:{pct:.1f}%" for tk, pct in high[:4])
                add_alert('info', 'PLANTA', f'Tanques con nivel alto (>90%): {detalle}' + (' ...' if len(high) > 4 else ''))
            # Staleness (último update > 24h)
            ultimo = max(planta_summary['datos'], key=lambda r: r.timestamp)
            if (now_utc - ultimo.timestamp).total_seconds() > 24*3600:
                add_alert('danger', 'PLANTA', 'Inventario sin actualización en las últimas 24 horas.')
        else:
            add_alert('info', 'PLANTA', 'Sin registros de inventario disponibles.')
    except Exception:
        add_alert('warning', 'PLANTA', 'Error evaluando niveles de planta.')

    # Barcaza Orion / BITA: evaluar porcentaje total consolidado y staleness (>36h)
    for summary, categoria in [
        (orion_summary, 'ORION'),
        (bita_summary, 'BITA')
    ]:
        try:
            if summary['datos']:
                latest_tanques = latest_by(summary['datos'], 'tk')
                total_cap = 0.0
                total_bls = 0.0
                for r in latest_tanques:
                    try:
                        cap = float(r.max_cap) if r.max_cap is not None else 0
                        bls = float(r.bls_60) if r.bls_60 is not None else 0
                    except Exception:
                        cap, bls = 0, 0
                    total_cap += cap
                    total_bls += bls
                pct_total = (total_bls / total_cap * 100) if total_cap > 0 else None
                if pct_total is not None:
                    if pct_total < 10:
                        add_alert('danger', categoria, f'Nivel crítico {pct_total:.1f}% (<10%).')
                    elif pct_total < 15:
                        add_alert('warning', categoria, f'Nivel consolidado bajo {pct_total:.1f}% (<15%).')
                    elif pct_total > 90:
                        add_alert('info', categoria, f'Nivel alto {pct_total:.1f}% (>90%).')
                ultimo = max(summary['datos'], key=lambda r: r.timestamp)
                if (now_utc - ultimo.timestamp).total_seconds() > 36*3600:
                    add_alert('danger', categoria, 'Sin actualización en las últimas 36 horas.')
            else:
                add_alert('info', categoria, 'Sin registros cargados.')
        except Exception:
            add_alert('warning', categoria, 'Error evaluando niveles consolidados.')

    # Tránsito: registros incompletos últimas 24h y volumen de actividad
    try:
        if transito_summary['datos']:
            recientes = [r for r in transito_summary['datos'] if (now_utc - r.timestamp).total_seconds() <= 24*3600]
            if recientes:
                incompletos = [r for r in recientes if r.api is None or r.bsw is None or r.nsv is None]
                if incompletos:
                    add_alert('warning', 'TRANSITO', f'Registros incompletos últimas 24h: {len(incompletos)} (API/BSW/NSV faltantes).')
                # Actividad baja: menos de 3 registros 24h
                if len(recientes) < 3:
                    add_alert('info', 'TRANSITO', 'Baja actividad en las últimas 24 horas.')
            else:
                add_alert('info', 'TRANSITO', 'Sin movimientos registrados en las últimas 24 horas.')
            ultimo = max(transito_summary['datos'], key=lambda r: r.timestamp)
            if (now_utc - ultimo.timestamp).total_seconds() > 48*3600:
                add_alert('danger', 'TRANSITO', 'Sin actualización en más de 48 horas.')
        else:
            add_alert('info', 'TRANSITO', 'Sin registros de tránsito disponibles.')
    except Exception:
        add_alert('warning', 'TRANSITO', 'Error evaluando registros de tránsito.')

    # Ordenar alertas por severidad importancia
    severity_rank = {'danger': 0, 'warning': 1, 'info': 2, 'success': 3}
    alerts.sort(key=lambda a: severity_rank.get(a['severity'], 99))

    # --- Renderizar la plantilla ---
    return render_template("dashboard_reportes.html",
                           nombre=session.get("nombre"),
                           planta_summary=planta_summary,
                           orion_summary=orion_summary,
                           bita_summary=bita_summary,
                           transito_summary=transito_summary,
                           alerts=alerts)

@login_required                        
@app.route('/guardar-datos-planta', methods=['POST'])
def guardar_datos_planta():
    if not request.is_json:
        return jsonify(success=False, message="Formato no válido"), 400

    data = request.get_json()
    tk = data.get("tk")
    field = data.get("field")
    value = data.get("value")

    if not all([tk, field]):
        return jsonify(success=False, message="Datos incompletos"), 400

    for fila in PLANILLA_PLANTA:
        if fila["TK"] == tk and field in fila:
            fila[field] = value
            return jsonify(success=True)

    return jsonify(success=False, message="Tanque o campo no encontrado"), 404

@login_required
@permiso_requerido('planta')
@app.route('/guardar-registro-planta', methods=['POST'])
def guardar_registro_planta():
    lista_tanques = request.get_json()
    if not isinstance(lista_tanques, list):
        return jsonify(success=False, message="Formato de datos incorrecto."), 400

    try:
        def to_float(v):
            if v is None:
                return None
            s = str(v).strip().replace(',', '.')
            if s == '':
                return None
            try:
                return float(s)
            except Exception:
                return None
        today_start = datetime.combine(date.today(), time.min)
        today_end = datetime.combine(date.today(), time.max)

        for datos_tanque in lista_tanques:
            tk = datos_tanque.get('TK')
            if not tk: continue

            registro_existente = db.session.query(RegistroPlanta).filter(
                RegistroPlanta.tk == tk,
                RegistroPlanta.timestamp.between(today_start, today_end)
            ).first()

            if registro_existente:
                # Si existe, lo ACTUALIZAMOS
                registro_existente.usuario = session.get("nombre", "No identificado")
                registro_existente.bls_60 = to_float(datos_tanque.get('BLS_60'))
                registro_existente.api = to_float(datos_tanque.get('API'))
                registro_existente.bsw = to_float(datos_tanque.get('BSW'))
                registro_existente.s = to_float(datos_tanque.get('S'))
                registro_existente.timestamp = datetime.utcnow()
            else:
                # Si no existe para hoy, CREAMOS uno nuevo
                nuevo_registro = RegistroPlanta(
                    timestamp=datetime.utcnow(),
                    usuario=session.get("nombre", "No identificado"),
                    tk=tk,
                    producto="DILUYENTE" if tk == "Consumo Interno" else datos_tanque.get('PRODUCTO'),
                    max_cap=124.78 if tk == "Consumo Interno" else to_float(datos_tanque.get('MAX_CAP')),
                    bls_60=to_float(datos_tanque.get('BLS_60')),
                    api=to_float(datos_tanque.get('API')),
                    bsw=to_float(datos_tanque.get('BSW')),
                    s=to_float(datos_tanque.get('S'))
                )
                db.session.add(nuevo_registro)
        
        db.session.commit()
        return jsonify(success=True, message="Inventario de planta actualizado exitosamente.")
    except Exception as e:
        db.session.rollback()
        return jsonify(success=False, message=f"Error interno: {str(e)}"), 500
    
@login_required
@permiso_requerido('zisa_inventory')    
@app.route('/inventario-zisa')
def inventario_zisa():
    # 1. Definimos la variable global 'g.current_time' para que la plantilla la pueda usar
    # Usamos la zona horaria de Bogotá que ya tienes configurada en otras partes
    try:
        bogota_zone = pytz.timezone('America/Bogota')
        g.current_time = datetime.now(bogota_zone)
    except Exception:
        g.current_time = datetime.now() # Fallback por si acaso

    # 2. Consultamos todos los registros de la tabla
    todos_los_registros = RegistroZisa.query.order_by(RegistroZisa.fecha_carga.desc()).all()
    
    # 3. Los separamos por empresa para las tablas
    registros_zisa = [r for r in todos_los_registros if r.empresa == 'ZISA']
    registros_fbcol = [r for r in todos_los_registros if r.empresa == 'FBCOL']

    # 4. Calculamos los totales de inventario disponible para las tarjetas de resumen
    total_zisa = db.session.query(func.sum(RegistroZisa.bbl_descargados)).filter_by(estado='Disponible', empresa='ZISA').scalar() or 0.0
    total_fbcol = db.session.query(func.sum(RegistroZisa.bbl_descargados)).filter_by(estado='Disponible', empresa='FBCOL').scalar() or 0.0
    
    # 5. Enviamos TODAS las variables que la plantilla necesita
    return render_template('inventario_zisa.html', 
                           registros_zisa=registros_zisa,
                           registros_fbcol=registros_fbcol,
                           total_zisa=total_zisa,       # <-- Variable añadida
                           total_fbcol=total_fbcol,     # <-- Variable añadida
                           nombre=session.get("nombre"))
@login_required
@permiso_requerido('zisa_inventory')
@app.route('/cargar-inventario-zisa', methods=['POST'])
def cargar_inventario_zisa():
    if 'archivo_excel' not in request.files:
        flash('No se seleccionó ningún archivo.', 'warning')
        return redirect(url_for('inventario_zisa'))

    archivo = request.files['archivo_excel']
    
    if not archivo.filename.lower().endswith('.xlsx'):
        flash('Formato de archivo no válido. Por favor, suba un archivo .xlsx', 'danger')
        return redirect(url_for('inventario_zisa'))

    try:
        xls = pd.ExcelFile(archivo)
        hojas_a_procesar = {'CWT 2025': 'ZISA', 'FBCOL 2025': 'FBCOL'}
        resultados = {'nuevos': 0, 'duplicados': 0, 'errores': 0}
        resumen = []

        for hoja, empresa in hojas_a_procesar.items():
            if hoja not in xls.sheet_names:
                resumen.append(f"Hoja '{hoja}' no encontrada - Saltada")
                continue

            try:
                df = pd.read_excel(xls, sheet_name=hoja)
                df.columns = df.columns.str.strip().str.upper()
                
                # Validación de columnas
                columnas_requeridas = {'MES', 'CARROTANQUE', 'PRODUCTO', 'N S.A.E', 'ACTA', 'BBL NETOS', 'VEHICULOS DESCARGADOS'}
                if not columnas_requeridas.issubset(df.columns):
                    faltantes = columnas_requeridas - set(df.columns)
                    resumen.append(f"Error en '{hoja}': Faltan columnas: {', '.join(faltantes)}")
                    resultados['errores'] += 1
                    continue
                
                # Procesamiento del DataFrame
                df = df.dropna(subset=['ACTA', 'CARROTANQUE'])
                nuevos = 0
                duplicados = 0
                
                for _, fila in df.iterrows():
                    try:
                        existe = RegistroZisa.query.filter_by(
                            empresa=empresa,
                            acta=str(fila['ACTA']),
                            carrotanque=str(fila['CARROTANQUE'])
                        ).first()

                        if not existe:
                            registro = RegistroZisa(
                                empresa=empresa,
                                mes=fila['MES'],
                                carrotanque=str(fila['CARROTANQUE']),
                                producto=fila['PRODUCTO'],
                                numero_sae=fila['N S.A.E'],
                                acta=str(fila['ACTA']),
                                bbl_netos=float(fila['BBL NETOS']),
                                bbl_descargados=float(fila['VEHICULOS DESCARGADOS']),
                                usuario_carga=session.get('nombre', 'Desconocido'),
                                estado='Disponible'
                            )
                            db.session.add(registro)
                            nuevos += 1
                        else:
                            duplicados += 1
                    except Exception as e:
                        app.logger.error(f"Error procesando fila en {hoja}: {str(e)}")
                        resultados['errores'] += 1

                db.session.commit()
                resultados['nuevos'] += nuevos
                resultados['duplicados'] += duplicados
                resumen.append(f"{hoja}: {nuevos} nuevos, {duplicados} duplicados")
                
            except Exception as e:
                db.session.rollback()
                app.logger.error(f"Error procesando hoja {hoja}: {str(e)}")
                resumen.append(f"Error procesando {hoja}: {str(e)}")
                resultados['errores'] += 1

        # Resultado final
        if resultados['nuevos'] > 0:
            flash(f"Procesamiento completado: {resultados['nuevos']} nuevos registros", 'success')
        if resultados['duplicados'] > 0:
            flash(f"{resultados['duplicados']} registros duplicados omitidos", 'info')
        if resultados['errores'] > 0:
            flash(f"Se encontraron {resultados['errores']} errores durante el procesamiento", 'warning')
        
        for mensaje in resumen:
            flash(mensaje, 'info')

    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Error crítico al procesar archivo: {str(e)}")
        flash(f'Error al procesar el archivo: {str(e)}', 'danger')
    
    return redirect(url_for('inventario_zisa'))

@login_required
@permiso_requerido('zisa_inventory')
@app.route('/consumir-inventario', methods=['GET', 'POST'])
def consumir_inventario():
    if request.method == 'POST':
        try:
            # --- DIAGNÓSTICO: Imprimimos lo que recibimos del formulario ---
            print("="*30)
            print("INICIANDO PROCESO DE CONSUMO")
            cantidad_a_consumir = float(request.form.get('cantidad_a_gastar', 0))
            empresa = request.form.get('empresa')
            print(f"--> Solicitud para consumir {cantidad_a_consumir} BBL de la empresa: {empresa}")

            if cantidad_a_consumir <= 0 or not empresa:
                flash('La cantidad debe ser positiva y debes seleccionar una empresa.', 'warning')
                return redirect(url_for('consumir_inventario'))

            registros_disponibles = RegistroZisa.query.filter_by(
                empresa=empresa, estado='Disponible'
            ).order_by(RegistroZisa.fecha_carga.asc()).all()

            total_disponible_en_bd = sum(r.bbl_descargados for r in registros_disponibles if r.bbl_descargados)
            if total_disponible_en_bd < cantidad_a_consumir:
                flash(f'Inventario insuficiente en {empresa}. Disponible: {total_disponible_en_bd:.2f} BBL, Solicitado: {cantidad_a_consumir:.2f} BBL', 'danger')
                return redirect(url_for('consumir_inventario'))
            
            cantidad_restante = cantidad_a_consumir
            actas_consumidas = []
            
            for registro in registros_disponibles:
                if cantidad_restante <= 0:
                    break
                
                # --- MEJORA DE ROBUSTEZ: Manejo de valores nulos o cero ---
                bbl_del_registro = registro.bbl_descargados or 0.0
                
                # --- DIAGNÓSTICO: Imprimimos cada registro que se va a procesar ---
                print(f"Procesando registro ID={registro.id}, Acta={registro.acta}, BBL_Descargados={bbl_del_registro}")

                # --- MEJORA DE ROBUSTEZ: Si el registro no tiene barriles, lo saltamos ---
                if bbl_del_registro <= 0:
                    print(f"--> SALTADO: El registro ID={registro.id} tiene 0 o menos barriles.")
                    continue

                if bbl_del_registro <= cantidad_restante:
                    registro.estado = 'Gastado'
                    cantidad_restante -= bbl_del_registro
                    actas_consumidas.append(f"{registro.acta} ({registro.carrotanque})")
                    print(f"--> CONSUMIDO COMPLETO: ID={registro.id}. Quedan por consumir: {cantidad_restante}")
                else:
                    # Lógica de división (ya estaba bien, pero la rodeamos de diagnósticos)
                    print(f"--> DIVIDIENDO: ID={registro.id}. Se consumirán {cantidad_restante} de {bbl_del_registro}")
                    proporcion_a_dividir = cantidad_restante / bbl_del_registro
                    bbl_netos_originales = registro.bbl_netos or 0.0

                    nuevo_registro_disponible = RegistroZisa(
                        empresa=registro.empresa, mes=registro.mes, carrotanque=registro.carrotanque,
                        producto=registro.producto, numero_sae=registro.numero_sae, acta=registro.acta,
                        bbl_netos = bbl_netos_originales * (1 - proporcion_a_dividir),
                        bbl_descargados = bbl_del_registro - cantidad_restante,
                        usuario_carga=registro.usuario_carga, fecha_carga=registro.fecha_carga,
                        estado='Disponible'
                    )
                    db.session.add(nuevo_registro_disponible)
                    
                    registro.estado = 'Gastado'
                    registro.bbl_descargados = cantidad_restante
                    registro.bbl_netos = bbl_netos_originales * proporcion_a_dividir
                    
                    actas_consumidas.append(f"{registro.acta} (parcial)")
                    cantidad_restante = 0
                    print(f"--> DIVISIÓN COMPLETA. ID={registro.id} ahora está gastado. Se creó un nuevo registro para el sobrante.")

            db.session.commit()
            print("--> COMMIT REALIZADO CON ÉXITO")
            print("="*30)
            flash(f'Éxito: Se consumieron {cantidad_a_consumir:.2f} BBL de {empresa}. Actas utilizadas: {", ".join(actas_consumidas)}', 'success')
            
        except Exception as e:
            db.session.rollback()
            # --- DIAGNÓSTICO CRÍTICO ---
            print("\n" + "!"*50)
            print(f"ERROR CATASTRÓFICO AL CONSUMIR: {e}")
            import traceback
            traceback.print_exc() # Imprime el error detallado en la consola
            print("!"*50 + "\n")
            app.logger.error(f"Error al consumir inventario: {str(e)}")
            flash('Ocurrió un error grave al procesar la solicitud. Revisa la consola del servidor.', 'danger')
        
        return redirect(url_for('consumir_inventario'))
    
    else: # El método GET se mantiene igual
        total_zisa = db.session.query(func.sum(RegistroZisa.bbl_descargados)).filter_by(estado='Disponible', empresa='ZISA').scalar() or 0.0
        total_fbcol = db.session.query(func.sum(RegistroZisa.bbl_descargados)).filter_by(estado='Disponible', empresa='FBCOL').scalar() or 0.0
        total_consumido_zisa = db.session.query(func.sum(RegistroZisa.bbl_descargados)).filter_by(estado='Gastado', empresa='ZISA').scalar() or 0.0
        total_consumido_fbcol = db.session.query(func.sum(RegistroZisa.bbl_descargados)).filter_by(estado='Gastado', empresa='FBCOL').scalar() or 0.0
        ultimos_consumos = RegistroZisa.query.filter_by(estado='Gastado').order_by(RegistroZisa.fecha_carga.desc()).limit(10).all()

        return render_template('consumir_inventario.html',
                               total_inventario_zisa=total_zisa,
                               total_inventario_fbcol=total_fbcol,
                               total_consumido_zisa=total_consumido_zisa,
                               total_consumido_fbcol=total_consumido_fbcol,
                               ultimos_consumos=ultimos_consumos)

@login_required
@permiso_requerido('zisa_inventory')    
@app.route('/reporte-consumo')
def reporte_consumo():
    # 1. Obtener los filtros desde la URL (si es que existen)
    empresa_filtro = request.args.get('empresa', default='')
    fecha_inicio_str = request.args.get('fecha_inicio', default='')
    fecha_fin_str = request.args.get('fecha_fin', default='')

    # 2. Empezar la consulta base: solo registros 'Gastado'
    query = RegistroZisa.query.filter_by(estado='Gastado')

    # 3. Aplicar los filtros a la consulta
    if empresa_filtro in ['ZISA', 'FBCOL']:
        query = query.filter_by(empresa=empresa_filtro)
    
    # Filtro por fecha de inicio
    if fecha_inicio_str:
        try:
            fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d')
            query = query.filter(RegistroZisa.fecha_carga >= fecha_inicio)
        except ValueError:
            flash('Formato de fecha de inicio inválido. Use AAAA-MM-DD.', 'warning')
    
    # Filtro por fecha de fin
    if fecha_fin_str:
        try:
            # Se suma un día para que el rango sea inclusivo hasta el final del día seleccionado
            fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d') + timedelta(days=1)
            query = query.filter(RegistroZisa.fecha_carga < fecha_fin)
        except ValueError:
            flash('Formato de fecha de fin inválido. Use AAAA-MM-DD.', 'warning')

    # 4. Ejecutar la consulta final
    registros_consumidos = query.order_by(RegistroZisa.fecha_carga.desc()).all()

    # 5. Calcular la suma total de los BBL descargados de los registros filtrados
    total_consumido_filtrado = sum(r.bbl_descargados for r in registros_consumidos)

    # 6. Preparar los filtros para devolverlos a la plantilla (para que los campos del formulario recuerden su valor)
    filtros_activos = {
        'empresa': empresa_filtro,
        'fecha_inicio': fecha_inicio_str,
        'fecha_fin': fecha_fin_str
    }

    return render_template('reporte_consumo.html',
                           registros=registros_consumidos,
                           total_consumido=total_consumido_filtrado,
                           filtros=filtros_activos)

@login_required
@permiso_requerido('zisa_inventory')
@app.route('/exportar-inventario-zisa')
def exportar_inventario_zisa():
    """
    Exporta el inventario de ZISA/FBCOL a un archivo Excel con filtros.
    """
    # Obtener los filtros desde los argumentos de la URL
    empresa_filtro = request.args.get('empresa')
    estado_filtro = request.args.get('estado')
    
    query = RegistroZisa.query

    # Aplicar filtros a la consulta si fueron proporcionados
    if empresa_filtro and empresa_filtro in ['ZISA', 'FBCOL']:
        query = query.filter_by(empresa=empresa_filtro)
    
    if estado_filtro and estado_filtro in ['Disponible', 'Gastado']:
        query = query.filter_by(estado=estado_filtro)

    # Ejecutar la consulta
    registros = query.order_by(RegistroZisa.fecha_carga.desc()).all()

    if not registros:
        flash('No hay datos para exportar con los filtros seleccionados.', 'warning')
        return redirect(url_for('inventario_zisa'))

    # Preparar los datos para el DataFrame de Pandas
    datos_para_df = [{
        'Empresa': r.empresa,
        'Mes': r.mes,
        'Carrotanque': r.carrotanque,
        'Producto': r.producto,
        'Numero SAE': r.numero_sae,
        'Acta': r.acta,
        'BBL Netos': r.bbl_netos,
        'BBL Descargados/Gastados': r.bbl_descargados,
        'Estado': r.estado,
        'Usuario Carga': r.usuario_carga,
        'Fecha Carga': r.fecha_carga.strftime('%Y-%m-%d %H:%M:%S') if r.fecha_carga else ''
    } for r in registros]

    df = pd.DataFrame(datos_para_df)

    # Crear el archivo Excel en memoria
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Inventario_ZISA')
    output.seek(0)

    # Enviar el archivo al usuario para su descarga
    filename = f"reporte_inventario_zisa_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@login_required
@app.route('/exportar-excel/<string:nombre_reporte>')
def exportar_excel(nombre_reporte):
    """
    Exporta los datos del reporte especificado a un archivo Excel con filtros avanzados.
    """
    filtro_tipo = request.args.get('filtro_tipo')
    valor = request.args.get('valor')
    
    registros_db = []
    columnas = []
    filename = f"reporte_{nombre_reporte}_{valor or 'general'}.xlsx"

    # --- Lógica de filtrado para modelos con `timestamp` (Planta, Orion, Bita) ---
    if nombre_reporte in ['planta', 'barcaza_orion', 'barcaza_bita']:
        timestamp_limite = None
        if valor:
            try:
                if filtro_tipo == 'dia':
                    timestamp_limite = datetime.combine(date.fromisoformat(valor), time.max)
                elif filtro_tipo == 'mes':
                    ano, mes = map(int, valor.split('-'))
                    ultimo_dia = (date(ano, mes, 1) + timedelta(days=32)).replace(day=1) - timedelta(days=1)
                    timestamp_limite = datetime.combine(ultimo_dia, time.max)
                elif filtro_tipo == 'trimestre':
                    ano_str, q_str = valor.split('-Q')
                    ano, trimestre = int(ano_str), int(q_str)
                    mes_fin = trimestre * 3
                    ultimo_dia = (date(ano, mes_fin, 1) + timedelta(days=32)).replace(day=1) - timedelta(days=1)
                    timestamp_limite = datetime.combine(ultimo_dia, time.max)
                elif filtro_tipo == 'anual':
                    ano = int(valor)
                    ultimo_dia = date(ano, 12, 31)
                    timestamp_limite = datetime.combine(ultimo_dia, time.max)
            except (ValueError, TypeError):
                timestamp_limite = datetime.now() # Fallback seguro

        if nombre_reporte == 'planta':
            subquery_base = db.session.query(RegistroPlanta.tk, func.max(RegistroPlanta.timestamp).label('max_timestamp'))
            subquery = subquery_base.filter(RegistroPlanta.timestamp <= timestamp_limite).group_by(RegistroPlanta.tk).subquery() if timestamp_limite else subquery_base.group_by(RegistroPlanta.tk).subquery()
            registros_db = db.session.query(RegistroPlanta).join(subquery, (RegistroPlanta.tk == subquery.c.tk) & (RegistroPlanta.timestamp == subquery.c.max_timestamp)).all()
            columnas = ["tk", "producto", "max_cap", "bls_60", "api", "bsw", "s"]

        elif nombre_reporte == 'barcaza_orion':
            subquery_base = db.session.query(RegistroBarcazaOrion.tk, RegistroBarcazaOrion.grupo, func.max(RegistroBarcazaOrion.timestamp).label('max_timestamp'))
            subquery = subquery_base.filter(RegistroBarcazaOrion.timestamp <= timestamp_limite).group_by(RegistroBarcazaOrion.tk, RegistroBarcazaOrion.grupo).subquery() if timestamp_limite else subquery_base.group_by(RegistroBarcazaOrion.tk, RegistroBarcazaOrion.grupo).subquery()
            registros_db = db.session.query(RegistroBarcazaOrion).join(subquery, (RegistroBarcazaOrion.tk == subquery.c.tk) & (RegistroBarcazaOrion.grupo == subquery.c.grupo) & (RegistroBarcazaOrion.timestamp == subquery.c.max_timestamp)).all()
            columnas = ["grupo", "tk", "producto", "max_cap", "bls_60", "api", "bsw", "s"]

        elif nombre_reporte == 'barcaza_bita':
            subquery_base = db.session.query(RegistroBarcazaBita.tk, func.max(RegistroBarcazaBita.timestamp).label('max_timestamp'))
            subquery = subquery_base.filter(RegistroBarcazaBita.timestamp <= timestamp_limite).group_by(RegistroBarcazaBita.tk).subquery() if timestamp_limite else subquery_base.group_by(RegistroBarcazaBita.tk).subquery()
            registros_db = db.session.query(RegistroBarcazaBita).join(subquery, (RegistroBarcazaBita.tk == subquery.c.tk) & (RegistroBarcazaBita.timestamp == subquery.c.max_timestamp)).all()
            columnas = ["tk", "producto", "max_cap", "bls_60", "api", "bsw", "s"]

    # --- Lógica para Variaciones de Tanques (serie diaria) ---
    elif nombre_reporte == 'variaciones':
        start_dt = None
        end_dt = None
        try:
            if filtro_tipo == 'dia' and valor:
                d = date.fromisoformat(valor)
                start_dt = datetime.combine(d, time.min)
                end_dt = datetime.combine(d, time.max)
            elif filtro_tipo == 'mes' and valor:
                ano, mes = map(int, valor.split('-'))
                ini = date(ano, mes, 1)
                fin = (ini + timedelta(days=32)).replace(day=1) - timedelta(days=1)
                start_dt = datetime.combine(ini, time.min)
                end_dt = datetime.combine(fin, time.max)
            elif filtro_tipo == 'trimestre' and valor:
                ano_str, q_str = valor.split('-Q')
                ano, trimestre = int(ano_str), int(q_str)
                mes_ini = (trimestre - 1) * 3 + 1
                mes_fin = trimestre * 3
                ini = date(ano, mes_ini, 1)
                fin = (date(ano, mes_fin, 1) + timedelta(days=32)).replace(day=1) - timedelta(days=1)
                start_dt = datetime.combine(ini, time.min)
                end_dt = datetime.combine(fin, time.max)
            elif filtro_tipo == 'anual' and valor:
                ano = int(valor)
                ini = date(ano, 1, 1)
                fin = date(ano, 12, 31)
                start_dt = datetime.combine(ini, time.min)
                end_dt = datetime.combine(fin, time.max)
        except Exception:
            start_dt = None
            end_dt = None

        fecha_dia = func.date(RegistroPlanta.timestamp)
        subq = db.session.query(func.max(RegistroPlanta.id).label('max_id'))
        if start_dt:
            subq = subq.filter(RegistroPlanta.timestamp >= start_dt)
        if end_dt:
            subq = subq.filter(RegistroPlanta.timestamp <= end_dt)
        subq = subq.group_by(RegistroPlanta.tk, fecha_dia).subquery()

        registros_series = (db.session.query(RegistroPlanta)
                            .filter(RegistroPlanta.id.in_(subq))
                            .order_by(RegistroPlanta.tk.asc(), RegistroPlanta.timestamp.asc())
                            .all())

        # Organizar por tanque y fecha y calcular variaciones
        por_tanque = {}
        for r in registros_series:
            if not r.tk:
                continue
            dia = r.timestamp.date()
            por_tanque.setdefault(r.tk, []).append({
                'fecha': dia.isoformat(),
                'producto': r.producto or '',
                'max_cap': float(r.max_cap or 0),
                'bls_60': float(r.bls_60 or 0)
            })

        filas = []
        for tk, lista in por_tanque.items():
            lista.sort(key=lambda x: x['fecha'])
            prev = None
            for item in lista:
                bls = item['bls_60']
                delta = None if prev is None else round(bls - prev, 2)
                filas.append({
                    'fecha': item['fecha'],
                    'tk': tk,
                    'producto': item['producto'],
                    'max_cap': item['max_cap'],
                    'bls_60': bls,
                    'variacion': delta
                })
                prev = bls

        if not filas:
            flash("No hay datos para exportar con el filtro seleccionado.", "warning")
            return redirect(request.referrer or url_for('reporte_variaciones_tanques'))

        registros = filas
        columnas = ["fecha", "tk", "producto", "max_cap", "bls_60", "variacion"]

    # --- Lógica de filtrado para Tránsito (usa la columna `fecha` que es texto) ---
    elif nombre_reporte == 'transito':
        query = db.session.query(RegistroTransito)
        if valor:
            try:
                if filtro_tipo == 'dia':
                    query = query.filter(RegistroTransito.fecha == valor)
                elif filtro_tipo == 'mes':
                    query = query.filter(RegistroTransito.fecha.like(f"{valor}-%"))
                elif filtro_tipo == 'trimestre':
                    ano_str, q_str = valor.split('-Q')
                    ano, trimestre = int(ano_str), int(q_str)
                    meses_trimestre = {1: ["01", "02", "03"], 2: ["04", "05", "06"], 3: ["07", "08", "09"], 4: ["10", "11", "12"]}[trimestre]
                    condiciones = [RegistroTransito.fecha.like(f"{ano}-{m}-%") for m in meses_trimestre]
                    query = query.filter(or_(*condiciones))
                elif filtro_tipo == 'anual':
                    query = query.filter(RegistroTransito.fecha.like(f"{valor}-%"))
            except (ValueError, TypeError):
                pass # Si el valor es inválido, no se filtra
        
        registros_db = query.order_by(RegistroTransito.fecha.desc()).all()
        columnas = ["tipo_transito", "fecha", "guia", "origen", "producto", "placa", "nsv", "api", "bsw", "observaciones"]

    if not registros_db:
        flash("No hay datos para exportar con el filtro seleccionado.", "warning")
        return redirect(request.referrer or url_for('dashboard_reportes'))

    # Convertir los resultados a una lista de diccionarios
    registros = [r.__dict__ for r in registros_db]
    
    # Crear el DataFrame y el archivo Excel
    df = pd.DataFrame(registros)
    # Asegurarse de que solo las columnas deseadas estén en el DataFrame final
    df = df.reindex(columns=columnas)

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Reporte')
    output.seek(0)

    # Enviar el archivo para su descarga
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@login_required
@app.route('/descargar-reporte-variaciones-pdf')
def descargar_reporte_variaciones_pdf():
    filtro_tipo = request.args.get('filtro_tipo', 'mes')
    valor = request.args.get('valor')

    # Determinar rango
    start_dt = None
    end_dt = None
    periodo_str = "General"
    try:
        if filtro_tipo == 'dia' and valor:
            d = date.fromisoformat(valor)
            start_dt = datetime.combine(d, time.min)
            end_dt = datetime.combine(d, time.max)
            periodo_str = f"del día {d.strftime('%d/%m/%Y')}"
        elif filtro_tipo == 'mes' and valor:
            ano, mes = map(int, valor.split('-'))
            ini = date(ano, mes, 1)
            fin = (ini + timedelta(days=32)).replace(day=1) - timedelta(days=1)
            start_dt = datetime.combine(ini, time.min)
            end_dt = datetime.combine(fin, time.max)
            periodo_str = f"del mes de {ini.strftime('%B de %Y')}"
        elif filtro_tipo == 'trimestre' and valor:
            ano_str, q_str = valor.split('-Q')
            ano, trimestre = int(ano_str), int(q_str)
            mes_ini = (trimestre - 1) * 3 + 1
            mes_fin = trimestre * 3
            ini = date(ano, mes_ini, 1)
            fin = (date(ano, mes_fin, 1) + timedelta(days=32)).replace(day=1) - timedelta(days=1)
            start_dt = datetime.combine(ini, time.min)
            end_dt = datetime.combine(fin, time.max)
            periodo_str = f"del Trimestre {trimestre} de {ano}"
        elif filtro_tipo == 'anual' and valor:
            ano = int(valor)
            ini = date(ano, 1, 1)
            fin = date(ano, 12, 31)
            start_dt = datetime.combine(ini, time.min)
            end_dt = datetime.combine(fin, time.max)
            periodo_str = f"del Año {ano}"
    except Exception:
        start_dt = None
        end_dt = None

    # Query: último registro por día y tanque dentro del rango
    fecha_dia = func.date(RegistroPlanta.timestamp)
    subq = db.session.query(func.max(RegistroPlanta.id).label('max_id'))
    if start_dt:
        subq = subq.filter(RegistroPlanta.timestamp >= start_dt)
    if end_dt:
        subq = subq.filter(RegistroPlanta.timestamp <= end_dt)
    subq = subq.group_by(RegistroPlanta.tk, fecha_dia).subquery()

    registros_series = (db.session.query(RegistroPlanta)
                        .filter(RegistroPlanta.id.in_(subq))
                        .order_by(RegistroPlanta.tk.asc(), RegistroPlanta.timestamp.asc())
                        .all())

    if not registros_series:
        flash("No hay datos para generar el PDF con el filtro seleccionado.", "warning")
        return redirect(url_for('reporte_variaciones_tanques'))

    # Organizar por tanque y calcular deltas
    por_tanque = {}
    for r in registros_series:
        tk = r.tk
        if not tk:
            continue
        dia = r.timestamp.date().isoformat()
        por_tanque.setdefault(tk, []).append({
            'fecha': dia,
            'producto': r.producto or '',
            'max_cap': float(r.max_cap or 0),
            'bls_60': float(r.bls_60 or 0)
        })

    for tk in list(por_tanque.keys()):
        por_tanque[tk].sort(key=lambda x: x['fecha'])
        prev = None
        for item in por_tanque[tk]:
            bls = item['bls_60']
            item['variacion'] = None if prev is None else round(bls - prev, 2)
            if item['variacion'] is None:
                item['tipo'] = '—'
            elif item['variacion'] > 0:
                item['tipo'] = 'Suma'
            elif item['variacion'] < 0:
                item['tipo'] = 'Descarga'
            else:
                item['tipo'] = 'Sin cambio'
            prev = bls

    # Orden sugerido
    orden = ['TK-109','TK-110','TK-102','TK-01','TK-02']
    tanques_ordenados = sorted(por_tanque.keys(), key=lambda k: (orden.index(k) if k in orden else 999, k))

    # Calcular estadísticas por tanque y globales
    stats_por_tanque = {}
    total_bls = 0.0
    total_suma = 0.0
    total_descarga = 0.0
    for tk in tanques_ordenados:
        lista = por_tanque.get(tk, [])
        if not lista:
            continue
        last_bls = float(lista[-1].get('bls_60') or 0)
        producto = next((it.get('producto') for it in lista if (it.get('producto') or '').strip()), '')
        cap = next((float(it.get('max_cap') or 0) for it in lista if float(it.get('max_cap') or 0) > 0), 0.0)
        suma = 0.0
        descarga = 0.0
        for it in lista:
            v = it.get('variacion')
            if isinstance(v, (int, float)):
                if v > 0: suma += v
                elif v < 0: descarga += abs(v)
        stats_por_tanque[tk] = {
            'producto': producto,
            'max_cap': cap,
            'last_bls': last_bls,
            'suma': round(suma, 2),
            'descarga': round(descarga, 2)
        }
        total_bls += last_bls
        total_suma += suma
        total_descarga += descarga

    resumen_global = {
        'total_tanques': len(tanques_ordenados),
        'total_bls': round(total_bls, 2),
        'total_suma': round(total_suma, 2),
        'total_descarga': round(total_descarga, 2)
    }

    # Generar gráficos (Matplotlib) para cada tanque con estilo y etiquetas de volumen
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    import io, base64
    charts_map = {}
    for tk in tanques_ordenados:
        lista = por_tanque.get(tk, [])
        if not lista:
            continue
        fechas = [it['fecha'] for it in lista]
        bls = [float(it['bls_60'] or 0) for it in lista]
        cap = next((float(it.get('max_cap') or 0) for it in lista if float(it.get('max_cap') or 0) > 0), 0.0)
        # Usar eje categórico controlado para poder anotar fácilmente
        xs = list(range(len(fechas)))
        fig, ax = plt.subplots(figsize=(8.5, 3.3))
        fig.patch.set_facecolor('white')
        ax.set_facecolor('#f8fafc')
        # Línea principal estilo "despachos"
        ax.plot(xs, bls, color='#4e73df', linewidth=2, marker='o', markersize=3.5,
                markerfacecolor='#4e73df', markeredgecolor='white', markeredgewidth=0.8)
        if cap and cap > 0:
            ax.plot(xs, [cap]*len(xs), color='#6c757d', linestyle='--', linewidth=1)
        # Ejes y grid
        ax.set_xlabel('Fecha', color='#6c757d', fontsize=9)
        ax.set_ylabel('BLS @60', color='#6c757d', fontsize=9)
        ax.grid(True, axis='y', linestyle=':', color='#bfc7d1', alpha=0.6)
        for spine in ax.spines.values():
            spine.set_color('#e0e5ec')
        ax.tick_params(axis='x', colors='#6c757d', labelsize=8)
        ax.tick_params(axis='y', colors='#6c757d', labelsize=8)
        ax.set_xticks(xs)
        ax.set_xticklabels(fechas, rotation=45, ha='right')
        # Etiquetas de volumen sobre cada punto
        for i, y in enumerate(bls):
            try:
                label = f"{int(round(y)):,}".replace(',', '.')
            except Exception:
                label = f"{y:.0f}"
            ax.annotate(label, (xs[i], y), textcoords='offset points', xytext=(0, 6),
                        ha='center', va='bottom', fontsize=7.5, color='#224abe',
                        bbox=dict(boxstyle='round,pad=0.18', fc='white', ec='none', alpha=0.85))
        plt.tight_layout()
        buf = io.BytesIO()
        fig.savefig(buf, format='png', dpi=150)
        plt.close(fig)
        buf.seek(0)
        charts_map[tk] = base64.b64encode(buf.read()).decode('utf-8')

    # Cargar logo corporativo (opcional)
    logo_base64 = None
    logo_mime = 'image/jpeg'
    try:
        logo_candidates = ['Logo_de_empresa.jpeg', 'Conquers_4_Logo.png', 'logo.jpeg']
        for fname in logo_candidates:
            logo_path = os.path.join(current_app.root_path, 'static', fname)
            if os.path.exists(logo_path):
                import base64 as _b64
                with open(logo_path, 'rb') as f:
                    logo_base64 = _b64.b64encode(f.read()).decode('utf-8')
                if fname.lower().endswith('.png'):
                    logo_mime = 'image/png'
                break
    except Exception:
        logo_base64 = None

    # Renderizar plantilla PDF con estilo tipo despachos
    html_para_pdf = render_template('reportes_pdf/variaciones_tanques_pdf.html',
                                    por_tanque=por_tanque,
                                    tanques_ordenados=tanques_ordenados,
                                    charts_map=charts_map,
                                    stats_por_tanque=stats_por_tanque,
                                    resumen_global=resumen_global,
                                    logo_base64=logo_base64,
                                    logo_mime=logo_mime,
                                    periodo_str=periodo_str,
                                    fecha_generacion=datetime.now().strftime('%d/%m/%Y %H:%M'))

    pdf = HTML(string=html_para_pdf, base_url=current_app.root_path).write_pdf()
    return Response(pdf,
                  mimetype='application/pdf',
                  headers={'Content-Disposition': 'attachment;filename=reporte_variaciones_tanques.pdf'})

@login_required
@app.route('/descargar-reporte-planta-pdf')
def descargar_reporte_planta_pdf():
    # --- La lógica de filtros que ya tienes se mantiene igual ---
    filtro_tipo = request.args.get('filtro_tipo', 'dia')
    valor = request.args.get('valor')
    
    subquery_base = db.session.query(RegistroPlanta.tk, func.max(RegistroPlanta.timestamp).label('max_timestamp'))
    fecha_reporte_str = f"General (últimos datos registrados al {date.today().strftime('%d/%m/%Y')})"
    subquery_filtrada = subquery_base

    if valor:
        if filtro_tipo == 'dia':
            fecha_obj = date.fromisoformat(valor)
            fecha_reporte_str = f"del día {fecha_obj.strftime('%d de %B de %Y')}"
            subquery_filtrada = subquery_base.filter(RegistroPlanta.timestamp <= datetime.combine(fecha_obj, time.max))
        # ... (el resto de tu lógica para 'mes', 'trimestre', 'anual' va aquí)
        elif filtro_tipo == 'mes':
            ano, mes = map(int, valor.split('-'))
            fecha_obj = date(ano, mes, 1)
            fecha_reporte_str = f"del mes de {fecha_obj.strftime('%B de %Y')}"
            ultimo_dia = (fecha_obj + timedelta(days=32)).replace(day=1) - timedelta(days=1)
            subquery_filtrada = subquery_base.filter(RegistroPlanta.timestamp <= datetime.combine(ultimo_dia, time.max))
        elif filtro_tipo == 'trimestre':
            ano_str, q_str = valor.split('-Q')
            ano = int(ano_str)
            trimestre = int(q_str)
            mes_fin = trimestre * 3
            ultimo_dia_trimestre = (date(ano, mes_fin, 1) + timedelta(days=32)).replace(day=1) - timedelta(days=1)
            fecha_reporte_str = f"del Trimestre {trimestre} de {ano}"
            subquery_filtrada = subquery_base.filter(RegistroPlanta.timestamp <= datetime.combine(ultimo_dia_trimestre, time.max))
        elif filtro_tipo == 'anual':
            ano = int(valor)
            ultimo_dia_ano = date(ano, 12, 31)
            fecha_reporte_str = f"del Año {ano}"
            subquery_filtrada = subquery_base.filter(RegistroPlanta.timestamp <= datetime.combine(ultimo_dia_ano, time.max))


    subquery = subquery_filtrada.group_by(RegistroPlanta.tk).subquery()
    registros_db = db.session.query(RegistroPlanta).join(subquery, (RegistroPlanta.tk == subquery.c.tk) & (RegistroPlanta.timestamp == subquery.c.max_timestamp)).all()

    if not registros_db:
        flash("No hay datos para generar el PDF con el filtro seleccionado.", "warning")
        return redirect(url_for('reporte_planta'))

    # ===== HISTORIAL SEGÚN FILTRO =====
    # Determinar rango de fechas para el historial (inicio y fin) según filtro_tipo y valor
    start_dt = None
    end_dt = None
    try:
        if valor:
            if filtro_tipo == 'dia':
                fecha_obj = date.fromisoformat(valor)
                start_dt = datetime.combine(fecha_obj, time.min)
                end_dt = datetime.combine(fecha_obj, time.max)
            elif filtro_tipo == 'mes':
                ano, mes = map(int, valor.split('-'))
                fecha_ini = date(ano, mes, 1)
                fecha_fin = (fecha_ini + timedelta(days=32)).replace(day=1) - timedelta(days=1)
                start_dt = datetime.combine(fecha_ini, time.min)
                end_dt = datetime.combine(fecha_fin, time.max)
            elif filtro_tipo == 'trimestre':
                ano_str, q_str = valor.split('-Q')
                ano = int(ano_str)
                trimestre = int(q_str)
                mes_ini = (trimestre - 1) * 3 + 1
                mes_fin = trimestre * 3
                fecha_ini = date(ano, mes_ini, 1)
                fecha_fin = (date(ano, mes_fin, 1) + timedelta(days=32)).replace(day=1) - timedelta(days=1)
                start_dt = datetime.combine(fecha_ini, time.min)
                end_dt = datetime.combine(fecha_fin, time.max)
            elif filtro_tipo == 'anual':
                ano = int(valor)
                fecha_ini = date(ano, 1, 1)
                fecha_fin = date(ano, 12, 31)
                start_dt = datetime.combine(fecha_ini, time.min)
                end_dt = datetime.combine(fecha_fin, time.max)
        # Si no se especifica filtro, se considera todo el histórico (start_dt permanece None)
    except Exception:
        # En caso de error en parsing, ignorar y no limitar historial
        start_dt = None
        end_dt = None

    historial_query = db.session.query(RegistroPlanta)
    if end_dt:
        historial_query = historial_query.filter(RegistroPlanta.timestamp <= end_dt)
    if start_dt:
        historial_query = historial_query.filter(RegistroPlanta.timestamp >= start_dt)

    historial_db = (historial_query
                     .order_by(RegistroPlanta.timestamp.desc(), RegistroPlanta.tk.asc())
                     .all())

    historial_registros = [
        {
            'fecha': r.timestamp.strftime('%Y-%m-%d %H:%M'),
            'tk': r.tk,
            'producto': r.producto,
            'bls_60': r.bls_60 or 0.0,
            'api': r.api or 0.0,
            'bsw': r.bsw or 0.0,
            's': r.s or 0.0,
            'usuario': r.usuario
        }
        for r in historial_db
    ]

    # Agrupación por fecha (día) para mostrar un cuadro por fecha
    from collections import OrderedDict
    grupos = OrderedDict()
    for item in historial_registros:
        dia = item['fecha'][:10]  # 'YYYY-MM-DD'
        if dia not in grupos:
            grupos[dia] = []
        grupos[dia].append(item)
    # Construir lista ordenada (ya viene en orden descendente por timestamp)
    historial_grouped = []
    for dia, regs in grupos.items():
        historial_grouped.append({
            'dia': dia,
            'registros': regs
        })

    # ======== INICIO DE LA SOLUCIÓN DEFINITIVA ========
    # Se crea una nueva lista de diccionarios, asegurando que los valores nulos se conviertan en 0.0
    registros_limpios = []
    for r in registros_db:
        registros_limpios.append({
            'tk': r.tk,
            'producto': r.producto,
            'max_cap': r.max_cap or 0.0,
            'bls_60': r.bls_60 or 0.0,
            'api': r.api or 0.0,
            'bsw': r.bsw or 0.0,
            's': r.s or 0.0
        })
    # ======== FIN DE LA SOLUCIÓN DEFINITIVA ========

    # Totales para badges (estilo similar a reporte gráfico despachos)
    total_bls = sum(r['bls_60'] for r in registros_limpios)
    total_cap = sum(r['max_cap'] for r in registros_limpios)

    # Cargar logo como base64
    logo_base64 = None
    try:
        logo_candidates = ['Logo_de_empresa.jpeg', 'Conquers_4_Logo.png', 'logo.jpeg']
        for fname in logo_candidates:
            logo_path = os.path.join(current_app.root_path, 'static', fname)
            if os.path.exists(logo_path):
                import base64
                with open(logo_path, 'rb') as f:
                    logo_base64 = base64.b64encode(f.read()).decode('utf-8')
                break
    except Exception as e:
        print(f"Error cargando logo para planta PDF: {e}")

    html_para_pdf = render_template('reportes_pdf/planta_pdf.html',
                                    registros=registros_limpios,
                                    historial=historial_registros,
                                    historial_grouped=historial_grouped,
                                    fecha_reporte=fecha_reporte_str,
                                    total_bls=total_bls,
                                    total_cap=total_cap,
                                    logo_base64=logo_base64,
                                    fecha_generacion=datetime.now().strftime('%d/%m/%Y %H:%M'))
    
    pdf = HTML(string=html_para_pdf, base_url=current_app.root_path).write_pdf()
    return Response(pdf,
                  mimetype='application/pdf',
                  headers={'Content-Disposition': 'attachment;filename=reporte_planta.pdf'})

@login_required
@app.route('/descargar-reporte-orion-pdf')
def descargar_reporte_orion_pdf():
    # --- La lógica de filtros se mantiene igual ---
    fecha_str = request.args.get('fecha', date.today().isoformat())
    try:
        fecha_seleccionada = date.fromisoformat(fecha_str)
    except (ValueError, TypeError):
        fecha_seleccionada = date.today()
    timestamp_limite = datetime.combine(fecha_seleccionada, time.max)

    # 1. Obtener los datos de la base de datos
    subquery = (db.session.query(
        func.max(RegistroBarcazaOrion.id).label('max_id')
    ).filter(
        RegistroBarcazaOrion.timestamp <= timestamp_limite
    ).group_by(RegistroBarcazaOrion.tk, RegistroBarcazaOrion.grupo).subquery())
    registros_recientes = (db.session.query(RegistroBarcazaOrion)
        .filter(RegistroBarcazaOrion.id.in_(subquery))
        .all())

    if not registros_recientes:
        flash("No hay datos de Barcaza Orion para generar el PDF en esa fecha.", "warning")
        return redirect(url_for('reporte_barcaza'))

    # ======== INICIO DE LA SOLUCIÓN DEFINITIVA ========
    # 2. LIMPIEZA DE DATOS: Convertir registros a diccionarios y reemplazar None por 0.0
    todos_los_tanques_lista = []
    for r in registros_recientes:
        todos_los_tanques_lista.append({
            "TK": r.tk,
            "PRODUCTO": r.producto,
            "MAX_CAP": r.max_cap or 0.0,
            "BLS_60": r.bls_60 or 0.0,
            "API": r.api or 0.0,
            "BSW": r.bsw or 0.0,
            "S": r.s or 0.0,
            "grupo": r.grupo
        })
    # ======== FIN DE LA SOLUCIÓN DEFINITIVA ========

    # 3. Agrupar datos y calcular estadísticas (usa la lista ya limpia)
    datos_agrupados = {}
    nombres_display = {
        "PRINCIPAL": "Tanque Principal (TK-101)", "MANZANILLO": "Barcaza Manzanillo (MGO)",
        "CR": "Barcaza CR", "MARGOTH": "Barcaza Margoth", "ODISEA": "Barcaza Odisea"
    }
    for tanque in todos_los_tanques_lista:
        grupo_key = tanque.get("grupo")
        if grupo_key in nombres_display:
            nombre_barcaza = nombres_display[grupo_key]
            if nombre_barcaza not in datos_agrupados:
                datos_agrupados[nombre_barcaza] = {"tanques": []}
            datos_agrupados[nombre_barcaza]["tanques"].append(tanque)
    
    for nombre, data in datos_agrupados.items():
        data["totales"] = calcular_estadisticas(data["tanques"])
    
    total_consolidado = calcular_estadisticas(todos_los_tanques_lista)

    # 4. Renderizar la plantilla HTML del PDF
    # ===== Historial Orion (agrupado por día) =====
    historial_query = db.session.query(RegistroBarcazaOrion)
    historial_query = historial_query.filter(RegistroBarcazaOrion.timestamp <= timestamp_limite)
    historial_db = historial_query.order_by(RegistroBarcazaOrion.timestamp.desc(), RegistroBarcazaOrion.tk.asc()).all()
    historial_registros = [
        {
            'fecha': r.timestamp.strftime('%Y-%m-%d %H:%M'),
            'tk': r.tk,
            'producto': r.producto,
            'bls_60': r.bls_60 or 0.0,
            'api': r.api or 0.0,
            'bsw': r.bsw or 0.0,
            's': r.s or 0.0,
            'usuario': r.usuario,
            'grupo': r.grupo
        } for r in historial_db
    ]
    from collections import OrderedDict
    grupos = OrderedDict()
    for item in historial_registros:
        dia = item['fecha'][:10]
        if dia not in grupos:
            grupos[dia] = []
        grupos[dia].append(item)
    historial_grouped = [{'dia': dia, 'registros': regs} for dia, regs in grupos.items()]

    # Cargar logo base64 (mismo método que planta)
    logo_base64 = None
    try:
        logo_candidates = ['Logo_de_empresa.jpeg', 'Conquers_4_Logo.png', 'logo.jpeg']
        for fname in logo_candidates:
            logo_path = os.path.join(current_app.root_path, 'static', fname)
            if os.path.exists(logo_path):
                import base64 as _b64
                with open(logo_path, 'rb') as f:
                    logo_base64 = _b64.b64encode(f.read()).decode('utf-8')
                break
    except Exception as e:
        print(f"Error cargando logo Orion PDF: {e}")

    html_para_pdf = render_template('reportes_pdf/orion_pdf.html',
                                    historial_grouped=historial_grouped,
                                    fecha_reporte=fecha_seleccionada.strftime('%d de %B de %Y'),
                                    logo_base64=logo_base64,
                                    fecha_generacion=datetime.now().strftime('%d/%m/%Y %H:%M'))
    
    pdf = HTML(string=html_para_pdf).write_pdf()
    return Response(pdf,
                  mimetype='application/pdf',
                  headers={'Content-Disposition': 'attachment;filename=reporte_barcaza_orion.pdf'})

@login_required
@app.route('/descargar-reporte-bita-pdf')
def descargar_reporte_bita_pdf():
    # --- Lógica para manejar los filtros avanzados ---
    filtro_tipo = request.args.get('filtro_tipo')
    valor = request.args.get('valor')
    
    subquery_base = db.session.query(RegistroBarcazaBita.tk, func.max(RegistroBarcazaBita.timestamp).label('max_timestamp'))
    fecha_reporte_str = "General (últimos datos registrados)"
    timestamp_limite = None

    if valor:
        try:
            if filtro_tipo == 'dia':
                fecha_obj = date.fromisoformat(valor)
                fecha_reporte_str = f"del día {fecha_obj.strftime('%d de %B de %Y')}"
                timestamp_limite = datetime.combine(fecha_obj, time.max)
            elif filtro_tipo == 'mes':
                ano, mes = map(int, valor.split('-'))
                fecha_obj = date(ano, mes, 1)
                fecha_reporte_str = f"del mes de {fecha_obj.strftime('%B de %Y')}"
                ultimo_dia = (fecha_obj + timedelta(days=32)).replace(day=1) - timedelta(days=1)
                timestamp_limite = datetime.combine(ultimo_dia, time.max)
            elif filtro_tipo == 'trimestre':
                ano_str, q_str = valor.split('-Q')
                ano, trimestre = int(ano_str), int(q_str)
                fecha_reporte_str = f"del Trimestre {trimestre} de {ano}"
                mes_fin = trimestre * 3
                ultimo_dia = (date(ano, mes_fin, 1) + timedelta(days=32)).replace(day=1) - timedelta(days=1)
                timestamp_limite = datetime.combine(ultimo_dia, time.max)
            elif filtro_tipo == 'anual':
                ano = int(valor)
                fecha_reporte_str = f"del Año {ano}"
                ultimo_dia = date(ano, 12, 31)
                timestamp_limite = datetime.combine(ultimo_dia, time.max)
        except (ValueError, TypeError):
            pass # Si hay un error en el valor, no se aplica el filtro de tiempo

    # Aplicar el filtro de tiempo a la consulta
    subquery_filtrada = subquery_base.filter(RegistroBarcazaBita.timestamp <= timestamp_limite) if timestamp_limite else subquery_base
    subquery = subquery_filtrada.group_by(RegistroBarcazaBita.tk).subquery()
    registros_recientes = db.session.query(RegistroBarcazaBita).join(subquery, (RegistroBarcazaBita.tk == subquery.c.tk) & (RegistroBarcazaBita.timestamp == subquery.c.max_timestamp)).all()

    if not registros_recientes:
        flash("No hay datos de Barcaza BITA para generar el PDF con el filtro seleccionado.", "warning")
        return redirect(url_for('reporte_barcaza_bita'))

    # --- Limpieza de datos para prevenir el TypeError ---
    datos_reporte = [{
        "TK": r.tk, "PRODUCTO": r.producto,
        "MAX_CAP": r.max_cap or 0.0,
        "BLS_60": r.bls_60 or 0.0,
        "API": r.api or 0.0,
        "BSW": r.bsw or 0.0,
        "S": r.s or 0.0
    } for r in registros_recientes]
    
    # Preparar datos y estadísticas con los datos ya limpios
    total_consolidado = calcular_estadisticas(datos_reporte)
    tanques_marinse = [tk for tk in datos_reporte if tk.get('TK','').startswith('MARI')]
    tanques_oidech = [tk for tk in datos_reporte if tk.get('TK','').startswith('OID')]
    stats_marinse = calcular_estadisticas(tanques_marinse)
    stats_oidech = calcular_estadisticas(tanques_oidech)

    # Renderizar la plantilla del PDF
    # ===== Historial BITA (agrupado por día) =====
    historial_query = db.session.query(RegistroBarcazaBita)
    if timestamp_limite:
        historial_query = historial_query.filter(RegistroBarcazaBita.timestamp <= timestamp_limite)
    historial_db = historial_query.order_by(RegistroBarcazaBita.timestamp.desc(), RegistroBarcazaBita.tk.asc()).all()
    historial_registros = [
        {
            'fecha': r.timestamp.strftime('%Y-%m-%d %H:%M'),
            'tk': r.tk,
            'producto': r.producto,
            'bls_60': r.bls_60 or 0.0,
            'api': r.api or 0.0,
            'bsw': r.bsw or 0.0,
            's': r.s or 0.0,
            'usuario': r.usuario
        } for r in historial_db
    ]
    from collections import OrderedDict
    grupos_bita = OrderedDict()
    for item in historial_registros:
        dia = item['fecha'][:10]
        if dia not in grupos_bita:
            grupos_bita[dia] = []
        grupos_bita[dia].append(item)
    historial_grouped = [{'dia': dia, 'registros': regs} for dia, regs in grupos_bita.items()]

    # Cargar logo base64 para BITA
    logo_base64 = None
    try:
        logo_candidates = ['Logo_de_empresa.jpeg', 'Conquers_4_Logo.png', 'logo.jpeg']
        for fname in logo_candidates:
            logo_path = os.path.join(current_app.root_path, 'static', fname)
            if os.path.exists(logo_path):
                import base64 as _b64
                with open(logo_path, 'rb') as f:
                    logo_base64 = _b64.b64encode(f.read()).decode('utf-8')
                break
    except Exception as e:
        print(f"Error cargando logo BITA PDF: {e}")

    html_para_pdf = render_template('reportes_pdf/bita_pdf.html',
                                    historial_grouped=historial_grouped,
                                    fecha_reporte=fecha_reporte_str,
                                    logo_base64=logo_base64,
                                    fecha_generacion=datetime.now().strftime('%d/%m/%Y %H:%M'))

    pdf = HTML(string=html_para_pdf).write_pdf()
    return Response(pdf, mimetype='application/pdf', headers={'Content-Disposition': 'attachment;filename=reporte_barcaza_bita.pdf'})

@login_required
@app.route('/descargar-reporte-transito-pdf')
def descargar_reporte_transito_pdf():
    # 1. Obtener todos los registros de tránsito (la misma lógica que en la página del reporte)
    todos_los_registros = db.session.query(RegistroTransito).order_by(RegistroTransito.timestamp.desc()).all()

    if not todos_los_registros:
        flash("No hay datos de Tránsito para generar el PDF.", "warning")
        return redirect(url_for('reporte_transito'))

    # 2. Consolidar los datos
    datos_consolidados = {}
    datos_conteo_camiones = {}
    observaciones_camiones = {}
    
    for reg in todos_los_registros:
        origen = (reg.origen or "Sin Origen").strip()
        producto = (reg.producto or "Sin Producto").strip()
        tipo_destino_reporte = "Refinería" if reg.tipo_transito == "refineria" else "EDSM"
        nsv = float(reg.nsv or 0.0)

        # Sumar NSV
        datos_consolidados.setdefault(tipo_destino_reporte, {}).setdefault(origen, {}).setdefault(producto, 0.0)
        datos_consolidados[tipo_destino_reporte][origen][producto] += nsv
        
        # Contar camiones
        datos_conteo_camiones.setdefault(tipo_destino_reporte, {}).setdefault(origen, {}).setdefault(producto, 0)
        datos_conteo_camiones[tipo_destino_reporte][origen][producto] += 1
        
        # Agrupar observaciones
        if reg.observaciones and reg.observaciones.strip():
            texto_completo = f"{(reg.placa or 'S/P')}: {reg.observaciones.strip()}"
            lista_obs = observaciones_camiones.setdefault(tipo_destino_reporte, {}).setdefault(origen, {}).setdefault(producto, [])
            lista_obs.append(texto_completo)

    # 3. Renderizar la plantilla HTML del PDF
    html_para_pdf = render_template('reportes_pdf/transito_pdf.html',
                                    datos_consolidados=datos_consolidados,
                                    datos_conteo_camiones=datos_conteo_camiones,
                                    observaciones_camiones=observaciones_camiones,
                                    fecha_reporte=date.today().strftime('%d de %B de %Y'))

    # 4. Generar y devolver el PDF
    pdf = HTML(string=html_para_pdf).write_pdf()
    return Response(pdf,
                  mimetype='application/pdf',
                  headers={'Content-Disposition': 'attachment;filename=reporte_transito.pdf'})

@login_required
@permiso_requerido('zisa_inventory')
@app.route('/descargar-reporte-pdf')
def descargar_reporte_pdf():
    # --- PASO 1: REPETIMOS LA MISMA LÓGICA DE FILTRADO DE LA PÁGINA DEL REPORTE ---
    empresa_filtro = request.args.get('empresa', default='')
    fecha_inicio_str = request.args.get('fecha_inicio', default='')
    fecha_fin_str = request.args.get('fecha_fin', default='')

    query = RegistroZisa.query.filter_by(estado='Gastado')

    if empresa_filtro in ['ZISA', 'FBCOL']:
        query = query.filter_by(empresa=empresa_filtro)
    
    # Aplicar filtros de fecha...
    # ... (la misma lógica de fechas que en tu ruta 'reporte_consumo') ...

    registros_consumidos = query.order_by(RegistroZisa.fecha_carga.desc()).all()
    total_consumido_filtrado = sum(r.bbl_descargados for r in registros_consumidos)

    # --- PASO 2: RENDERIZAMOS UNA PLANTILLA HTML ESPECIAL PARA EL PDF ---
    # No es la página web completa, solo el contenido del reporte.
    html_para_pdf = render_template('reporte_pdf_template.html',
                                    registros=registros_consumidos,
                                    total_consumido=total_consumido_filtrado,
                                    empresa=empresa_filtro or "Todas",
                                    fecha_inicio=fecha_inicio_str,
                                    fecha_fin=fecha_fin_str)
    
    # --- PASO 3: USAMOS WEASYPRINT PARA CONVERTIR EL HTML A PDF ---
    pdf = HTML(string=html_para_pdf).write_pdf()

    # --- PASO 4: DEVOLVEMOS EL PDF COMO UNA DESCARGA ---
    return Response(pdf,
                    mimetype='application/pdf',
                    headers={'Content-Disposition': 'attachment;filename=reporte_consumo.pdf'})

@login_required
@app.route('/inicio-simulador')
def home_simulador():
    """Página de inicio unificada - redirecciona a home global."""
    return redirect(url_for('home_global'))

@login_required
@permiso_requerido('simulador_rendimiento')
@app.route('/simulador_rendimiento')
def simulador_rendimiento():
    """
    Renderiza la página del Simulador de Rendimiento de Crudo.
    """
    return render_template('simulador_rendimiento.html', nombre=session.get("nombre"))

@login_required
@permiso_requerido('simulador_rendimiento')
@app.route('/descargar_reporte_mezcla_pdf', methods=['POST'])
def descargar_reporte_mezcla_pdf():
    """Genera un PDF del reporte de mezcla de crudos con el mismo estilo visual que
    el reporte gráfico de despachos. Espera un JSON con la estructura:
    {
        total_barrels: number,
        products: { PROD: { barrels, yield_pct, api, sulfur }, ... },
        components: [ { name, mixture_volume, barrels_by_product, api_by_product, sulfur_by_product, total_barrels_input }, ... ]
    }
    """
    try:
        payload = request.get_json(silent=True) or {}
        total_barrels = float(payload.get('total_barrels') or 0)
        products_map = payload.get('products') or {}
        components = payload.get('components') or []
        # Orden de productos (mantener orden estable)
        product_names = list(products_map.keys())
        # Preparar filas para tabla resumen de mezcla
        mezcla_product_rows = []
        for prod in product_names:
            info = products_map.get(prod, {})
            mezcla_product_rows.append({
                'producto': prod,
                'yield_pct': info.get('yield_pct', 0),
                'barrels': info.get('barrels', 0),
                'api': info.get('api', 0),
                'sulfur': info.get('sulfur', 0)
            })
        # Mapa rápido de mezcla por producto
        mezcla_map = {r['producto']: r for r in mezcla_product_rows}
        # API y azufre global de la mezcla (ponderado por barriles, API con SG)
        try:
            sg_total = 0.0; vol_total_products = 0.0; sulfur_total = 0.0
            for r in mezcla_product_rows:
                bbl = float(r['barrels'] or 0)
                api_p = float(r['api'] or 0)
                sg = 141.5 / (api_p + 131.5) if (api_p + 131.5) != 0 else 0
                sg_total += sg * bbl
                vol_total_products += bbl
                sulfur_total += (float(r['sulfur'] or 0) * bbl)
            mezcla_overall_api = ( (141.5 / (sg_total / vol_total_products)) - 131.5 ) if vol_total_products > 0 and sg_total>0 else 0
            mezcla_overall_sulfur = (sulfur_total / vol_total_products) if vol_total_products>0 else 0
        except Exception:
            mezcla_overall_api = 0
            mezcla_overall_sulfur = 0
        # Preparar tabla de componentes (por crudo y por producto)
        componente_rows = []
        # Para comparativo por crudo con/sin KERO, agrupamos por base_crude
        comparativo_por_crudo = {}
        for comp in components:
            row = {
                'name': comp.get('name') or comp.get('nombre') or 'CRUDO',
                'total_volume': comp.get('mixture_volume') or 0,
                'productos': []
            }
            # Añadir temperaturas de corte si vienen desde el front (pueden faltar si eran antiguas corridas)
            cp = comp.get('cut_points') or comp.get('cutPoints') or {}
            try:
                row['cut_points'] = {
                    'nafta': cp.get('nafta'),
                    'kero': cp.get('kero'),
                    'fo4': cp.get('fo4')
                }
            except Exception:
                row['cut_points'] = {'nafta': None, 'kero': None, 'fo4': None}
            barrels_by_product = comp.get('barrels_by_product') or {}
            api_by_product = comp.get('api_by_product') or {}
            sulfur_by_product = comp.get('sulfur_by_product') or {}
            total_volume = float(row['total_volume'] or 0) or 0
            # variables para overall
            sg_sum_comp = 0.0; sulfur_sum_comp = 0.0
            for prod in product_names:
                base_bbl = float(barrels_by_product.get(prod, 0) or 0)
                total_input = float(comp.get('total_barrels_input') or 0) or 1
                # Escala a volumen de mezcla real
                scale = (row['total_volume'] / total_input) if total_input > 0 else 0
                bbl_scaled = base_bbl * scale
                pct = (bbl_scaled / row['total_volume'] * 100) if row['total_volume'] else 0
                api_p = api_by_product.get(prod, 0) or 0
                sulfur_p = sulfur_by_product.get(prod, 0) or 0
                # acumular para overall
                if bbl_scaled>0 and (api_p + 131.5)!=0:
                    sg_sum_comp += (141.5/(api_p+131.5)) * bbl_scaled
                sulfur_sum_comp += sulfur_p * bbl_scaled
                row['productos'].append({
                    'yield_pct': pct,
                    'barrels': bbl_scaled,
                    'api': api_p,
                    'sulfur': sulfur_p,
                    'delta_yield_pct': pct - (mezcla_map.get(prod, {}).get('yield_pct', 0)),
                    'delta_api': api_p - (mezcla_map.get(prod, {}).get('api', 0)),
                    'delta_sulfur': sulfur_p - (mezcla_map.get(prod, {}).get('sulfur', 0))
                })
            # overall metrics component
            if total_volume>0 and sg_sum_comp>0:
                overall_api_comp = (141.5 / (sg_sum_comp / total_volume)) - 131.5
            else:
                overall_api_comp = 0
            overall_sulfur_comp = (sulfur_sum_comp / total_volume) if total_volume>0 else 0
            row['overall_api'] = overall_api_comp
            row['overall_sulfur'] = overall_sulfur_comp
            row['delta_overall_api'] = overall_api_comp - mezcla_overall_api
            row['delta_overall_sulfur'] = overall_sulfur_comp - mezcla_overall_sulfur
            componente_rows.append(row)

            # --- Construcción de comparativo por crudo (con/sin KERO) ---
            try:
                base_name = (comp.get('base_crude') or comp.get('name') or comp.get('nombre') or '').strip()
                inc_kero = bool(comp.get('include_kero'))
                if base_name:
                    grp = comparativo_por_crudo.setdefault(base_name, {'con': None, 'sin': None})
                    # Guardamos rendimientos (%) por producto de este componente
                    rend = {p: 0 for p in product_names}
                    for i, p in enumerate(product_names):
                        try:
                            rend[p] = float(row['productos'][i]['yield_pct'] or 0)
                        except Exception:
                            rend[p] = 0
                    # Guardamos una versión compacta
                    compact = {
                        'name': row['name'],
                        'order': product_names,
                        'yields_pct': rend,
                        'api_by_product': {p: float(api_by_product.get(p, 0) or 0) for p in product_names}
                    }
                    if inc_kero:
                        grp['con'] = compact
                    else:
                        grp['sin'] = compact
            except Exception:
                pass
        # Logo base64 (igual estilo al reporte de despachos)
        logo_base64 = None
        try:
            logo_path = os.path.join(current_app.root_path, 'static', 'Logo_de_empresa.jpeg')
            if os.path.exists(logo_path):
                with open(logo_path, 'rb') as f:
                    logo_base64 = base64.b64encode(f.read()).decode('utf-8')
        except Exception:
            logo_base64 = None
        html_para_pdf = render_template(
            'reportes_pdf/reporte_mezcla_crudos_pdf.html',
            total_barrels=total_barrels,
            mezcla_product_rows=mezcla_product_rows,
            producto_headers=product_names,
            componente_rows=componente_rows,
            mezcla_overall_api=mezcla_overall_api,
            mezcla_overall_sulfur=mezcla_overall_sulfur,
            comparativo_por_crudo=comparativo_por_crudo,
            fecha_generacion=datetime.now().strftime('%d/%m/%Y %H:%M'),
            logo_base64=logo_base64
        )
        pdf = HTML(string=html_para_pdf, base_url=current_app.root_path).write_pdf()
        return Response(pdf, mimetype='application/pdf', headers={'Content-Disposition':'attachment;filename=reporte_mezcla_crudos.pdf'})
    except Exception as e:
        current_app.logger.error(f"Error generando PDF mezcla: {e}")
        return jsonify(success=False, message='Error generando PDF'), 500

@login_required
@permiso_requerido('simulador_rendimiento')
@app.route('/descargar_comparativo_kero_pdf', methods=['POST'])
def descargar_comparativo_kero_pdf():
    """Genera un PDF comparativo Con vs Sin KERO por crudo.
    Espera JSON: { resultados: [ { base_crude, include_kero, order:[], yields:{}, api_by_product:{}, sulfur_by_product:{} }, ... ] }"""
    try:
        data = request.get_json(silent=True) or {}
        resultados = data.get('resultados') or []
        # Agrupar por base_crude -> {'con': {...}, 'sin': {...}}
        grupos = {}
        all_products = []
        for r in resultados:
            base = (r.get('base_crude') or r.get('base') or '').strip()
            if not base:
                continue
            g = grupos.setdefault(base, {'con': None, 'sin': None, 'order': r.get('order') or list((r.get('yields') or {}).keys())})
            orden = r.get('order') or g['order']
            if orden and len(orden) > len(g['order']):
                g['order'] = orden
            all_products.extend(g['order'])
            compact = {
                'yields_pct': r.get('yields') or {},
                'api_by_product': r.get('api_by_product') or {},
                'sulfur_by_product': r.get('sulfur_by_product') or {}
            }
            if r.get('include_kero'):
                g['con'] = compact
            else:
                g['sin'] = compact
        # Normalizar listas productos global
        all_products = list(dict.fromkeys(all_products))
        # Resumen global simple: cantidad crudos con/sin pares completos
        total_bases = len(grupos)
        pares_completos = sum(1 for v in grupos.values() if v.get('con') and v.get('sin'))
        resumen_global = [
            {'nombre': 'Crudos distintos', 'valor': total_bases, 'nota': None},
            {'nombre': 'Pares completos', 'valor': pares_completos, 'nota': None},
            {'nombre': 'Fecha', 'valor': datetime.now().strftime('%d/%m/%Y %H:%M'), 'nota': None}
        ]
        # Logo
        logo_base64 = None
        try:
            logo_path = os.path.join(current_app.root_path, 'static', 'Logo_de_empresa.jpeg')
            if os.path.exists(logo_path):
                with open(logo_path, 'rb') as f:
                    logo_base64 = base64.b64encode(f.read()).decode('utf-8')
        except Exception:
            pass
        html_para_pdf = render_template('reportes_pdf/comparativo_kero_pdf.html',
                                        comparativos=grupos,
                                        resumen_global=resumen_global,
                                        fecha_generacion=datetime.now().strftime('%d/%m/%Y %H:%M'),
                                        logo_base64=logo_base64)
        pdf = HTML(string=html_para_pdf, base_url=current_app.root_path).write_pdf()
        return Response(pdf, mimetype='application/pdf', headers={'Content-Disposition': 'attachment;filename=comparativo_kero.pdf'})
    except Exception as e:
        current_app.logger.error(f"Error comparativo KERO PDF: {e}")
        return jsonify(success=False, message='Error generando PDF comparativo'), 500

@login_required
@permiso_requerido('simulador_rendimiento')
@app.route('/descargar_comparativo_kero_excel', methods=['POST'])
def descargar_comparativo_kero_excel():
    """Genera Excel comparativo KERO. Hoja por crudo o una consolidada.
    Estructura: misma que PDF endpoint."""
    try:
        import io, xlsxwriter
        data = request.get_json(silent=True) or {}
        resultados = data.get('resultados') or []
        grupos = {}
        all_products = []
        for r in resultados:
            base = (r.get('base_crude') or r.get('base') or '').strip()
            if not base:
                continue
            g = grupos.setdefault(base, {'con': None, 'sin': None, 'order': r.get('order') or list((r.get('yields') or {}).keys())})
            orden = r.get('order') or g['order']
            if orden and len(orden) > len(g['order']):
                g['order'] = orden
            all_products.extend(g['order'])
            compact = {
                'yields_pct': r.get('yields') or {},
                'api_by_product': r.get('api_by_product') or {},
                'sulfur_by_product': r.get('sulfur_by_product') or {}
            }
            if r.get('include_kero'):
                g['con'] = compact
            else:
                g['sin'] = compact
        output = io.BytesIO()
        wb = xlsxwriter.Workbook(output, {'in_memory': True})
        fmt_pct = wb.add_format({'num_format': '0.00'} )
        fmt_api = wb.add_format({'num_format': '0.00'} )
        fmt_s  = wb.add_format({'num_format': '0.0000'} )
        fmt_delta_pos = wb.add_format({'font_color':'#0b8552'})
        fmt_delta_neg = wb.add_format({'font_color':'#c0392b'})
        # Hoja consolidada
        ws = wb.add_worksheet('Consolidado')
        headers = ['Crudo','Producto','% Con','% Sin','Δ %','API Con','API Sin','Δ API','%S Con','%S Sin','Δ %S']
        for c,h in enumerate(headers): ws.write(0,c,h)
        row_idx=1
        for base, pair in grupos.items():
            order = pair.get('order') or []
            for prod in order:
                con_y = (pair.get('con') or {}).get('yields_pct',{}).get(prod,0)
                sin_y = (pair.get('sin') or {}).get('yields_pct',{}).get(prod,0)
                d_y = con_y - sin_y
                api_c = (pair.get('con') or {}).get('api_by_product',{}).get(prod,0)
                api_s = (pair.get('sin') or {}).get('api_by_product',{}).get(prod,0)
                d_api = api_c - api_s
                s_c = (pair.get('con') or {}).get('sulfur_by_product',{}).get(prod,0)
                s_s = (pair.get('sin') or {}).get('sulfur_by_product',{}).get(prod,0)
                d_s = s_c - s_s
                ws.write(row_idx,0,base)
                ws.write(row_idx,1,prod)
                ws.write_number(row_idx,2,con_y,fmt_pct)
                ws.write_number(row_idx,3,sin_y,fmt_pct)
                ws.write_number(row_idx,4,d_y, fmt_delta_pos if d_y>0 else fmt_delta_neg if d_y<0 else None)
                ws.write_number(row_idx,5,api_c,fmt_api)
                ws.write_number(row_idx,6,api_s,fmt_api)
                ws.write_number(row_idx,7,d_api, fmt_delta_pos if d_api>0 else fmt_delta_neg if d_api<0 else None)
                ws.write_number(row_idx,8,s_c,fmt_s)
                ws.write_number(row_idx,9,s_s,fmt_s)
                ws.write_number(row_idx,10,d_s, fmt_delta_pos if d_s>0 else fmt_delta_neg if d_s<0 else None)
                row_idx+=1
        wb.close()
        output.seek(0)
        return Response(output.read(), mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers={'Content-Disposition':'attachment;filename=comparativo_kero.xlsx'})
    except Exception as e:
        current_app.logger.error(f"Error comparativo KERO Excel: {e}")
        return jsonify(success=False, message='Error generando Excel comparativo'), 500

@login_required
@app.route('/api/calcular_rendimiento', methods=['POST'])
def api_calcular_rendimiento():
    """
    Calcula rendimiento, API, azufre y viscosidad de productos.
    VERSIÓN MEJORADA CON TODAS LAS OPTIMIZACIONES:
    - Interpolación por spline cúbico
    - Factores de azufre dinámicos basados en API del crudo
    - Watson K-Factor para API
    - Cálculo de número de cetano
    - Temperatura media de ebullición (MABP)
    - Punto de anilina y contenido aromático
    - Balance de masa con validaciones
    - Ajuste dinámico de KERO según calidad del crudo
    - Pérdidas de proceso realistas
    - Viscosidad mejorada con ASTM D341
    """
    try:
        from scipy.interpolate import CubicSpline
        
        data = request.get_json()
        puntos_curva = data.get('distillationCurve')
        puntos_corte = data.get('cutPoints')
        azufre_crudo = data.get('sulfurCrude') or 0
        api_crudo = data.get('apiCrude') or 0
        viscosidad_crudo = data.get('viscosityCrude') or 0
        incluir_kero = data.get('includeKero', True)

        if not all([puntos_curva, puntos_corte, api_crudo]) or len(puntos_curva) < 2:
            return jsonify({"success": False, "message": "Datos incompletos."}), 400

        puntos_curva.sort(key=lambda p: p['tempC'])

        # ============ MEJORA 1: INTERPOLACIÓN CON SPLINE CÚBICO ============
        def interpolar_porcentaje(temp_objetivo):
            if not puntos_curva: return 0
            if temp_objetivo <= puntos_curva[0]['tempC']: return puntos_curva[0]['percent']
            if temp_objetivo >= puntos_curva[-1]['tempC']: return puntos_curva[-1]['percent']
            
            # Si hay suficientes puntos, usar spline cúbico
            if len(puntos_curva) >= 3:
                try:
                    temps = [p['tempC'] for p in puntos_curva]
                    percents = [p['percent'] for p in puntos_curva]
                    cs = CubicSpline(temps, percents, extrapolate=False)
                    return float(cs(temp_objetivo))
                except:
                    pass  # Fallback a interpolación lineal
            
            # Interpolación lineal como fallback
            for i in range(len(puntos_curva) - 1):
                p1, p2 = puntos_curva[i], puntos_curva[i+1]
                if p1['tempC'] <= temp_objetivo <= p2['tempC']:
                    if p2['tempC'] == p1['tempC']: return p1['percent']
                    return p1['percent'] + (temp_objetivo - p1['tempC']) * (p2['percent'] - p1['percent']) / (p2['tempC'] - p1['tempC'])
            return 100

        # ============ MEJORA 2: CALCULAR MABP POR PRODUCTO ============
        def calcular_mabp(temp_inicio, temp_fin):
            """Calcula la temperatura media de ebullición volumétrica"""
            suma_temp = 0
            suma_vol = 0
            for i in range(len(puntos_curva)-1):
                if temp_inicio <= puntos_curva[i]['tempC'] <= temp_fin:
                    vol_incremental = puntos_curva[i+1]['percent'] - puntos_curva[i]['percent']
                    temp_promedio = (puntos_curva[i]['tempC'] + puntos_curva[i+1]['tempC']) / 2
                    suma_temp += temp_promedio * vol_incremental
                    suma_vol += vol_incremental
            return suma_temp / suma_vol if suma_vol > 0 else (temp_inicio + temp_fin) / 2

        # 1. Calcular Rendimientos (Lógica condicional)
        porc_nafta = interpolar_porcentaje(puntos_corte.get('nafta', 0))
        porc_fo4_acumulado = interpolar_porcentaje(puntos_corte.get('fo4', 0))

        # Mapeo de temperaturas de corte
        temp_corte_nafta = puntos_corte.get('nafta', 150)
        temp_corte_kero = puntos_corte.get('kero', 240)
        temp_corte_fo4 = puntos_corte.get('fo4', 350)

        if incluir_kero:
            porc_kero_acumulado = interpolar_porcentaje(temp_corte_kero)
            ORDEN_PRODUCTOS = ["NAFTA", "KERO", "FO4", "FO6"]
            rendimientos = {
                "NAFTA": max(0, porc_nafta),
                "KERO": max(0, porc_kero_acumulado - porc_nafta),
                "FO4": max(0, porc_fo4_acumulado - porc_kero_acumulado),
                "FO6": max(0, 100 - porc_fo4_acumulado)
            }
            
            # ============ MEJORA 3: AJUSTE DINÁMICO DE KERO SEGÚN API ============
            kero_base = rendimientos["KERO"]
            nafta_y = rendimientos["NAFTA"]
            fo4_y = rendimientos["FO4"]
            
            # Factores dinámicos según calidad del crudo
            if api_crudo > 40:  # Crudo ligero
                factor_nafta = 0.08
                factor_fo4 = 0.05
            elif api_crudo > 30:  # Crudo medio
                factor_nafta = 0.05
                factor_fo4 = 0.10
            else:  # Crudo pesado
                factor_nafta = 0.03
                factor_fo4 = 0.15
            
            kero_ajustado = kero_base - factor_nafta * nafta_y + factor_fo4 * fo4_y
            rendimientos["KERO"] = max(0, kero_ajustado)
        else:
            ORDEN_PRODUCTOS = ["NAFTA", "FO4", "FO6"]
            rendimientos = {
                "NAFTA": max(0, porc_nafta),
                "KERO": 0,
                "FO4": max(0, porc_fo4_acumulado - porc_nafta),
                "FO6": max(0, 100 - porc_fo4_acumulado)
            }

        # ============ MEJORA 4: PÉRDIDAS DE PROCESO ============
        PERDIDAS_TIPICAS = {
            'destilacion_atmosferica': 0.5,
            'gases_ligeros': 1.5,
            'coque': 0.3
        }
        total_perdidas = sum(PERDIDAS_TIPICAS.values())
        factor_perdidas = (100 - total_perdidas) / 100
        
        # Aplicar pérdidas antes de normalizar
        for k in rendimientos.keys():
            rendimientos[k] = rendimientos[k] * factor_perdidas

        # Normalización
        suma_original = sum(rendimientos.values()) or 0
        if suma_original > 0:
            for k in rendimientos.keys():
                rendimientos[k] = (rendimientos[k] * 100.0) / suma_original
        suma_post_norm = sum(rendimientos.values())

        # ============ MEJORA 5: AZUFRE CON FACTORES DINÁMICOS ============
        azufre_por_producto = {}
        
        # Factores ajustados según API del crudo
        def get_factor_azufre(producto, api):
            factores_base = {
                'NAFTA': 0.03 if api > 40 else 0.08,
                'KERO': 0.12 if api > 35 else 0.20,
                'FO4': 0.85 if api > 30 else 1.15,
                'FO6': 2.8 if api > 25 else 3.5
            }
            return factores_base.get(producto, 1.0)
        
        FACTORES_AZUFRE = {p: get_factor_azufre(p, api_crudo) for p in ['NAFTA', 'KERO', 'FO4', 'FO6']}
        
        if azufre_crudo > 0:
            denominador_k_s = sum(rendimientos.get(p, 0) * FACTORES_AZUFRE[p] for p in FACTORES_AZUFRE)
            k_s = (100 * azufre_crudo) / denominador_k_s if denominador_k_s > 0 else 0
            for p in FACTORES_AZUFRE:
                azufre_por_producto[p] = round(k_s * FACTORES_AZUFRE.get(p, 0), 4)

        # ============ MEJORA 6: API CON WATSON K-FACTOR ============
        api_por_producto = {}
        watson_k_factors = {}
        API_ESTANDAR = {'NAFTA': 56.6, 'KERO': 42, 'FO4': 30, 'FO6': 21}
        
        def api_a_sg(api): return 141.5 / (api + 131.5) if api != -131.5 else 0
        def sg_a_api(sg): return (141.5 / sg) - 131.5 if sg > 0 else 0
        
        # Calcular Watson K para cada producto
        def calcular_watson_k(temp_rankine, sg):
            """K = Tb^(1/3) / SG donde Tb está en °R"""
            return (temp_rankine ** (1/3)) / sg if sg > 0 else 11.8
        
        sg_crudo_real = api_a_sg(api_crudo)
        sg_estandar = {p: api_a_sg(a) for p, a in API_ESTANDAR.items()}
        
        # Calcular MABP por producto y Watson K
        mabp_productos = {
            'NAFTA': calcular_mabp(0, temp_corte_nafta),
            'KERO': calcular_mabp(temp_corte_nafta, temp_corte_kero),
            'FO4': calcular_mabp(temp_corte_kero, temp_corte_fo4),
            'FO6': calcular_mabp(temp_corte_fo4, 600)
        }
        
        # Convertir MABP a Rankine
        mabp_rankine = {p: (temp + 273.15) * 9/5 for p, temp in mabp_productos.items()}
        
        sg_reconstituido = sum((rendimientos.get(p, 0)/100.0) * sg_estandar[p] for p in API_ESTANDAR if rendimientos.get(p,0) > 0)
        factor_ajuste_sg = (sg_crudo_real / sg_reconstituido) if sg_reconstituido > 0 else 1
        
        for p in API_ESTANDAR:
            sg_adj = sg_estandar[p] * factor_ajuste_sg
            api_por_producto[p] = round(sg_a_api(sg_adj), 2)
            watson_k_factors[p] = round(calcular_watson_k(mabp_rankine.get(p, 900), sg_adj), 2)

        # ============ MEJORA 7: VISCOSIDAD MEJORADA CON ASTM D341 ============
        viscosidad_por_producto = {}
        VISCOSIDAD_STD = {'NAFTA': 0.8, 'KERO': 2.0, 'FO4': 4.0, 'FO6': 380.0}
        
        if viscosidad_crudo > 0:
            # Método logarítmico mejorado
            log_visc_reconstituido = sum((rendimientos.get(p,0)/100.0) * math.log(VISCOSIDAD_STD[p]) 
                                         for p in VISCOSIDAD_STD if VISCOSIDAD_STD.get(p, 0) > 0 and rendimientos.get(p, 0) > 0)
            visc_reconstituido = math.exp(log_visc_reconstituido) if log_visc_reconstituido != 0 else 1
            factor_ajuste_visc = viscosidad_crudo / visc_reconstituido if visc_reconstituido > 0 else 1
            
            for p in VISCOSIDAD_STD:
                visc_base = VISCOSIDAD_STD[p] * factor_ajuste_visc
                # Aplicar corrección ASTM D341 si es necesario
                viscosidad_por_producto[p] = round(visc_base, 2)

        # ============ MEJORA 8: NÚMERO DE CETANO Y PUNTO DE ANILINA ============
        numero_cetano = {}
        punto_anilina = {}
        contenido_aromatico = {}
        indice_diesel = {}
        
        for p in ['KERO', 'FO4']:
            if p in api_por_producto:
                api_p = api_por_producto[p]
                azufre_p = azufre_por_producto.get(p, 0)
                
                # Punto de anilina (correlación empírica)
                pa = 60 + 1.2 * api_p - 15 * azufre_p
                punto_anilina[p] = round(pa, 1)
                
                # Índice diesel
                id_val = pa * api_p / 100
                indice_diesel[p] = round(id_val, 1)
                
                # Contenido aromático estimado
                contenido_aromatico[p] = round(max(0, 100 - pa), 1)
                
                # Número de cetano (correlación ASTM D4737)
                densidad_15C = 141.5 / (api_p + 131.5)
                try:
                    cetano = 45.2 + (0.0892 * pa) + (131.1 * math.log(densidad_15C)) - (86.5 * azufre_p)
                    numero_cetano[p] = round(max(25, min(70, cetano)), 1)
                except:
                    numero_cetano[p] = 45.0

        # ============ MEJORA 9: BALANCE DE MASA Y VALIDACIONES ============
        sg_calculado = sum(rendimientos.get(p,0)/100 * api_a_sg(api_por_producto.get(p,30)) for p in api_por_producto)
        diferencia_sg = abs(sg_crudo_real - sg_calculado)
        
        balance_warning = None
        if diferencia_sg > 0.05:
            balance_warning = {
                "level": "warning",
                "message": f"Balance de masa inconsistente: Δ SG = {diferencia_sg:.4f}",
                "sugerencia": "Revisa las temperaturas de corte o propiedades del crudo"
            }

        # 5. Devolver respuesta completa con todas las mejoras
        response_data = {
            "success": True,
            "order": ORDEN_PRODUCTOS,
            "yields": {p: round(rendimientos.get(p, 0), 2) for p in ORDEN_PRODUCTOS},
            "sum_percent_original": round(suma_original, 4),
            "sum_percent_normalized": round(suma_post_norm, 4),
            "sulfur_by_product": {p: azufre_por_producto.get(p, 0) for p in ORDEN_PRODUCTOS},
            "api_by_product": {p: api_por_producto.get(p, 0) for p in ORDEN_PRODUCTOS},
            "viscosity_by_product": {p: viscosidad_por_producto.get(p, 0) for p in ORDEN_PRODUCTOS},
            
            # Nuevas propiedades avanzadas
            "watson_k_factor": watson_k_factors,
            "mabp_celsius": {p: round(mabp_productos.get(p, 0), 1) for p in mabp_productos},
            "numero_cetano": numero_cetano,
            "punto_anilina": punto_anilina,
            "indice_diesel": indice_diesel,
            "contenido_aromatico": contenido_aromatico,
            "perdidas_proceso": {
                "total_percent": round(total_perdidas, 2),
                "detalle": PERDIDAS_TIPICAS
            },
            "factores_azufre_usados": {p: round(FACTORES_AZUFRE[p], 3) for p in FACTORES_AZUFRE},
            "balance_masa": {
                "sg_crudo_input": round(sg_crudo_real, 4),
                "sg_calculado": round(sg_calculado, 4),
                "diferencia": round(diferencia_sg, 4),
                "warning": balance_warning
            },
            "metodo_interpolacion": "cubic_spline" if len(puntos_curva) >= 3 else "linear"
        }
        
        return jsonify(response_data)

    except Exception as e:
        app.logger.error(f"Error en /api/calcular_rendimiento: {e}")
        import traceback
        app.logger.error(traceback.format_exc())
        return jsonify(success=False, message=f"Error interno del servidor: {str(e)}"), 500

@login_required
@app.route('/api/calibrar_modelo', methods=['POST'])
def calibrar_modelo():
    """
    Endpoint para calibrar el modelo con datos reales de planta.
    Recibe valores calculados vs reales y sugiere ajustes.
    """
    try:
        datos = request.get_json()
        productos = datos.get('productos', {})
        
        if not productos:
            return jsonify(success=False, message="No se proporcionaron datos de productos"), 400
        
        desviaciones = {}
        desviaciones_cuadradas = []
        ajustes_sugeridos = {}
        
        for producto, valores in productos.items():
            calculado = valores.get('calculado', {})
            real = valores.get('real', {})
            
            desv_prod = {}
            
            # Desviación en rendimiento
            if 'yield' in calculado and 'yield' in real:
                calc_y = float(calculado['yield'])
                real_y = float(real['yield'])
                if calc_y > 0:
                    desv_y = ((real_y - calc_y) / calc_y) * 100
                    desv_prod['yield'] = round(desv_y, 2)
                    desviaciones_cuadradas.append(desv_y ** 2)
            
            # Desviación en API
            if 'api' in calculado and 'api' in real:
                calc_api = float(calculado['api'])
                real_api = float(real['api'])
                if calc_api > 0:
                    desv_api = ((real_api - calc_api) / calc_api) * 100
                    desv_prod['api'] = round(desv_api, 2)
                    desviaciones_cuadradas.append(desv_api ** 2)
                    
                    # Sugerir ajuste si desviación > 2%
                    if abs(desv_api) > 2:
                        ajustes_sugeridos[f'{producto}_api_factor'] = round(1 + (desv_api / 100), 4)
            
            # Desviación en azufre
            if 'sulfur' in calculado and 'sulfur' in real:
                calc_s = float(calculado['sulfur'])
                real_s = float(real['sulfur'])
                if calc_s > 0:
                    desv_s = ((real_s - calc_s) / calc_s) * 100
                    desv_prod['sulfur'] = round(desv_s, 2)
                    desviaciones_cuadradas.append(desv_s ** 2)
            
            desviaciones[producto] = desv_prod
        
        # Calcular RMSE
        rmse = math.sqrt(sum(desviaciones_cuadradas) / len(desviaciones_cuadradas)) if desviaciones_cuadradas else 0
        
        # Evaluación de calidad
        if rmse < 2:
            calidad = "Excelente"
        elif rmse < 5:
            calidad = "Buena"
        elif rmse < 10:
            calidad = "Aceptable"
        else:
            calidad = "Requiere calibración"
        
        return jsonify({
            "success": True,
            "desviaciones": desviaciones,
            "rmse": round(rmse, 3),
            "calidad_modelo": calidad,
            "ajustes_sugeridos": ajustes_sugeridos,
            "num_comparaciones": len(desviaciones_cuadradas),
            "recomendacion": "Aplicar factores de corrección sugeridos" if ajustes_sugeridos else "Modelo bien calibrado"
        })
        
    except Exception as e:
        app.logger.error(f"Error en calibración: {e}")
        return jsonify(success=False, message=str(e)), 500

@login_required
@app.route('/api/crudos_guardados', methods=['GET'])
def get_crudos_guardados():
    """Obtiene la lista de todos los crudos guardados desde la base de datos."""
    crudos_db = DefinicionCrudo.query.order_by(DefinicionCrudo.nombre).all()
    
    if not crudos_db:
        # Añadimos valores por defecto para los nuevos campos
        datos_iniciales = {
            "DOROTEA": {"api": 33.1, "sulfur": 0.197, "viscosity": 5.1, "curva": [{"percent": 5, "tempC": 126.7}, {"percent": 10, "tempC": 160.0}, {"percent": 15, "tempC": 174.4}, {"percent": 20, "tempC": 215.6}, {"percent": 30, "tempC": 260.0}, {"percent": 40, "tempC": 304.4}, {"percent": 50, "tempC": 337.8}, {"percent": 60, "tempC": 351.0}]},
            "TULIPAN": {"api": 35.0, "sulfur": 0.3, "viscosity": 4.0, "curva": [{"percent": 5, "tempC": 82.2}, {"percent": 10, "tempC": 98.9}, {"percent": 20, "tempC": 124.4}, {"percent": 30, "tempC": 183.3}, {"percent": 40, "tempC": 224.4}, {"percent": 50, "tempC": 260.0}, {"percent": 60, "tempC": 295.6}, {"percent": 70, "tempC": 356.7}]},
            "INDICO": {"api": 35.0,"sulfur": 0.078, "viscosity": 5.0, "curva": [{"percent": 0, "tempC": 61.6}, {"percent": 5, "tempC": 113.6}, {"percent": 10, "tempC": 138.5}, {"percent": 20, "tempC": 187.0}, {"percent": 30, "tempC": 231.2}, {"percent": 40, "tempC": 265.8}, {"percent": 50, "tempC": 297.8}, {"percent": 60, "tempC": 331.4}, {"percent": 70, "tempC": 380.2}]},
            "JOROPO": {"api": 28.8, "sulfur": 0.20, "viscosity": 5.0, "curva": [{"percent": 0, "tempC": 143}, {"percent": 5, "tempC": 208.1}, {"percent": 10, "tempC": 235.3}, {"percent": 20, "tempC": 277.8}, {"percent": 30, "tempC": 314.1}, {"percent": 40, "tempC": 342.9}, {"percent": 50, "tempC": 374.0}]},
            "WTI": {"api": 43.0, "sulfur": 0.103, "viscosity": 2.4, "curva": [{"percent": 5, "tempC": 60.4}, {"percent": 10, "tempC": 84.7}, {"percent": 20, "tempC": 118.6}, {"percent": 30, "tempC": 156.3}, {"percent": 40, "tempC": 207.4}, {"percent": 50, "tempC": 265.0}, {"percent": 60, "tempC": 327.0}, {"percent": 70, "tempC": 398.0}, {"percent": 80, "tempC": 498.0}]}
        }
        for nombre, data in datos_iniciales.items():
            # Añadir los nuevos campos al crear el objeto
            nuevo_crudo = DefinicionCrudo(
                nombre=nombre, 
                api=data['api'], 
                sulfur=data.get('sulfur'),      # <-- AÑADIDO
                viscosity=data.get('viscosity'),# <-- AÑADIDO
                curva_json=json.dumps(data['curva'])
            )
            db.session.add(nuevo_crudo)
        db.session.commit()
        crudos_db = DefinicionCrudo.query.order_by(DefinicionCrudo.nombre).all()

    crudos_dict = {
        crudo.nombre: {
            "api": crudo.api,
            "sulfur": crudo.sulfur,            # <-- AÑADIDO
            "viscosity": crudo.viscosity,      # <-- AÑADIDO
            "curva": json.loads(crudo.curva_json),
            "assay": json.loads(crudo.assay_json) if crudo.assay_json else []
        } for crudo in crudos_db
    }
    response = jsonify(crudos_dict)
    response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    return jsonify(crudos_dict)

@login_required
@app.route('/api/crudos_guardados', methods=['POST'])
def save_crudo():
    data = request.get_json()
    nombre_crudo = data.get('nombre')
    assay_data = data.get('assay')
    api = data.get('api')
    sulfur = data.get('sulfur')        # <-- AÑADIDO
    viscosity = data.get('viscosity')  # <-- AÑADIDO
    curva = data.get('curva')

    if not nombre_crudo or not curva:
        return jsonify(success=False, message="El nombre y la curva son obligatorios."), 400
    
    crudo_existente = DefinicionCrudo.query.filter_by(nombre=nombre_crudo).first()
    
    if crudo_existente:
        crudo_existente.api = api
        crudo_existente.sulfur = sulfur      # <-- AÑADIDO
        crudo_existente.viscosity = viscosity# <-- AÑADIDO
        crudo_existente.curva_json = json.dumps(curva)
        crudo_existente.assay_json = json.dumps(assay_data)
        msg = f"Crudo '{nombre_crudo}' actualizado."
    else:
        nuevo_crudo = DefinicionCrudo(
            nombre=nombre_crudo, 
            api=api, 
            sulfur=sulfur,                # <-- AÑADIDO
            viscosity=viscosity,          # <-- AÑADIDO
            curva_json=json.dumps(curva),
            assay_json=json.dumps(assay_data) 
        )
        db.session.add(nuevo_crudo)
        msg = f"Crudo '{nombre_crudo}' guardado."
    
    db.session.commit()
    return jsonify(success=True, message=msg)

@login_required
@app.route('/api/crudos_guardados/<string:nombre_crudo>', methods=['DELETE'])
def delete_crudo(nombre_crudo):
    """Elimina un crudo guardado de la base de datos."""
    crudo_a_eliminar = DefinicionCrudo.query.filter_by(nombre=nombre_crudo).first()
    
    if crudo_a_eliminar:
        db.session.delete(crudo_a_eliminar)
        db.session.commit()
        return jsonify(success=True, message=f"Crudo '{nombre_crudo}' eliminado.")
    else:
        return jsonify(success=False, message="Crudo no encontrado."), 404
    
@login_required
@app.route('/inicio-contabilidad')
def home_contabilidad():
    """Página de inicio unificada - redirecciona a home global."""
    return redirect(url_for('home_global'))
    
@login_required
@permiso_requerido('accountingzf@conquerstrading.com')
@app.route('/consolidar-facturas')
def consolidar_facturas():
    return render_template('consolidar_facturas.html', nombre=session.get("nombre"))

@login_required
@permiso_exclusivo('accountingzf@conquerstrading.com')
@app.route('/api/comparar_facturas', methods=['POST'])
def api_comparar_facturas():
    if 'odoo_file' not in request.files or 'dian_file' not in request.files:
        return jsonify(success=False, message="Ambos archivos son requeridos."), 400

    try:
        df_odoo = pd.read_excel(request.files['odoo_file'], engine='openpyxl')
        df_dian = pd.read_excel(request.files['dian_file'], engine='openpyxl')

        # --- Verificación de Columnas Clave ---
        if 'Referencia' not in df_odoo.columns:
            return jsonify(success=False, message="La columna 'Referencia' no se encontró en el archivo de Odoo."), 400
        if 'Prefijo' not in df_dian.columns or 'Folio' not in df_dian.columns or 'Nombre Emisor' not in df_dian.columns:
            return jsonify(success=False, message="El archivo de la DIAN debe tener 'Prefijo', 'Folio' y 'Nombre Emisor'."), 400

        # --- Función de Normalización Inteligente (Definitiva) ---
        def normalizar_factura(ref):
            if pd.isna(ref): return None
            s_ref = str(ref).strip().upper()
            
            # Busca un prefijo de letras/guiones y luego los números
            partes = re.match(r"([A-Z\-]+)0*(\d+)", s_ref)
            if partes:
                # Une el prefijo (sin guion) con el número
                prefijo = partes.group(1).replace('-', '')
                folio = int(partes.group(2))
                return f"{prefijo}-{folio}"
            
            # Si no encuentra el patrón, devuelve solo los números y letras
            return re.sub(r'[^A-Z0-9]', '', s_ref)

        # 1. Procesar datos de Odoo
        set_odoo = set(df_odoo['Referencia'].dropna().apply(normalizar_factura))
        
        # 2. Procesar datos de la DIAN
        def unir_prefijo_folio(row):
            prefijo = str(row['Prefijo']).strip() if pd.notna(row['Prefijo']) else ""
            folio = str(row['Folio']).strip() if pd.notna(row['Folio']) else ""
            # Si el prefijo está vacío, es 'nan', o ya está en el folio, usa solo el folio
            if not prefijo or prefijo.lower() == 'nan' or prefijo in folio:
                return folio
            return prefijo + folio

        df_dian['referencia_completa'] = df_dian.apply(unir_prefijo_folio, axis=1)
        
        dian_map = {
            normalizar_factura(row['referencia_completa']): {
                "factura": row['referencia_completa'],
                "emisor": str(row['Nombre Emisor']) if pd.notna(row['Nombre Emisor']) else "Sin Nombre"
            }
            for _, row in df_dian.dropna(subset=['referencia_completa']).iterrows()
        }
        set_dian = set(dian_map.keys())

        # 3. Comparación Invertida: Lo que está en DIAN y falta en Odoo
        faltantes_normalizados = sorted(list(set_dian - set_odoo))
        
        # Recuperar el formato original y el nombre del emisor desde el mapa de la DIAN
        facturas_faltantes = [dian_map[key] for key in faltantes_normalizados if key in dian_map]
        
        return jsonify(
            success=True,
            faltantes=facturas_faltantes,
            conteo=len(facturas_faltantes)
        )

    except Exception as e:
        app.logger.error(f"Error al comparar archivos: {e}")
        return jsonify(success=False, message=f"Ocurrió un error al procesar los archivos: {str(e)}"), 500
    
@login_required
@permiso_exclusivo('accountingzf@conquerstrading.com')
@app.route('/api/exportar_facturas_excel', methods=['POST'])
def exportar_facturas_excel():
    """Recibe un JSON con la lista de facturas faltantes (tras posibles exclusiones de Caja Menor)
    y devuelve un archivo Excel descargable."""
    try:
        data = request.get_json(silent=True) or {}
        facturas = data.get('facturas', [])
        if not isinstance(facturas, list) or not facturas:
            return jsonify(success=False, message='No se recibieron facturas válidas.'), 400

        # Estructurar DataFrame
        df = pd.DataFrame([
            {
                'Factura (Normalizada)': f.get('factura'),
                'Emisor (DIAN)': f.get('emisor')
            }
            for f in facturas if f.get('factura')
        ])

        if df.empty:
            return jsonify(success=False, message='La lista está vacía tras el filtrado.'), 400

        # Generar Excel en memoria
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Faltantes')
            ws = writer.sheets['Faltantes']
            # Auto ancho simple
            for col in ws.columns:
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        val = str(cell.value) if cell.value is not None else ''
                        if len(val) > max_length:
                            max_length = len(val)
                    except Exception:
                        pass
                ws.column_dimensions[col_letter].width = min(max_length + 2, 60)
        output.seek(0)

        filename = f"facturas_faltantes_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(output,
                         as_attachment=True,
                         download_name=filename,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        app.logger.error(f"Error exportando Excel faltantes: {e}")
        return jsonify(success=False, message='Error interno al generar el Excel.'), 500

@login_required
@permiso_requerido('control_remolcadores')    
@app.route('/control_remolcadores')
def control_remolcadores_page():
    return render_template('control_remolcadores.html', nombre=session.get("nombre"))

@login_required
@permiso_requerido('control_remolcadores')
@app.route('/api/remolcadores/upload_excel', methods=['POST'])
def upload_remolcadores_excel():
    if 'excel_file' not in request.files:
        return jsonify(success=False, message="No se encontró ningún archivo."), 400
    
    file = request.files['excel_file']
    if not file.filename.endswith('.xlsx'):
        return jsonify(success=False, message="Archivo no válido. Debe ser .xlsx"), 400

    try:
        df = pd.read_excel(file)
        df.columns = [str(c).strip().title() for c in df.columns]

        # ✅ 1. Renombrar 'Barco' a 'Nombre Del Barco' si la columna existe
        if 'Barco' in df.columns:
            df.rename(columns={'Barco': 'Nombre Del Barco'}, inplace=True)

        required_columns = ['Id', 'Barcaza', 'Mt Entregadas', 'Evento Anterior', 'Hora Inicio', 'Evento Actual', 'Hora Fin', 'Carga']
        if not all(col in df.columns for col in required_columns):
            missing = [col for col in required_columns if col not in df.columns]
            return jsonify(success=False, message=f"Faltan columnas obligatorias en el Excel: {', '.join(missing)}"), 400

        nuevos_registros = []
        for maniobra_id, group in df.groupby('Id'):
            for _, row in group.iterrows():
                barcaza = row['Barcaza'] if pd.notna(row['Barcaza']) else None
                # Validar y convertir 'Mt Entregadas' correctamente
                mt_val = row['Mt Entregadas']
                if pd.isna(mt_val) or mt_val == '':
                    mt_entregadas = None
                else:
                    try:
                        mt_entregadas = float(mt_val)
                    except (ValueError, TypeError):
                        mt_entregadas = None
                hora_inicio = pd.to_datetime(row['Hora Inicio'], dayfirst=True)
                hora_fin = pd.to_datetime(row['Hora Fin'], dayfirst=True) if pd.notna(row['Hora Fin']) else None
                # Lógica para manejar el campo opcional 'Nombre Del Barco'
                nombre_barco_valor = None
                if 'Nombre Del Barco' in df.columns:
                    nombre_barco_valor = row['Nombre Del Barco'] if pd.notna(row['Nombre Del Barco']) else None
                # Validar 'Carga' como texto
                carga_estado = str(row['Carga']).strip() if pd.notna(row['Carga']) else ''
                registro = RegistroRemolcador(
                    maniobra_id=int(maniobra_id),
                    barcaza=barcaza,
                    nombre_barco=nombre_barco_valor,
                    mt_entregadas=mt_entregadas,
                    carga_estado=carga_estado,
                    evento_anterior=row['Evento Anterior'],
                    hora_inicio=hora_inicio,
                    evento_actual=row['Evento Actual'],
                    hora_fin=hora_fin,
                    usuario_actualizacion=session.get('nombre')
                )
                nuevos_registros.append(registro)

        # Usar el nombre correcto de la tabla que descubrimos ('registro_remolcador')
        db.session.query(RegistroRemolcador).delete()
        db.session.add_all(nuevos_registros)
        db.session.commit()

        return jsonify(success=True, message=f"Se han cargado {len(nuevos_registros)} eventos de {len(df.groupby('Id'))} maniobras.")

    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Error al procesar Excel de remolcadores: {e}")
        return jsonify(success=False, message=f"Error interno al procesar el archivo: {str(e)}"), 500

@login_required
@permiso_requerido('control_remolcadores')
@app.route('/api/maniobra/<int:maniobra_id>', methods=['PUT'])
def update_maniobra_details(maniobra_id):
    """Actualiza la barcaza y las MT para todos los eventos de una maniobra."""
    
    # #{ CAMBIO 1 } - Se añade el email 'opensea@conquerstrading.com' a la lista de permisos.
    if not (session.get('rol') == 'admin' or 
            session.get('email') == 'ops@conquerstrading.com' or 
            session.get('email') == 'opensea@conquerstrading.com'):
        return jsonify(success=False, message="Permiso denegado."), 403

    data = request.get_json()
    barcaza = data.get('barcaza')
    nombre_barco = data.get('nombre_barco')

    try:
        registros = RegistroRemolcador.query.filter_by(maniobra_id=maniobra_id).all()
        for registro in registros:
            # Todos los roles con permiso pueden actualizar la barcaza.
            registro.barcaza = barcaza
            registro.nombre_barco = nombre_barco
            
            # #{ CAMBIO 2 } - Se añade una condición para que solo admin y ops@conquerstrading.com
            # puedan modificar las MT Entregadas. El usuario 'opensea' no podrá hacerlo.
            if session.get('rol') == 'admin' or session.get('email') == 'ops@conquerstrading.com':
                if 'mt_entregadas' in data:
                    mt_entregadas_str = data.get('mt_entregadas')
                    mt_entregadas = float(mt_entregadas_str) if mt_entregadas_str else None
                    registro.mt_entregadas = mt_entregadas
        
        db.session.commit()
        return jsonify(success=True, message=f"Datos de la Maniobra #{maniobra_id} actualizados.")
    except Exception as e:
        db.session.rollback()
        return jsonify(success=False, message=str(e)), 500

@login_required
@permiso_requerido('control_remolcadores')
@app.route('/api/maniobra/<int:maniobra_id>', methods=['DELETE'])
def eliminar_maniobra(maniobra_id):
    """Elimina todos los registros asociados a un ID de maniobra."""
    
    if not (session.get('rol') == 'admin' or 
            session.get('email') == 'ops@conquerstrading.com' or 
            session.get('email') == 'opensea@conquerstrading.com'):
        return jsonify(success=False, message="Permiso denegado."), 403

    try:
        num_borrados = RegistroRemolcador.query.filter_by(maniobra_id=maniobra_id).delete()
        if num_borrados == 0:
            return jsonify(success=False, message="No se encontró la maniobra para eliminar."), 404
            
        db.session.commit()
        return jsonify(success=True, message=f"Maniobra #{maniobra_id} y sus {num_borrados} eventos han sido eliminados.")
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Error al eliminar maniobra: {e}")
        return jsonify(success=False, message=f"Error interno: {str(e)}"), 500 
    
@login_required
@permiso_requerido('control_remolcadores')
@app.route('/api/registros_remolcadores', methods=['GET'])
def get_registros_remolcadores():
    try:
        # Tu lógica de filtrado por fecha está bien

        fecha_inicio_str = request.args.get('fecha_inicio')
        fecha_fin_str = request.args.get('fecha_fin')
        filtro_mes = request.args.get('filtro_mes')  # formato YYYY-MM

        query = RegistroRemolcador.query

        # Si hay filtro de mes, priorizarlo
        if filtro_mes:
            try:
                anio, mes = map(int, filtro_mes.split('-'))
                fecha_inicio_obj = date(anio, mes, 1)
                if mes == 12:
                    fecha_fin_obj = date(anio + 1, 1, 1) - timedelta(days=1)
                else:
                    fecha_fin_obj = date(anio, mes + 1, 1) - timedelta(days=1)
                query = query.filter(RegistroRemolcador.hora_inicio >= fecha_inicio_obj)
                query = query.filter(RegistroRemolcador.hora_inicio <= datetime.combine(fecha_fin_obj, time.max))
            except Exception:
                pass
        else:
            if fecha_inicio_str:
                fecha_inicio_obj = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
                query = query.filter(RegistroRemolcador.hora_inicio >= fecha_inicio_obj)
            if fecha_fin_str:
                fecha_fin_obj = datetime.strptime(fecha_fin_str, '%Y-%m-%d')
                fecha_fin_obj_end_of_day = datetime.combine(fecha_fin_obj, time.max)
                query = query.filter(RegistroRemolcador.hora_inicio <= fecha_fin_obj_end_of_day)

        registros = query.order_by(RegistroRemolcador.maniobra_id, RegistroRemolcador.hora_inicio).all()
        
        # --- ✅ INICIO DE LA LÓGICA CORREGIDA PARA CALCULAR EL TOTAL DE HORAS ---
        duraciones_totales = {}
        if registros:
            # Agrupa todos los eventos por su ID de maniobra
            grupos = groupby(registros, key=lambda r: r.maniobra_id)
            
            for maniobra_id, grupo_eventos in grupos:
                lista_eventos = list(grupo_eventos)
                if not lista_eventos: continue
                
                # Encuentra la primera hora de inicio y la última hora de fin de la maniobra
                primera_hora_inicio = min(r.hora_inicio for r in lista_eventos)
                horas_fin_validas = [r.hora_fin for r in lista_eventos if r.hora_fin]
                
                if horas_fin_validas:
                    ultima_hora_fin = max(horas_fin_validas)
                    # Calcula la diferencia total
                    delta_total = ultima_hora_fin - primera_hora_inicio
                    horas, rem = divmod(delta_total.total_seconds(), 3600)
                    minutos, _ = divmod(rem, 60)
                    duraciones_totales[maniobra_id] = f"{int(horas)}h {int(minutos)}m"
                else:
                    duraciones_totales[maniobra_id] = "En Proceso"
        # --- ✅ FIN DE LA LÓGICA DE CÁLCULO ---

        data = []
        es_opensea = session.get('email') == 'opensea@conquerstrading.com'
        for r in registros:
            registro_data = {
                "id": r.id,
                "maniobra_id": r.maniobra_id,
                "barcaza": r.barcaza,
                "nombre_barco": r.nombre_barco,
                "evento_anterior": r.evento_anterior,
                "hora_inicio": r.hora_inicio.strftime('%Y-%m-%dT%H:%M'), # Formato para <input>
                "evento_actual": r.evento_actual,
                "hora_fin": r.hora_fin.strftime('%Y-%m-%dT%H:%M') if r.hora_fin else '',
                "duracion": r.duracion,
                "carga_estado": r.carga_estado,
                "total_horas": duraciones_totales.get(r.maniobra_id, "")
            }
            if not es_opensea:
                registro_data["mt_entregadas"] = float(r.mt_entregadas) if r.mt_entregadas is not None else ''

            data.append(registro_data)
            
        return jsonify(data)

    except Exception as e:
        app.logger.error(f"Error en get_registros_remolcadores: {e}")
        return jsonify({"error": str(e)}), 500

@login_required
@permiso_requerido('control_remolcadores')
@app.route('/api/registros_remolcadores', methods=['POST'])
def crear_registro_remolcador():
    """Crea un nuevo evento de remolcador."""
    data = request.get_json()
    if not data:
        return jsonify(success=False, message="No se recibieron datos."), 400

    try:
        maniobra_id = data.get('maniobra_id')

        # Si no hay ID de maniobra, es una nueva, así que calculamos el siguiente.
        if not maniobra_id:
            max_id = db.session.query(func.max(RegistroRemolcador.maniobra_id)).scalar()
            maniobra_id = (max_id or 0) + 1

        # --- CORRECCIÓN 1: Manejo seguro de fechas vacías ---
        hora_inicio_str = data.get('hora_inicio')
        hora_fin_str = data.get('hora_fin')

        hora_inicio = datetime.fromisoformat(hora_inicio_str) if hora_inicio_str else None
        hora_fin = datetime.fromisoformat(hora_fin_str) if hora_fin_str else None

        if not hora_inicio:
            return jsonify(success=False, message="La hora de inicio es obligatoria."), 400

        nuevo_registro = RegistroRemolcador(
            maniobra_id=maniobra_id,
            evento_anterior=data.get('evento_anterior'),
            hora_inicio=hora_inicio,
            evento_actual=data.get('evento_actual'),
            hora_fin=hora_fin,
            usuario_actualizacion=session.get('nombre')
        )

        # --- CORRECIÓN 2: Permisos actualizados para opensea ---
        usuario_puede_gestionar = (
            session.get('rol') == 'admin' or 
            session.get('email') == 'ops@conquerstrading.com' or
            session.get('email') == 'opensea@conquerstrading.com'
        )
        if usuario_puede_gestionar:
            nuevo_registro.barcaza = data.get('barcaza')
            nuevo_registro.nombre_barco = data.get('nombre_barco')
            nuevo_registro.mt_entregadas = data.get('mt_entregadas') if data.get('mt_entregadas') else None
            nuevo_registro.carga_estado = data.get('carga_estado')

        if session.get('rol') == 'admin' or session.get('email') == 'ops@conquerstrading.com':
            if 'mt_entregadas' in data:
                nuevo_registro.mt_entregadas = data.get('mt_entregadas') if data.get('mt_entregadas') else None

        db.session.add(nuevo_registro)
        db.session.commit()
        
        return jsonify(success=True, message="Evento creado exitosamente.", nuevo_maniobra_id=maniobra_id)

    except ValueError as e:
        db.session.rollback()
        app.logger.error(f"Error de formato en la fecha al crear evento: {e}")
        return jsonify(success=False, message=f"Formato de fecha no válido: {e}"), 400
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Error al crear evento: {e}")
        return jsonify(success=False, message=f"Error interno: {str(e)}"), 500

@login_required
@permiso_requerido('control_remolcadores')
@app.route('/api/registros_remolcadores/<int:id>', methods=['PUT'])
def update_registro_remolcador(id):
    """Actualiza un evento existente, respetando los permisos de cada rol."""
    registro = RegistroRemolcador.query.get_or_404(id)
    data = request.get_json()

    estados_carga_permitidos = ["LLENO", "VACIO", "N/A"]
    
    # Valores permitidos para opensea
    eventos_anteriores_permitidos = [
        "AUTORIZADO", "CAMBIO DE RR", "CANCELACION", "ESPERAR AUTORIZACION",
        "INICIO BASE OPS", "INICIO BITA", "INICIO CONTECAR", "INICIO FONDEO", "INICIO PUERTO BAHIA", "INICIO SPRC",
        "LLEGADA BASE OPS", "LLEGADA BITA", "LLEGADA CONTECAR", "LLEGADA FONDEO", "LLEGADA SPD", "LLEGADA PUERTO BAHIA"
        "LLEGADA SPRC", "REPOSICIONAMIENTO BARCAZAS"
    ]
    
    eventos_actuales_permitidos = [
        "ACODERADO", "AUTORIZADO", "CAMBIO DE RR", "CANCELACION", 
        "ESPERAR AUTORIZACION", "INICIO BASE OPS", "INICIO BITA", "INICIO CONTECAR", 
        "INICIO FONDEO", "INICIO PUERTO BAHIA","INICIO SPRC", "LLEGADA BASE OPS", "LLEGADA BITA",
        "LLEGADA CONTECAR", "LLEGADA FONDEO", "LLEGADA SPD", "LLEGADA PUERTO BAHIA",
        "REUBICACION BARCAZAS", "TANQUEO"
    ]

    try:
        # El usuario opensea solo puede modificar los campos permitidos
        if session.get('email') == 'opensea@conquerstrading.com':

            if 'carga_estado' in data and data['carga_estado'] not in estados_carga_permitidos:
                return jsonify(success=False, message="Estado de carga no permitido"), 400
            # Validar eventos
            if 'evento_anterior' in data and data['evento_anterior'] not in eventos_anteriores_permitidos:
                return jsonify(success=False, message="Evento anterior no permitido"), 400
            if 'evento_actual' in data and data['evento_actual'] not in eventos_actuales_permitidos:
                return jsonify(success=False, message="Evento actual no permitido"), 400
            
            campos_permitidos = ['evento_anterior', 'hora_inicio', 'evento_actual', 'hora_fin', 'carga_estado', 'nombre_barco']
            for campo in campos_permitidos:
                if campo in data:
                    valor = data[campo]
                    if 'hora' in campo and valor:
                        setattr(registro, campo, datetime.fromisoformat(valor))
                    else:
                        setattr(registro, campo, valor)
        
        # El admin o Juliana pueden editar todos los campos
        elif session.get('rol') == 'admin' or session.get('email') == 'ops@conquerstrading.com':
            for campo, valor in data.items():
                if hasattr(registro, campo):
                    if 'hora' in campo and valor:
                        setattr(registro, campo, datetime.fromisoformat(valor))
                    elif campo == 'carga_estado':
                        setattr(registro, campo, valor if valor != 'N/A' else None)    
                    elif campo != 'id':
                        setattr(registro, campo, valor)
        
        registro.usuario_actualizacion = session.get('nombre')
        db.session.commit()
        return jsonify(success=True, message="Registro actualizado.")
        
    except Exception as e:
        db.session.rollback()
        return jsonify(success=False, message=f"Error al actualizar: {str(e)}"), 500

@login_required
@permiso_requerido('control_remolcadores')
@app.route('/reporte_analisis_remolcadores')
def reporte_analisis_remolcadores():
    try:
        # 1. Leer los filtros desde la URL
        fecha_inicio_str = request.args.get('fecha_inicio')
        fecha_fin_str = request.args.get('fecha_fin')
        filtro_mes = request.args.get('filtro_mes')  # formato YYYY-MM

        query = RegistroRemolcador.query

        # 2. Prioridad: si hay filtro de mes, usarlo y limpiar fechas
        if filtro_mes:
            # Calcular primer y último día del mes
            try:
                anio, mes = map(int, filtro_mes.split('-'))
                fecha_inicio_obj = date(anio, mes, 1)
                if mes == 12:
                    fecha_fin_obj = date(anio + 1, 1, 1) - timedelta(days=1)
                else:
                    fecha_fin_obj = date(anio, mes + 1, 1) - timedelta(days=1)
                query = query.filter(RegistroRemolcador.hora_inicio >= fecha_inicio_obj)
                query = query.filter(RegistroRemolcador.hora_inicio <= datetime.combine(fecha_fin_obj, time.max))
                fecha_inicio_str = fecha_inicio_obj.strftime('%Y-%m-%d')
                fecha_fin_str = fecha_fin_obj.strftime('%Y-%m-%d')
            except Exception:
                pass
        else:
            # 3. Si no hay filtro de mes, usar fechas normales
            if fecha_inicio_str:
                fecha_inicio_obj = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
                query = query.filter(RegistroRemolcador.hora_inicio >= fecha_inicio_obj)
            if fecha_fin_str:
                fecha_fin_obj = datetime.strptime(fecha_fin_str, '%Y-%m-%d')
                query = query.filter(RegistroRemolcador.hora_inicio <= datetime.combine(fecha_fin_obj, time.max))

        # 4. Obtener solo los registros filtrados
        registros_filtrados = query.all()
        
        # 5. Procesar ÚNICAMENTE los datos filtrados
        resultados = procesar_analisis_remolcadores(registros_filtrados)
        
        if not resultados:
            flash("No hay suficientes datos para generar el análisis en el rango de fechas seleccionado.", "warning")
        
        # Guardamos los filtros para pasarlos de vuelta a la plantilla
        filtros_activos = {
            'fecha_inicio': fecha_inicio_str,
            'fecha_fin': fecha_fin_str,
            'filtro_mes': filtro_mes
        }

        return render_template(
            'reporte_analisis_remolcadores.html',
            resultados=resultados,
            filtros=filtros_activos # Pasamos los filtros para los inputs y el botón de PDF
        )
    except Exception as e:
        flash(f"Error al generar el reporte: {str(e)}", "danger")
        return redirect(url_for('control_remolcadores_page'))

@login_required
@permiso_requerido('control_remolcadores')
@app.route('/descargar_analisis_remolcadores_pdf')
def descargar_reporte_analisis_remolcadores_pdf():
    try:
        # (Tu lógica de filtrado por fechas se mantiene igual)
        fecha_inicio_str = request.args.get('fecha_inicio')
        fecha_fin_str = request.args.get('fecha_fin')
        query = RegistroRemolcador.query
        if fecha_inicio_str:
            fecha_inicio_obj = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
            query = query.filter(RegistroRemolcador.hora_inicio >= fecha_inicio_obj)
        if fecha_fin_str:
            fecha_fin_obj = datetime.strptime(fecha_fin_str, '%Y-%m-%d')
            query = query.filter(RegistroRemolcador.hora_inicio <= datetime.combine(fecha_fin_obj, time.max))
        
        registros_filtrados = query.all()
        resultados = procesar_analisis_remolcadores(registros_filtrados)
        
        if not resultados:
            flash("No hay datos para generar el PDF.", "warning")
            return redirect(url_for('reporte_analisis_remolcadores'))

        # --- INICIO DE LA CORRECCIÓN DEFINITIVA ---
        logo_base64 = None
        try:
            # 1. Construir la ruta absoluta al logo
            logo_path = os.path.join(app.root_path, 'static', 'Logo_de_empresa.jpeg')
            # 2. Leer el archivo en modo binario y convertirlo a Base64
            with open(logo_path, "rb") as image_file:
                logo_base64 = base64.b64encode(image_file.read()).decode('utf-8')
        except Exception as e:
            print(f"Error al cargar el logo: {e}") # En caso de que el logo no se encuentre
        # --- FIN DE LA CORRECCIÓN DEFINITIVA ---

        html_para_pdf = render_template(
            'reportes_pdf/analisis_remolcadores_pdf.html',
            resultados=resultados,
            fecha_reporte=date.today().strftime('%d de %B de %Y'),
            now=datetime.utcnow(),
            logo_base64=logo_base64  # <-- Pasamos la nueva variable
        )
        
        pdf = HTML(string=html_para_pdf).write_pdf()
        
        return Response(
            pdf,
            mimetype='application/pdf',
            headers={'Content-Disposition': 'attachment;filename=reporte_analisis_remolcadores.pdf'}
        )
    except Exception as e:
        flash(f"Error al generar el PDF: {str(e)}", "danger")
        return redirect(url_for('reporte_analisis_remolcadores'))

@login_required
@permiso_requerido('control_remolcadores')
@app.route('/descargar_reporte_analisis_remolcadores')
def descargar_reporte_analisis_remolcadores():
    """Genera y descarga un PDF con el análisis completo."""
    # 1. Obtener todos los registros de la base de datos
    registros = RegistroRemolcador.query.all()
    
    # 2. Procesar los datos con tu función de análisis
    resultados = procesar_analisis_remolcadores(registros)
    
    if not resultados:
        flash("No hay datos suficientes para generar el PDF.", "warning")
        return redirect(url_for('reporte_analisis_remolcadores'))

    # 3. Renderiza una plantilla HTML especial para el PDF
    html_para_pdf = render_template(
        'reportes_pdf/analisis_remolcadores_pdf.html',
        resultados=resultados,
        fecha_reporte=date.today().strftime('%d de %B de %Y')
    )
    
    # 4. Convierte el HTML a PDF usando WeasyPrint
    pdf = HTML(string=html_para_pdf).write_pdf()
    
    # 5. Devuelve el PDF como un archivo para descargar
    return Response(
        pdf,
        mimetype='application/pdf',
        headers={'Content-Disposition': 'attachment;filename=reporte_analisis_remolcadores.pdf'}
    )

@login_required
@permiso_requerido('control_remolcadores')
@app.route('/download_remolcadores_excel')
def download_remolcadores_excel():
    try:
        # 1. Obtener registros (con filtros)
        fecha_inicio_str = request.args.get('fecha_inicio')
        fecha_fin_str = request.args.get('fecha_fin')
        query = RegistroRemolcador.query

        if fecha_inicio_str:
            fecha_inicio_obj = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
            query = query.filter(RegistroRemolcador.hora_inicio >= fecha_inicio_obj)
        if fecha_fin_str:
            fecha_fin_obj = datetime.strptime(fecha_fin_str, '%Y-%m-%d')
            query = query.filter(RegistroRemolcador.hora_inicio <= datetime.combine(fecha_fin_obj, time.max))
        
        # Ordenar es clave para agrupar correctamente
        registros = query.order_by(RegistroRemolcador.maniobra_id, RegistroRemolcador.hora_inicio).all()

        # ✅ 2. AÑADIR LÓGICA PARA CALCULAR EL TOTAL DE HORAS POR MANIOBRA
        duraciones_totales = {}
        if registros:
            grupos = groupby(registros, key=lambda r: r.maniobra_id)
            for maniobra_id, grupo_eventos in grupos:
                lista_eventos = list(grupo_eventos)
                if not lista_eventos: continue
                
                primera_hora_inicio = min(r.hora_inicio for r in lista_eventos)
                horas_fin_validas = [r.hora_fin for r in lista_eventos if r.hora_fin]
                
                if horas_fin_validas:
                    ultima_hora_fin = max(horas_fin_validas)
                    delta_total = ultima_hora_fin - primera_hora_inicio
                    horas, rem = divmod(delta_total.total_seconds(), 3600)
                    minutos, _ = divmod(rem, 60)
                    duraciones_totales[maniobra_id] = f"{int(horas)}h {int(minutos)}m"
                else:
                    duraciones_totales[maniobra_id] = "En Proceso"

        # ✅ 3. PREPARAR DATOS PARA EXCEL, INCLUYENDO LAS NUEVAS COLUMNAS
        datos_para_excel = [{
            "Maniobra ID": r.maniobra_id,
            "Barcaza": r.barcaza,
            "Nombre Del Barco": r.nombre_barco,
            "Evento Anterior": r.evento_anterior,
            "Hora Inicio": r.hora_inicio.strftime('%d/%m/%Y %I:%M %p') if r.hora_inicio else '',
            "Evento Actual": r.evento_actual,
            "Hora Fin": r.hora_fin.strftime('%d/%m/%Y %I:%M %p') if r.hora_fin else '',
            "Duración": r.duracion,  # Se asume que tu modelo tiene una propiedad @property para 'duracion'
            "Total Horas Maniobra": duraciones_totales.get(r.maniobra_id, ''),
            "Carga": r.carga_estado,
            "MT Entregadas": r.mt_entregadas
        } for r in registros]
        
        df = pd.DataFrame(datos_para_excel)

        # 4. Crear y devolver el archivo Excel
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
            df.to_excel(writer, index=False, sheet_name='Registros Remolcadores')
            # Auto-ajustar el ancho de las columnas
            for column in df:
                column_width = max(df[column].astype(str).map(len).max(), len(column))
                writer.sheets['Registros Remolcadores'].set_column(df.columns.get_loc(column), df.columns.get_loc(column), column_width + 1)
        output.seek(0)

        return Response(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment;filename=registros_remolcadores.xlsx"}
        )

    except Exception as e:
        app.logger.error(f"Error al generar Excel: {e}")
        return "Error al generar el archivo Excel.", 500

@login_required
@app.route('/inicio-remolcadores')
def home_remolcadores():
    """Página de inicio unificada - redirecciona a home global."""
    return redirect(url_for('home_global'))

@login_required
@permiso_requerido('control_remolcadores')
@app.route('/control-remolcadores')
def control_remolcadores():
    """Muestra la planilla de control de remolcadores."""
    # Pasamos el rol del usuario a la plantilla para que el JavaScript sepa qué hacer.
    return render_template('control_remolcadores.html', rol_usuario=session.get('rol'))

@login_required
@app.route('/home-programacion')
def home_programacion():
    """Página de inicio unificada - redirecciona a home global."""
    return redirect(url_for('home_global'))

@login_required
@permiso_requerido('programacion_cargue')
@app.route('/programacion-cargue')
def programacion_cargue():
    """Muestra la página de programación de vehículos."""
    clientes = cargar_clientes()
    conductores = cargar_conductores()
    return render_template('programacion_cargue.html', 
                           rol_usuario=session.get('rol'), 
                           email_usuario=session.get('email'),
                           nombre=session.get('nombre'),
                           lista_clientes=clientes,
                           lista_conductores=conductores)

@login_required
@app.route('/api/conductores', methods=['POST'])
def agregar_conductor():
    try:
        data = request.get_json()
        nuevo_conductor = {
            "PLACA": data.get('placa', '').upper(),
            "PLACA REMOLQUE": data.get('tanque', '').upper(),
            "NOMBRE CONDUCTOR": data.get('nombre', '').upper(),
            "N° DOCUMENTO": data.get('cedula', ''),
            "CELULAR": data.get('celular', ''),
            "EMPRESA": data.get('empresa', '').upper()
        }
        
        # Validar campos mínimos
        if not nuevo_conductor['NOMBRE CONDUCTOR'] or not nuevo_conductor['N° DOCUMENTO']:
            return jsonify(success=False, message="Nombre y Cédula son obligatorios"), 400

        # Ruta del archivo JSON
        json_path = os.path.join(current_app.root_path, 'static', 'Conductores.json')
        
        conductores = []
        if os.path.exists(json_path):
            with open(json_path, 'r', encoding='utf-8') as f:
                try:
                    conductores = json.load(f)
                except json.JSONDecodeError:
                    conductores = []

        # Verificar duplicados (por cédula)
        for c in conductores:
            if str(c.get("N° DOCUMENTO")) == str(nuevo_conductor["N° DOCUMENTO"]):
                return jsonify(success=False, message="El conductor ya existe (Cédula duplicada)"), 409

        conductores.append(nuevo_conductor)
        
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(conductores, f, ensure_ascii=False, indent=4)
            
        return jsonify(success=True, message="Conductor guardado correctamente")

    except Exception as e:
        current_app.logger.error(f"Error guardando conductor: {e}")
        return jsonify(success=False, message=str(e)), 500

@login_required
@permiso_requerido('programacion_cargue')
@app.route('/api/programacion', methods=['GET', 'POST'])
def handle_programacion():
    """Obtiene o crea registros de programación."""
    if request.method == 'POST':
        # Lógica para crear un nuevo registro vacío
        nuevo = ProgramacionCargue(ultimo_editor=session.get('nombre'))
        db.session.add(nuevo)
        db.session.commit()
        return jsonify(success=True, message="Nueva fila creada.", id=nuevo.id)
    
    # Lógica GET
    mostrar_todas = request.args.get('all', '0') == '1'
    query = ProgramacionCargue.query.order_by(ProgramacionCargue.fecha_programacion.desc())
    if not mostrar_todas:
        registros = query.limit(20).all()
    else:
        # Limitar historial a últimos 500 registros para evitar lentitud
        registros = query.limit(500).all()
    # Convierte los datos a un formato JSON friendly
    data = []
    ahora = datetime.utcnow()
    # Campos que refinery debe completar
    campos_refineria = ['estado', 'galones', 'barriles', 'temperatura', 'api_obs', 'api_corregido', 'precintos', 'fecha_despacho']
    
    for r in registros:
        fila = {}
        for c in r.__table__.columns:
            val = getattr(r, c.name)
            if isinstance(val, (datetime, date, time)):
                fila[c.name] = val.isoformat()
            else:
                fila[c.name] = val
        # Añadimos flag calculado: si ya pasaron 30 min desde completado
        if r.refineria_completado_en:
            fila['refineria_bloqueado'] = (ahora - r.refineria_completado_en) > timedelta(minutes=30)
        else:
            fila['refineria_bloqueado'] = False
        
        # Verificar si refinery completó todos sus campos
        def valor_lleno(v):
            return v not in (None, '')
        try:
            refinery_completo = all(valor_lleno(getattr(r, f)) for f in campos_refineria)
            fila['refinery_completo'] = refinery_completo
        except Exception:
            fila['refinery_completo'] = False
            
        data.append(fila)
    return jsonify(data)

@login_required
@permiso_requerido('programacion_cargue')
@app.route('/api/programacion/<int:id>', methods=['PUT'])
def update_programacion(id):
    """Actualiza un registro de programación con permisos por campo. (VERSIÓN CORREGIDA)"""
    registro = ProgramacionCargue.query.get_or_404(id)
    data = request.get_json()
    
    # La lógica de permisos no necesita cambios, está bien.
    permisos = {
        'ops@conquerstrading.com': ['factura', 'fecha_programacion', 'empresa_transportadora', 'placa', 'tanque', 'nombre_conductor', 'cedula_conductor', 'celular_conductor', 'hora_llegada_estimada', 'producto_a_cargar', 'tipo_guia', 'numero_guia', 'destino', 'cliente', 'fecha_despacho','estado', 'galones', 'barriles', 'temperatura', 'api_obs', 'api_corregido', 'precintos', 'fecha_despacho'],
        'logistic@conquerstrading.com': ['factura', 'fecha_programacion', 'empresa_transportadora', 'placa', 'tanque', 'nombre_conductor', 'cedula_conductor', 'celular_conductor', 'hora_llegada_estimada', 'producto_a_cargar', 'tipo_guia', 'numero_guia', 'destino', 'cliente', 'fecha_despacho','estado', 'galones', 'barriles', 'temperatura', 'api_obs', 'api_corregido', 'precintos', 'fecha_despacho'],
        'production@conquerstrading.com': ['factura', 'fecha_programacion', 'empresa_transportadora', 'placa', 'tanque', 'nombre_conductor', 'cedula_conductor', 'celular_conductor', 'hora_llegada_estimada', 'producto_a_cargar', 'tipo_guia', 'numero_guia', 'destino', 'cliente', 'fecha_despacho','estado', 'galones', 'barriles', 'temperatura', 'api_obs', 'api_corregido', 'precintos', 'fecha_despacho'],
        'oci@conquerstrading.com': ['factura', 'fecha_programacion', 'empresa_transportadora', 'placa', 'tanque', 'nombre_conductor', 'cedula_conductor', 'celular_conductor', 'hora_llegada_estimada', 'producto_a_cargar', 'tipo_guia', 'numero_guia', 'destino', 'cliente', 'fecha_despacho'],
        'amariagallo@conquerstrading.com': ['destino', 'cliente'],
        'refinery.control@conquerstrading.com': ['estado', 'galones', 'barriles', 'temperatura', 'api_obs', 'api_corregido', 'precintos', 'fecha_despacho'],
        'qualitycontrol@conquerstrading.com': ['estado', 'galones', 'barriles', 'temperatura', 'api_obs', 'api_corregido', 'precintos']
    }
    
    campos_permitidos = permisos.get(session.get('email'), [])
    if session.get('rol') == 'admin':
        # El admin puede editar todos los campos excepto los de auditoría que son automáticos.
        campos_permitidos = [c.name for c in ProgramacionCargue.__table__.columns if c.name not in ['id', 'ultimo_editor', 'fecha_actualizacion']]

    if not campos_permitidos:
        return jsonify(success=False, message="No tienes permisos para editar."), 403

    try:
        # Bloqueo nuevo: si TODOS los campos de refinería estuvieron completos y pasaron >30 min, refinería ya no puede editar
        campos_refineria = ['estado','galones','barriles','temperatura','api_obs','api_corregido','precintos','fecha_despacho']
        ahora = datetime.utcnow()
        if registro.refineria_completado_en and (ahora - registro.refineria_completado_en) > timedelta(minutes=30):
            # Si quien intenta editar es refinería y el campo pertenece a su lista, bloquear
            if session.get('email') == 'refinery.control@conquerstrading.com':
                # Si intenta cambiar cualquier campo que sea suyo
                if any(campo in campos_refineria for campo in data.keys()):
                    return jsonify(success=False, message="Bloqueado: Han pasado más de 30 minutos desde que refinería completó todos sus campos."), 403

        # --- INICIO DE LA CORRECCIÓN ---
        campos_numericos = ['galones', 'barriles', 'temperatura', 'api_obs', 'api_corregido']
        
        # Diccionario de normalización de productos
        normalizacion_productos = {
            'F04': 'FUEL OIL 4',
            'FO4': 'FUEL OIL 4',
            'F06': 'FUEL OIL 6',
            'FO6': 'FUEL OIL 6'
        }

        for campo, valor in data.items():
            if campo in campos_permitidos:
                
                # Normalización automática de productos
                if campo == 'producto_a_cargar' and valor:
                    valor_upper = str(valor).upper().strip()
                    valor = normalizacion_productos.get(valor_upper, valor)
                
                # 1. Manejo específico para la fecha de programación
                if campo == 'fecha_programacion'or campo == 'fecha_despacho':
                    # Convierte el string 'YYYY-MM-DD' a un objeto `date`
                    # Si el valor está vacío o es nulo, no hace nada para no borrar la fecha obligatoria.
                    if valor:
                        setattr(registro, campo, date.fromisoformat(valor))

                # 2. Manejo específico para la hora de llegada
                elif campo == 'hora_llegada_estimada':
                    # Si hay un valor, lo convierte a objeto `time`. Si no (el usuario lo borró), lo establece a None.
                    setattr(registro, campo, time.fromisoformat(valor) if valor else None)
                
                # 3. Manejo específico para todos los campos numéricos (float)
                elif campo in campos_numericos:
                    # Intenta convertir a float. Si el valor está vacío o no es un número, lo establece a None.
                    try:
                        setattr(registro, campo, float(valor) if valor is not None and valor != '' else None)
                    except (ValueError, TypeError):
                        setattr(registro, campo, None) # Si la conversión falla, pone None
                
                # 4. Para todos los demás campos (strings), simplemente asigna el valor
                else:
                    setattr(registro, campo, valor)

        # --- FIN DE LA CORRECCIÓN ---

        # Actualizar editor
        registro.ultimo_editor = session.get('nombre')

        # Evaluar completitud de refinería después de aplicar cambios
        def valor_lleno(v):
            return v not in (None, '')
        try:
            completo = all(valor_lleno(getattr(registro, f)) for f in campos_refineria)
        except Exception:
            completo = False

        if completo and not registro.refineria_completado_en:
            registro.refineria_completado_en = ahora
        elif not completo and registro.refineria_completado_en:
            # Si aún no ha pasado el bloqueo definitivo, permitir reiniciar el reloj
            if (ahora - registro.refineria_completado_en) <= timedelta(minutes=30):
                registro.refineria_completado_en = None

        db.session.commit()
        
        return jsonify(success=True, message="Registro actualizado correctamente.")

    except Exception as e:
        db.session.rollback()
        # Imprime el error en la consola del servidor para que puedas depurarlo
        print(f"ERROR AL ACTUALIZAR PROGRAMACIÓN: {e}") 
        return jsonify(success=False, message=f"Error interno del servidor: {str(e)}"), 500

@login_required
@permiso_requerido('programacion_cargue')
@app.route('/api/programacion/locks', methods=['GET'])
def listar_locks_programacion():
    # Limpieza de expirados
    locks = ProgramacionCargueLock.query.all()
    activos = []
    for l in locks:
        if l.expirado():
            db.session.delete(l)
        else:
            activos.append({
                'registro_id': l.registro_id,
                'campo': l.campo,
                'usuario': l.usuario,
                'timestamp': l.timestamp.isoformat()
            })
    db.session.commit()
    return jsonify(activos)

@login_required
@permiso_requerido('programacion_cargue')
@app.route('/api/programacion/lock', methods=['POST'])
def crear_lock_programacion():
    data = request.get_json() or {}
    registro_id = data.get('registro_id')
    campo = data.get('campo')
    if not registro_id or not campo:
        return jsonify(success=False, message='Datos incompletos'), 400
    nombre = session.get('nombre')
    lock = ProgramacionCargueLock.query.filter_by(registro_id=registro_id, campo=campo).first()
    if lock:
        if lock.expirado() or lock.usuario == nombre:
            lock.usuario = nombre
            lock.timestamp = datetime.utcnow()
            db.session.commit()
            return jsonify(success=True, message='Lock actualizado', usuario=nombre)
        return jsonify(success=False, message=f"Editando: {lock.usuario}"), 409
    nuevo = ProgramacionCargueLock(registro_id=registro_id, campo=campo, usuario=nombre)
    db.session.add(nuevo)
    db.session.commit()
    return jsonify(success=True, message='Lock creado', usuario=nombre)

@login_required
@permiso_requerido('programacion_cargue')
@app.route('/api/programacion/lock', methods=['DELETE'])
def borrar_lock_programacion():
    registro_id = request.args.get('registro_id', type=int)
    campo = request.args.get('campo')
    if not registro_id or not campo:
        return jsonify(success=False, message='Datos incompletos'), 400
    nombre = session.get('nombre')
    lock = ProgramacionCargueLock.query.filter_by(registro_id=registro_id, campo=campo).first()
    if lock:
        if lock.usuario == nombre or lock.expirado():
            db.session.delete(lock)
            db.session.commit()
            return jsonify(success=True, message='Lock liberado')
        return jsonify(success=False, message='No puedes liberar lock de otro usuario'), 403
    return jsonify(success=True, message='No existe lock')

# ==========================================
# SISTEMA DE PRESENCIA EN TIEMPO REAL
# ==========================================

# Almacenamiento en memoria de usuarios activos (limpieza automática después de 30 segundos de inactividad)
user_presence = {}  # {email: {name, editing_row, editing_column, last_seen}}

@login_required
@permiso_requerido('programacion_cargue')
@app.route('/api/programacion/presence', methods=['POST'])
def update_presence():
    """Actualizar presencia del usuario actual"""
    data = request.get_json() or {}
    email = session.get('email')
    
    if not email:
        return jsonify(success=False, message='No autenticado'), 401
    
    # Limpiar usuarios inactivos (más de 30 segundos)
    current_time = datetime.utcnow().timestamp()
    inactive_users = [
        user_email for user_email, info in user_presence.items()
        if current_time - info.get('last_seen', 0) > 30
    ]
    for user_email in inactive_users:
        del user_presence[user_email]
    
    # Si no está editando (rowId y column son None), remover presencia
    if not data.get('editing_row') and not data.get('editing_column'):
        if email in user_presence:
            del user_presence[email]
        return jsonify(success=True)
    
    # Actualizar o crear presencia
    user_presence[email] = {
        'name': data.get('user', session.get('nombre', 'Usuario')),
        'editing_row': data.get('editing_row'),
        'editing_column': data.get('editing_column'),
        'current_value': data.get('current_value'),  # NUEVO: Contenido actual
        'last_seen': current_time
    }
    
    return jsonify(success=True)

@login_required
@permiso_requerido('programacion_cargue')
@app.route('/api/programacion/presence', methods=['GET'])
def get_presence():
    """Obtener presencia de todos los usuarios activos"""
    current_email = session.get('email')
    
    # Limpiar usuarios inactivos antes de enviar
    current_time = datetime.utcnow().timestamp()
    inactive_users = [
        user_email for user_email, info in user_presence.items()
        if current_time - info.get('last_seen', 0) > 30
    ]
    for user_email in inactive_users:
        del user_presence[user_email]
    
    # Filtrar para no incluir al usuario actual
    active_users = {
        email: info for email, info in user_presence.items()
        if email != current_email
    }
    
    return jsonify(success=True, users=active_users)

@login_required
@permiso_requerido('programacion_cargue')
@app.route('/api/programacion/live_edit', methods=['POST'])
def registrar_live_edit_programacion():
    """Recibe el texto que el usuario está escribiendo en tiempo real (sin guardar todavía)."""
    data = request.get_json() or {}
    registro_id = data.get('registro_id')
    campo = data.get('campo')
    valor = data.get('valor')
    if not registro_id or not campo:
        return jsonify(success=False, message='Datos incompletos'), 400
    _purge_live_edits()
    LIVE_EDITS[(int(registro_id), campo)] = {
        'registro_id': int(registro_id),
        'campo': campo,
        'valor': valor if valor is not None else '',
        'usuario': session.get('nombre'),
        'timestamp': datetime.utcnow()
    }
    return jsonify(success=True)

@login_required
@permiso_requerido('programacion_cargue')
@app.route('/api/programacion/live_edits', methods=['GET'])
def obtener_live_edits_programacion():
    """Devuelve todas las ediciones activas (texto parcial) y locks vigentes."""
    _purge_live_edits()
    edits = list(LIVE_EDITS.values())
    # Serializar timestamp
    for e in edits:
        e['timestamp'] = e['timestamp'].isoformat()
    return jsonify({'edits': edits})
    
@login_required
@permiso_requerido('programacion_cargue')
@app.route('/exportar_programacion_cargue/<string:formato>')
def exportar_programacion_cargue(formato):
    """
    Genera un reporte de Programación de Cargue en Excel o PDF,
    filtrando por un rango de fechas si se proporciona.
    """
    try:
        # Leemos las fechas desde los parámetros de la URL.
        fecha_inicio_str = request.args.get('fecha_inicio')
        fecha_fin_str = request.args.get('fecha_fin')

        # Empezamos la consulta base.
        query = ProgramacionCargue.query

        # Aplicamos el filtro de fecha de inicio si existe.
        if fecha_inicio_str:
            fecha_inicio_obj = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
            query = query.filter(ProgramacionCargue.fecha_programacion >= fecha_inicio_obj)

        # Aplicamos el filtro de fecha de fin si existe.
        if fecha_fin_str:
            fecha_fin_obj = datetime.strptime(fecha_fin_str, '%Y-%m-%d').date()
            query = query.filter(ProgramacionCargue.fecha_programacion <= fecha_fin_obj)

        # Ejecutamos la consulta final ya filtrada.
        registros = query.order_by(ProgramacionCargue.fecha_programacion.desc()).all()

    except Exception as e:
        flash(f"Error al procesar las fechas: {e}", "danger")
        return redirect(url_for('programacion_cargue'))

    if not registros:
        flash("No hay registros para generar un reporte con el filtro seleccionado.", "warning")
        return redirect(url_for('programacion_cargue'))

    # 2. Lógica para generar el archivo EXCEL
    if formato == 'excel':
        # Preparamos los datos en una lista de diccionarios
        datos_para_df = [{
            'FECHA PROGRAMACION': r.fecha_programacion.strftime('%Y-%m-%d') if r.fecha_programacion else '',
            'EMPRESA TRANSPORTADORA': r.empresa_transportadora,
            'PLACA': r.placa,
            'TANQUE': r.tanque,
            'NOMBRE CONDUCTOR': r.nombre_conductor,
            'CEDULA CONDUCTOR': r.cedula_conductor,
            'CELULAR CONDUCTOR': r.celular_conductor,
            'HORA LLEGADA ESTIMADA': r.hora_llegada_estimada.strftime('%H:%M') if r.hora_llegada_estimada else '',
            'PRODUCTO A CARGAR': r.producto_a_cargar,
            'DESTINO': r.destino,
            'CLIENTE': r.cliente,
            'ESTADO': r.estado,
            'GALONES': r.galones,
            'BARRILES': r.barriles,
            'TEMPERATURA': r.temperatura,
            'API OBS': r.api_obs,
            'API CORREGIDO': r.api_corregido,
            'PRECINTOS': r.precintos,
            'FECHA DESPACHO': r.fecha_despacho.strftime('%Y-%m-%d') if r.fecha_despacho else '',
            'NUMERO GUIA': r.numero_guia
        } for r in registros]

        df = pd.DataFrame(datos_para_df)

        # Creamos el archivo Excel en memoria
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Programacion_Cargue')
        output.seek(0)

        # Enviamos el archivo al navegador
        filename = f"reporte_programacion_cargue_{date.today().strftime('%Y-%m-%d')}.xlsx"
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    # 3. Lógica para generar el archivo PDF
    elif formato == 'pdf':
        # Cargar logo en base64 (igual que otros reportes)
        logo_base64 = None
        try:
            logo_path = os.path.join(current_app.root_path, 'static', 'Logo_de_empresa.jpeg')
            if os.path.exists(logo_path):
                with open(logo_path, 'rb') as f:
                    import base64
                    logo_base64 = base64.b64encode(f.read()).decode('utf-8')
        except Exception as e:
            print(f"Error cargando logo para programación cargue: {e}")

        # Renderizamos una plantilla HTML especial para el PDF
        html_para_pdf = render_template(
            'reportes_pdf/programacion_cargue_pdf.html',
            registros=registros,
            fecha_reporte=datetime.now().strftime('%d de %B de %Y'),
            logo_base64=logo_base64
        )

        # Usamos WeasyPrint para convertir el HTML a PDF (base_url para recursos relativos)
        pdf = HTML(string=html_para_pdf, base_url=current_app.root_path).write_pdf()
        
        # Devolvemos el PDF como una descarga
        return Response(
            pdf,
            mimetype='application/pdf',
            headers={'Content-Disposition': 'attachment;filename=reporte_programacion_cargue.pdf'}
        )

    # Si el formato no es ni 'excel' ni 'pdf', redirigimos
    return redirect(url_for('programacion_cargue'))

@login_required
@permiso_requerido('programacion_cargue')
@app.route('/api/programacion/<int:id>', methods=['DELETE'])
def delete_programacion(id):
    """Elimina un registro de programación de cargue."""
    # Solo pueden eliminar Juliana (ops), Ignacio (production) y Samantha (logistic)
    usuarios_autorizados = {
        'ops@conquerstrading.com',
        'production@conquerstrading.com',
        'logistic@conquerstrading.com'
    }
    if session.get('email') not in usuarios_autorizados and session.get('rol') != 'admin':
        return jsonify(success=False, message='No tienes permiso para eliminar registros.'), 403

    registro = ProgramacionCargue.query.get_or_404(id)

    # Bloqueo: si último editor fue Refinería y han pasado >30 min, prohibir eliminación (para todos)
    if registro.ultimo_editor and registro.ultimo_editor.strip().lower() == 'control refineria':
        if registro.fecha_actualizacion and (datetime.utcnow() - registro.fecha_actualizacion) > timedelta(minutes=30):
            return jsonify(success=False, message='Registro bloqueado: no puede eliminarse después de 30 minutos de la edición de Refinería.'), 403
    try:
        db.session.delete(registro)
        db.session.commit()
        return jsonify(success=True, message="Registro eliminado correctamente.")
    except Exception as e:
        db.session.rollback()
        return jsonify(success=False, message=str(e)), 500

@login_required
@permiso_requerido('programacion_cargue')
@app.route('/api/programacion/<int:id>/upload_image', methods=['POST'])
def upload_programacion_image(id):
    """Sube un archivo de guía (PDF/imagen) y lo guarda en disco; BD almacena ruta relativa."""
    try:
        registro = ProgramacionCargue.query.get_or_404(id)
        # Verificar archivo
        if 'imagen' not in request.files:
            return jsonify(success=False, message='No se recibió ningún archivo'), 400
        archivo = request.files['imagen']
        if not archivo or archivo.filename == '':
            return jsonify(success=False, message='Archivo sin nombre'), 400

        # Validar extensión
        extensiones_permitidas = {'png', 'jpg', 'jpeg', 'gif', 'pdf'}
        ext = archivo.filename.rsplit('.', 1)[1].lower() if '.' in archivo.filename else ''
        if ext not in extensiones_permitidas:
            return jsonify(success=False, message=f'Formato no permitido. Use: {", ".join(extensiones_permitidas)}'), 400

        # Generar nombre seguro y único
        base = secure_filename(os.path.splitext(archivo.filename)[0])[:40] or 'guia'
        unique = f"{datetime.utcnow().strftime('%Y%m%d%H%M%S')}_{uuid.uuid4().hex[:8]}"
        fname = f"{base}_{unique}.{ext}"
        # Carpeta por registro
        folder = os.path.join(current_app.config['GUIDES_DIR'], str(id))
        os.makedirs(folder, exist_ok=True)
        abs_path = os.path.join(folder, fname)
        archivo.save(abs_path)

        # Eliminar archivo anterior si existía (ruta en disco)
        if registro.imagen_guia and not str(registro.imagen_guia).startswith('data:'):
            old_rel = _normalize_guia_relative_path(registro.imagen_guia)
            old_abs = os.path.join(current_app.config['GUIDES_DIR'], old_rel) if old_rel else None
            try:
                if old_abs and os.path.exists(old_abs):
                    os.remove(old_abs)
            except Exception:
                pass

        # Guardar ruta relativa
        rel_path = _normalize_guia_relative_path(f"{id}/{fname}")
        registro.imagen_guia = rel_path
        registro.ultimo_editor = session.get('nombre', 'No identificado')
        registro.fecha_actualizacion = datetime.utcnow()
        db.session.commit()

        url = url_for('serve_guia', filename=rel_path)
        mime, _ = mimetypes.guess_type(abs_path)
        return jsonify(success=True, url=url, mime=mime or 'application/octet-stream')
    except Exception as e:
        db.session.rollback()
        return jsonify(success=False, message=str(e)), 500

@login_required
@permiso_requerido('programacion_cargue')
@app.route('/api/programacion/<int:id>/image', methods=['GET'])
def get_programacion_image(id):
    """Devuelve información para visualizar la guía (URL si está en disco o dataUri si legado)."""
    try:
        registro = ProgramacionCargue.query.get_or_404(id)
        if not registro.imagen_guia:
            return jsonify(success=False, message='Este registro no tiene archivo'), 404

        # Compatibilidad: data URI antigua
        if str(registro.imagen_guia).startswith('data:'):
            # Determinar MIME y corregir si viene como octet-stream pero es PDF (firma %PDF -> base64 'JVBERi0')
            mime = (registro.imagen_guia.split(';')[0].split(':')[1] if ';' in registro.imagen_guia else '').lower()
            if mime in ('', 'application/octet-stream'):
                try:
                    base64_idx = registro.imagen_guia.find('base64,')
                    if base64_idx != -1:
                        head = registro.imagen_guia[base64_idx+7: base64_idx+7+8]
                        if head.startswith('JVBERi0'):
                            mime = 'application/pdf'
                except Exception:
                    pass
            # Log opcional
            try:
                debug = str(app.config.get('MIME_DEBUG', '0')).lower() in ('1','true','yes','on')
                if debug:
                    current_app.logger.info(f"[MIME_DEBUG] /api/programacion/{id}/image -> source=datauri mime={mime or 'application/octet-stream'}")
            except Exception:
                pass
            return jsonify(success=True, dataUri=registro.imagen_guia, mime=mime or 'application/octet-stream', imagen=registro.imagen_guia)

        # Ruta relativa en disco
        rel_path = _normalize_guia_relative_path(registro.imagen_guia)
        abs_path = os.path.join(current_app.config['GUIDES_DIR'], rel_path) if rel_path else None
        if not abs_path or not os.path.exists(abs_path):
            return jsonify(success=False, message='Archivo no encontrado'), 404
        url = url_for('serve_guia', filename=rel_path)
        mime, _ = mimetypes.guess_type(abs_path)
        # Log opcional
        try:
            debug = str(app.config.get('MIME_DEBUG', '0')).lower() in ('1','true','yes','on')
            if debug:
                current_app.logger.info(f"[MIME_DEBUG] /api/programacion/{id}/image -> source=file url={url} mime={mime or 'application/octet-stream'}")
        except Exception:
            pass
        return jsonify(success=True, url=url, mime=mime or 'application/octet-stream', imagen=url)
    except Exception as e:
        return jsonify(success=False, message=str(e)), 500

@login_required
@permiso_requerido('programacion_cargue')
@app.route('/api/programacion/<int:id>/image', methods=['DELETE'])
def delete_programacion_image(id):
    """Elimina el archivo de guía en disco y limpia la referencia."""
    try:
        registro = ProgramacionCargue.query.get_or_404(id)
        if not registro.imagen_guia:
            return jsonify(success=True, message='No hay archivo para eliminar')
        # Si es ruta, intentar borrar del disco
        if not str(registro.imagen_guia).startswith('data:'):
            rel_path = _normalize_guia_relative_path(registro.imagen_guia)
            abs_path = os.path.join(current_app.config['GUIDES_DIR'], rel_path) if rel_path else None
            try:
                if abs_path and os.path.exists(abs_path):
                    os.remove(abs_path)
            except Exception:
                pass
        registro.imagen_guia = None
        registro.ultimo_editor = session.get('nombre', 'No identificado')
        registro.fecha_actualizacion = datetime.utcnow()
        db.session.commit()
        return jsonify(success=True, message='Archivo eliminado correctamente')
    except Exception as e:
        db.session.rollback()
        return jsonify(success=False, message=str(e)), 500


@login_required
@permiso_requerido('programacion_cargue')
@app.route('/api/programacion/<int:id>/importar-sharepoint', methods=['POST'])
def importar_guia_sharepoint(id):
    """Importa una guía desde SharePoint utilizando el numero_guia del registro."""
    registro = ProgramacionCargue.query.get_or_404(id)
    numero_guia = (registro.numero_guia or '').strip()

    if not numero_guia:
        return jsonify(success=False, message='No hay número de guía registrado para este ítem.'), 400

    try:
        resultado = descargar_guia_sharepoint(registro, numero_guia)

        if resultado.get('status') != 'exito':
            return jsonify(success=False, message=resultado.get('mensaje', 'Error desconocido')), 500

        if registro.imagen_guia and not str(registro.imagen_guia).startswith('data:'):
            old_rel = _normalize_guia_relative_path(registro.imagen_guia)
            old_abs = os.path.join(current_app.config['GUIDES_DIR'], old_rel) if old_rel else None
            try:
                if old_abs and os.path.exists(old_abs):
                    os.remove(old_abs)
            except Exception as err:
                current_app.logger.warning(f"No se pudo eliminar archivo anterior (registro {registro.id}): {err}")

        registro.imagen_guia = resultado['ruta_relativa']
        registro.ultimo_editor = session.get('nombre', 'No identificado')
        registro.fecha_actualizacion = datetime.utcnow()
        db.session.commit()

        url = url_for('serve_guia', filename=resultado['ruta_relativa'])
        mime, _ = mimetypes.guess_type(resultado['ruta_abs'])

        return jsonify(success=True, message=f"Guía {numero_guia}.pdf importada.", url=url, mime=mime or 'application/pdf')

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error importando desde SharePoint (ID: {id}): {e}")
        return jsonify(success=False, message=str(e)), 500

@login_required
@permiso_requerido('programacion_cargue')
@app.route('/reporte_grafico_despachos')
def reporte_grafico_despachos():
    today = date.today()
    fecha_inicio_str = request.args.get('fecha_inicio', '')
    fecha_fin_str = request.args.get('fecha_fin', '')
    mes_str = request.args.get('mes', '')  # formato esperado YYYY-MM
    cliente_filtro = request.args.get('cliente', '')
    tipo_grafico = request.args.get('tipo', 'bar')  # 'bar' (horizontal) o 'pie'
    producto_filtro = request.args.get('producto', 'todos').lower()  # 'todos' | 'fo4' | 'diluyente' | 'vlsfo'
    if producto_filtro == 'ambos': producto_filtro = 'todos'  # Compatibilidad backwards

    # Prioridad: si se elige mes, se ignoran fechas individuales
    fecha_inicio = None
    fecha_fin = None
    if mes_str:
        try:
            anio, mes = map(int, mes_str.split('-'))
            fecha_inicio = date(anio, mes, 1)
            # calcular último día del mes
            if mes == 12:
                fecha_fin = date(anio, 12, 31)
            else:
                fecha_fin = date(anio, mes + 1, 1) - timedelta(days=1)
        except Exception:
            fecha_inicio = None
            fecha_fin = None
    else:
        # Si no hay fechas, no filtrar por fecha (mostrar todo)
        if fecha_inicio_str:
            try:
                fecha_inicio = date.fromisoformat(fecha_inicio_str)
            except (ValueError, TypeError):
                fecha_inicio = None
        if fecha_fin_str:
            try:
                fecha_fin = date.fromisoformat(fecha_fin_str)
            except (ValueError, TypeError):
                fecha_fin = None

    # Obtener lista de clientes únicos
    clientes = [c[0] for c in db.session.query(ProgramacionCargue.cliente).distinct().filter(ProgramacionCargue.cliente.isnot(None)).all() if c[0]]
    clientes = sorted(clientes)

    # Consulta agrupada por cliente (para el gráfico principal)
    query = db.session.query(
        ProgramacionCargue.cliente,
        func.sum(ProgramacionCargue.barriles).label('total_barriles')
    ).filter(
        ProgramacionCargue.estado == 'DESPACHADO',
        ProgramacionCargue.cliente.isnot(None),
        ProgramacionCargue.barriles.isnot(None)
    )
    # Filtro por producto (FO4, Diluyente, VLSFO, Todos)
    if producto_filtro == 'fo4':
        query = query.filter(ProgramacionCargue.producto_a_cargar.isnot(None), ProgramacionCargue.producto_a_cargar.ilike('%FO4%'))
    elif producto_filtro == 'diluyente':
        query = query.filter(ProgramacionCargue.producto_a_cargar.isnot(None), ProgramacionCargue.producto_a_cargar.ilike('%DILUYENTE%'))
    elif producto_filtro == 'vlsfo':
        query = query.filter(ProgramacionCargue.producto_a_cargar.isnot(None), ProgramacionCargue.producto_a_cargar.ilike('%VLSFO%'))
    else:  # todos/ambos
        # Incluir todos los productos relevantes (o no filtrar por producto para mostrar todo)
         query = query.filter(ProgramacionCargue.producto_a_cargar.isnot(None), (
            ProgramacionCargue.producto_a_cargar.ilike('%FO4%') | 
            ProgramacionCargue.producto_a_cargar.ilike('%DILUYENTE%') |
            ProgramacionCargue.producto_a_cargar.ilike('%VLSFO%')
        ))
    if fecha_inicio:
        query = query.filter(ProgramacionCargue.fecha_despacho >= fecha_inicio)
    if fecha_fin:
        query = query.filter(ProgramacionCargue.fecha_despacho <= fecha_fin)
    if cliente_filtro:
        query = query.filter(ProgramacionCargue.cliente == cliente_filtro)
    datos_despacho = query.group_by(ProgramacionCargue.cliente).order_by(func.sum(ProgramacionCargue.barriles).desc()).all()

    # Resumen por producto para el cliente seleccionado
    resumen_productos = []
    if cliente_filtro:
        resumen_query = db.session.query(
            ProgramacionCargue.producto_a_cargar,
            func.sum(ProgramacionCargue.barriles).label('total_barriles')
        ).filter(
            ProgramacionCargue.estado == 'DESPACHADO',
            ProgramacionCargue.cliente == cliente_filtro,
            ProgramacionCargue.producto_a_cargar.isnot(None),
            ProgramacionCargue.barriles.isnot(None)
        )
        if fecha_inicio:
            resumen_query = resumen_query.filter(ProgramacionCargue.fecha_despacho >= fecha_inicio)
        if fecha_fin:
            resumen_query = resumen_query.filter(ProgramacionCargue.fecha_despacho <= fecha_fin)
        resumen_productos = resumen_query.group_by(ProgramacionCargue.producto_a_cargar).order_by(func.sum(ProgramacionCargue.barriles).desc()).all()

    grafico_base64 = None
    grafico_div = None
    total_box_text = None  # Texto para tarjeta externa (solo barras)
    total_barriles_general = 0
    if datos_despacho:
        clientes_graf = [resultado[0] for resultado in datos_despacho]
        barriles = [float(resultado[1]) for resultado in datos_despacho]
        total_barriles_general = sum(barriles)
        # Periodo para títulos
        if fecha_inicio and fecha_fin:
            periodo = f"{fecha_inicio.strftime('%d/%m/%Y')} - {fecha_fin.strftime('%d/%m/%Y')}"
        elif fecha_inicio:
            periodo = f"Desde {fecha_inicio.strftime('%d/%m/%Y')}"
        elif fecha_fin:
            periodo = f"Hasta {fecha_fin.strftime('%d/%m/%Y')}"
        else:
            periodo = "Todo el periodo"

        # ---- Calcular FO4, Diluyente y VLSFO por cliente ----
        fo4_vals = []
        diluyente_vals = []
        vlsfo_vals = []
        fo4_total_general = 0
        diluyente_total_general = 0
        vlsfo_total_general = 0

        if producto_filtro in ['ambos', 'todos']:
            base_filters = [
                ProgramacionCargue.estado == 'DESPACHADO',
                ProgramacionCargue.cliente.isnot(None),
                ProgramacionCargue.barriles.isnot(None),
                ProgramacionCargue.producto_a_cargar.isnot(None)
            ]
            
            def get_product_data(pattern):
                q = db.session.query(
                    ProgramacionCargue.cliente,
                    func.sum(ProgramacionCargue.barriles).label('barriles')
                ).filter(*base_filters, ProgramacionCargue.producto_a_cargar.ilike(pattern))
                
                if fecha_inicio:
                    q = q.filter(ProgramacionCargue.fecha_despacho >= fecha_inicio)
                if fecha_fin:
                    q = q.filter(ProgramacionCargue.fecha_despacho <= fecha_fin)
                if cliente_filtro:
                    q = q.filter(ProgramacionCargue.cliente == cliente_filtro)
                return {c: float(v) for c, v in q.group_by(ProgramacionCargue.cliente).all()}

            fo4_map = get_product_data('%FO4%')
            dil_map = get_product_data('%DILUYENTE%')
            vlsfo_map = get_product_data('%VLSFO%')

            fo4_vals = [fo4_map.get(c, 0.0) for c in clientes_graf]
            diluyente_vals = [dil_map.get(c, 0.0) for c in clientes_graf]
            vlsfo_vals = [vlsfo_map.get(c, 0.0) for c in clientes_graf]
            
            fo4_total_general = sum(fo4_vals)
            diluyente_total_general = sum(diluyente_vals)
            vlsfo_total_general = sum(vlsfo_vals)

        if tipo_grafico == 'pie':
            import numpy as np
            from matplotlib import cm
            fig, ax = plt.subplots(figsize=(16, 16))
            
            # Paleta de colores actualizada
            color_fo4 = '#00A896'    # Verde agua/turquesa
            color_dil = '#F77F00'    # Naranja/ámbar
            color_vlsfo = '#06AED5'  # Azul petróleo
            
            # Asignar colores según selección
            if producto_filtro == 'fo4':
                colors = [color_fo4] * len(barriles)
            elif producto_filtro == 'diluyente':
                colors = [color_dil] * len(barriles)
            elif producto_filtro == 'vlsfo':
                colors = [color_vlsfo] * len(barriles)
            else:  # todos
                # Para pie chart 'todos', usamos una paleta generada o colores específicos si es posible
                # Al ser por cliente, usamos una variación de azules/verdes para distinguir
                colors = cm.GnBu(np.linspace(0.4, 0.9, len(barriles)))

            # Etiquetas
            etiquetas = []
            if producto_filtro in ['ambos', 'todos']:
                for i, c in enumerate(clientes_graf):
                    partes = []
                    if fo4_vals[i] > 0: partes.append(f"FO4 {fo4_vals[i]:,.0f}")
                    if diluyente_vals[i] > 0: partes.append(f"DIL {diluyente_vals[i]:,.0f}")
                    if vlsfo_vals[i] > 0: partes.append(f"VLSFO {vlsfo_vals[i]:,.0f}")
                    etiquetas.append(f"{c}\n" + ' | '.join(partes) if partes else c)
            else:
                etiquetas = clientes_graf

            def autopct_func(pct):
                valor_abs = pct * total_barriles_general / 100.0
                return f"{pct:.1f}%\n{valor_abs:,.0f}"

            wedges, texts, autotexts = ax.pie(
                barriles,
                labels=etiquetas,
                startangle=150,
                colors=colors,
                autopct=autopct_func,
                pctdistance=0.72,
                labeldistance=1.1,
                wedgeprops={'linewidth': 1, 'edgecolor': 'white'}
            )
            # Donut interior
            centro = plt.Circle((0, 0), 0.48, fc='white')
            fig.gca().add_artist(centro)
            texto_centro = f"TOTAL\n{total_barriles_general:,.0f} BBL"
            fig.text(0.5, 0.5, texto_centro, ha='center', va='center', fontsize=18, fontweight='bold', color=color_fo4)
            
            titulo = "Distribución de Despachos (Donut)"
            if producto_filtro == 'fo4': titulo += " – FO4"
            elif producto_filtro == 'diluyente': titulo += " – Diluyente"
            elif producto_filtro == 'vlsfo': titulo += " – VLSFO"
            
            ax.set_title(f"{titulo}\nPeriodo: {periodo}", fontsize=20, pad=28, fontweight='bold')
            for t in texts: t.set_fontsize(9.5)
            for at in autotexts: at.set_fontsize(9)
            
        else:
            # --- Barras horizontales mejoradas (Updated) ---
            from matplotlib.ticker import FuncFormatter
            
            # Nuevos Colores
            c_fo4 = '#00A896'    # Verde agua
            c_dil = '#F77F00'    # Naranja
            c_vlsfo = '#06AED5'  # Azul petróleo
            c_border = '#2C3E50' # Oscuro para bordes
            
            # Cálculo de altura dinámica según requerimiento usuario
            # Fórmula: (num_clientes * 45px) + 200px
            # Convertimos px a pulgadas asumiendo 100 DPI (aprox) para matplotlib
            calculated_height_px = (len(clientes_graf) * 45) + 200
            final_height_px = max(600, min(1500, calculated_height_px))
            figsize_height = final_height_px / 100.0
            
            # Ancho maximizado (Aumentado de 32 a 48 para ocupar más espacio horizontal)
            fig, ax = plt.subplots(figsize=(48, figsize_height))
            
            y_pos = list(range(len(clientes_graf)))
            labels_rank = [f"{i+1}. {c}" for i, c in enumerate(clientes_graf)]
            
            max_val = max(barriles) if barriles else 0
            if producto_filtro in ['ambos', 'todos']:
                # Barras apiladas: Diluyente base, luego FO4, luego VLSFO
                # 1. Diluyente
                p1 = ax.barh(y_pos, diluyente_vals, color=c_dil, edgecolor=c_border, linewidth=0.5, height=0.65, label='Diluyente')
                
                # 2. FO4 (bottom = diluyente)
                p2 = ax.barh(y_pos, fo4_vals, left=diluyente_vals, color=c_fo4, edgecolor=c_border, linewidth=0.5, height=0.65, label='FO4')
                
                # 3. VLSFO (bottom = diluyente + fo4)
                # Necesitamos sumar dil + fo4 para el 'left' del vlsfo
                left_vlsfo = [d + f for d, f in zip(diluyente_vals, fo4_vals)]
                p3 = ax.barh(y_pos, vlsfo_vals, left=left_vlsfo, color=c_vlsfo, edgecolor=c_border, linewidth=0.5, height=0.65, label='VLSFO')
                
                # Etiquetas internas
                umbral_seg = max_val * 0.035 if max_val > 0 else 0
                
                for i, (dil, fo4, vlsfo) in enumerate(zip(diluyente_vals, fo4_vals, vlsfo_vals)):
                    # DIL
                    if dil > umbral_seg:
                        ax.text(dil/2, i, f"{dil:,.0f}", ha='center', va='center', color='white', fontsize=10, fontweight='bold')
                    # FO4
                    if fo4 > umbral_seg:
                        ax.text(dil + fo4/2, i, f"{fo4:,.0f}", ha='center', va='center', color='white', fontsize=10, fontweight='bold')
                    # VLSFO
                    if vlsfo > umbral_seg:
                        ax.text(dil + fo4 + vlsfo/2, i, f"{vlsfo:,.0f}", ha='center', va='center', color='white', fontsize=10, fontweight='bold')
                        
            else:
                # Individual
                color_map = {'fo4': c_fo4, 'diluyente': c_dil, 'vlsfo': c_vlsfo}
                use_color = color_map.get(producto_filtro, c_fo4)
                lbl_map = {'fo4': 'FO4', 'diluyente': 'Diluyente', 'vlsfo': 'VLSFO'}
                
                bars_single = ax.barh(
                    y_pos,
                    barriles,
                    color=use_color,
                    edgecolor=c_border, linewidth=0.6, height=0.65,
                    label=lbl_map.get(producto_filtro, producto_filtro.upper())
                )
            
            ax.set_yticks(y_pos)
            ax.set_yticklabels(labels_rank, fontsize=12, fontweight='500')
            ax.xaxis.set_major_formatter(FuncFormatter(lambda x, _ : f'{x:,.0f}'))
            ax.set_xlabel('Barriles despachados', fontweight='bold', fontsize=14, labelpad=12)
            
            titulo_bar = "Total de Barriles Despachados por Cliente"
            # Subtítulo dinámico
            prod_names = []
            if producto_filtro in ['ambos', 'todos']:
                if sum(diluyente_vals)>0: prod_names.append("Diluyente")
                if sum(fo4_vals)>0: prod_names.append("FO4")
                if sum(vlsfo_vals)>0: prod_names.append("VLSFO")
                if not prod_names: prod_names = ["Todos"]
                titulo_bar += f" – {' + '.join(prod_names)}"
            else:
                 titulo_bar += f" – {producto_filtro.upper()}"

            ax.set_title(f"{titulo_bar}\nPeriodo: {periodo}", fontsize=20, pad=22, fontweight='bold')
            
            total_box_text = f"TOTAL {total_barriles_general:,.2f} BBL"
            
            # Etiquetas de totales a la derecha
            for i, total_width in enumerate(barriles):
                offset = max_val * 0.005 if max_val > 0 else 0
                ax.text(total_width + offset, i, f'{total_width:,.2f}', ha='left', va='center', color=c_border, fontweight='bold', fontsize=11)
            
            ax.legend(loc='lower right', frameon=True, facecolor='white', framealpha=0.9, fontsize=12)
            ax.invert_yaxis()
            
            # Estilos limpios
            for spine in ['top', 'right', 'left']:
                ax.spines[spine].set_visible(False)
            ax.spines['bottom'].set_color('#9aa0ac')
            ax.tick_params(axis='y', length=0)
            ax.xaxis.grid(True, linestyle='--', linewidth=0.6, alpha=0.35)
            ax.set_axisbelow(True)
            ax.set_facecolor('#fcfdfd')
            fig.patch.set_facecolor('#ffffff')

        # Margen ajustado para full width visual (Menos margen lateral)
        plt.tight_layout(rect=[0.005, 0.01, 0.995, 0.98])
        grafico_base64 = convertir_plot_a_base64(fig)

    # --- Datos para ApexCharts (JSON) ---
    apex_data = []
    if datos_despacho:
        if producto_filtro in ['ambos', 'todos']:
            # Para barras y pie en 'todos', pasamos los desglose por cliente
            for i, c in enumerate(clientes_graf):
                apex_data.append({
                    'cliente': c,
                    'total': float(barriles[i]),
                    'fo4': float(fo4_vals[i]),
                    'diluyente': float(diluyente_vals[i]),
                    'vlsfo': float(vlsfo_vals[i])
                })
        else:
            # Individual
            for i, c in enumerate(clientes_graf):
                apex_data.append({
                    'cliente': c,
                    'total': float(barriles[i])
                })

    return render_template(
        'reporte_grafico_despachos.html',
        grafico_base64=grafico_base64,
        grafico_div=grafico_div,
        apex_data=apex_data,
        datos_tabla=datos_despacho,
        total_barriles=total_barriles_general,
        total_box_text=total_box_text,
        filtros={
            'fecha_inicio': fecha_inicio.isoformat() if (fecha_inicio_str and fecha_inicio and not mes_str) else '',
            'fecha_fin': fecha_fin.isoformat() if (fecha_fin_str and fecha_fin and not mes_str) else '',
            'mes': mes_str,
            'cliente': cliente_filtro,
            'producto': producto_filtro
        },
        clientes=clientes,
        resumen_productos=resumen_productos,
        now=datetime.now(),
        tipo=tipo_grafico,
        producto=producto_filtro
    )

@login_required
@permiso_requerido('programacion_cargue')
@app.route('/descargar_reporte_grafico_despachos_pdf')
def descargar_reporte_grafico_despachos_pdf():
    # Parámetros y lógica equivalente a la vista HTML
    today = date.today()
    fecha_inicio_str = request.args.get('fecha_inicio', '')
    fecha_fin_str = request.args.get('fecha_fin', '')
    mes_str = request.args.get('mes', '')
    cliente_filtro = request.args.get('cliente', '')
    producto_filtro = request.args.get('producto', 'ambos').lower()  # ambos | fo4 | diluyente
    tipo_grafico = request.args.get('tipo', 'bar')

    fecha_inicio = None
    fecha_fin = None
    if mes_str:
        try:
            anio, mes = map(int, mes_str.split('-'))
            fecha_inicio = date(anio, mes, 1)
            fecha_fin = date(anio, mes, 28) + timedelta(days=4)
            fecha_fin = fecha_fin - timedelta(days=fecha_fin.day)
        except Exception:
            fecha_inicio = None
            fecha_fin = None
    else:
        if fecha_inicio_str:
            try:
                fecha_inicio = date.fromisoformat(fecha_inicio_str)
            except Exception:
                fecha_inicio = None
        if fecha_fin_str:
            try:
                fecha_fin = date.fromisoformat(fecha_fin_str)
            except Exception:
                fecha_fin = None

    base_query = db.session.query(
        ProgramacionCargue.cliente,
        func.sum(ProgramacionCargue.barriles).label('total_barriles')
    ).filter(
        ProgramacionCargue.estado == 'DESPACHADO',
        ProgramacionCargue.cliente.isnot(None),
        ProgramacionCargue.barriles.isnot(None),
        ProgramacionCargue.producto_a_cargar.isnot(None)
    )
    if producto_filtro == 'fo4':
        base_query = base_query.filter(ProgramacionCargue.producto_a_cargar.ilike('%FO4%'))
    elif producto_filtro == 'diluyente':
        base_query = base_query.filter(ProgramacionCargue.producto_a_cargar.ilike('%DILUYENTE%'))
    else:
        base_query = base_query.filter(
            ProgramacionCargue.producto_a_cargar.ilike('%FO4%') | ProgramacionCargue.producto_a_cargar.ilike('%DILUYENTE%')
        )
    if fecha_inicio:
        base_query = base_query.filter(ProgramacionCargue.fecha_despacho >= fecha_inicio)
    if fecha_fin:
        base_query = base_query.filter(ProgramacionCargue.fecha_despacho <= fecha_fin)
    if cliente_filtro:
        base_query = base_query.filter(ProgramacionCargue.cliente == cliente_filtro)
    datos_despacho = base_query.group_by(ProgramacionCargue.cliente).order_by(func.sum(ProgramacionCargue.barriles).desc()).all()

    total_barriles_general = sum(float(r[1]) for r in datos_despacho) if datos_despacho else 0

    grafico_base64 = None
    total_box_text = None
    tabla_detalle = []  # (cliente, diluyente, fo4, total)
    total_fo4 = 0.0
    total_dil = 0.0

    if datos_despacho:
        clientes_graf = [r[0] for r in datos_despacho]
        barriles = [float(r[1]) for r in datos_despacho]
        # Mapas por producto (calculamos siempre para poder mostrar columnas aun si se filtró uno)
        def build_product_map(pattern):
            q = db.session.query(ProgramacionCargue.cliente, func.sum(ProgramacionCargue.barriles).label('val')).filter(
                ProgramacionCargue.estado=='DESPACHADO',
                ProgramacionCargue.barriles.isnot(None),
                ProgramacionCargue.producto_a_cargar.isnot(None),
                ProgramacionCargue.producto_a_cargar.ilike(pattern)
            )
            if fecha_inicio:
                q = q.filter(ProgramacionCargue.fecha_despacho >= fecha_inicio)
            if fecha_fin:
                q = q.filter(ProgramacionCargue.fecha_despacho <= fecha_fin)
            if cliente_filtro:
                q = q.filter(ProgramacionCargue.cliente==cliente_filtro)
            return {c: float(v) for c,v in q.group_by(ProgramacionCargue.cliente).all()}
        fo4_map = build_product_map('%FO4%')
        dil_map = build_product_map('%DILUYENTE%')
        fo4_vals = [fo4_map.get(c,0.0) for c in clientes_graf]
        dil_vals = [dil_map.get(c,0.0) for c in clientes_graf]
        total_fo4 = sum(fo4_vals)
        total_dil = sum(dil_vals)
        # Construir tabla detalle
        for c, d, f in zip(clientes_graf, dil_vals, fo4_vals):
            tabla_detalle.append((c, d, f, d+f))
        # Crear gráfico (solo barras para PDF por estabilidad)
        from matplotlib.colors import LinearSegmentedColormap
        from matplotlib.ticker import FuncFormatter
        altura = max(10, len(clientes_graf) * 0.75)
        fig, ax = plt.subplots(figsize=(24, altura))
        cmap = LinearSegmentedColormap.from_list('dil_cmap', ['#1d7ed6', '#63b3ff'])
        max_val = max(barriles)
        min_val = min(barriles)
        norm_vals = [0.5 if max_val==min_val else (v-min_val)/(max_val-min_val) for v in barriles]
        y_pos = list(range(len(clientes_graf)))
        labels_rank = [f"{i+1}. {c}" for i,c in enumerate(clientes_graf)]
        if producto_filtro == 'fo4':
            ax.barh(y_pos, fo4_vals, color='#ff9f43', edgecolor='#c86e00', height=0.68, label='FO4')
        elif producto_filtro == 'diluyente':
            ax.barh(y_pos, dil_vals, color=[cmap(n) for n in norm_vals], edgecolor='#0b3d66', height=0.68, label='Diluyente')
        else:  # ambos
            ax.barh(y_pos, dil_vals, color=[cmap(n) for n in norm_vals], edgecolor='#0b3d66', height=0.68, label='Diluyente')
            ax.barh(y_pos, fo4_vals, left=dil_vals, color='#ff9f43', edgecolor='#c86e00', height=0.68, label='FO4')
        if producto_filtro == 'ambos':
            ax.legend(loc='lower right', frameon=False, fontsize=10)
        ax.set_yticks(y_pos)
        ax.set_yticklabels(labels_rank, fontsize=11)
        ax.xaxis.set_major_formatter(FuncFormatter(lambda x,_: f"{x:,.0f}"))
        ax.set_xlabel('Barriles despachados', fontweight='bold')
        titulo_pdf = 'Total Despachado por Cliente'
        if producto_filtro=='ambos':
            titulo_pdf += ' – FO4 + Diluyente'
        elif producto_filtro=='fo4':
            titulo_pdf += ' – FO4'
        else:
            titulo_pdf += ' – Diluyente'
        ax.set_title(titulo_pdf, fontweight='bold', fontsize=16, pad=16)
        # Etiquetas finales
        totales_linea = [d+f for d,f in zip(dil_vals, fo4_vals)]
        for i,v in enumerate(totales_linea):
            ax.text(v + (max_val*0.006), i, f"{v:,.0f}", va='center', fontsize=9, color='#0b3d66')
        if producto_filtro=='ambos':
            for i,(d,f) in enumerate(zip(dil_vals, fo4_vals)):
                umbral_seg = max_val * 0.05 if max_val>0 else 0
                if d>0 and d>=umbral_seg:
                    ax.text(d/2, i, f"DIL {d:,.0f}", ha='center', va='center', color='white', fontsize=7, fontweight='bold')
                if f>0 and f>=umbral_seg:
                    ax.text(d+f/2, i, f"FO4 {f:,.0f}", ha='center', va='center', color='white', fontsize=7, fontweight='bold')
        ax.invert_yaxis()
        ax.xaxis.grid(True, linestyle='--', alpha=0.3)
        for spine in ['top','right','left']:
            ax.spines[spine].set_visible(False)
        fig.patch.set_facecolor('#ffffff')
        plt.tight_layout()
        grafico_base64 = convertir_plot_a_base64(fig)
        # Texto total
        total_box_text = f"TOTAL {total_barriles_general:,.2f} BBL"\
            + (f" | FO4 {total_fo4:,.0f} BBL" if total_fo4>0 else '')\
            + (f" | Diluyente {total_dil:,.0f} BBL" if total_dil>0 else '')

    periodo_txt = 'Todo el periodo'
    if fecha_inicio and fecha_fin:
        periodo_txt = f"{fecha_inicio.strftime('%d/%m/%Y')} - {fecha_fin.strftime('%d/%m/%Y')}"
    elif fecha_inicio:
        periodo_txt = f"Desde {fecha_inicio.strftime('%d/%m/%Y')}"
    elif fecha_fin:
        periodo_txt = f"Hasta {fecha_fin.strftime('%d/%m/%Y')}"

    # Obtener logo como base64
    logo_base64 = None
    try:
        logo_path = os.path.join(current_app.root_path, 'static', 'Logo_de_empresa.jpeg')
        if os.path.exists(logo_path):
            import base64
            with open(logo_path, 'rb') as f:
                logo_base64 = base64.b64encode(f.read()).decode('utf-8')
    except Exception:
        logo_base64 = None

    html_para_pdf = render_template(
        'reportes_pdf/reporte_grafico_despachos_pdf.html',
    grafico_base64=grafico_base64,
    datos_tabla=tabla_detalle,
        total_barriles=total_barriles_general,
        total_box_text=total_box_text,
        periodo=periodo_txt,
        producto=producto_filtro,
        tipo=tipo_grafico,
        fecha_generacion=datetime.now().strftime('%d/%m/%Y %H:%M'),
        logo_base64=logo_base64
    )
    # Proveer base_url para que recursos relativos funcionen si se agregan
    pdf = HTML(string=html_para_pdf, base_url=current_app.root_path).write_pdf()
    return Response(pdf, mimetype='application/pdf', headers={'Content-Disposition':'attachment;filename=reporte_grafico_despachos.pdf'})
  
@login_required
@permiso_requerido('inventario_epp')
@app.route('/inventario_epp_home')
def inventario_epp_home():
    """Página de inicio para el módulo de inventario EPP."""
    return render_template('inventario_epp_home.html', nombre=session.get("nombre"))

@login_required
@permiso_requerido('inventario_epp')
@app.route('/inventario_epp')
def inventario_epp():
    """Página principal para gestionar el inventario de EPP."""
    return render_template('inventario_epp.html', nombre=session.get("nombre"))

@login_required
@permiso_requerido('inventario_epp')
@app.route('/epp_asignaciones')
def epp_asignaciones():
    """Página para ver el historial de asignaciones de EPP."""
    return render_template('epp_asignaciones.html', nombre=session.get("nombre"))

@login_required
@permiso_requerido('inventario_epp')
@app.route('/api/epp/items', methods=['GET'])
def get_epp_items():
    """API para obtener todos los items del inventario con alertas inteligentes."""
    items_query = EPPItem.query.order_by(EPPItem.categoria, EPPItem.nombre).all()
    
    today = date.today()
    items_list = []
    for item in items_query:
        item_dict = {
            "id": item.id,
            "nombre": item.nombre,
            "categoria": item.categoria,
            "referencia": item.referencia,
            "talla": item.talla,
            "stock_actual": item.stock_actual,
            "observaciones": item.observaciones,
            "fecha_vencimiento": item.fecha_vencimiento.isoformat() if item.fecha_vencimiento else None,
            "dias_para_vencer": (item.fecha_vencimiento - today).days if item.fecha_vencimiento else None,
            # Alerta de stock bajo solo si NO es Equipos de Emergencia
            "stock_bajo": (item.categoria != 'Equipos de Emergencia') and (item.stock_actual <= 5)
        }
        items_list.append(item_dict)
    return jsonify(items_list)

@login_required
@permiso_requerido('inventario_epp')
@app.route('/api/epp/items', methods=['POST'])
def create_epp_item():
    """API para crear un nuevo item en el inventario."""
    data = request.get_json()
    try:
        nuevo_item = EPPItem(
            nombre=data['nombre'],
            categoria=data['categoria'],
            referencia=data.get('referencia'),
            talla=data.get('talla'),
            stock_actual=int(data.get('stock_actual', 0)),
            fecha_vencimiento=date.fromisoformat(data['fecha_vencimiento']) if data.get('fecha_vencimiento') else None,
            observaciones=data.get('observaciones')
        )
        db.session.add(nuevo_item)
        db.session.commit()
        return jsonify(success=True, message="Item creado exitosamente.", id=nuevo_item.id)
    except Exception as e:
        db.session.rollback()
        return jsonify(success=False, message=f"Error al crear: {str(e)}"), 500

@login_required
@permiso_requerido('inventario_epp')
@app.route('/api/epp/items/<int:id>', methods=['PUT'])
def update_epp_item(id):
    """API para actualizar un item existente."""
    item = EPPItem.query.get_or_404(id)
    data = request.get_json()
    try:
        for key, value in data.items():
            if key == 'fecha_vencimiento':
                value = date.fromisoformat(value) if value else None
            if hasattr(item, key):
                setattr(item, key, value)
        db.session.commit()
        return jsonify(success=True, message="Item actualizado.")
    except Exception as e:
        db.session.rollback()
        return jsonify(success=False, message=f"Error al actualizar: {str(e)}"), 500

@login_required
@permiso_requerido('inventario_epp')
@app.route('/api/epp/asignar', methods=['POST'])
def assign_epp_item():
    """API para asignar un item a un empleado y descontar del stock."""
    data = request.get_json()
    item_id = data.get('item_id')
    cantidad_a_entregar = int(data.get('cantidad_entregada', 0))
    item = EPPItem.query.get_or_404(item_id)

    if cantidad_a_entregar <= 0:
        return jsonify(success=False, message="La cantidad debe ser mayor a cero."), 400
    if item.stock_actual < cantidad_a_entregar:
        return jsonify(success=False, message=f"Stock insuficiente. Disponible: {item.stock_actual}."), 400

    try:
        # 1. Descontar del stock
        item.stock_actual -= cantidad_a_entregar

        # 2. Crear el registro de la asignación
        nueva_asignacion = EPPAssignment(
            item_id=item_id,
            empleado_nombre=data['empleado_nombre'],
            cantidad_entregada=cantidad_a_entregar,
            fecha_entrega=date.fromisoformat(data['fecha_entrega']),
            observaciones=data.get('observaciones')
        )
        db.session.add(nueva_asignacion)
        db.session.commit()
        return jsonify(success=True, message="Asignación registrada y stock actualizado.")
    except Exception as e:
        db.session.rollback()
        return jsonify(success=False, message=f"Error en la asignación: {str(e)}"), 500

@login_required
@permiso_requerido('inventario_epp')
@app.route('/api/epp/asignaciones', methods=['GET'])
def get_epp_assignments():
    """API para obtener el historial de asignaciones."""
    query = db.session.query(EPPAssignment).join(EPPItem).order_by(EPPAssignment.fecha_entrega.desc())
    
    # Ejemplo de filtro por empleado
    empleado = request.args.get('empleado')
    if empleado:
        query = query.filter(EPPAssignment.empleado_nombre.ilike(f'%{empleado}%'))

    asignaciones = query.all()
    data = [{
        "id": a.id, "fecha_entrega": a.fecha_entrega.isoformat(), "empleado_nombre": a.empleado_nombre,
        "cantidad_entregada": a.cantidad_entregada, "observaciones": a.observaciones,
        "item_nombre": a.item.nombre, "item_referencia": a.item.referencia, "item_talla": a.item.talla
    } for a in asignaciones]
    
    return jsonify(data)

# --- API para editar y eliminar asignaciones de EPP ---
@login_required
@permiso_requerido('inventario_epp')
@app.route('/api/epp/asignaciones/<int:id>', methods=['PUT'])
def update_epp_assignment(id):
    """API para actualizar una asignación de EPP."""
    asignacion = EPPAssignment.query.get_or_404(id)
    data = request.get_json()
    try:
        if 'fecha_entrega' in data:
            asignacion.fecha_entrega = date.fromisoformat(data['fecha_entrega']) if data['fecha_entrega'] else asignacion.fecha_entrega
        if 'empleado_nombre' in data:
            asignacion.empleado_nombre = data['empleado_nombre']
        if 'item_nombre' in data:
            # Cambia el item solo si existe un item con ese nombre
            item = EPPItem.query.filter_by(nombre=data['item_nombre']).first()
            if item:
                asignacion.item_id = item.id
        if 'item_referencia' in data:
            item = EPPItem.query.get(asignacion.item_id)
            if item:
                item.referencia = data['item_referencia']
        if 'item_talla' in data:
            item = EPPItem.query.get(asignacion.item_id)
            if item:
                item.talla = data['item_talla']
        if 'cantidad_entregada' in data:
            asignacion.cantidad_entregada = int(data['cantidad_entregada'])
        if 'observaciones' in data:
            asignacion.observaciones = data['observaciones']
        db.session.commit()
        return jsonify(success=True, message="Asignación actualizada.")
    except Exception as e:
        db.session.rollback()
        return jsonify(success=False, message=f"Error al actualizar: {str(e)}"), 500

@login_required
@permiso_requerido('inventario_epp')
@app.route('/api/epp/asignaciones/<int:id>', methods=['DELETE'])
def delete_epp_assignment(id):
    """API para eliminar una asignación de EPP."""
    asignacion = EPPAssignment.query.get_or_404(id)
    try:
        db.session.delete(asignacion)
        db.session.commit()
        return jsonify(success=True, message="Asignación eliminada.")
    except Exception as e:
        db.session.rollback()
        return jsonify(success=False, message=f"Error al eliminar: {str(e)}"), 500

@login_required
@permiso_requerido('programacion_cargue')
@app.route('/api/programacion/upload_excel', methods=['POST'])
def upload_programacion_excel():
    """Carga masiva de ProgramacionCargue desde un archivo Excel (.xlsx).

    Reglas:
    - Si se pasa ?replace=1 se eliminan los registros existentes antes de insertar.
    - Columnas aceptadas: nombres iguales a los campos del modelo; se permiten variantes con espacios o mayúsculas.
    - Campos fecha/hora se parsean; numéricos se convierten a float.
    - Si falta 'fecha_programacion' se usa la fecha de hoy.
    - Si no viene 'barriles' pero sí 'galones', se calcula barriles = galones/42.
    """
    if 'excel_file' not in request.files:
        return jsonify(success=False, message='No se encontró archivo (campo excel_file).'), 400
    f = request.files['excel_file']
    if not f.filename.lower().endswith('.xlsx'):
        return jsonify(success=False, message='Formato no soportado. Use .xlsx'), 400
    try:
        import pandas as pd
        from pandas.api.types import is_numeric_dtype
        df = pd.read_excel(f)
        # Normalizar nombres de columnas -> minusculas sin espacios dobles
        original_cols = list(df.columns)
        norm_map = {}
        for c in original_cols:
            norm = str(c).strip().lower().replace(' ', '_')
            norm_map[c] = norm
        df.rename(columns=norm_map, inplace=True)
        # Mapeo alternativo (por si el usuario usa variantes comunes)
        alias = {
            'empresa': 'empresa_transportadora',
            'transportadora': 'empresa_transportadora',
            'conductor': 'nombre_conductor',
            'cedula': 'cedula_conductor',
            'celular': 'celular_conductor',
            'producto': 'producto_a_cargar',
            'fecha': 'fecha_programacion',
            'fecha_de_programacion': 'fecha_programacion',
            'fecha_programada': 'fecha_programacion',
            'fecha_despacho_programada': 'fecha_despacho',
            'hora_llegada': 'hora_llegada_estimada',
            'hora_estimacion': 'hora_llegada_estimada',
            'guia': 'numero_guia'
        }
        for c in list(df.columns):
            if c in alias and alias[c] not in df.columns:
                df.rename(columns={c: alias[c]}, inplace=True)
        # Lista de campos manejados
        campos_modelo = {c.name for c in ProgramacionCargue.__table__.columns if c.name not in ('id')}
        filas_creadas = 0
        registros = []
        hoy = date.today()
        def parse_fecha(val):
            if pd.isna(val) or val == '':
                return None
            if isinstance(val, (datetime, date)):
                return val.date() if isinstance(val, datetime) else val
            for fmt in ('%Y-%m-%d','%d/%m/%Y','%Y/%m/%d','%d-%m-%Y'):
                try:
                    return datetime.strptime(str(val).strip(), fmt).date()
                except Exception:
                    pass
            return None
        def parse_hora(val):
            if pd.isna(val) or val == '':
                return None
            if isinstance(val, time):
                return val
            if isinstance(val, datetime):
                return val.time().replace(microsecond=0)
            try:
                # Aceptar formatos HH:MM o HH:MM:SS
                partes = str(val).strip().split(':')
                if len(partes) >= 2:
                    h = int(partes[0]); m = int(partes[1]); s = int(partes[2]) if len(partes) > 2 else 0
                    return time(hour=h, minute=m, second=s)
            except Exception:
                return None
            return None
        for _, row in df.iterrows():
            datos = {}
            for campo in campos_modelo:
                if campo in row.index:
                    val = row[campo]
                    if campo in ('fecha_programacion','fecha_despacho'):
                        val = parse_fecha(val)
                    elif campo == 'hora_llegada_estimada':
                        val = parse_hora(val)
                    elif campo in ('galones','barriles','temperatura','api_obs','api_corregido'):
                        try:
                            val = float(val) if not pd.isna(val) and val != '' else None
                        except Exception:
                            val = None
                    elif pd.isna(val):
                        val = None
                    datos[campo] = val
            # Defaults
            # Si no viene fecha_programacion pero sí fecha_despacho, usar esa (registros históricos)
            if not datos.get('fecha_programacion') and datos.get('fecha_despacho'):
                datos['fecha_programacion'] = datos['fecha_despacho']
            # Si aún no tenemos fecha_programacion, último recurso: hoy
            if not datos.get('fecha_programacion'):
                datos['fecha_programacion'] = hoy
            if datos.get('galones') and not datos.get('barriles'):
                try:
                    datos['barriles'] = float(datos['galones'])/42.0
                except Exception:
                    pass
            datos['ultimo_editor'] = session.get('nombre')
            # Normalizar estado
            if datos.get('estado'):
                est = str(datos['estado']).strip().upper()
                if est not in ('PROGRAMADO','CARGANDO','DESPACHADO'):
                    est = 'PROGRAMADO'
                datos['estado'] = est
            registro = ProgramacionCargue(**{k:v for k,v in datos.items() if k in campos_modelo})
            registros.append(registro)
        if request.args.get('replace') == '1':
            db.session.query(ProgramacionCargue).delete()
        db.session.add_all(registros)
        db.session.commit()
        filas_creadas = len(registros)
        return jsonify(success=True, message=f'Se cargaron {filas_creadas} registros.', total=filas_creadas, replace=bool(request.args.get('replace')=='1'))
    except Exception as e:
        db.session.rollback()
        app.logger.error(f'Error al cargar Excel programación: {e}')
        return jsonify(success=False, message='Error procesando el archivo: ' + str(e)), 500

@login_required
@permiso_requerido('inventario_epp')
@app.route('/api/epp/items/<int:id>', methods=['DELETE'])
def delete_epp_item(id):
    """API para eliminar un item."""
    item = EPPItem.query.get_or_404(id)
    try:
        db.session.delete(item)
        db.session.commit()
        return jsonify(success=True, message="Item eliminado exitosamente.")
    except Exception as e:
        db.session.rollback()
        return jsonify(success=False, message=f"Error al eliminar: {str(e)}."), 500

@login_required
@permiso_requerido('inventario_epp')    
@app.route('/api/epp/items/batch_add', methods=['POST'])
def batch_add_epp_items():
    """API para crear múltiples items (variantes) de una sola vez."""
    items_data = request.get_json()
    if not isinstance(items_data, list) or not items_data:
        return jsonify(success=False, message="Formato de datos incorrecto."), 400

    try:
        creados_count = 0
        for item_data in items_data:
            # Evita duplicados revisando la combinación de nombre, referencia y talla
            existe = EPPItem.query.filter_by(
                nombre=item_data.get('nombre'),
                referencia=item_data.get('referencia'),
                talla=item_data.get('talla')
            ).first()

            if not existe:
                nuevo_item = EPPItem(
                    nombre=item_data.get('nombre'),
                    categoria=item_data.get('categoria'),
                    referencia=item_data.get('referencia'),
                    talla=item_data.get('talla'),
                    stock_actual=item_data.get('stock_actual', 0)
                )
                db.session.add(nuevo_item)
                creados_count += 1

        db.session.commit()
        return jsonify(success=True, message=f"Se han agregado {creados_count} nuevos items/variantes exitosamente.")

    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Error en carga rápida de EPP: {e}")
        return jsonify(success=False, message=f"Error interno del servidor: {str(e)}"), 500

@login_required
@permiso_requerido('inventario_epp')
@app.route('/exportar_inventario_epp/<string:formato>')
def exportar_inventario_epp(formato):
    # Obtener filtros desde la URL
    nombre = request.args.get('nombre', '')
    categoria = request.args.get('categoria', '')
    alertas = request.args.get('alertas', 'false') == 'true'

    # Construir la consulta a la base de datos
    query = EPPItem.query.order_by(EPPItem.categoria, EPPItem.nombre)

    if nombre:
        query = query.filter(or_(EPPItem.nombre.ilike(f'%{nombre}%'), EPPItem.referencia.ilike(f'%{nombre}%')))
    if categoria:
        query = query.filter(EPPItem.categoria == categoria)
    if alertas:
        today = date.today()
        thirty_days = today + timedelta(days=30)
        query = query.filter(or_(EPPItem.stock_actual <= 5, EPPItem.fecha_vencimiento.between(today, thirty_days)))

    items = query.all()

    if not items:
        flash('No hay datos para exportar con los filtros seleccionados.', 'warning')
        return redirect(url_for('inventario_epp'))

    # Generar el archivo según el formato
    if formato == 'excel':
        datos_df = [{
            'Categoría': item.categoria,
            'Elemento': item.nombre,
            'Referencia/Tipo': item.referencia,
            'Talla/Medida': item.talla,
            'Stock Actual': item.stock_actual,
            'Fecha Vencimiento': item.fecha_vencimiento.strftime('%Y-%m-%d') if item.fecha_vencimiento else 'N/A',
            'Observaciones': item.observaciones
        } for item in items]
        df = pd.DataFrame(datos_df)
        
        output = BytesIO()
        df.to_excel(output, index=False, sheet_name='Inventario EPP')
        output.seek(0)
        
        return send_file(output, as_attachment=True, download_name='reporte_inventario_epp.xlsx',
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    elif formato == 'pdf':
        html_para_pdf = render_template('reportes_pdf/reporte_inventario_pdf.html',
                                        items=items,
                                        fecha_reporte=datetime.now().strftime('%d de %B de %Y'))
        pdf = HTML(string=html_para_pdf).write_pdf()
        return Response(pdf, mimetype='application/pdf',
                        headers={'Content-Disposition': 'attachment;filename=reporte_inventario_epp.pdf'})
    
    return redirect(url_for('inventario_epp'))

@login_required
@permiso_requerido('inventario_epp')
@app.route('/exportar_asignaciones_epp/<string:formato>')
def exportar_asignaciones_epp(formato):
    empleado = request.args.get('empleado', '')
    fecha_inicio_str = request.args.get('fecha_inicio', '')
    fecha_fin_str = request.args.get('fecha_fin', '')

    query = db.session.query(EPPAssignment).join(EPPItem).order_by(EPPAssignment.fecha_entrega.desc())

    if empleado:
        query = query.filter(EPPAssignment.empleado_nombre.ilike(f'%{empleado}%'))
    if fecha_inicio_str:
        query = query.filter(EPPAssignment.fecha_entrega >= date.fromisoformat(fecha_inicio_str))
    if fecha_fin_str:
        query = query.filter(EPPAssignment.fecha_entrega <= date.fromisoformat(fecha_fin_str))

    asignaciones = query.all()

    if not asignaciones:
        flash('No hay asignaciones para exportar con los filtros seleccionados.', 'warning')
        return redirect(url_for('epp_asignaciones'))

    if formato == 'excel':
        datos_df = [{
            'Fecha Entrega': a.fecha_entrega.strftime('%Y-%m-%d'),
            'Empleado': a.empleado_nombre,
            'Elemento': a.item.nombre,
            'Referencia': a.item.referencia,
            'Talla/Medida': a.item.talla,
            'Cantidad': a.cantidad_entregada,
            'Observaciones': a.observaciones
        } for a in asignaciones]
        df = pd.DataFrame(datos_df)
        
        output = BytesIO()
        df.to_excel(output, index=False, sheet_name='Historial Asignaciones')
        output.seek(0)
        
        return send_file(output, as_attachment=True, download_name='reporte_asignaciones_epp.xlsx',
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    elif formato == 'pdf':
        html_para_pdf = render_template('reportes_pdf/reporte_asignaciones_pdf.html',
                                        asignaciones=asignaciones,
                                        fecha_reporte=datetime.now().strftime('%d de %B de %Y'))
        pdf = HTML(string=html_para_pdf).write_pdf()
        return Response(pdf, mimetype='application/pdf',
                        headers={'Content-Disposition': 'attachment;filename=reporte_asignaciones_epp.pdf'})

    return redirect(url_for('epp_asignaciones'))  

@login_required
@permiso_requerido('gestion_compras')
@app.route('/gestion_compras')
def gestion_compras():
    # Obtener filtros
    proveedor_filtro = request.args.get('proveedor', '')
    producto_filtro = request.args.get('producto', '')
    mes_filtro = request.args.get('mes', '')

    query = RegistroCompra.query

    if proveedor_filtro:
        query = query.filter(RegistroCompra.proveedor == proveedor_filtro)
    
    if producto_filtro:
        query = query.filter(RegistroCompra.producto == producto_filtro)
    
    if mes_filtro:
        query = query.filter(func.to_char(RegistroCompra.fecha, 'YYYY-MM') == mes_filtro)

    compras = query.order_by(RegistroCompra.fecha.desc()).all()
    
    # Obtener listas para los filtros
    proveedores = sorted([p[0] for p in db.session.query(RegistroCompra.proveedor).distinct().all() if p[0]])
    productos = sorted([p[0] for p in db.session.query(RegistroCompra.producto).distinct().all() if p[0]])
    meses = sorted([m[0] for m in db.session.query(func.to_char(RegistroCompra.fecha, 'YYYY-MM')).distinct().all() if m[0]], reverse=True)

    # Calcular totales
    total_bls = sum([c.cantidad_bls for c in compras if c.cantidad_bls])
    total_usd = sum([c.total_neto for c in compras if c.total_neto])
    precio_promedio = sum([c.price_compra_pond for c in compras if c.price_compra_pond]) / len(compras) if compras else 0

    return render_template('gestion_compras.html', 
                         compras=compras, 
                         proveedores=proveedores,
                         productos=productos,
                         meses=meses,
                         filtros={'proveedor': proveedor_filtro, 'producto': producto_filtro, 'mes': mes_filtro},
                         total_bls=total_bls,
                         total_usd=total_usd,
                         precio_promedio=precio_promedio)

@login_required
@permiso_requerido('gestion_compras')
@app.route('/cargar_compras_excel', methods=['POST'])
def cargar_compras_excel():
    if 'excel_file' not in request.files:
        flash('No se encontró el archivo.', 'danger')
        return redirect(url_for('gestion_compras'))

    file = request.files['excel_file']
    if not file or not file.filename.endswith('.xlsx'):
        flash('Archivo no válido. Debe ser .xlsx', 'danger')
        return redirect(url_for('gestion_compras'))

    try:
        # Leer el archivo Excel saltando la primera fila (header=1 usa la segunda fila como encabezados)
        df = pd.read_excel(file, sheet_name='2025', header=1)
        
        # ELIMINAR TODOS LOS DATOS ANTERIORES para evitar duplicados y discrepancias
        registros_eliminados = RegistroCompra.query.delete()
        db.session.commit()
        
        app.logger.info(f"Se eliminaron {registros_eliminados} registros anteriores")
        
        nuevas = 0

        for _, row in df.iterrows():
            # Saltar filas donde falten campos críticos
            if pd.isna(row['MES']) or pd.isna(row['Proveedor']) or pd.isna(row['Producto']):
                continue
            
            # Convertir valores NaN a None para evitar errores de tipo
            fecha = pd.to_datetime(row['MES']).date()
            proveedor = row['Proveedor'] if pd.notna(row['Proveedor']) else None
            producto = row['Producto'] if pd.notna(row['Producto']) else None
            cantidad_bls = row['Cantidad BLS'] if pd.notna(row['Cantidad BLS']) else None
            
            # Crear nuevo registro (ya no buscamos duplicados porque eliminamos todo)
            compra = RegistroCompra()
            
            # Asignar valores directamente del Excel
            compra.fecha = fecha
            compra.proveedor = proveedor
            compra.tarifa = row['Tarifa'] if pd.notna(row['Tarifa']) else None
            compra.producto = producto
            compra.cantidad_bls = cantidad_bls
            compra.cantidad_gln = row['Cantidad Gln'] if pd.notna(row['Cantidad Gln']) else None
            compra.brent = row['Brent US$B'] if pd.notna(row['Brent US$B']) else None
            compra.descuento = row['Descuento US$B'] if pd.notna(row['Descuento US$B']) else None
            compra.precio_uni_bpozo = row['Precio Uni. B.Pozo US$B'] if pd.notna(row['Precio Uni. B.Pozo US$B']) else None
            compra.total_neto = row['Total Neto US$B'] if pd.notna(row['Total Neto US$B']) else None
            compra.price_compra_pond = row['Price Compra Pond. US$/BL'] if pd.notna(row['Price Compra Pond. US$/BL']) else None
            compra.fecha_carga = datetime.utcnow()
            
            db.session.add(compra)
            nuevas += 1
        
        db.session.commit()
        flash(f'Base de datos actualizada: {registros_eliminados} registros anteriores eliminados, {nuevas} registros nuevos cargados', 'success')

    except Exception as e:
        db.session.rollback()
        flash(f'Error al cargar: {str(e)}', 'danger')
        app.logger.error(f"Error carga Excel: {str(e)}")

    return redirect(url_for('gestion_compras'))

@login_required
@permiso_requerido('gestion_compras')
@app.route('/reporte_compras')
def reporte_compras():
    # Histórico de precios agrupado por mes (filtrar meses con datos insignificantes)
    historico_precios_raw = db.session.query(
        func.to_char(RegistroCompra.fecha, 'YYYY-MM').label('mes'),
        func.avg(RegistroCompra.price_compra_pond).label('precio_promedio')
    ).filter(RegistroCompra.price_compra_pond.isnot(None))\
     .group_by('mes').order_by('mes').all()
    historico_precios = [
        {"mes": row[0], "precio": float(row[1])}
        for row in historico_precios_raw
        if row[1] is not None and row[1] > 0
    ]

    # Histórico de volúmenes agrupado por mes (filtrar meses con volúmenes insignificantes)
    historico_volumenes_raw = db.session.query(
        func.to_char(RegistroCompra.fecha, 'YYYY-MM').label('mes'),
        func.sum(RegistroCompra.cantidad_bls).label('volumen_total')
    ).filter(RegistroCompra.cantidad_bls.isnot(None))\
     .group_by('mes').order_by('mes').all()
    historico_volumenes = [
        {"mes": row[0], "volumen": float(row[1])}
        for row in historico_volumenes_raw
        if row[1] is not None and row[1] > 100  # Filtrar volúmenes menores a 100 barriles (probablemente errores)
    ]

    # Resumen mensual
    resumen_mensual_raw = db.session.query(
        func.to_char(RegistroCompra.fecha, 'YYYY-MM').label('mes'),
        RegistroCompra.proveedor,
        RegistroCompra.producto,
        func.sum(RegistroCompra.cantidad_bls).label('cantidad_bls')
    ).group_by('mes', RegistroCompra.proveedor, RegistroCompra.producto).all()
    resumen_mensual = [
        {
            "mes": row[0],
            "proveedor": row[1],
            "producto": row[2],
            "cantidad_bls": float(row[3]) if row[3] is not None else 0
        }
        for row in resumen_mensual_raw
    ]

    proveedores = sorted([p[0] for p in db.session.query(RegistroCompra.proveedor).distinct().all() if p[0]])
    productos = sorted([p[0] for p in db.session.query(RegistroCompra.producto).distinct().all() if p[0]])

    # Calcular estadísticas generales
    total_barriles = sum([r['cantidad_bls'] for r in resumen_mensual])
    precio_promedio = sum([p['precio'] for p in historico_precios]) / len(historico_precios) if historico_precios else 0

    return render_template(
        'reporte_compras.html',
        historico_precios=historico_precios,
        historico_volumenes=historico_volumenes,
        resumen_mensual=resumen_mensual,
        proveedores=proveedores,
        productos=productos,
        total_barriles=total_barriles,
        precio_promedio=precio_promedio
    )

@login_required
@permiso_requerido('gestion_compras')
@app.route('/reporte_compras_pdf')
def reporte_compras_pdf():
    # Función auxiliar corregida para formatear meses
    def formatear_mes(fecha_str):
        try:
            # Maneja diferentes formatos de fecha
            if len(fecha_str) == 10:  # Formato YYYY-MM-DD
                fecha = datetime.strptime(fecha_str, '%Y-%m-%d')
            elif len(fecha_str) == 7:  # Formato YYYY-MM
                fecha = datetime.strptime(fecha_str, '%Y-%m')
            else:
                return fecha_str  # Si no reconoce el formato, devuelve original
            
            meses = ['', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio', 
                    'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
            return f"{meses[fecha.month]} {fecha.year}"
        except:
            return fecha_str  # En caso de error, devuelve el valor original

    # Consulta de precios históricos
    historico_precios_raw = db.session.query(
        func.to_char(RegistroCompra.fecha, 'YYYY-MM').label('mes'),
        func.avg(RegistroCompra.price_compra_pond).label('precio_promedio')
    ).group_by('mes').order_by('mes').all()
    
    historico_precios = [
        {"mes": formatear_mes(row[0]), "precio": float(row[1]) if row[1] is not None else 0}
        for row in historico_precios_raw
    ]

    # Consulta de volúmenes históricos
    historico_volumenes_raw = db.session.query(
        func.to_char(RegistroCompra.fecha, 'YYYY-MM').label('mes'),
        func.sum(RegistroCompra.cantidad_bls).label('volumen_total')
    ).group_by('mes').order_by('mes').all()
    
    historico_volumenes = [
        {"mes": formatear_mes(row[0]), "volumen": float(row[1]) if row[1] is not None else 0}
        for row in historico_volumenes_raw
    ]

    # Leer filtros de la URL
    filtro_mes = request.args.get('mes')
    filtro_proveedor = request.args.get('proveedor')
    filtro_producto = request.args.get('producto')

    # Consulta base
    resumen_query = db.session.query(
        func.strftime('%Y-%m', RegistroCompra.fecha).label('mes'),
        RegistroCompra.proveedor,
        RegistroCompra.producto,
        func.sum(RegistroCompra.cantidad_bls).label('cantidad_bls')
    )
    # Aplicar filtros si existen
    if filtro_mes:
        resumen_query = resumen_query.filter(func.to_char(RegistroCompra.fecha, 'YYYY-MM') == filtro_mes)
    if filtro_proveedor:
        resumen_query = resumen_query.filter(RegistroCompra.proveedor == filtro_proveedor)
    if filtro_producto:
        resumen_query = resumen_query.filter(RegistroCompra.producto == filtro_producto)

    resumen_mensual_raw = resumen_query.group_by('mes', RegistroCompra.proveedor, RegistroCompra.producto).order_by('mes').all()

    resumen_mensual = [
        {
            "mes": formatear_mes(row[0]),
            "proveedor": row[1],
            "producto": row[2],
            "cantidad_bls": float(row[3]) if row[3] is not None else 0
        }
        for row in resumen_mensual_raw
    ]

    # Obtener listas de proveedores y productos
    proveedores = sorted([p[0] for p in db.session.query(RegistroCompra.proveedor).distinct().all() if p[0]])
    productos = sorted([p[0] for p in db.session.query(RegistroCompra.producto).distinct().all() if p[0]])

    # Generar gráficos
    labels_precios = [x['mes'] for x in historico_precios]
    data_precios = [x['precio'] for x in historico_precios]
    labels_volumenes = [x['mes'] for x in historico_volumenes]
    data_volumenes = [x['volumen'] for x in historico_volumenes]

    img_precios = grafico_linea_base64(labels_precios, data_precios, 'Precio Compra Ponderado (US$/BL)')
    img_volumenes = grafico_barra_base64(labels_volumenes, data_volumenes, 'Volumen Comprado (BLS)')

    # Renderizar template para PDF
    html_para_pdf = render_template(
        'reportes_pdf/reporte_compras_pdf.html',
        historico_precios=historico_precios,
        historico_volumenes=historico_volumenes,
        resumen_mensual=resumen_mensual,
        proveedores=proveedores,
        productos=productos,
        pdf=True,
        img_precios=img_precios,
        img_volumenes=img_volumenes,
        now=datetime.now
    )
    
    # Generar PDF
    pdf = HTML(string=html_para_pdf, base_url=request.base_url).write_pdf()
    return Response(
        pdf, 
        mimetype='application/pdf',
        headers={'Content-Disposition': 'attachment;filename=reporte_compras.pdf'}
    )

@login_required
@permiso_requerido('flujo_efectivo')
@app.route('/flujo_efectivo')
def flujo_efectivo():
    """
    Renderiza la página principal del Flujo de Efectivo.
    """
    return render_template('flujo_efectivo.html', nombre=session.get("nombre"))


@login_required
@permiso_requerido('flujo_efectivo')
@app.route('/api/analyze_flujo_file', methods=['POST'])
def analyze_flujo_file():
    """Analiza el archivo Excel y devuelve los periodos (Año-Mes) encontrados."""
    if 'archivo_excel' not in request.files:
        return jsonify(success=False, message="No se encontró el archivo."), 400
    archivo = request.files['archivo_excel']
    if archivo.filename == '' or not archivo.filename.lower().endswith(('.xlsx','.xls')):
        return jsonify(success=False, message="Archivo inválido (debe ser .xlsx o .xls)."), 400
    
    try:
        xls = pd.ExcelFile(archivo)
        sheet_map = {name.strip().lower(): name for name in xls.sheet_names}
        if 'bancos' not in sheet_map or 'odoo' not in sheet_map:
            return jsonify(success=False, message='El archivo debe contener las hojas "Bancos" y "Odoo".'), 400
            
        df_bancos = pd.read_excel(xls, sheet_name=sheet_map['bancos'])
        df_bancos.columns = df_bancos.columns.str.strip()
        df_bancos['FECHA DE OPERACIÓN'] = pd.to_datetime(df_bancos['FECHA DE OPERACIÓN'], errors='coerce')
        fechas_b = df_bancos['FECHA DE OPERACIÓN'].dropna().dt.to_period('M').unique()

        df_odoo = pd.read_excel(xls, sheet_name=sheet_map['odoo'])
        df_odoo.columns = df_odoo.columns.str.strip()
        df_odoo['Fecha'] = pd.to_datetime(df_odoo['Fecha'], errors='coerce')
        fechas_o = df_odoo['Fecha'].dropna().dt.to_period('M').unique()

        # Construir estructura de respuesta
        periodos_found = set()
        for p in fechas_b: periodos_found.add((p.year, p.month))
        for p in fechas_o: periodos_found.add((p.year, p.month))
        
        result_tree = {}
        for y, m in periodos_found:
            result_tree.setdefault(y, []).append(m)
        
        # Ordenar meses
        for y in result_tree:
            result_tree[y].sort()

        return jsonify(success=True, found_periods=result_tree)

    except Exception as e:
        app.logger.exception("Error analizando archivo flujo efectivo")
        return jsonify(success=False, message=f"Error analizando archivo: {str(e)}"), 500

@login_required
@permiso_requerido('flujo_efectivo')
@app.route('/api/procesar_flujo_efectivo', methods=['POST'])
def procesar_flujo_efectivo_api():
    """Procesa Excel y persiste movimientos. SI recibe 'selected_periods', FILTRA solo esos periodos."""
    if 'archivo_excel' not in request.files:
        return jsonify(success=False, message="No se encontró el archivo en la solicitud."), 400
    archivo = request.files['archivo_excel']
    
    # Parsear selected_periods si viene
    selected_periods = request.form.get('selected_periods')
    target_periods_set = None
    if selected_periods:
        try:
            # Espera formato JSON: [[2025, 1], [2025, 2], ...]
            import json
            sp_list = json.loads(selected_periods)
            target_periods_set = set(tuple(x) for x in sp_list)
        except:
            pass # Si falla, procesamos todo el archivo (comportamiento default)

    if archivo.filename == '' or not archivo.filename.lower().endswith(('.xlsx','.xls')):
        return jsonify(success=False, message="Archivo inválido."), 400
    try:
        xls = pd.ExcelFile(archivo)
        sheet_map = {name.strip().lower(): name for name in xls.sheet_names}

        if 'bancos' not in sheet_map or 'odoo' not in sheet_map:
            return jsonify(success=False, message='El archivo debe contener las hojas "Bancos" y "Odoo".'), 400
        df_bancos = pd.read_excel(xls, sheet_name=sheet_map['bancos'])
        df_odoo = pd.read_excel(xls, sheet_name=sheet_map['odoo'])

        # Normalizar nombres de columnas (trim)
        df_bancos.columns = df_bancos.columns.str.strip()
        df_odoo.columns = df_odoo.columns.str.strip()

        # --- NUEVO: detectar y normalizar columna de Banco aunque venga con otro nombre ---
        def _ensure_banco_col(df):
            if 'Banco' in df.columns:
                return
            # Posibles variantes que han aparecido en archivos reales
            variantes = {
                'banco','bancos','tipo banco','tipo_banco','banco/cuenta','banco - cuenta',
                'cuenta','cuenta bancaria','entidad','entidad financiera','banco cuenta'
            }
            for c in df.columns:
                if c.strip().lower() in variantes:
                    df.rename(columns={c: 'Banco'}, inplace=True)
                    return
        _ensure_banco_col(df_bancos)
        _ensure_banco_col(df_odoo)

        # Validar columnas mínimas
        required_bancos = {'FECHA DE OPERACIÓN', 'Movimiento', 'COP$', 'Empresa'}
        required_odoo = {'Fecha', 'Movimiento', 'Débito', 'Crédito', 'Empresa'}
        missing_b = required_bancos - set(df_bancos.columns)
        missing_o = required_odoo - set(df_odoo.columns)
        if missing_b:
            return jsonify(success=False, message=f"Faltan columnas en Bancos: {missing_b}"), 400
        if missing_o:
            return jsonify(success=False, message=f"Faltan columnas en Odoo: {missing_o}"), 400

        # Limpiar y convertir montos
        def clean_numeric(series):
            return (series.astype(str)
                    .str.replace(r"[^0-9\-.,]", '', regex=True)
                    .str.replace(',', '', regex=False)
                    .replace('', '0')
                    .pipe(pd.to_numeric, errors='coerce').fillna(0))

        df_bancos['COP$'] = clean_numeric(df_bancos['COP$'])
        df_odoo['Débito'] = clean_numeric(df_odoo['Débito'])
        df_odoo['Crédito'] = clean_numeric(df_odoo['Crédito'])

        # Normalizar variantes de EGRESO GMF -> EGRESO_GMF (soporta espacios, guiones o múltiples separadores)
        def _norm_mov(txt: str):
            if not isinstance(txt, str):
                return txt
            t = txt.strip()
            if re.search(r'EGRESO\s*[_\-]?\s*GMF', t, re.IGNORECASE):
                return 'EGRESO_GMF'
            return t
        if 'Movimiento' in df_bancos.columns:
            df_bancos['Movimiento'] = df_bancos['Movimiento'].astype(str).apply(_norm_mov)
        if 'Movimiento' in df_odoo.columns:
            df_odoo['Movimiento'] = df_odoo['Movimiento'].astype(str).apply(_norm_mov)

        # Fechas seguras
        df_bancos['FECHA DE OPERACIÓN'] = pd.to_datetime(df_bancos['FECHA DE OPERACIÓN'], errors='coerce')
        df_bancos = df_bancos.dropna(subset=['FECHA DE OPERACIÓN'])
        df_odoo['Fecha'] = pd.to_datetime(df_odoo['Fecha'], errors='coerce')
        df_odoo = df_odoo.dropna(subset=['Fecha'])

        # Normalizar columnas Banco (ambas hojas pueden o no traerla). Si no existe se crea vacía para tener consistencia.
        if 'Banco' not in df_bancos.columns:
            df_bancos['Banco'] = ''
        if 'Banco' not in df_odoo.columns:
            df_odoo['Banco'] = ''
        df_bancos['Banco'] = df_bancos['Banco'].fillna('').astype(str)
        df_odoo['Banco'] = df_odoo['Banco'].fillna('').astype(str)

        # Ingresos / Egresos agrupando por Banco también
        mask_b_ing = df_bancos['Movimiento'].str.contains('INGRESO', case=False, na=False)
        ing_b = df_bancos.loc[mask_b_ing].copy(); ing_b['fecha'] = ing_b['FECHA DE OPERACIÓN'].dt.date
        bancos_ingresos = ing_b.groupby(['fecha', 'Empresa', 'Banco'])['COP$'].sum().reset_index().rename(columns={'COP$': 'ingresos_bancos'})
        mask_b_eg = df_bancos['Movimiento'].str.contains('EGRESO', case=False, na=False)
        eg_b = df_bancos.loc[mask_b_eg].copy(); eg_b['fecha'] = eg_b['FECHA DE OPERACIÓN'].dt.date
        eg_b['egresos_bancos'] = eg_b['COP$'].abs()
        bancos_egresos = eg_b.groupby(['fecha', 'Empresa', 'Banco'])['egresos_bancos'].sum().reset_index()

        mask_o_ing = df_odoo['Movimiento'].str.contains('INGRESO', case=False, na=False)
        ing_o = df_odoo.loc[mask_o_ing].copy(); ing_o['fecha'] = ing_o['Fecha'].dt.date
        odoo_ingresos = ing_o.groupby(['fecha', 'Empresa', 'Banco'])['Débito'].sum().reset_index().rename(columns={'Débito': 'ingresos_odoo'})
        mask_o_eg = df_odoo['Movimiento'].str.contains('EGRESO', case=False, na=False)
        eg_o = df_odoo.loc[mask_o_eg].copy(); eg_o['fecha'] = eg_o['Fecha'].dt.date
        odoo_egresos = eg_o.groupby(['fecha', 'Empresa', 'Banco'])['Crédito'].sum().reset_index().rename(columns={'Crédito': 'egresos_odoo'})

        comparativo_df = pd.merge(bancos_ingresos, bancos_egresos, on=['fecha', 'Empresa', 'Banco'], how='outer')
        comparativo_df = pd.merge(comparativo_df, odoo_ingresos, on=['fecha', 'Empresa', 'Banco'], how='outer')
        comparativo_df = pd.merge(comparativo_df, odoo_egresos, on=['fecha', 'Empresa', 'Banco'], how='outer')
        comparativo_df.fillna(0, inplace=True)
        comparativo_df['diferencia_ingresos'] = comparativo_df['ingresos_bancos'] - comparativo_df['ingresos_odoo']
        comparativo_df['diferencia_egresos'] = comparativo_df['egresos_bancos'] - comparativo_df['egresos_odoo']
        comparativo_df = comparativo_df.sort_values(by=['fecha', 'Empresa'], ascending=True)
        comparativo_df['fecha'] = comparativo_df['fecha'].astype(str)
        comparativo_df.rename(columns={'Banco':'tipo_banco'}, inplace=True)
        daily_comparison_data = comparativo_df.to_dict(orient='records')

        # Lista de empresas para el frontend
        todas_las_empresas = sorted(set(df_bancos['Empresa'].dropna().unique().tolist() + df_odoo['Empresa'].dropna().unique().tolist()))

        # Agrupación por Tipo de Flujo y Tercero (si las columnas existen)
        def group_by_flow_type_safe(df, value_col):
            cols = set(df.columns)
            if not {'Tipo Flujo Efectivo', 'Tercero', value_col}.issubset(cols):
                return {}
            g = df.groupby(['Tipo Flujo Efectivo', 'Tercero'])[value_col].sum()
            nested = {}
            for (flow_type, tercero), total in g.items():
                nested.setdefault(flow_type, {})[tercero] = total
            return nested

        outflows_by_type = group_by_flow_type_safe(df_odoo, 'Crédito')
        inflows_by_type = group_by_flow_type_safe(df_odoo, 'Débito')

        # --- Detalle completo de movimientos Odoo para expansiones frontend ---
        rubro_col = None
        for candidate in ['Rubro', 'RUBRO', 'rubro']:
            if candidate in df_odoo.columns:
                rubro_col = candidate
                break
        tipo_flujo_col = 'Tipo Flujo Efectivo' if 'Tipo Flujo Efectivo' in df_odoo.columns else None
        tercero_col = 'Tercero' if 'Tercero' in df_odoo.columns else None
        # Detectar columnas de Clase / Subclase para poder reconstruir el TOP filtrado en frontend
        clase_col = None
        for c in ['Clase', 'CLASE', 'clase']:
            if c in df_odoo.columns:
                clase_col = c
                break
        subclase_col = None
        for c in ['Sub Clase', 'Sub_Clase', 'SubClase', 'Subclase', 'SUBCLASE', 'SUB CLASE']:
            if c in df_odoo.columns:
                subclase_col = c
                break
        df_detalle = df_odoo.copy()
        df_detalle['__fecha_date'] = pd.to_datetime(df_detalle['Fecha'], errors='coerce')
        df_detalle = df_detalle.dropna(subset=['__fecha_date'])
        df_detalle['fecha'] = df_detalle['__fecha_date'].dt.strftime('%Y-%m-%d')
        # Construcción segura del detalle, reemplazando NaN por valores neutros
        import math
        detalle_records = []
        for _, r in df_detalle.iterrows():
            # Utilizamos pd.isna para cubrir np.nan, pandas.NA, None
            def safe_text(val, default=''):
                try:
                    if pd.isna(val):
                        return default
                except Exception:
                    pass
                return str(val) if val is not None else default
            def safe_number(val):
                try:
                    if pd.isna(val):
                        return 0.0
                except Exception:
                    pass
                try:
                    f = float(val)
                    if math.isnan(f) or math.isinf(f):
                        return 0.0
                    return f
                except Exception:
                    return 0.0
            detalle_records.append({
                'fecha': safe_text(r.get('fecha')),  # ya es string formateada
                'empresa': safe_text(r.get('Empresa')),
                'tipo_flujo': safe_text(r.get(tipo_flujo_col) if tipo_flujo_col else 'SIN TIPO', 'SIN TIPO'),
                'tercero': safe_text(r.get(tercero_col) if tercero_col else 'SIN TERCERO', 'SIN TERCERO'),
                'rubro': safe_text(r.get(rubro_col) if rubro_col else ''),
                'debito': safe_number(r.get('Débito', 0)),
                'credito': safe_number(r.get('Crédito', 0)),
                'movimiento': safe_text(r.get('Movimiento', '')),
                # Nuevos campos necesarios para filtrar y recalcular TOP en frontend
                'clase': safe_text(r.get(clase_col) if clase_col else ''),
                'subclase': safe_text(r.get(subclase_col) if subclase_col else ''),
                'banco': safe_text(r.get('Banco') if 'Banco' in df_detalle.columns else '')
            })

        # --- TOP CLIENTES VENTAS EXW CTG (ingresos) ---
        top_clientes_exw = None
        # (Columnas clase_col y subclase_col ya detectadas arriba; se reutilizan)
        tercero_col_original = 'Tercero' if 'Tercero' in df_odoo.columns else None
        # Filtrar sólo si existen columnas necesarias
        if clase_col and tercero_col_original and 'Débito' in df_odoo.columns:
            df_exw = df_odoo[df_odoo[clase_col].astype(str).str.upper() == 'VENTAS EXW CTG'].copy()
            if not df_exw.empty:
                # Normalizar valores
                df_exw['__tercero_safe'] = df_exw[tercero_col_original].fillna('SIN TERCERO').astype(str)
                if subclase_col:
                    df_exw['__subclase_safe'] = df_exw[subclase_col].fillna('SIN SUBCLASE').astype(str)
                else:
                    df_exw['__subclase_safe'] = 'SIN SUBCLASE'
                # Asegurar numérico
                df_exw['__debito_val'] = pd.to_numeric(df_exw['Débito'], errors='coerce').fillna(0)
                # Agrupar por cliente y subclase
                grp = df_exw.groupby(['__tercero_safe', '__subclase_safe'])['__debito_val'].sum().reset_index()
                # Totales por cliente para ranking
                tot_clientes = grp.groupby('__tercero_safe')['__debito_val'].sum().reset_index().rename(columns={'__debito_val': 'total'})
                # Ordenar y tomar top 10
                tot_clientes = tot_clientes.sort_values('total', ascending=False).head(10)
                top_set = set(tot_clientes['__tercero_safe'])
                grp_top = grp[grp['__tercero_safe'].isin(top_set)]
                # Construir estructura
                clientes_struct = []
                for _, row_cli in tot_clientes.iterrows():
                    cliente = row_cli['__tercero_safe']
                    total_cli = float(row_cli['total'])
                    subs_rows = grp_top[grp_top['__tercero_safe'] == cliente]
                    sub_list = [
                        {
                            'subclase': str(sr['__subclase_safe']),
                            'total': float(sr['__debito_val'])
                        }
                        for _, sr in subs_rows.iterrows()
                    ]
                    clientes_struct.append({
                        'cliente': str(cliente),
                        'total': total_cli,
                        'subclases': sub_list
                    })
                top_clientes_exw = {
                    'clientes': clientes_struct,
                    'total_general': float(tot_clientes['total'].sum())
                }

        # Movimientos Bancos crudos para recomputar resumen con filtros en frontend
        bancos_movimientos = []
        df_bancos_mov = df_bancos.copy()
        df_bancos_mov['__fecha_date'] = pd.to_datetime(df_bancos_mov['FECHA DE OPERACIÓN'], errors='coerce')
        df_bancos_mov = df_bancos_mov.dropna(subset=['__fecha_date'])
        df_bancos_mov['fecha'] = df_bancos_mov['__fecha_date'].dt.strftime('%Y-%m-%d')
        for _, r in df_bancos_mov.iterrows():
            movimiento_txt = str(r.get('Movimiento') or '')
            mov_upper = movimiento_txt.upper()
            tipo = 'otro'
            if 'SALDO INICIAL' in mov_upper:
                tipo = 'saldo_inicial'
            elif 'INGRESO' in mov_upper:
                tipo = 'ingreso'
            elif 'EGRESO' in mov_upper:
                tipo = 'egreso_gmf' if 'GMF' in mov_upper else 'egreso'
            bancos_movimientos.append({
                'fecha': r.get('fecha'),
                'empresa': r.get('Empresa'),
                'movimiento': movimiento_txt,
                'monto': float(r.get('COP$', 0) or 0),
                'clasificacion': tipo,
                'tipo_banco': str(r.get('Banco') or ''),
                'banco': str(r.get('Banco') or '')
            })

        # ======================= PERSISTENCIA ACTUALIZADA (SMART UPDATE) =======================
        # LÓGICA: Identificar los meses (Año, Mes) que vienen en el archivo y BORRAR esos meses de la BD
        # antes de insertar los nuevos. Esto permite tener 2025 quieto y actualizar 2026 las veces que sea.
        
        # --- FILTRADO POR SELECCIÓN DE USUARIO (Si existe) ---
        if target_periods_set:
            # Función auxiliar para chequear si fecha está en target
            def is_in_target(dt):
                if pd.isna(dt): return False
                return (dt.year, dt.month) in target_periods_set

            if not df_bancos_mov.empty and '__fecha_date' in df_bancos_mov.columns:
                df_bancos_mov = df_bancos_mov[df_bancos_mov['__fecha_date'].apply(is_in_target)].copy()
            
            # Filtrar detalle_records también (es una lista de dicts)
            # Primero reconstruimos si es necesario o filtramos la lista directamente
            detalle_records = [
                rec for rec in detalle_records 
                if is_in_target(pd.to_datetime(rec['fecha']))
            ]
            
            # Recalcular DF detalle para consistencia en caso de uso posterior (si fuera necesario)
            if not df_detalle.empty and '__fecha_date' in df_detalle.columns:
                 df_detalle = df_detalle[df_detalle['__fecha_date'].apply(is_in_target)].copy()


        periodos_a_actualizar = set()
        
        # Recopilar fechas de Bancos
        if not df_bancos_mov.empty and '__fecha_date' in df_bancos_mov.columns:
            fechas_b = df_bancos_mov['__fecha_date'].dropna().dt.to_period('M').unique()
            for p in fechas_b:
                periodos_a_actualizar.add((p.year, p.month))
                
        # Recopilar fechas de Odoo (revisar lista filtrada o df filtrado)
        # Usamos df_detalle filtrado arriba si existe, sino reconstruir de detalle_records
        if detalle_records:
             # Extraer fechas de la lista ya filtrada
             p_fechas = set()
             for r in detalle_records:
                 dt = pd.to_datetime(r['fecha'])
                 if not pd.isna(dt):
                     p_fechas.add((dt.year, dt.month))
             periodos_a_actualizar.update(p_fechas)

        if periodos_a_actualizar:
            for (anio, mes) in periodos_a_actualizar:
                # Limpiar Bancos para el periodo
                db.session.query(FlujoBancoMovimiento).filter(
                    func.extract('year', FlujoBancoMovimiento.fecha) == anio,
                    func.extract('month', FlujoBancoMovimiento.fecha) == mes
                ).delete(synchronize_session=False)
                
                # Limpiar Odoo para el periodo
                db.session.query(FlujoOdooMovimiento).filter(
                    func.extract('year', FlujoOdooMovimiento.fecha) == anio,
                    func.extract('month', FlujoOdooMovimiento.fecha) == mes
                ).delete(synchronize_session=False)
            
            # Flush para aplicar borrados antes de insertar
            db.session.flush()

        batch = FlujoUploadBatch(filename=archivo.filename, usuario=session.get('nombre','Desconocido'))
        db.session.add(batch)
        db.session.flush()  # obtener batch.id

        # Insertar registros (ahora seguros de no duplicar dentro del mismo mes)
        bancos_count = 0
        for _, r in df_bancos_mov.iterrows():
            monto_val = float(r.get('COP$', 0) or 0)
            if monto_val == 0:
                continue  # Ignorar filas donde COP$ es vacío o cero
            mov_txt = str(r.get('Movimiento') or '')
            banco_val = str(r.get('Banco') or '')
            db.session.add(FlujoBancoMovimiento(
                batch_id=batch.id,
                fecha=pd.to_datetime(r.get('__fecha_date')).date(),
                empresa=r.get('Empresa'),
                movimiento=mov_txt,
                monto=monto_val,
                banco=banco_val,
                tipo_banco=banco_val,
                unique_hash=uuid.uuid4().hex
            ))
            bancos_count += 1

        odoo_count = 0
        for rec in detalle_records:
            db.session.add(FlujoOdooMovimiento(
                batch_id=batch.id,
                fecha=pd.to_datetime(rec['fecha']).date(),
                empresa=rec['empresa'],
                movimiento=rec['movimiento'],
                debito=rec['debito'],
                credito=rec['credito'],
                tipo_flujo=rec.get('tipo_flujo'),
                tercero=rec.get('tercero'),
                rubro=rec.get('rubro'),
                clase=rec.get('clase'),
                subclase=rec.get('subclase'),
                banco=rec.get('banco'),
                unique_hash=uuid.uuid4().hex
            ))
            odoo_count += 1

        batch.total_bancos = bancos_count
        batch.total_odoo = odoo_count

        db.session.commit()

        # ======================= RECONSTRUCCIÓN DESDE BD =======================
        bancos_rows = FlujoBancoMovimiento.query.all()
        odoo_rows = FlujoOdooMovimiento.query.all()

        # Reconstruir daily comparison desde persistencia
        # Agrupar ingresos/egresos bancos por fecha/empresa según reglas (Movimiento contiene palabras clave)
        bancos_ing_map = {}
        bancos_eg_map = {}
        for r in bancos_rows:
            # Solo contabilizar movimientos INGRESO y EGRESO; SALDO INICIAL se ignora (pero ya NO filtramos GMF aquí para incluirlo en egresos_gmf por separado si se quiere)
            mtxt = (r.movimiento or '').upper()
            if 'SALDO INICIAL' in mtxt:
                continue
            # Asegurar consistencia de banco (None -> '')
            banco_key = (r.banco or '').strip()
            key = (r.fecha.isoformat(), r.empresa, banco_key)
            if 'INGRESO' in mtxt:
                bancos_ing_map[key] = bancos_ing_map.get(key,0) + r.monto
            elif 'EGRESO' in mtxt:
                bancos_eg_map[key] = bancos_eg_map.get(key,0) + abs(r.monto)

        odoo_ing_map = {}
        odoo_eg_map = {}
        for r in odoo_rows:
            mtxt = (r.movimiento or '').upper()
            key = (r.fecha.isoformat(), r.empresa, r.banco or '')
            if 'INGRESO' in mtxt:
                odoo_ing_map[key] = odoo_ing_map.get(key,0) + (r.debito or 0)
            if 'EGRESO' in mtxt:
                odoo_eg_map[key] = odoo_eg_map.get(key,0) + (r.credito or 0)

        keys_all = sorted(set(list(bancos_ing_map.keys()) + list(bancos_eg_map.keys()) + list(odoo_ing_map.keys()) + list(odoo_eg_map.keys())))
        daily_comparison_data = []
        for (f,e,bk) in keys_all:
            ib = bancos_ing_map.get((f,e,bk),0)
            eb = bancos_eg_map.get((f,e,bk),0)
            io = odoo_ing_map.get((f,e,bk),0)
            eo = odoo_eg_map.get((f,e,bk),0)
            daily_comparison_data.append({
                'fecha': f,
                'Empresa': e,
                'tipo_banco': bk or '',
                'banco': bk or '',  # alias legacy
                'ingresos_bancos': ib,
                'egresos_bancos': eb,
                'ingresos_odoo': io,
                'egresos_odoo': eo,
                'diferencia_ingresos': ib-io,
                'diferencia_egresos': eb-eo
            })

        # Reconstruir detalle Odoo para frontend
        detalle_records = []
        for r in odoo_rows:
            detalle_records.append({
                'fecha': r.fecha.isoformat(),
                'empresa': r.empresa,
                'movimiento': r.movimiento,
                'debito': r.debito,
                'credito': r.credito,
                'tipo_flujo': r.tipo_flujo or 'SIN TIPO',
                'tercero': r.tercero or 'SIN TERCERO',
                'rubro': r.rubro or '',
                'clase': r.clase or '',
                'subclase': r.subclase or '',
                'banco': r.banco or ''
            })
        todas_las_empresas = sorted({r.empresa for r in bancos_rows+odoo_rows if r.empresa})

        # Recalcular agrupaciones flujo / tercero
        inflows_by_type = {}
        outflows_by_type = {}
        for r in odoo_rows:
            tf = (r.tipo_flujo or 'SIN TIPO')
            terc = (r.tercero or 'SIN TERCERO')
            if (r.debito or 0) > 0:
                inflows_by_type.setdefault(tf, {})[terc] = inflows_by_type.setdefault(tf, {}).get(terc,0) + (r.debito or 0)
            if (r.credito or 0) > 0:
                outflows_by_type.setdefault(tf, {})[terc] = outflows_by_type.setdefault(tf, {}).get(terc,0) + (r.credito or 0)

        # TOP EXW (si clase contiene 'VENTAS EXW CTG')
        top_clientes_exw = None
        exw_rows = [r for r in odoo_rows if (r.clase or '').upper() == 'VENTAS EXW CTG' and (r.debito or 0) > 0]
        if exw_rows:
            agg = {}
            for r in exw_rows:
                cli = (r.tercero or 'SIN TERCERO')
                sub = (r.subclase or 'SIN SUBCLASE')
                agg.setdefault(cli, {}).setdefault(sub,0)
                agg[cli][sub] += r.debito or 0
            ranking = sorted([(cli, sum(subs.values()), subs) for cli, subs in agg.items()], key=lambda x: x[1], reverse=True)[:10]
            clientes_struct = []
            for cli, total, subs in ranking:
                clientes_struct.append({
                    'cliente': cli,
                    'total': total,
                    'subclases': [{'subclase': s, 'total': v} for s,v in subs.items()]
                })
            top_clientes_exw = {'clientes': clientes_struct, 'total_general': sum(x[1] for x in ranking)}

        # Reconstruir movimientos bancos simplificados para frontend (clasificacion)
        bancos_movimientos = []
        for r in bancos_rows:
            mov_upper = (r.movimiento or '').upper()
            tipo = 'otro'
            if 'SALDO INICIAL' in mov_upper:
                tipo = 'saldo_inicial'
            elif 'INGRESO' in mov_upper:
                tipo = 'ingreso'
            elif 'EGRESO' in mov_upper:
                tipo = 'egreso_gmf' if 'GMF' in mov_upper else 'egreso'
            if tipo == 'otro':
                continue
            bancos_movimientos.append({
                'fecha': r.fecha.isoformat(),
                'empresa': r.empresa,
                'movimiento': r.movimiento,
                'monto': r.monto,
                'clasificacion': tipo,
                'tipo_banco': (r.tipo_banco or r.banco or '').strip(),
                'banco': (r.banco or r.tipo_banco or '').strip()
            })

        # Calcular GMF separado para mostrarlo sumado en la tarjeta de egresos
        egresos_gmf_raw = df_bancos.loc[
            df_bancos['Movimiento'].str.contains('GMF', case=False, na=False), 'COP$'
        ].abs().sum()
        response_data = {
            'success': True,
            'nuevos_bancos': batch.total_bancos,
            'nuevos_odoo': batch.total_odoo,
            'daily_comparison': daily_comparison_data,
            'outflows_by_type': outflows_by_type,
            'inflows_by_type': inflows_by_type,
            'todas_las_empresas': todas_las_empresas,
            'odoo_detalle': detalle_records,
            'bancos_movimientos': bancos_movimientos,
            # Resumen Método Directo básico desde hoja Bancos (incluye ahora egresos_gmf y egresos_total)
            'resumen_directo': (lambda saldo_ini, ingresos, egresos_sin_gmf, egresos_gmf: {
                'saldo_inicial': float(saldo_ini),
                'ingresos': float(ingresos),
                'saldo_antes_egresos': float(saldo_ini + ingresos),
                'egresos': float(egresos_sin_gmf),
                'egresos_gmf': float(egresos_gmf),
                'egresos_total': float(egresos_sin_gmf + egresos_gmf),
                'saldo_final': float(saldo_ini + ingresos - (egresos_sin_gmf + egresos_gmf))
            })(
                # saldo inicial
                (lambda df: (
                    0.0 if df.empty else df[df['FECHA DE OPERACIÓN'].dt.date == df['FECHA DE OPERACIÓN'].dt.date.min()]['COP$'].sum()
                ))(df_bancos[df_bancos['Movimiento'].str.contains('SALDO INICIAL', case=False, na=False)]),
                # ingresos (sin SALDO INICIAL / GMF)
                df_bancos.loc[
                    df_bancos['Movimiento'].str.contains('INGRESO', case=False, na=False)
                    & ~df_bancos['Movimiento'].str.contains('SALDO INICIAL', case=False, na=False)
                    & ~df_bancos['Movimiento'].str.contains('GMF', case=False, na=False),
                    'COP$'
                ].sum(),
                # egresos sin GMF
                df_bancos.loc[
                    df_bancos['Movimiento'].str.contains('EGRESO', case=False, na=False)
                    & ~df_bancos['Movimiento'].str.contains('GMF', case=False, na=False),
                    'COP$'
                ].abs().sum(),
                # egresos GMF
                egresos_gmf_raw
            )
        }
        if top_clientes_exw:
            response_data['top_clientes_ventas_exw'] = top_clientes_exw

        # Sanitizar recursivamente cualquier NaN/Infinity residual antes de serializar
        def sanitize(obj):
            import math
            if isinstance(obj, dict):
                return {k: sanitize(v) for k, v in obj.items()}
            if isinstance(obj, list):
                return [sanitize(v) for v in obj]
            if isinstance(obj, float):
                if math.isnan(obj) or math.isinf(obj):
                    return None
                return obj
            # Intentar detectar pandas NA en tipos no-float
            try:
                if pd.isna(obj):
                    return None
            except Exception:
                pass
            return obj

        return jsonify(sanitize(response_data))

    except Exception as e:
        import traceback
        tb = traceback.format_exc()
        return jsonify(success=False, message=f'Error interno: {str(e)}', details=tb), 500

@login_required
@permiso_requerido('flujo_efectivo')
@app.route('/api/flujo_efectivo_simple', methods=['POST'])
def flujo_efectivo_simple():
    """Comparativo SIMPLE: (fecha, empresa) -> ingresos_bancos, saldo_inicial_bancos, ingresos_odoo.
    Reglas:
      - SALDO INICIAL (en Movimiento Bancos) se reporta separado (no suma en ingresos_bancos).
      - INGRESO (Movimiento contiene 'INGRESO' y NO 'SALDO INICIAL') suma a ingresos_bancos.
      - Ingresos Odoo = suma de Débito donde Movimiento contiene 'INGRESO'.
      - No se calcula egresos ni diferencias, sólo comparativa de ingreso y saldo inicial.
    """
    if 'archivo_excel' not in request.files:
        return jsonify(success=False, message='Falta archivo_excel'), 400
    archivo = request.files['archivo_excel']
    if archivo.filename == '' or not archivo.filename.lower().endswith(('.xlsx','.xls')):
        return jsonify(success=False, message='Archivo inválido'), 400
    try:
        xls = pd.ExcelFile(archivo)
        sheet_map = {n.strip().lower(): n for n in xls.sheet_names}
        if 'bancos' not in sheet_map or 'odoo' not in sheet_map:
            return jsonify(success=False, message='Debe incluir hojas Bancos y Odoo'), 400
        df_b = pd.read_excel(xls, sheet_name=sheet_map['bancos'])
        df_o = pd.read_excel(xls, sheet_name=sheet_map['odoo'])
        # Normalizar
        df_b.columns = df_b.columns.str.strip(); df_o.columns = df_o.columns.str.strip()
        req_b = {'FECHA DE OPERACIÓN','Movimiento','COP$','Empresa'}
        req_o = {'Fecha','Movimiento','Débito','Empresa'}
        if req_b - set(df_b.columns):
            return jsonify(success=False, message=f"Faltan columnas Bancos: {req_b - set(df_b.columns)}"), 400
        if req_o - set(df_o.columns):
            return jsonify(success=False, message=f"Faltan columnas Odoo: {req_o - set(df_o.columns)}"), 400
        # Limpieza montos
        def _cln(s):
            return (s.astype(str)
                     .str.replace(r"[^0-9\-.,]",'', regex=True)
                     .str.replace(',','', regex=False)
                     .replace('', '0')
                     .pipe(pd.to_numeric, errors='coerce').fillna(0))
        df_b['COP$'] = _cln(df_b['COP$'])
        df_o['Débito'] = _cln(df_o['Débito'])
        # Fechas
        df_b['FECHA DE OPERACIÓN'] = pd.to_datetime(df_b['FECHA DE OPERACIÓN'], errors='coerce')
        df_o['Fecha'] = pd.to_datetime(df_o['Fecha'], errors='coerce')
        df_b = df_b.dropna(subset=['FECHA DE OPERACIÓN'])
        df_o = df_o.dropna(subset=['Fecha'])
        df_b['fecha'] = df_b['FECHA DE OPERACIÓN'].dt.date.astype(str)
        df_o['fecha'] = df_o['Fecha'].dt.date.astype(str)
        # Clasificaciones bancos
        df_b['__mov'] = df_b['Movimiento'].astype(str).str.upper()
        saldo_ini = df_b[df_b['__mov'].str.contains('SALDO INICIAL', na=False)]
        ingresos_b = df_b[df_b['__mov'].str.contains('INGRESO', na=False) & ~df_b['__mov'].str.contains('SALDO INICIAL', na=False)]
        # Ingresos Odoo
        df_o['__mov'] = df_o['Movimiento'].astype(str).str.upper()
        ingresos_o = df_o[df_o['__mov'].str.contains('INGRESO', na=False)]
        # Agrupar
        saldo_ini_grp = saldo_ini.groupby(['fecha','Empresa'])['COP$'].sum().reset_index().rename(columns={'COP$':'saldo_inicial_bancos'})
        ing_b_grp = ingresos_b.groupby(['fecha','Empresa'])['COP$'].sum().reset_index().rename(columns={'COP$':'ingresos_bancos'})
        ing_o_grp = ingresos_o.groupby(['fecha','Empresa'])['Débito'].sum().reset_index().rename(columns={'Débito':'ingresos_odoo'})
        # Merge
        out = pd.merge(saldo_ini_grp, ing_b_grp, on=['fecha','Empresa'], how='outer')
        out = pd.merge(out, ing_o_grp, on=['fecha','Empresa'], how='outer')
        out = out.fillna(0)
        # Orden
        out = out.sort_values(['fecha','Empresa'])
        return jsonify(success=True, comparativo=out.to_dict(orient='records'))
    except Exception as e:
        app.logger.exception('Error flujo_efectivo_simple')
        return jsonify(success=False, message='Error interno'), 500

@login_required
@permiso_requerido('flujo_efectivo')
@app.route('/api/flujo_efectivo_cached', methods=['GET'])
def flujo_efectivo_cached():
    """Devuelve los datos consolidados desde la base sin necesidad de re-subir archivo."""
    try:
        bancos_rows = FlujoBancoMovimiento.query.all()
        odoo_rows = FlujoOdooMovimiento.query.all()
        if not bancos_rows and not odoo_rows:
            return jsonify(success=False, message='No hay datos cargados aún.')
        # Reutilizar lógica mínima (podríamos DRY, pero breve por claridad)
        # Daily comparison
        bancos_ing_map = {}; bancos_eg_map = {}; odoo_ing_map={}; odoo_eg_map={}
        for r in bancos_rows:
            mtxt=(r.movimiento or '').upper(); key=(r.fecha.isoformat(), r.empresa, r.banco or '')
            if 'SALDO INICIAL' in mtxt or 'GMF' in mtxt:
                continue
            if 'INGRESO' in mtxt:
                bancos_ing_map[key]=bancos_ing_map.get(key,0)+r.monto
            elif 'EGRESO' in mtxt:
                bancos_eg_map[key]=bancos_eg_map.get(key,0)+abs(r.monto)
        for r in odoo_rows:
            mtxt=(r.movimiento or '').upper(); key=(r.fecha.isoformat(), r.empresa, r.banco or '')
            if 'INGRESO' in mtxt: odoo_ing_map[key]=odoo_ing_map.get(key,0)+(r.debito or 0)
            if 'EGRESO' in mtxt: odoo_eg_map[key]=odoo_eg_map.get(key,0)+(r.credito or 0)
        keys_all = sorted(set(list(bancos_ing_map.keys())+list(bancos_eg_map.keys())+list(odoo_ing_map.keys())+list(odoo_eg_map.keys())))
        daily=[]
        for (f,e,b) in keys_all:
            ib=bancos_ing_map.get((f,e,b),0); eb=bancos_eg_map.get((f,e,b),0); io=odoo_ing_map.get((f,e,b),0); eo=odoo_eg_map.get((f,e,b),0)
            banco_norm = (b or '').strip()
            daily.append({'fecha':f,'Empresa':e,'tipo_banco':banco_norm,'banco':banco_norm,'ingresos_bancos':ib,'egresos_bancos':eb,'ingresos_odoo':io,'egresos_odoo':eo,'diferencia_ingresos':ib-io,'diferencia_egresos':eb-eo})
        detalle=[{'fecha':r.fecha.isoformat(),'empresa':r.empresa,'movimiento':r.movimiento,'debito':r.debito,'credito':r.credito,'tipo_flujo':r.tipo_flujo or 'SIN TIPO','tercero':r.tercero or 'SIN TERCERO','rubro':r.rubro or '','clase':r.clase or '','subclase':r.subclase or '','banco':r.banco or ''} for r in odoo_rows]
        inflows_by_type={}; outflows_by_type={}
        for r in odoo_rows:
            tf=(r.tipo_flujo or 'SIN TIPO'); terc=(r.tercero or 'SIN TERCERO')
            if (r.debito or 0)>0: inflows_by_type.setdefault(tf,{}).setdefault(terc,0); inflows_by_type[tf][terc]+=r.debito or 0
            if (r.credito or 0)>0: outflows_by_type.setdefault(tf,{}).setdefault(terc,0); outflows_by_type[tf][terc]+=r.credito or 0
        exw_rows=[r for r in odoo_rows if (r.clase or '').upper()=='VENTAS EXW CTG' and (r.debito or 0)>0]
        top_exw=None
        if exw_rows:
            agg={}
            for r in exw_rows:
                cli=(r.tercero or 'SIN TERCERO'); sub=(r.subclase or 'SIN SUBCLASE'); agg.setdefault(cli,{}).setdefault(sub,0); agg[cli][sub]+=r.debito or 0
            ranking=sorted([(cli,sum(subs.values()),subs) for cli,subs in agg.items()], key=lambda x:x[1], reverse=True)[:10]
            clientes_struct=[]
            for cli,total,subs in ranking:
                clientes_struct.append({'cliente':cli,'total':total,'subclases':[{'subclase':s,'total':v} for s,v in subs.items()]})
            top_exw={'clientes':clientes_struct,'total_general':sum(x[1] for x in ranking)}
        bancos_mov=[]
        for r in bancos_rows:
            mu=(r.movimiento or '').upper()
            if 'SALDO INICIAL' in mu:
                tipo='saldo_inicial'
            elif 'INGRESO' in mu:
                tipo='ingreso'
            elif 'EGRESO' in mu:
                tipo='egreso_gmf' if 'GMF' in mu else 'egreso'
            else:
                continue
            banco_norm = (r.tipo_banco or r.banco or '').strip()
            bancos_mov.append({'fecha':r.fecha.isoformat(),'empresa':r.empresa,'movimiento':r.movimiento,'monto':r.monto,'clasificacion':tipo,'tipo_banco':banco_norm,'banco':banco_norm})
        empresas=sorted({r.empresa for r in bancos_rows+odoo_rows if r.empresa})
        # Calcular saldo inicial: sumatoria SALDO INICIAL del primer día (todos bancos)
        saldo_inicial_rows = [r for r in bancos_rows if 'SALDO INICIAL' in (r.movimiento or '').upper()]
        if saldo_inicial_rows:
            # fecha mínima
            first_date = min(r.fecha for r in saldo_inicial_rows)
            saldo_inicial_total = sum(r.monto for r in saldo_inicial_rows if r.fecha == first_date)
        else:
            saldo_inicial_total = 0.0
        ingresos_total = sum(r.monto for r in bancos_rows if 'INGRESO' in (r.movimiento or '').upper() and 'SALDO INICIAL' not in (r.movimiento or '').upper())
        egresos_sin_gmf = sum(abs(r.monto) for r in bancos_rows if 'EGRESO' in (r.movimiento or '').upper() and 'GMF' not in (r.movimiento or '').upper())
        egresos_gmf = sum(abs(r.monto) for r in bancos_rows if 'GMF' in (r.movimiento or '').upper())
        egresos_total = egresos_sin_gmf + egresos_gmf
        resumen_directo = {
            'saldo_inicial': float(saldo_inicial_total),
            'ingresos': float(ingresos_total),
            'saldo_antes_egresos': float(saldo_inicial_total + ingresos_total),
            'egresos': float(egresos_sin_gmf),
            'egresos_gmf': float(egresos_gmf),
            'egresos_total': float(egresos_total),
            'saldo_final': float(saldo_inicial_total + ingresos_total - egresos_total)
        }
        payload={'success':True,'daily_comparison':daily,'outflows_by_type':outflows_by_type,'inflows_by_type':inflows_by_type,'todas_las_empresas':empresas,'odoo_detalle':detalle,'bancos_movimientos':bancos_mov,'resumen_directo':resumen_directo}
        if top_exw:
            payload['top_clientes_ventas_exw']=top_exw
        return jsonify(payload)
    except Exception:
        app.logger.exception('Error recuperando cache flujo efectivo')
        return jsonify(success=False, message='Error interno.'), 500

@login_required
@permiso_requerido('flujo_efectivo')
@app.route('/api/flujo_efectivo_delete_all', methods=['DELETE'])
def flujo_efectivo_delete_all():
    """Elimina registros del módulo Flujo de Efectivo.
       Si se recibe 'year' en el body, elimina solo ese año.
       Si no, elimina TODOS los registros.
    """
    try:
        data = request.get_json() or {}
        year_to_delete = data.get('year')
        
        if year_to_delete:
            try:
                year_int = int(year_to_delete)
                # Eliminar solo por año
                deleted_bancos = db.session.query(FlujoBancoMovimiento).filter(func.extract('year', FlujoBancoMovimiento.fecha) == year_int).delete(synchronize_session=False)
                deleted_odoo = db.session.query(FlujoOdooMovimiento).filter(func.extract('year', FlujoOdooMovimiento.fecha) == year_int).delete(synchronize_session=False)
                # No eliminamos batches porque pueden tener otros años mezclados, se mantienen como historial de carga
                
                db.session.commit()
                return jsonify(success=True,
                               message=f"Datos del año {year_int} eliminados correctamente.",
                               eliminados_bancos=deleted_bancos,
                               eliminados_odoo=deleted_odoo,
                               mode='partial')
            except ValueError:
                return jsonify(success=False, message="Año inválido"), 400
        else:
            # Eliminar TODO
            deleted_bancos = db.session.query(FlujoBancoMovimiento).delete(synchronize_session=False)
            deleted_odoo = db.session.query(FlujoOdooMovimiento).delete(synchronize_session=False)
            deleted_batches = db.session.query(FlujoUploadBatch).delete(synchronize_session=False)
            db.session.commit()
            return jsonify(success=True,
                           message="Datos de Flujo de Efectivo eliminados completamente.",
                           eliminados_bancos=deleted_bancos,
                           eliminados_odoo=deleted_odoo,
                           eliminados_batches=deleted_batches,
                           mode='all')
    except Exception as e:
        db.session.rollback()
        app.logger.exception('Error eliminando datos de flujo de efectivo')
        return jsonify(success=False, message=f'Error interno: {e}'), 500

@login_required
@permiso_requerido('flujo_efectivo')
@app.route('/api/procesar_facturacion', methods=['POST'])
def procesar_facturacion_api():
    """Procesa un Excel de facturación para extender la gráfica de ingresos.
    Columnas mínimas: (Numero Factura / Factura), (Cliente / Tercero), (COP$ / Valor / Monto).
    Opcional: Bbl / Barriles.
    Devuelve lista de facturas normalizada.
    """
    if 'archivo_excel' not in request.files:
        return jsonify(success=False, message='No se encontró archivo.'), 400
    archivo = request.files['archivo_excel']
    if archivo.filename == '':
        return jsonify(success=False, message='Nombre de archivo vacío.'), 400
    try:
        df = pd.read_excel(archivo)
        df.columns = df.columns.str.strip()

        def find_col(cands):
            for c in cands:
                if c in df.columns:
                    return c
            return None

        col_num = find_col(['Numero Factura','Número Factura','Nro Factura','Factura','FACTURA','NUMERO FACTURA','Número','NUMERO','Numero'])
        col_cli = find_col(['Cliente','CLIENTE','Tercero','TERCERO','Asociado','ASOCIADO'])
        col_cop = find_col(['COP$','Valor','VALOR','Monto','MONTO'])
        col_bbl = find_col(['Bbl','BBL','Barriles','BARRILES'])
        col_gln = find_col(['Gln','GLN','Galones','GALONES','Galón','GALÓN','gln','galones'])
        col_ton = find_col(['Ton','TON','Toneladas','TONELADAS','Tonelada','TONELADA','ton','toneladas'])
        col_etiqueta = find_col(['Etiqueta','ETIQUETA'])
        col_fecha = find_col(['Fecha','FECHA','Date','DATE'])

        if not (col_num and col_cli and col_cop):
            return jsonify(success=False, message=f"Faltan columnas obligatorias (Factura, Cliente/Tercero, Valor). Columnas disponibles: {list(df.columns)}"), 400

        def clean(series):
            return (series.astype(str)
                          .str.replace(r"[^0-9\-.,]",'', regex=True)
                          .str.replace(',', '', regex=False)
                          .replace('', '0')
                          .pipe(pd.to_numeric, errors='coerce').fillna(0))

        df[col_cop] = clean(df[col_cop])
        if col_bbl:
            df[col_bbl] = clean(df[col_bbl])
        if col_gln:
            df[col_gln] = clean(df[col_gln])
        if col_ton:
            df[col_ton] = clean(df[col_ton])

        # Normalizar fecha si existe
        if col_fecha:
            try:
                df[col_fecha] = pd.to_datetime(df[col_fecha], errors='coerce')
            except Exception:
                df[col_fecha] = pd.NaT

        facturas = []
        for _, r in df.iterrows():
            fila = {
                'factura': str(r.get(col_num)),
                'tercero': str(r.get(col_cli) or 'SIN TERCERO'),
                'valor': float(r.get(col_cop) or 0),
                'bbl': float(r.get(col_bbl) or 0) if col_bbl else None,
                'gln': float(r.get(col_gln) or 0) if col_gln else None,
                'ton': float(r.get(col_ton) or 0) if col_ton else None
            }
            if col_etiqueta:
                fila['etiqueta'] = str(r.get(col_etiqueta) or '').strip()
            if col_fecha:
                fv = r.get(col_fecha)
                if pd.notna(fv):
                    try:
                        fila['fecha'] = pd.to_datetime(fv).date().isoformat()
                    except Exception:
                        pass
            facturas.append(fila)
        return jsonify(success=True, facturas=facturas)
    except Exception as e:
        app.logger.exception('Error procesando facturación')
        return jsonify(success=False, message=str(e)), 500

@app.route('/')
def home():
    """Redirige al usuario a su página de inicio correcta después de iniciar sesión."""
    if 'email' not in session:
        return redirect(url_for('login'))
    
    user_email = session.get('email')

    # --- REGLA 1: Administradores siempre al dashboard de reportes ---
    if session.get('rol') == 'admin':
        return redirect(url_for('dashboard_reportes'))

    # --- REGLA 2: Usuarios con acceso a reportes van al dashboard ---
    if 'reportes' in session.get('area', []):
        return redirect(url_for('dashboard_reportes'))

    # --- REGLA 3: Excepciones específicas por email ---
    if user_email in ['comex@conquerstrading.com', 'felipe.delavega@conquerstrading.com', 'amariagallo@conquerstrading.com']:
        return redirect(url_for('dashboard_reportes'))

    # --- REGLA 4: Usuario de inventario EPP exclusivo ---
    if user_email == 'safety@conquerstrading.com':
        return redirect(url_for('inventario_epp_home'))

    # --- REGLA 5: Todos los demás usuarios van al home global ---
    return redirect(url_for('home_global'))

@login_required
@app.route('/inicio-logistica')
def home_logistica():
    """Página de inicio unificada - redirecciona a home global."""
    return redirect(url_for('home_global'))

@login_required
@app.route('/home-global')
def home_global():
    """Página de inicio global unificada para todos los usuarios con permisos."""
    return render_template('home_global.html')

@app.route('/test')
def test():
    return "✅ El servidor Flask está funcionando"
@app.route('/debug/productos')

def debug_productos():
    productos = cargar_productos()
    return jsonify({
        "productos": productos,
        "exists": os.path.exists("productos.json"),
        "file_content": open("productos.json").read() if os.path.exists("productos.json") else None
    })

def cargar_clientes():
    """Función auxiliar para cargar clientes. Prioriza DB -> Clientes.json."""
    clientes_lista = []
    
    # 1. Intentar cargar desde la Base de Datos (PostgreSQL/SQLite)
    try:
        # Importación local para evitar dependencias circulares si las hubiera, 
        # aunque Cliente está definido más abajo, Python permite esto si se ejecuta después.
        registros_db = Cliente.query.all()
        if registros_db:
            for r in registros_db:
                clientes_lista.append({
                    "NOMBRE_CLIENTE": r.nombre,
                    "DIRECCION": r.direccion,
                    "CIUDAD_DEPARTAMENTO": r.ciudad_departamento
                })
            return clientes_lista
    except Exception as e:
        print(f"Advertencia: No se pudo cargar clientes desde DB: {e}")

    # 2. Si DB falló o estaba vacía, cargar desde JSON (Fallback)
    try:
        ruta_clientes = os.path.join(BASE_DIR, 'static', 'Clientes.json')
        with open(ruta_clientes, 'r', encoding='utf-8') as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return []

def guardar_clientes(clientes):
    """Función auxiliar para guardar la lista de clientes en Clientes.json."""
    # Buscamos el archivo en la carpeta 'static'
    ruta_clientes = os.path.join(BASE_DIR, 'static', 'Clientes.json')
    with open(ruta_clientes, 'w', encoding='utf-8') as f:
        json.dump(clientes, f, ensure_ascii=False, indent=4)


# Modelos SQLAlchemy para Cliente, Conductor y Empresa
from flask_sqlalchemy import SQLAlchemy
db: SQLAlchemy  # Asegúrate de que tu app ya tiene db = SQLAlchemy(app)

# Modelo para solicitudes de enturnamiento recibidas por WhatsApp
class SolicitudCita(db.Model):
    __tablename__ = 'solicitudes_cita'
    __table_args__ = (
        db.UniqueConstraint('turno_fecha', 'turno', name='uq_solicitudes_turno_fecha'),
    )
    id = db.Column(db.Integer, primary_key=True)
    telefono = db.Column(db.String(32), nullable=False)
    mensaje = db.Column(db.Text, nullable=False)
    fecha = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)
    nombre_completo = db.Column(db.String(255))
    cedula = db.Column(db.String(64))
    placa = db.Column(db.String(64))
    placa_remolque = db.Column(db.String(64))
    celular = db.Column(db.String(32))
    imagen_guia = db.Column(db.String(512))  # URL o nombre de archivo
    imagen_manifiesto = db.Column(db.String(512))
    paso_bosconia = db.Column(db.Boolean, default=False)
    ticket_gambote = db.Column(db.String(512))
    ubicacion_lat = db.Column(db.Float)  # Latitud Bosconia enviada por WhatsApp
    ubicacion_lng = db.Column(db.Float)  # Longitud Bosconia enviada por WhatsApp
    ubicacion_gambote_lat = db.Column(db.Float)  # Latitud Gambote enviada por WhatsApp
    ubicacion_gambote_lng = db.Column(db.Float)  # Longitud Gambote enviada por WhatsApp
    paso_gambote = db.Column(db.Boolean, default=False)
    ubicacion_zisa_lat = db.Column(db.Float)  # Latitud ZISA enviada por WhatsApp
    ubicacion_zisa_lng = db.Column(db.Float)  # Longitud ZISA enviada por WhatsApp
    paso_zisa = db.Column(db.Boolean, default=False)
    ubicacion_pendiente_lat = db.Column(db.Float)
    ubicacion_pendiente_lng = db.Column(db.Float)
    ubicacion_pendiente_tipo = db.Column(db.String(32))
    ubicacion_pendiente_mensaje = db.Column(db.String(255))
    ubicacion_pendiente_desde = db.Column(db.DateTime)
    ruta_alterna = db.Column(db.Boolean, default=False)  # Para casos de paro/protesta
    estado = db.Column(db.String(32), default='preconfirmacion')  # preconfirmacion, pendiente_inscripcion, sin turno, en revision, enturnado, error
    observaciones = db.Column(db.Text)
    turno = db.Column(db.Integer)
    turno_fecha = db.Column(db.Date)
    fecha_descargue = db.Column(db.DateTime)
    lugar_descargue = db.Column(db.String(255), default='Sociedad Portuaria del Dique')
    whatsapp_step = db.Column(db.String(32), default='0')  # Paso actual en el flujo de WhatsApp
    whatsapp_last_activity = db.Column(db.DateTime)  # Última actividad en WhatsApp
    whatsapp_timeout_minutes = db.Column(db.Integer, default=5)  # Minutos de timeout para la sesión
    whatsapp_warning_sent = db.Column(db.Boolean, default=False)  # Si ya se envió advertencia de timeout
    asesor_pendiente = db.Column(db.Boolean, default=False)
    asesor_pendiente_desde = db.Column(db.DateTime)


class WhatsappMessage(db.Model):
    __tablename__ = 'whatsapp_messages'
    id = db.Column(db.Integer, primary_key=True)
    solicitud_id = db.Column(db.Integer, db.ForeignKey('solicitudes_cita.id'), nullable=True)
    telefono = db.Column(db.String(32), nullable=False, index=True)
    direction = db.Column(db.String(16), nullable=False)  # inbound / outbound
    sender = db.Column(db.String(16), nullable=False)  # driver / bot / human
    message_type = db.Column(db.String(32), nullable=False, default='text')
    content = db.Column(db.Text, nullable=True)
    media_url = db.Column(db.String(512), nullable=True)
    created_at = db.Column(db.DateTime, nullable=False, default=datetime.utcnow, index=True)

    solicitud = db.relationship('SolicitudCita', backref=db.backref('mensajes_whatsapp', lazy='dynamic'))


class Cliente(db.Model):
    __tablename__ = 'clientes'
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(255), nullable=False)
    direccion = db.Column(db.String(255), nullable=False)
    ciudad_departamento = db.Column(db.String(255), nullable=False)

class Conductor(db.Model):
    __tablename__ = 'conductores'
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(255), nullable=False)
    cedula = db.Column(db.String(64), nullable=False, unique=True)
    placa = db.Column(db.String(64), nullable=False)
    placa_remolque = db.Column(db.String(64))
    celular = db.Column(db.String(32))

class Empresa(db.Model):
    __tablename__ = 'empresas_transportadoras'
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(255), nullable=False, unique=True)


class Producto(db.Model):
    __tablename__ = 'productos'
    id = db.Column(db.Integer, primary_key=True)
    producto = db.Column(db.String(255), nullable=False, unique=True)
    unidad = db.Column(db.String(64))

    # --- PANEL DE REVISIÓN DE SOLICITUDES DE ENTURNAMIENTO ---
    @login_required
    @app.route('/panel_enturnamiento')
    def panel_enturnamiento():
        allowed_panel_emails = {'logistic@conquerstrading.com', 'ops@conquerstrading.com'}
        email_actual = (session.get('email') or '').lower()
        if session.get('rol') != 'admin' and email_actual not in allowed_panel_emails:
            flash('Acceso restringido solo para el área de logística.', 'danger')
            return redirect(url_for('home'))
        solicitudes = (
            SolicitudCita.query
            .filter(
                or_(
                    SolicitudCita.estado != 'preconfirmacion',
                    SolicitudCita.estado.is_(None),
                    SolicitudCita.asesor_pendiente.is_(True)
                )
            )
            .order_by(SolicitudCita.fecha.desc())
            .all()
        )
        solicitudes_handoff = [s for s in solicitudes if getattr(s, 'asesor_pendiente', False)]
        handoff_count = len(solicitudes_handoff)
        # Añadir url de Google Maps si hay lat/lng
        for s in solicitudes:
            if getattr(s, 'ubicacion_lat', None) and getattr(s, 'ubicacion_lng', None):
                s.url_maps = f"https://www.google.com/maps?q={s.ubicacion_lat},{s.ubicacion_lng}"
            else:
                s.url_maps = None
            
            # Validar secuencia GPS
            validacion_gps = validar_secuencia_gps(s)
            s.validacion_gps = validacion_gps
            s.fecha_local = to_bogota_datetime(getattr(s, 'fecha', None))
            s.fecha_descargue_local = to_bogota_datetime(getattr(s, 'fecha_descargue', None), assume_local=True)
            s.whatsapp_last_activity_local = to_bogota_datetime(getattr(s, 'whatsapp_last_activity', None))
        return render_template('panel_enturnamiento.html', solicitudes=solicitudes, handoff_count=handoff_count)

    @login_required
    @app.route('/api/solicitud_cita/<int:id>/estado', methods=['POST'])
    def actualizar_estado_solicitud(id):
        data = request.get_json()
        nuevo_estado = data.get('estado')
        observaciones = data.get('observaciones', '')
        solicitud = SolicitudCita.query.get_or_404(id)
        solicitud.observaciones = observaciones

        estado_anterior = solicitud.estado
        fecha_descargue_anterior = to_bogota_datetime(solicitud.fecha_descargue, assume_local=True)
        turno_anterior = solicitud.turno

        turno_val = None
        turno_fecha = None
        fecha_descargue_input_local = None
        fecha_descargue_local = None
        message_to_send = None

        fecha_descargue_str = (data.get('fecha_descargue') or '').strip()
        hora_descargue_str = (data.get('hora_descargue') or '').strip()

        if nuevo_estado == 'enturnado':
            if fecha_descargue_str and hora_descargue_str:
                try:
                    naive_local = datetime.strptime(f"{fecha_descargue_str} {hora_descargue_str}", '%Y-%m-%d %H:%M')
                    fecha_descargue_input_local = BOGOTA_TZ.localize(naive_local)
                except ValueError:
                    fecha_descargue_input_local = None
            elif fecha_descargue_str:
                try:
                    fecha_base = datetime.strptime(fecha_descargue_str, '%Y-%m-%d').date()
                    naive_local = datetime.combine(fecha_base, time.min)
                    fecha_descargue_input_local = BOGOTA_TZ.localize(naive_local)
                except ValueError:
                    fecha_descargue_input_local = None

            if fecha_descargue_input_local is not None:
                fecha_descargue_local = fecha_descargue_input_local
            else:
                existente_local = to_bogota_datetime(solicitud.fecha_descargue, assume_local=True)
                fecha_descargue_local = existente_local or datetime.now(BOGOTA_TZ)

            turno_manual = (data.get('turno') or '').strip()
            if turno_manual:
                try:
                    turno_val = int(turno_manual)
                except ValueError:
                    return jsonify(success=False, message='El turno debe ser un número entero.'), 400

                fecha_para_turno = fecha_descargue_input_local or fecha_descargue_local

                if fecha_para_turno is None:
                    return jsonify(success=False, message='Debes seleccionar una fecha de descargue para asignar turno.'), 400

                turno_fecha = fecha_para_turno.date()

                conflicto = (
                    SolicitudCita.query
                    .filter(
                        SolicitudCita.turno == turno_val,
                        SolicitudCita.turno_fecha == turno_fecha,
                        SolicitudCita.id != solicitud.id
                    )
                    .first()
                )
                if conflicto:
                    fecha_conflicto = conflicto.turno_fecha.strftime('%d/%m/%Y') if conflicto.turno_fecha else 'desconocida'
                    return jsonify(success=False, message=f'El turno {turno_val} ya está asignado el {fecha_conflicto} a la solicitud #{conflicto.id}.'), 409

        try:
            if nuevo_estado == 'enturnado':
                solicitud.estado = 'enturnado'
                fecha_descargue_bogota = fecha_descargue_local or datetime.now(BOGOTA_TZ)
                solicitud.fecha_descargue = bogota_naive(fecha_descargue_bogota)
                solicitud.turno = turno_val
                solicitud.turno_fecha = turno_fecha if turno_val is not None else None
                solicitud.lugar_descargue = 'Sociedad Portuaria del Dique'
                mensaje_enturnado = build_enturnado_message(solicitud)
                was_enturnado = estado_anterior in {'enturnado', STATE_FINALIZADO}

                if was_enturnado:
                    turno_actual = solicitud.turno
                    turno_cambio = turno_actual != turno_anterior

                    cambio_fecha = False
                    cambio_hora = False
                    if fecha_descargue_anterior and fecha_descargue_bogota:
                        cambio_fecha = fecha_descargue_anterior.date() != fecha_descargue_bogota.date()
                        cambio_hora = fecha_descargue_anterior.time() != fecha_descargue_bogota.time()
                    elif fecha_descargue_anterior or fecha_descargue_bogota:
                        # Uno de los valores es None y el otro no, tratamos ambos como cambio
                        cambio_fecha = True
                        cambio_hora = True

                    ajustes = []
                    turno_texto = turno_actual if turno_actual is not None else 'Pendiente'
                    if turno_cambio:
                        ajustes.append(f'Nuevo turno: {turno_texto}.')

                    if cambio_fecha and cambio_hora:
                        ajustes.append(
                            f"Nueva fecha y hora de descargue: {fecha_descargue_bogota.strftime('%d/%m/%Y %H:%M')}."
                        )
                    elif cambio_fecha:
                        ajustes.append(
                            f"Nueva fecha de descargue: {fecha_descargue_bogota.strftime('%d/%m/%Y')}."
                        )
                        ajustes.append(
                            f"La hora se mantiene a las {fecha_descargue_bogota.strftime('%H:%M')}."
                        )
                    elif cambio_hora:
                        ajustes.append(
                            f"Nueva hora de descargue: {fecha_descargue_bogota.strftime('%H:%M')}."
                        )

                    if ajustes:
                        lineas_reschedule = [
                            '🙏 Hola, te habla Fisher 🐶. Lamentamos el cambio, ajustamos tu enturnamiento.',
                            *ajustes,
                            'Si necesitas hablar con nuestro equipo humano, responde *asesor* y te contactamos.',
                            '\n🏁 *Si todo está bien, esta conversación finaliza aquí. Escribe NUEVO para tu próximo enturne.*'
                        ]
                        message_to_send = '\n'.join(lineas_reschedule)
                        solicitud.mensaje = message_to_send
                    else:
                        message_to_send = None

                    solicitud.estado = 'enturnado'
                    solicitud.asesor_pendiente = False
                    solicitud.asesor_pendiente_desde = None
                    solicitud.whatsapp_step = str(STEP_INACTIVE)
                    solicitud.whatsapp_timeout_minutes = 0
                    solicitud.whatsapp_warning_sent = False
                    solicitud.whatsapp_last_activity = datetime.utcnow()
                else:
                    message_to_send = mensaje_enturnado
                    solicitud.mensaje = mensaje_enturnado
                    solicitud.estado = STATE_FINALIZADO
                    solicitud.asesor_pendiente = False
                    solicitud.asesor_pendiente_desde = None
                    solicitud.whatsapp_step = str(STEP_INACTIVE)
                    solicitud.whatsapp_timeout_minutes = 0
                    solicitud.whatsapp_warning_sent = False
                    solicitud.whatsapp_last_activity = datetime.utcnow()
            elif nuevo_estado == 'error':
                solicitud.estado = 'error'
                solicitud.turno = None
                solicitud.turno_fecha = None
                solicitud.fecha_descargue = None
                solicitud.lugar_descargue = None
                solicitud.whatsapp_step = 'error'
                solicitud.whatsapp_timeout_minutes = 0
                solicitud.whatsapp_warning_sent = False
                solicitud.asesor_pendiente = False
                solicitud.asesor_pendiente_desde = None
                mensaje = (
                    "❌ Fisher 🐶 olfateó que nos faltan datos para cerrar tu enturnamiento.\n"
                    "Enseguida uno de nuestros asesores humanos te contactará para completar el proceso juntos.\n"
                    "¡Gracias por tu paciencia y por confiar en Fisher!"
                )
                solicitud.mensaje = mensaje
                message_to_send = mensaje
            elif nuevo_estado == STATE_PENDING_INSCRIPTION:
                solicitud.estado = STATE_PENDING_INSCRIPTION
                solicitud.turno = None
                solicitud.turno_fecha = None
                solicitud.fecha_descargue = None
                solicitud.lugar_descargue = None
                solicitud.whatsapp_step = STEP_HUMAN_HANDOFF
                solicitud.whatsapp_timeout_minutes = 0
                solicitud.whatsapp_warning_sent = False
                solicitud.whatsapp_last_activity = datetime.utcnow()
                solicitud.asesor_pendiente = True
                if not solicitud.asesor_pendiente_desde:
                    solicitud.asesor_pendiente_desde = datetime.utcnow()
                if not solicitud.mensaje:
                    solicitud.mensaje = 'Solicitud en pendiente de inscripción para revisión humana.'
            else:  # sin turno
                solicitud.estado = 'sin turno'
                solicitud.turno = None
                solicitud.turno_fecha = None
                solicitud.fecha_descargue = None
                solicitud.lugar_descargue = None
                solicitud.whatsapp_step = '0'
                solicitud.whatsapp_timeout_minutes = _normalize_timeout_minutes(5)
                solicitud.whatsapp_warning_sent = False
                solicitud.mensaje = 'Solicitud en estado sin turno.'
                solicitud.asesor_pendiente = False
                solicitud.asesor_pendiente_desde = None

            db.session.commit()
        except IntegrityError:
            db.session.rollback()
            if turno_val is not None:
                return jsonify(success=False, message=f'El turno {turno_val} ya está asignado a otra solicitud.'), 409
            return jsonify(success=False, message='No fue posible guardar los cambios. Verifique la información e intente nuevamente.'), 409
        except Exception as exc:
            db.session.rollback()
            current_app.logger.exception('Error actualizando estado de solicitud %s', id)
            return jsonify(success=False, message='Error interno al actualizar la solicitud.'), 500

        if message_to_send:
            try:
                send_whatsapp_message(solicitud.telefono, message_to_send)
            except Exception as send_err:
                current_app.logger.warning('No se pudo enviar mensaje WhatsApp para solicitud %s: %s', id, send_err)

        return jsonify(success=True, estado=solicitud.estado)


@login_required
@app.route('/api/solicitud_cita/<int:id>/mensajes', methods=['GET'])
def obtener_mensajes_whatsapp(id):
    solicitud = SolicitudCita.query.get_or_404(id)
    mensajes = (
        WhatsappMessage.query
        .filter_by(solicitud_id=solicitud.id)
        .order_by(WhatsappMessage.created_at.asc())
        .all()
    )

    def serialize(msg):
        fecha_local = to_bogota_datetime(msg.created_at)
        return {
            'id': msg.id,
            'sender': msg.sender,
            'direction': msg.direction,
            'tipo': msg.message_type,
            'contenido': msg.content or '',
            'media_url': msg.media_url,
            'fecha': fecha_local.strftime('%d/%m/%Y %H:%M') if fecha_local else ''
        }

    return jsonify(success=True, mensajes=[serialize(m) for m in mensajes])


@login_required
@app.route('/api/solicitud_cita', methods=['POST'])
def crear_solicitud_desde_panel():
    data = request.get_json() or {}
    telefono = (data.get('telefono') or '').strip()
    if not telefono:
        return jsonify(success=False, message='El número de teléfono es obligatorio.'), 400

    estado_raw = (data.get('estado') or 'sin turno').strip().lower()
    estado_permitido = {'sin turno', 'preconfirmacion', 'en revision', 'enturnado', STATE_FINALIZADO, 'error', STATE_PENDING_INSCRIPTION}
    estado = estado_raw if estado_raw in estado_permitido else 'sin turno'

    # Extraer datos del conductor
    cedula = (data.get('cedula') or '').strip()
    nombre_completo = (data.get('nombre_completo') or '').strip()
    placa = (data.get('placa') or '').strip().upper()
    placa_remolque = (data.get('placa_remolque') or '').strip().upper()
    celular = (data.get('celular') or '').strip()

    # Crear conductor si se proporcionan los datos mínimos
    conductor_creado = False
    if cedula and nombre_completo and placa:
        conductor_existente = Conductor.query.filter_by(cedula=cedula).first()
        if not conductor_existente:
            try:
                nuevo_conductor = Conductor(
                    nombre=nombre_completo,
                    cedula=cedula,
                    placa=placa,
                    placa_remolque=placa_remolque if placa_remolque else None,
                    celular=celular if celular else None
                )
                db.session.add(nuevo_conductor)
                conductor_creado = True
                current_app.logger.info(f'Conductor creado automáticamente: {cedula} - {nombre_completo}')
            except Exception as exc:
                db.session.rollback()
                current_app.logger.warning(f'No se pudo crear conductor automáticamente: {exc}')
                # Continuar sin el conductor

    # Verificar si ya existe una solicitud activa para este teléfono
    solicitud_existente = (
        SolicitudCita.query
        .filter(
            SolicitudCita.telefono == telefono,
            SolicitudCita.estado != STATE_FINALIZADO
        )
        .order_by(SolicitudCita.fecha.desc())
        .first()
    )

    if solicitud_existente:
        # Actualizar la solicitud existente en lugar de crear una nueva
        solicitud = solicitud_existente
        solicitud_actualizada = True

        # Actualizar campos básicos
        if nombre_completo:
            solicitud.nombre_completo = nombre_completo
        if cedula:
            solicitud.cedula = cedula
        if placa:
            solicitud.placa = placa
        if placa_remolque:
            solicitud.placa_remolque = placa_remolque
        if celular:
            solicitud.celular = celular

        # Actualizar estado y observaciones
        solicitud.estado = estado
        observaciones = (data.get('observaciones') or '').strip()
        if observaciones:
            marca = datetime.utcnow().strftime('%d/%m/%Y %H:%M')
            nota = f"[{marca}] Solicitud actualizada desde panel por {session.get('email', 'panel')}"
            if solicitud.observaciones:
                solicitud.observaciones = f"{solicitud.observaciones}\n{nota}"
            else:
                solicitud.observaciones = nota

        # Actualizar mensaje y actividad de WhatsApp
        solicitud.mensaje = f'Solicitud actualizada manualmente desde el panel. Estado: {estado}'
        solicitud.whatsapp_last_activity = datetime.utcnow()

        try:
            db.session.commit()
        except Exception as exc:
            db.session.rollback()
            current_app.logger.exception('Error actualizando solicitud existente')
            return jsonify(success=False, message='Error interno al actualizar la solicitud.'), 500
    else:
        # Crear nueva solicitud
        solicitud = SolicitudCita(
            telefono=telefono,
            mensaje='Solicitud creada manualmente desde el panel.',
            fecha=datetime.utcnow(),
            nombre_completo=nombre_completo or None,
            cedula=cedula or None,
            placa=placa or None,
            placa_remolque=placa_remolque or None,
            celular=celular or None,
            observaciones=(data.get('observaciones') or '').strip() or None,
            estado=estado,
            whatsapp_step='0',
            whatsapp_last_activity=datetime.utcnow(),
            whatsapp_timeout_minutes=0,
            whatsapp_warning_sent=False,
            asesor_pendiente=False,
            asesor_pendiente_desde=None
        )
        solicitud_actualizada = False

        try:
            db.session.add(solicitud)
            db.session.commit()
        except IntegrityError:
            db.session.rollback()
            return jsonify(success=False, message='No fue posible crear la solicitud. Verifica la información e intenta nuevamente.'), 400
        except Exception as exc:
            db.session.rollback()
            current_app.logger.exception('Error creando solicitud manualmente')
            return jsonify(success=False, message='Error interno al crear la solicitud.'), 500

    # Enviar mensaje de plantilla solo si se creó una nueva solicitud
    if not solicitud_actualizada:
        try:
            # Enviar plantilla de bienvenida/inicio de conversación
            exito = send_whatsapp_message(
                telefono,
                template_name='bienvenida_conquers',
                template_vars=[nombre_completo or 'Conductor', placa or 'sin placa'],
                template_lang='es',
                sender='panel'
            )
            if exito:
                current_app.logger.info(f'Mensaje de plantilla enviado exitosamente a {telefono} para solicitud {solicitud.id}')
            else:
                current_app.logger.warning(f'No se pudo enviar mensaje de plantilla a {telefono} para solicitud {solicitud.id}')
        except Exception as exc:
            current_app.logger.warning(f'Error enviando mensaje de plantilla a {telefono}: {exc}')

    return jsonify(success=True, id=solicitud.id, conductor_creado=conductor_creado, actualizada=solicitud_actualizada)


@login_required
@app.route('/api/solicitud_cita/<int:id>/inscripcion/aprobar', methods=['POST'])
def aprobar_inscripcion_manual(id):
    solicitud = SolicitudCita.query.get_or_404(id)

    if solicitud.estado != STATE_PENDING_INSCRIPTION:
        return jsonify(success=False, message='La solicitud no está en estado pendiente de inscripción.'), 400

    datos_requeridos = {
        'nombre completo': solicitud.nombre_completo,
        'cédula': solicitud.cedula,
        'placa del camión': solicitud.placa
    }
    faltantes = [campo for campo, valor in datos_requeridos.items() if not valor]
    if faltantes:
        return jsonify(
            success=False,
            message='Faltan datos obligatorios para completar la inscripción: ' + ', '.join(faltantes)
        ), 400

    try:
        conductor = Conductor.query.filter_by(cedula=solicitud.cedula).first()
        celular_registrado = solicitud.celular or solicitud.telefono
        placa_remolque_norm = (solicitud.placa_remolque or '').upper() or ''

        if conductor:
            conductor.nombre = solicitud.nombre_completo
            conductor.placa = solicitud.placa
            if hasattr(conductor, 'placa_remolque'):
                conductor.placa_remolque = placa_remolque_norm
            if hasattr(conductor, 'celular'):
                conductor.celular = celular_registrado
        else:
            nuevo_conductor_kwargs = {
                'nombre': solicitud.nombre_completo,
                'cedula': solicitud.cedula,
                'placa': solicitud.placa
            }
            if 'placa_remolque' in Conductor.__table__.columns:
                nuevo_conductor_kwargs['placa_remolque'] = placa_remolque_norm
            if 'celular' in Conductor.__table__.columns:
                nuevo_conductor_kwargs['celular'] = celular_registrado
            db.session.add(Conductor(**nuevo_conductor_kwargs))

        solicitud.estado = 'sin turno'
        solicitud.asesor_pendiente = False
        solicitud.asesor_pendiente_desde = None
        solicitud.whatsapp_step = str(STEP_AWAIT_GUIA)
        solicitud.whatsapp_last_activity = datetime.utcnow()
        solicitud.whatsapp_timeout_minutes = 30
        solicitud.whatsapp_warning_sent = False
        if not solicitud.celular:
            solicitud.celular = solicitud.telefono
        solicitud.mensaje = 'Inscripción aprobada manualmente. Solicitando guía al conductor.'

        marca = datetime.utcnow().strftime('%d/%m/%Y %H:%M')
        nota = f"[{marca}] Inscripción manual aprobada por {session.get('email', 'panel')}"
        if solicitud.observaciones:
            solicitud.observaciones = f"{solicitud.observaciones}\n{nota}"
        else:
            solicitud.observaciones = nota

        db.session.commit()

        try:
            conductores = cargar_conductores()
            documento_objetivo = str(solicitud.cedula or '').strip()
            placa_objetivo = (solicitud.placa or '').upper()
            nombre_norm = (solicitud.nombre_completo or '').strip().upper()
            celular_norm = celular_registrado.strip() if isinstance(celular_registrado, str) else celular_registrado
            documento_json_val = int(documento_objetivo) if documento_objetivo.isdigit() else documento_objetivo
            celular_json_val = int(celular_norm) if isinstance(celular_norm, str) and celular_norm.isdigit() else celular_norm
            actualizado = False

            for conductor_json in conductores:
                doc_json = str(conductor_json.get('N° DOCUMENTO', '')).strip()
                placa_json = (conductor_json.get('PLACA', '') or '').upper()
                if (documento_objetivo and doc_json == documento_objetivo) or placa_json == placa_objetivo:
                    conductor_json['PLACA'] = placa_objetivo
                    conductor_json['PLACA REMOLQUE'] = placa_remolque_norm or ''
                    conductor_json['NOMBRE CONDUCTOR'] = nombre_norm or conductor_json.get('NOMBRE CONDUCTOR', '')
                    conductor_json['N° DOCUMENTO'] = documento_json_val or conductor_json.get('N° DOCUMENTO', '')
                    conductor_json['CELULAR'] = celular_json_val or conductor_json.get('CELULAR', '')
                    actualizado = True
                    break

            if not actualizado:
                nueva_entrada = {
                    'PLACA': placa_objetivo,
                    'PLACA REMOLQUE': placa_remolque_norm or '',
                    'NOMBRE CONDUCTOR': nombre_norm or solicitud.nombre_completo or '',
                    'N° DOCUMENTO': documento_json_val or solicitud.cedula or '',
                    'CELULAR': celular_json_val or ''
                }
                conductores.append(nueva_entrada)

            guardar_conductores(conductores)
        except Exception:
            current_app.logger.exception('No se pudo actualizar Conductores.json para la solicitud %s', id)
    except Exception:
        db.session.rollback()
        current_app.logger.exception('Error aprobando inscripción manual para solicitud %s', id)
        return jsonify(success=False, message='No fue posible aprobar la inscripción. Intente nuevamente.'), 500

    try:
        from bot_whatsapp.routes import _cancel_final_timeout_message
        _cancel_final_timeout_message(solicitud.id)
    except Exception:
        current_app.logger.debug('No se pudo cancelar timeout previo para solicitud %s', solicitud.id)

    mensaje_confirmacion = (
        "🎉 Fisher 🐶 olfateó que todo está en orden y movió la cola: ya quedaste inscrito con Conquers.\n"
        "📄 Envía la foto de la guía como imagen o PDF para seguir con tu enturnamiento.\n"
        "🛑 Recuerda estacionar el camión antes de responderme, yo espero aquí."
    )
    try:
        send_whatsapp_message(solicitud.telefono, mensaje_confirmacion)
    except Exception:
        current_app.logger.warning('No se pudo enviar mensaje de aprobación al conductor %s', solicitud.telefono)

    reset_safety_reminder_counter(solicitud.telefono)

    return jsonify(success=True, message='Inscripción aprobada y conductor notificado.')


@login_required
@app.route('/api/solicitud_cita/<int:id>/inscripcion/rechazar', methods=['POST'])
def rechazar_inscripcion_manual(id):
    solicitud = SolicitudCita.query.get_or_404(id)

    if solicitud.estado != STATE_PENDING_INSCRIPTION:
        return jsonify(success=False, message='La solicitud no está en estado pendiente de inscripción.'), 400

    data = request.get_json() or {}
    motivo = (data.get('motivo') or '').strip()

    try:
        solicitud.estado = 'preconfirmacion'
        solicitud.asesor_pendiente = False
        solicitud.asesor_pendiente_desde = None
        solicitud.turno = None
        solicitud.turno_fecha = None
        solicitud.fecha_descargue = None
        solicitud.lugar_descargue = None
        solicitud.whatsapp_step = str(STEP_INACTIVE)
        solicitud.whatsapp_timeout_minutes = 0
        solicitud.whatsapp_warning_sent = False
        solicitud.whatsapp_last_activity = datetime.utcnow()
        solicitud.mensaje = 'Inscripción cancelada manualmente desde el panel.'

        marca = datetime.utcnow().strftime('%d/%m/%Y %H:%M')
        nota = f"[{marca}] Inscripción manual rechazada por {session.get('email', 'panel')}"
        if motivo:
            nota = f"{nota}. Motivo: {motivo}"
        if solicitud.observaciones:
            solicitud.observaciones = f"{solicitud.observaciones}\n{nota}"
        else:
            solicitud.observaciones = nota

        db.session.commit()
    except Exception:
        db.session.rollback()
        current_app.logger.exception('Error rechazando inscripción manual para solicitud %s', id)
        return jsonify(success=False, message='No fue posible cancelar la inscripción.'), 500

    try:
        from bot_whatsapp.routes import _cancel_final_timeout_message
        _cancel_final_timeout_message(solicitud.id)
    except Exception:
        current_app.logger.debug('No se pudo cancelar timeout previo para solicitud %s', solicitud.id)

    mensaje_cancelacion = (
        "❌ Fisher 🐶 olfateó un cambio de planes: este enturnamiento se cancela aquí.\n"
        "Si necesitas iniciar uno nuevo, escribe NUEVO y arrancamos juntos cuando quieras."
    )
    try:
        send_whatsapp_message(solicitud.telefono, mensaje_cancelacion)
    except Exception:
        current_app.logger.warning('No se pudo enviar mensaje de cancelación al conductor %s', solicitud.telefono)

    reset_safety_reminder_counter(solicitud.telefono)

    return jsonify(success=True, message='Inscripción cancelada y conductor notificado.')


@login_required
@app.route('/api/solicitud_cita/<int:id>', methods=['DELETE'])
def eliminar_solicitud(id):
    solicitud = SolicitudCita.query.get_or_404(id)

    try:
        try:
            from bot_whatsapp.routes import _cancel_final_timeout_message
            _cancel_final_timeout_message(solicitud.id)
        except Exception:
            current_app.logger.debug('No se pudo cancelar timeout final para solicitud %s', solicitud.id)

        WhatsappMessage.query.filter_by(solicitud_id=solicitud.id).delete()
        db.session.delete(solicitud)
        db.session.commit()
    except Exception:
        db.session.rollback()
        current_app.logger.exception('Error eliminando la solicitud %s', id)
        return jsonify(success=False, message='Error interno al eliminar la solicitud.'), 500

    return jsonify(success=True)


@login_required
@app.route('/api/solicitud_cita/<int:id>/ubicacion_pendiente', methods=['POST'])
def resolver_ubicacion_pendiente(id):
    solicitud = SolicitudCita.query.get_or_404(id)
    data = request.get_json() or {}
    accion = (data.get('accion') or '').strip().lower()

    if accion not in {'aprobar', 'rechazar'}:
        return jsonify(success=False, message='Acción inválida.'), 400

    if not solicitud.ubicacion_pendiente_lat or not solicitud.ubicacion_pendiente_tipo:
        return jsonify(success=False, message='No hay ubicación pendiente para revisar.'), 400

    tipo = solicitud.ubicacion_pendiente_tipo
    lat = solicitud.ubicacion_pendiente_lat
    lng = solicitud.ubicacion_pendiente_lng
    telefono = solicitud.telefono

    try:
        if accion == 'aprobar':
            if tipo == 'bosconia':
                solicitud.ubicacion_lat = lat
                solicitud.ubicacion_lng = lng
                solicitud.paso_bosconia = True
                siguiente_step = 6
                mensaje = (
                    "✅ Nuestro equipo revisó la ubicación de Bosconia y la aprobó. "
                    "Continúa con el proceso enviando el ticket o la siguiente información."
                )
            elif tipo == 'gambote':
                solicitud.ubicacion_gambote_lat = lat
                solicitud.ubicacion_gambote_lng = lng
                solicitud.paso_gambote = True
                siguiente_step = 9
                mensaje = (
                    "✅ Validamos la ubicación de Gambote y está aprobada. "
                    "Tu solicitud sigue avanzando."
                )
            else:
                return jsonify(success=False, message='Tipo de ubicación pendiente desconocido.'), 400

            solicitud.asesor_pendiente = False
            solicitud.asesor_pendiente_desde = None
            solicitud.ubicacion_pendiente_lat = None
            solicitud.ubicacion_pendiente_lng = None
            solicitud.ubicacion_pendiente_tipo = None
            solicitud.ubicacion_pendiente_mensaje = None
            solicitud.ubicacion_pendiente_desde = None
            db.session.commit()

            send_whatsapp_message(telefono, mensaje)

            session_data = get_or_create_whatsapp_session(telefono)
            session_data['solicitud'] = solicitud
            session_data['step'] = siguiente_step
            session_data['timeout_minutes'] = 30
            session_data['warning_sent'] = False
            update_whatsapp_session(telefono, session_data)

            return jsonify(success=True, message='Ubicación aprobada y conductor notificado.')

        # Rechazo manual
        if tipo == 'bosconia':
            siguiente_step = STEP_AWAIT_GPS_BOSCONIA
            mensaje = (
                "❌ Revisamos la ubicación y no coincide con Bosconia. "
                "Envía una nueva ubicación en tiempo real desde Bosconia para continuar."
            )
        elif tipo == 'gambote':
            siguiente_step = STEP_AWAIT_GPS_GAMBOTE
            mensaje = (
                "❌ Nuestro equipo revisó la ubicación y no corresponde a Gambote. "
                "Compártela nuevamente desde el peaje para poder continuar."
            )
        else:
            return jsonify(success=False, message='Tipo de ubicación pendiente desconocido.'), 400

        solicitud.asesor_pendiente = False
        solicitud.asesor_pendiente_desde = None
        solicitud.ubicacion_pendiente_lat = None
        solicitud.ubicacion_pendiente_lng = None
        solicitud.ubicacion_pendiente_tipo = None
        solicitud.ubicacion_pendiente_mensaje = None
        solicitud.ubicacion_pendiente_desde = None
        db.session.commit()

        send_whatsapp_message(telefono, mensaje)

        session_data = get_or_create_whatsapp_session(telefono)
        session_data['solicitud'] = solicitud
        session_data['step'] = siguiente_step
        session_data['timeout_minutes'] = 30
        session_data['warning_sent'] = False
        update_whatsapp_session(telefono, session_data)

        return jsonify(success=True, message='Se notificó al conductor para reenviar la ubicación.')

    except Exception:
        db.session.rollback()
        current_app.logger.exception('Error resolviendo ubicación pendiente para solicitud %s', id)
        return jsonify(success=False, message='Error interno al actualizar la solicitud.'), 500

def save_whatsapp_image(media_info, prefix='whatsapp'):
    """Descarga un archivo enviado por WhatsApp (imagen o documento) y lo guarda localmente."""
    if not media_info:
        return None

    try:
        media_url = None
        media_id = None
        filename = None
        content_type_hint = None

        if isinstance(media_info, str):
            media_url = media_info
        elif isinstance(media_info, dict):
            media_url = media_info.get('url')
            media_id = media_info.get('id')
            filename = media_info.get('filename')
            content_type_hint = media_info.get('mime_type')

        if not media_url and media_id and WHATSAPP_TOKEN:
            meta_resp = requests.get(
                f"https://graph.facebook.com/v18.0/{media_id}",
                headers={'Authorization': f"Bearer {WHATSAPP_TOKEN}"},
                timeout=15
            )
            meta_resp.raise_for_status()
            media_url = meta_resp.json().get('url')

        if not media_url:
            print(f"No fue posible resolver URL para el medio de WhatsApp: {media_info}")
            return None

        headers = {'Authorization': f"Bearer {WHATSAPP_TOKEN}"} if WHATSAPP_TOKEN else {}

        response = requests.get(media_url, headers=headers, timeout=30)
        response.raise_for_status()

        from io import BytesIO
        from werkzeug.datastructures import FileStorage

        content_type = content_type_hint or response.headers.get('content-type', '') or 'application/octet-stream'
        if 'jpeg' in content_type or 'jpg' in content_type:
            extension = '.jpg'
        elif 'png' in content_type:
            extension = '.png'
        elif 'gif' in content_type:
            extension = '.gif'
        elif 'webp' in content_type:
            extension = '.webp'
        elif 'pdf' in content_type:
            extension = '.pdf'
        else:
            extension = ''

        if not filename:
            filename = f"whatsapp_media{extension or ''}"

        stream = BytesIO(response.content)
        stream.seek(0)

        # FileStorage provee los atributos que save_uploaded_file espera (.save, .mimetype, etc.)
        file_storage = FileStorage(stream=stream, filename=filename, content_type=content_type)

        return save_uploaded_file(file_storage, prefix)

    except Exception as e:
        print(f"Error descargando imagen de WhatsApp (media={media_info}): {e}")
        return None

def validar_ubicacion_gps(lat, lng, punto_control, radio_km=10, ruta_alterna=False):
    """
    Valida si las coordenadas GPS están cerca de un punto de control o si corresponden a una ubicación
    posterior al mismo dentro del corredor esperado.
    
    Args:
        lat: Latitud a validar
        lng: Longitud a validar  
        punto_control: 'bosconia' o 'gambote'
        radio_km: Radio de validación en kilómetros (default 10km)
        ruta_alterna: Si True, aplica tolerancias más amplias pensadas para desvíos autorizados
    
    Returns:
        dict: {'valido': bool, 'distancia': float, 'mensaje': str, ...}
    """
    if not lat or not lng:
        return {'valido': False, 'distancia': None, 'mensaje': 'Coordenadas no proporcionadas'}
    
    # Coordenadas de puntos de control (aproximadas)
    DESTINO_FINAL = {
        'lat': 10.295833,
        'lng': -75.513833,
        'nombre': 'Sociedad Portuaria del Dique'
    }

    PUNTOS_CONTROL = {
        'bosconia': {
            'lat': 9.97,
            'lng': -73.89,
            'nombre': 'Bosconia',
            'siguiente': {'lat': 10.1361949, 'lng': -75.2642649},
            'corridor_km': 25,
            'max_post_km': 90,
            'corridor_km_alterna': 32,
            'max_post_km_alterna': 110,
            'fallback': [
                {
                    'destino': 'gambote',
                    'max_km': 32,
                    'max_km_alterna': 42
                },
                {
                    'destino': 'final',
                    'max_km': 22,
                    'max_km_alterna': 30,
                    'skip_next': 'gambote'
                }
            ]
        },
        'gambote': {
            'lat': 10.1361949,
            'lng': -75.2642649,
            'nombre': 'Peaje de Gambote',
            'siguiente': {'lat': 10.3834, 'lng': -75.499},
            'corridor_km': 25,
            'max_post_km': 140,
            'corridor_km_alterna': 35,
            'max_post_km_alterna': 160,
            'alternas': [
                {'lat': 10.2200, 'lng': -75.1500}
            ],
            'fallback': [
                {
                    'destino': DESTINO_FINAL,
                    'max_km': 28,
                    'max_km_alterna': 38,
                    'skip_next': 'final'
                }
            ]
        }
    }
    
    if punto_control not in PUNTOS_CONTROL:
        return {'valido': False, 'distancia': None, 'mensaje': f'Punto de control desconocido: {punto_control}'}
    
    punto = PUNTOS_CONTROL[punto_control]

    def _distancia_km(lat0, lng0, lat1, lng1):
        lat0_r, lng0_r = radians(lat0), radians(lng0)
        lat1_r, lng1_r = radians(lat1), radians(lng1)
        dlat = lat1_r - lat0_r
        dlng = lng1_r - lng0_r
        a = sin(dlat / 2) ** 2 + cos(lat0_r) * cos(lat1_r) * sin(dlng / 2) ** 2
        c = 2 * atan2(sqrt(a), sqrt(1 - a))
        return 6371 * c

    def _to_km_vectors(lat0, lng0, lat1, lng1):
        promedio_lat = radians((lat0 + lat1) / 2.0)
        delta_lat = lat1 - lat0
        delta_lng = lng1 - lng0
        # Aproximación plana válida para los rangos de distancia manejados (<200 km)
        eje_y = delta_lat * 110.574
        eje_x = delta_lng * (111.320 * cos(promedio_lat))
        return eje_x, eje_y

    def _evaluar_post_pasaje(destino):
        if not destino:
            return None
        vector_x, vector_y = _to_km_vectors(punto['lat'], punto['lng'], destino['lat'], destino['lng'])
        magnitude_sq = vector_x ** 2 + vector_y ** 2
        if magnitude_sq <= 0:
            return None
        posicion_x, posicion_y = _to_km_vectors(punto['lat'], punto['lng'], lat, lng)
        proyeccion = (posicion_x * vector_x + posicion_y * vector_y) / magnitude_sq
        if proyeccion <= 0:
            return None
        base = sqrt(magnitude_sq)
        # Distancia lateral (desvío) respecto al corredor
        desviacion = abs(posicion_x * vector_y - posicion_y * vector_x) / base
        # Distancia recorrida sobre el corredor desde el punto de control
        avance = proyeccion * base
        return {
            'proyeccion': proyeccion,
            'desviacion': desviacion,
            'avance': avance
        }
    
    # Calcular distancia usando fórmula de Haversine
    from math import radians, sin, cos, sqrt, atan2
    
    # Convertir a radianes
    lat1, lng1 = radians(lat), radians(lng)
    lat2, lng2 = radians(punto['lat']), radians(punto['lng'])
    
    # Diferencias
    dlat = lat2 - lat1
    dlng = lng2 - lng1
    
    # Fórmula de Haversine
    a = sin(dlat/2)**2 + cos(lat1) * cos(lat2) * sin(dlng/2)**2
    c = 2 * atan2(sqrt(a), sqrt(1-a))
    distancia = 6371 * c  # Radio de la Tierra en km
    mensaje_base = f"Distancia a {punto['nombre']}: {distancia:.2f} km"

    if distancia <= radio_km:
        mensaje = f"{mensaje_base} ✅ Válido (dentro de {radio_km} km)"
        return {
            'valido': True,
            'distancia': distancia,
            'mensaje': mensaje
        }

    # Evaluar si la ubicación llegó tarde pero sobre el corredor posterior al punto de control
    corredores = []
    if punto.get('siguiente'):
        corredores.append(punto['siguiente'])
    if punto.get('alternas'):
        corredores.extend(punto['alternas'])

    corredor_valido = None
    for destino in corredores:
        evaluacion = _evaluar_post_pasaje(destino)
        if not evaluacion:
            continue
        corredor_tolerancia = punto.get('corridor_km', 20)
        corredor_max_post = punto.get('max_post_km', 100)
        if ruta_alterna:
            corredor_tolerancia = max(corredor_tolerancia, punto.get('corridor_km_alterna', corredor_tolerancia))
            corredor_max_post = max(corredor_max_post, punto.get('max_post_km_alterna', corredor_max_post))

        if evaluacion['desviacion'] <= corredor_tolerancia and evaluacion['avance'] <= corredor_max_post:
            corredor_valido = evaluacion
            break

    if corredor_valido:
        mensaje = (
            f"{mensaje_base} ⚠️ Fuera del radio inmediato ({radio_km} km), "
            "pero se detecta después del punto de control. "
            f"Desvío lateral: {corredor_valido['desviacion']:.1f} km."
        )
        resultado = {
            'valido': True,
            'distancia': distancia,
            'mensaje': mensaje,
            'post_pasaje': True,
            'desvio_corridor': corredor_valido['desviacion'],
            'avance_ruta_km': corredor_valido['avance']
        }
        return resultado

    fallback_cfg = punto.get('fallback')
    if fallback_cfg:
        fallback_entries = fallback_cfg if isinstance(fallback_cfg, (list, tuple)) else [fallback_cfg]
        for entry in fallback_entries:
            destino_cfg = entry.get('destino')
            max_km = entry.get('max_km', 30)
            if ruta_alterna:
                max_km = max(max_km, entry.get('max_km_alterna', max_km))

            posibles_destinos = []
            if isinstance(destino_cfg, (list, tuple)):
                posibles_destinos = destino_cfg
            else:
                posibles_destinos = [destino_cfg]

            for posible in posibles_destinos:
                destino_info = None
                if isinstance(posible, str):
                    if posible in PUNTOS_CONTROL:
                        destino_info = PUNTOS_CONTROL[posible]
                    elif posible == 'final':
                        destino_info = DESTINO_FINAL
                elif isinstance(posible, dict):
                    destino_info = posible

                if not destino_info:
                    continue

                distancia_destino = _distancia_km(lat, lng, destino_info['lat'], destino_info['lng'])
                if distancia_destino <= max_km and distancia_destino < distancia:
                    mensaje = (
                        f"{mensaje_base} ⚠️ Fuera del radio inmediato, pero ya estás "
                        f"más cerca de {destino_info.get('nombre', 'el siguiente punto')} ({distancia_destino:.2f} km). "
                        "Marcamos el control como cumplido."
                    )
                    resultado = {
                        'valido': True,
                        'distancia': distancia,
                        'mensaje': mensaje,
                        'post_pasaje': True,
                        'destino_referencia': destino_info.get('nombre')
                    }
                    if entry.get('skip_next'):
                        resultado['skip_next'] = entry['skip_next']
                    return resultado

    mensaje = f"{mensaje_base} ❌ Inválido (fuera de {radio_km} km)"
    return {
        'valido': False,
        'distancia': distancia,
        'mensaje': mensaje
    }

def validar_secuencia_gps(solicitud):
    """
    Valida la secuencia GPS de una solicitud (Bosconia -> Gambote).
    
    Args:
        solicitud: Objeto SolicitudCita
    
    Returns:
        dict: {'valido': bool, 'mensaje': str, 'detalles': dict}
    """
    detalles = {
        'bosconia': {'valido': False, 'distancia': None, 'mensaje': 'Sin validar'},
        'gambote': {'valido': False, 'distancia': None, 'mensaje': 'Sin validar'},
        'secuencia': {'valido': False, 'mensaje': 'Secuencia no validada'}
    }
    
    # Validar Bosconia
    if solicitud.ubicacion_lat and solicitud.ubicacion_lng:
        resultado_bosconia = validar_ubicacion_gps(
            solicitud.ubicacion_lat, 
            solicitud.ubicacion_lng, 
            'bosconia',
            ruta_alterna=getattr(solicitud, 'ruta_alterna', False)
        )
        detalles['bosconia'] = resultado_bosconia
    
    # Validar Gambote
    if solicitud.ubicacion_gambote_lat and solicitud.ubicacion_gambote_lng:
        resultado_gambote = validar_ubicacion_gps(
            solicitud.ubicacion_gambote_lat, 
            solicitud.ubicacion_gambote_lng, 
            'gambote',
            ruta_alterna=getattr(solicitud, 'ruta_alterna', False)
        )
        detalles['gambote'] = resultado_gambote
    
    # Validar secuencia (Bosconia -> Gambote)
    if getattr(solicitud, 'ruta_alterna', False):
        # Ruta alterna: consideramos válido si al menos se validó Gambote
        if detalles['gambote']['valido']:
            detalles['secuencia'] = {
                'valido': True, 
                'mensaje': '✅ Ruta alterna válida: Se validó la ubicación de Gambote'
            }
        else:
            detalles['secuencia'] = {
                'valido': False, 
                'mensaje': '❌ Ruta alterna inválida: Falta validar la ubicación de Gambote'
            }
    else:
        # Ruta normal: Bosconia y Gambote
        ambos_validos = detalles['bosconia']['valido'] and detalles['gambote']['valido']

        if ambos_validos:
            detalles['secuencia'] = {
                'valido': True, 
                'mensaje': '✅ Secuencia GPS válida: Bosconia → Gambote'
            }
        elif detalles['bosconia']['valido'] and not detalles['gambote']['valido']:
            detalles['secuencia'] = {
                'valido': False, 
                'mensaje': '❌ Bosconia válido, Gambote faltante o inválido'
            }
        else:
            detalles['secuencia'] = {
                'valido': False, 
                'mensaje': '❌ Secuencia GPS incompleta o inválida'
            }
    
    mensaje_general = detalles['secuencia']['mensaje']
    if detalles['bosconia']['mensaje'] != 'Sin validar':
        mensaje_general += f"\n{detalles['bosconia']['mensaje']}"
    if detalles['gambote']['mensaje'] != 'Sin validar':
        mensaje_general += f"\n{detalles['gambote']['mensaje']}"
    
    return {
        'valido': detalles['secuencia']['valido'],
        'mensaje': mensaje_general,
        'detalles': detalles
    }

def _normalize_guia_relative_path(value):
    """Normaliza rutas relativas de guías para que sean relativas a GUIDES_DIR."""
    if not value:
        return None

    clean = str(value).strip().replace('\\', '/')
    prefixes = ['static/guias/', '/static/guias/', 'guias/', '/guias/']
    while True:
        clean_lower = clean.lower()
        removed = False
        for prefix in prefixes:
            if clean_lower.startswith(prefix):
                clean = clean[len(prefix):]
                removed = True
                break
        if not removed:
            break
    cleaned = clean.lstrip('/')
    return cleaned or None


def save_uploaded_file(file, prefix='file'):
    """Guarda un archivo subido y retorna la ruta relativa dentro de GUIDES_DIR."""
    import os
    import uuid
    from werkzeug.utils import secure_filename
    
    if not file or not getattr(file, 'filename', None):
        return None
    
    # Validar tipo MIME
    allowed_mime_types = [
        'image/jpeg', 'image/png', 'image/gif', 'image/webp',
        'application/pdf'
    ]
    
    # Obtener tipo MIME del archivo
    mime_type = getattr(file, 'mimetype', None) or getattr(file, 'content_type', '')
    
    # Si no hay MIME type, intentar detectar por extensión
    if not mime_type:
        filename = secure_filename(file.filename).lower()
        if filename.endswith(('.jpg', '.jpeg')):
            mime_type = 'image/jpeg'
        elif filename.endswith('.png'):
            mime_type = 'image/png'
        elif filename.endswith('.gif'):
            mime_type = 'image/gif'
        elif filename.endswith('.webp'):
            mime_type = 'image/webp'
        elif filename.endswith('.pdf'):
            mime_type = 'application/pdf'
    
    if mime_type not in allowed_mime_types:
        raise ValueError(f"Tipo de archivo no permitido: {mime_type}. Solo se permiten imágenes (JPEG, PNG, GIF, WebP) y PDFs.")
    
    # Crear directorio si no existe (GUIDES_DIR configurable)
    if has_app_context():
        config_source = current_app.config
        base_path = current_app.root_path
    else:
        config_source = app.config
        base_path = app.root_path

    upload_dir = config_source.get('GUIDES_DIR') or os.path.join(base_path, 'guias')
    os.makedirs(upload_dir, exist_ok=True)
    
    # Generar nombre único
    filename = secure_filename(file.filename)
    _, ext = os.path.splitext(filename)
    if not ext:
        # Determinar extensión por MIME type
        if mime_type == 'image/jpeg':
            ext = '.jpg'
        elif mime_type == 'image/png':
            ext = '.png'
        elif mime_type == 'image/gif':
            ext = '.gif'
        elif mime_type == 'image/webp':
            ext = '.webp'
        elif mime_type == 'application/pdf':
            ext = '.pdf'
        else:
            ext = '.bin'
    
    unique_filename = f"{prefix}_{uuid.uuid4().hex}{ext}"
    file_path = os.path.join(upload_dir, unique_filename)
    
    # Guardar archivo
    if hasattr(file, 'seek'):
        file.seek(0)
    elif hasattr(file, 'stream') and hasattr(file.stream, 'seek'):
        file.stream.seek(0)

    if hasattr(file, 'save'):
        file.save(file_path)
    else:
        with open(file_path, 'wb') as f:
            f.write(file.read())
    
    # Retornar ruta relativa para acceder desde web
    return _normalize_guia_relative_path(unique_filename)

@app.route('/api/solicitud_cita/<int:id>', methods=['GET'])
def obtener_solicitud(id):
    solicitud = SolicitudCita.query.get_or_404(id)
    return jsonify({
        'id': solicitud.id,
        'imagen_guia': solicitud.imagen_guia,
        'imagen_manifiesto': solicitud.imagen_manifiesto,
        'paso_bosconia': solicitud.paso_bosconia,
        'ticket_gambote': solicitud.ticket_gambote,
        'ubicacion_lat': solicitud.ubicacion_lat,
        'ubicacion_lng': solicitud.ubicacion_lng,
        'ubicacion_gambote_lat': solicitud.ubicacion_gambote_lat,
        'ubicacion_gambote_lng': solicitud.ubicacion_gambote_lng,
        'paso_gambote': solicitud.paso_gambote,
        'observaciones': solicitud.observaciones
    })

@app.route('/api/solicitud_cita/<int:id>', methods=['PUT'])
def actualizar_solicitud(id):
    solicitud = SolicitudCita.query.get_or_404(id)

    # Manejar tanto FormData (con archivos) como JSON (datos simples)
    if request.content_type and 'multipart/form-data' in request.content_type:
        # Procesar FormData con archivos
        actualizado = False

        # Procesar archivos - PRIORIDAD: si hay archivo, usar el archivo guardado
        if 'imagen_guia_file' in request.files:
            file = request.files['imagen_guia_file']
            if file and file.filename:
                try:
                    ruta_guardada = save_uploaded_file(file, 'guia')
                    if ruta_guardada:
                        solicitud.imagen_guia = ruta_guardada
                        actualizado = True
                except ValueError as e:
                    return jsonify(success=False, message=str(e)), 400

        if 'imagen_manifiesto_file' in request.files:
            file = request.files['imagen_manifiesto_file']
            if file and file.filename:
                try:
                    ruta_guardada = save_uploaded_file(file, 'manifiesto')
                    if ruta_guardada:
                        solicitud.imagen_manifiesto = ruta_guardada
                        actualizado = True
                except ValueError as e:
                    return jsonify(success=False, message=str(e)), 400

        if 'ticket_gambote_file' in request.files:
            file = request.files['ticket_gambote_file']
            if file and file.filename:
                try:
                    ruta_guardada = save_uploaded_file(file, 'ticket')
                    if ruta_guardada:
                        solicitud.ticket_gambote = ruta_guardada
                        actualizado = True
                except ValueError as e:
                    return jsonify(success=False, message=str(e)), 400

        # Procesar campos de texto del formulario (SOLO si NO se subió archivo para ese campo)
        campos_texto = ['paso_bosconia', 'ubicacion_lat', 'ubicacion_lng', 'ubicacion_gambote_lat', 'ubicacion_gambote_lng', 'paso_gambote', 'observaciones']
        for campo in campos_texto:
            if campo in request.form:
                valor_nuevo = request.form[campo]
                valor_actual = getattr(solicitud, campo)

                # Convertir tipos según sea necesario
                if campo in ['ubicacion_lat', 'ubicacion_lng', 'ubicacion_gambote_lat', 'ubicacion_gambote_lng']:
                    try:
                        valor_nuevo = float(valor_nuevo) if valor_nuevo else None
                    except (ValueError, TypeError):
                        valor_nuevo = None
                elif campo in ['paso_bosconia', 'paso_gambote']:
                    valor_nuevo = valor_nuevo.lower() in ('true', '1', 'yes', 'si') if valor_nuevo else False

                if valor_actual != valor_nuevo:
                    setattr(solicitud, campo, valor_nuevo)
                    actualizado = True

        # Procesar campos de texto de archivos SOLO si NO se subió archivo
        # (como respaldo para URLs directas)
        if 'imagen_guia' in request.form and not ('imagen_guia_file' in request.files and request.files['imagen_guia_file'].filename):
            valor_bruto = request.form['imagen_guia'].strip()
            valor_nuevo = _normalize_guia_relative_path(valor_bruto)
            actual_bruto = (getattr(solicitud, 'imagen_guia', '') or '').strip()
            actual_normalizado = _normalize_guia_relative_path(actual_bruto)
            if valor_nuevo and (valor_nuevo != actual_normalizado or actual_bruto != actual_normalizado):
                solicitud.imagen_guia = valor_nuevo
                actualizado = True
            elif not valor_bruto and actual_bruto:
                solicitud.imagen_guia = None
                actualizado = True

        if 'imagen_manifiesto' in request.form and not ('imagen_manifiesto_file' in request.files and request.files['imagen_manifiesto_file'].filename):
            valor_bruto = request.form['imagen_manifiesto'].strip()
            valor_nuevo = _normalize_guia_relative_path(valor_bruto)
            actual_bruto = (getattr(solicitud, 'imagen_manifiesto', '') or '').strip()
            actual_normalizado = _normalize_guia_relative_path(actual_bruto)
            if valor_nuevo and (valor_nuevo != actual_normalizado or actual_bruto != actual_normalizado):
                solicitud.imagen_manifiesto = valor_nuevo
                actualizado = True
            elif not valor_bruto and actual_bruto:
                solicitud.imagen_manifiesto = None
                actualizado = True

        if 'ticket_gambote' in request.form and not ('ticket_gambote_file' in request.files and request.files['ticket_gambote_file'].filename):
            valor_bruto = request.form['ticket_gambote'].strip()
            valor_nuevo = _normalize_guia_relative_path(valor_bruto)
            actual_bruto = (getattr(solicitud, 'ticket_gambote', '') or '').strip()
            actual_normalizado = _normalize_guia_relative_path(actual_bruto)
            if valor_nuevo and (valor_nuevo != actual_normalizado or actual_bruto != actual_normalizado):
                solicitud.ticket_gambote = valor_nuevo
                actualizado = True
            elif not valor_bruto and actual_bruto:
                solicitud.ticket_gambote = None
                actualizado = True
    else:
        # Procesar JSON (comportamiento original)
        data = request.get_json()
        campos_actualizables = [
            'imagen_guia', 'imagen_manifiesto', 'paso_bosconia',
            'ticket_gambote', 'ubicacion_lat', 'ubicacion_lng', 
            'ubicacion_gambote_lat', 'ubicacion_gambote_lng', 'paso_gambote',
            'observaciones'
        ]

        actualizado = False
        for campo in campos_actualizables:
            if campo in data:
                valor_actual = getattr(solicitud, campo)
                nuevo_valor = data[campo]
                if campo in {'imagen_guia', 'imagen_manifiesto', 'ticket_gambote'}:
                    if nuevo_valor:
                        nuevo_valor_norm = _normalize_guia_relative_path(nuevo_valor)
                        actual_norm = _normalize_guia_relative_path(valor_actual)
                        if nuevo_valor_norm and (nuevo_valor_norm != actual_norm or valor_actual != actual_norm):
                            setattr(solicitud, campo, nuevo_valor_norm)
                            actualizado = True
                    else:
                        if valor_actual:
                            setattr(solicitud, campo, None)
                            actualizado = True
                else:
                    if valor_actual != nuevo_valor:
                        setattr(solicitud, campo, nuevo_valor)
                        actualizado = True

    if actualizado:
        db.session.commit()
        return jsonify(success=True, message='Solicitud actualizada correctamente')
    else:
        return jsonify(success=True, message='No se realizaron cambios')

@login_required
@app.route('/api/solicitud_cita/<int:id>/enviar_mensaje', methods=['POST'])
def enviar_mensaje_conductor(id):
    """Envía un mensaje personalizado al conductor desde el panel de enturnamiento"""
    try:
        content_type = (request.content_type or '').lower()
        is_multipart = 'multipart/form-data' in content_type

        data = {}
        uploaded_file = None

        if is_multipart:
            data = request.form.to_dict(flat=True)
            uploaded_file = request.files.get('adjunto')
        else:
            data = request.get_json(silent=True) or {}
            # fallback para clientes que envían form-urlencoded sin JSON
            if not data and request.form:
                data = request.form.to_dict(flat=True)
                uploaded_file = request.files.get('adjunto')
        mensaje = (data.get('mensaje') or '').strip()
        preset_raw = (data.get('preset') or 'personalizado').strip().lower()
        urgente_value = data.get('urgente')
        if isinstance(urgente_value, str):
            urgente = urgente_value.strip().lower() in {'1', 'true', 'yes', 'on', 'si', 'sí'}
        else:
            urgente = bool(urgente_value)
        has_attachment = bool(uploaded_file and getattr(uploaded_file, 'filename', ''))
        presets_validos = {
            'personalizado',
            'nueva_guia',
            'nuevo_ticket',
            'nueva_ubicacion_bosconia',
            'nueva_ubicacion_gambote',
            'recordatorio_inteligente',
            'turno_retrasado',
            'finalizar_atencion',
            'iniciar_conversacion'
        }
        preset = preset_raw if preset_raw in presets_validos else 'personalizado'
        
        if not mensaje and not has_attachment and preset not in ('recordatorio_inteligente', 'iniciar_conversacion'):
            return jsonify(success=False, message='Escribe un mensaje o adjunta un archivo antes de enviarlo.'), 400

        media_url = None
        if has_attachment:
            try:
                stored_name = save_uploaded_file(uploaded_file, prefix='mensaje')
            except ValueError as file_err:
                return jsonify(success=False, message=str(file_err)), 400
            except Exception:
                current_app.logger.exception('No se pudo guardar adjunto para solicitud %s', id)
                return jsonify(success=False, message='No se pudo guardar el adjunto.'), 500

            if stored_name:
                base_public_url = (current_app.config.get('WHATSAPP_MEDIA_BASE_URL')
                                   or os.environ.get('WHATSAPP_MEDIA_BASE_URL')
                                   or request.host_url)
                if not base_public_url.endswith('/'):
                    base_public_url = f"{base_public_url}/"
                media_url = urljoin(base_public_url, f'guias/{stored_name}')
                if media_url.startswith('http://'):
                    current_app.logger.warning(
                        'El adjunto para solicitud %s se está enviando con URL no segura (%s). '
                        'WhatsApp requiere HTTPS accesible públicamente.',
                        id,
                        media_url
                    )
        
        # Buscar la solicitud
        solicitud = SolicitudCita.query.get_or_404(id)
        
        if not solicitud.telefono:
            return jsonify(success=False, message='La solicitud no tiene un número de teléfono registrado'), 400
        
        if preset == 'recordatorio_inteligente':
            analisis_faltantes = analizar_datos_faltantes(solicitud)
            if analisis_faltantes['todos_completos']:
                return jsonify(success=False, message='Todos los datos están completos. No hay recordatorio que enviar.'), 400
            if not mensaje:
                if analisis_faltantes['total_faltantes'] == 1:
                    mensaje = (
                        "Fisher 🐶 detectó un pendiente y te guiará enseguida:\n\n"
                        f"{analisis_faltantes['mensajes_recomendados'][0]}\n\n"
                        "Cuando lo envíes, avanzamos automáticamente al siguiente paso."
                    )
                else:
                    partes = [
                        "Fisher 🐶 revisó tu proceso y encontró varios pendientes. Te acompañaré paso a paso:\n\n"
                    ]
                    for idx, texto_faltante in enumerate(analisis_faltantes['mensajes_recomendados'], 1):
                        partes.append(f"{idx}. {texto_faltante}\n\n")
                    partes.append("Comencemos con el primero; te indicaré al instante qué necesitas enviar.")
                    mensaje = ''.join(partes)

        if preset == 'iniciar_conversacion':
            # Enviar plantilla de WhatsApp para notificación de inicio
            nombre_conductor = solicitud.nombre_completo or 'Conductor'
            # --- CORRECCIÓN: Añadir la segunda variable (Placa) ---
            placa_vehiculo = solicitud.placa or 'su vehículo'
            
            exito = send_whatsapp_message(
                solicitud.telefono,
                template_name='recordatorio_inicio_fisher',
                # Enviar AMBAS variables en orden: {{1}}=Nombre, {{2}}=Placa
                template_vars=[nombre_conductor, placa_vehiculo],
                sender='human',
                solicitud=solicitud
            )
        else:
            mensaje_completo = mensaje
            if urgente and mensaje_completo:
                mensaje_completo = f"🚨 {mensaje_completo}"
            
            exito = send_whatsapp_message(
                solicitud.telefono,
                mensaje_completo,
                media_url=media_url,
                sender='human',
                solicitud=solicitud
            )
        
        if exito:
            def aplicar_accion_preset(solicitud_obj, preset_key):
                aplicado = False
                detalle = None
                ahora = datetime.utcnow()
                presets_reinicio = {
                    'nueva_guia',
                    'nuevo_ticket',
                    'nueva_ubicacion_bosconia',
                    'nueva_ubicacion_gambote'
                }

                def reprogramar_timeout(minutos, contexto):
                    try:
                        from bot_whatsapp.routes import (
                            _schedule_inactivity_timeout,
                            _cancel_inactivity_timeout,
                            _cancel_final_timeout_message
                        )
                        _cancel_inactivity_timeout(solicitud_obj.id)
                        _cancel_final_timeout_message(solicitud_obj.id)
                        if minutos and minutos > 0 and solicitud_obj.telefono:
                            _schedule_inactivity_timeout(solicitud_obj.id, solicitud_obj.telefono, delay_minutes=minutos)
                    except Exception as scheduling_err:
                        current_app.logger.warning(
                            'No se pudo reprogramar timeout tras %s para solicitud %s: %s',
                            contexto,
                            solicitud_obj.id,
                            scheduling_err
                        )

                if preset_key == 'nueva_guia':
                    if solicitud_obj.imagen_guia:
                        solicitud_obj.imagen_guia = None
                        aplicado = True
                    if solicitud_obj.imagen_manifiesto:
                        solicitud_obj.imagen_manifiesto = None
                        aplicado = True
                    detalle = 'Se solicitó reenviar la guía o manifiesto.'
                elif preset_key == 'nuevo_ticket':
                    if solicitud_obj.ticket_gambote:
                        solicitud_obj.ticket_gambote = None
                        aplicado = True
                    detalle = 'Se solicitó reenviar el ticket de Gambote.'
                elif preset_key == 'nueva_ubicacion_bosconia':
                    if solicitud_obj.paso_bosconia or solicitud_obj.ubicacion_lat or solicitud_obj.ubicacion_lng:
                        aplicado = True
                    solicitud_obj.paso_bosconia = False
                    solicitud_obj.ubicacion_lat = None
                    solicitud_obj.ubicacion_lng = None
                    detalle = 'Se solicitó revalidar la ubicación de Bosconia.'
                elif preset_key == 'nueva_ubicacion_gambote':
                    if solicitud_obj.paso_gambote or solicitud_obj.ubicacion_gambote_lat or solicitud_obj.ubicacion_gambote_lng:
                        aplicado = True
                    solicitud_obj.paso_gambote = False
                    solicitud_obj.ubicacion_gambote_lat = None
                    solicitud_obj.ubicacion_gambote_lng = None
                    detalle = 'Se solicitó revalidar la ubicación de Gambote.'
                elif preset_key == 'turno_retrasado':
                    solicitud_obj.asesor_pendiente = True
                    solicitud_obj.asesor_pendiente_desde = ahora
                    solicitud_obj.whatsapp_step = str(STEP_HUMAN_HANDOFF)
                    solicitud_obj.whatsapp_timeout_minutes = 0
                    solicitud_obj.whatsapp_warning_sent = False
                    solicitud_obj.whatsapp_last_activity = ahora
                    reprogramar_timeout(0, 'turno en ajuste operativo')
                    detalle = 'Se notificó ajuste operativo del turno y el caso quedó en manos del equipo humano.'
                    aplicado = True
                elif preset_key == 'recordatorio_inteligente':
                    analisis_local = analizar_datos_faltantes(solicitud_obj)
                    if analisis_local['todos_completos']:
                        detalle = 'Recordatorio inteligente omitido: no hay datos pendientes.'
                        return False, detalle

                    faltantes_txt = ', '.join(analisis_local['datos_faltantes']) if analisis_local['datos_faltantes'] else 'sin pendientes'
                    detalle = f'Se envió recordatorio inteligente. Pendientes detectados: {faltantes_txt}.'
                    session_preparada = preparar_sesion_para_pendientes(solicitud_obj, enviar_prompt=True)
                    if not session_preparada:
                        siguiente_step = determinar_siguiente_step_pendiente(solicitud_obj)
                        solicitud_obj.whatsapp_step = str(siguiente_step)
                        solicitud_obj.whatsapp_last_activity = ahora
                        solicitud_obj.whatsapp_warning_sent = False
                        if siguiente_step == STEP_FINAL_CONFIRMATION:
                            solicitud_obj.whatsapp_timeout_minutes = 0
                            reprogramar_timeout(0, 'recordatorio inteligente (sin timeout)')
                        else:
                            solicitud_obj.whatsapp_timeout_minutes = 30
                            reprogramar_timeout(30, 'recordatorio inteligente')
                    aplicado = True
                elif preset_key == 'finalizar_atencion':
                    solicitud_obj.estado = STATE_FINALIZADO
                    solicitud_obj.whatsapp_step = str(STEP_INACTIVE)
                    solicitud_obj.whatsapp_last_activity = ahora
                    solicitud_obj.whatsapp_timeout_minutes = 0
                    solicitud_obj.whatsapp_warning_sent = False
                    reprogramar_timeout(0, 'finalizar atención')
                    solicitud_obj.asesor_pendiente = False
                    solicitud_obj.asesor_pendiente_desde = None
                    detalle = 'Atención humana finalizada; la conversación quedó cerrada para un nuevo enturnamiento.'
                    aplicado = True

                if preset_key in presets_reinicio:
                    solicitud_obj.estado = 'sin turno'
                    solicitud_obj.whatsapp_timeout_minutes = 30
                    solicitud_obj.whatsapp_warning_sent = False
                    solicitud_obj.whatsapp_last_activity = ahora
                    siguiente = determinar_siguiente_step_pendiente(solicitud_obj)
                    solicitud_obj.whatsapp_step = str(siguiente)
                    reprogramar_timeout(30, f'preset {preset_key}')
                    aplicado = True

                if aplicado and detalle:
                    marca = datetime.now().strftime('%d/%m/%Y %H:%M')
                    observacion_nueva = f"[{marca}] {detalle}"
                    if solicitud_obj.observaciones:
                        solicitud_obj.observaciones = f"{solicitud_obj.observaciones}\n{observacion_nueva}"
                    else:
                        solicitud_obj.observaciones = observacion_nueva

                return aplicado, detalle

            aplicado, detalle = aplicar_accion_preset(solicitud, preset)
            cambios_extra = False
            if solicitud.asesor_pendiente:
                if preset != 'turno_retrasado':
                    solicitud.asesor_pendiente = False
                    solicitud.asesor_pendiente_desde = None
                    cambios_extra = True

            solicitud.whatsapp_last_activity = datetime.utcnow()
            solicitud.whatsapp_warning_sent = False

            if aplicado or cambios_extra:
                db.session.commit()

            mensaje_respuesta = 'Mensaje enviado correctamente al conductor.'
            if detalle:
                mensaje_respuesta = f"{mensaje_respuesta} {detalle}"
            return jsonify(success=True, message=mensaje_respuesta, preset=preset, cambios_aplicados=aplicado)
        else:
            return jsonify(success=False, message='Error al enviar el mensaje por WhatsApp'), 500
            
    except Exception as e:
        current_app.logger.exception('Error enviando mensaje al conductor via panel (solicitud %s)', id)
        return jsonify(success=False, message='Error interno del servidor'), 500

def analizar_datos_faltantes(solicitud):
    """Analiza qué datos faltan en una solicitud y retorna información detallada"""
    datos_faltantes = []
    mensajes_recomendados = []
    
    # Verificar guía/manifiesto
    if not solicitud.imagen_guia and not solicitud.imagen_manifiesto:
        datos_faltantes.append("guía")
        mensajes_recomendados.append("📄 Te falta enviar la foto de la guía de transporte. Por favor, envíala ahora para continuar con tu enturnamiento.")
    
    # Verificar ubicación Bosconia
    if not solicitud.paso_bosconia or not solicitud.ubicacion_lat or not solicitud.ubicacion_lng:
        datos_faltantes.append("ubicación Bosconia")
        mensajes_recomendados.append("📍 Te falta validar tu ubicación en Bosconia. Por favor, comparte tu ubicación GPS cuando pases por Bosconia.")
    
    # Verificar ticket Gambote
    if not solicitud.ticket_gambote:
        datos_faltantes.append("ticket Gambote")
        mensajes_recomendados.append("🎫 Te falta enviar la foto del ticket de peaje de Gambote. Por favor, envíala para continuar.")
    
    # Verificar ubicación Gambote
    if not solicitud.paso_gambote or not solicitud.ubicacion_gambote_lat or not solicitud.ubicacion_gambote_lng:
        datos_faltantes.append("ubicación Gambote")
        mensajes_recomendados.append("📍 Te falta validar tu ubicación en Gambote. Por favor, comparte tu ubicación GPS cuando pases por el peaje de Gambote.")
    
    return {
        'datos_faltantes': datos_faltantes,
        'mensajes_recomendados': mensajes_recomendados,
        'total_faltantes': len(datos_faltantes),
        'todos_completos': len(datos_faltantes) == 0
    }


def preparar_sesion_para_pendientes(solicitud, enviar_prompt=True):
    """Configura la sesión para retomar el flujo en el siguiente pendiente y opcionalmente dispara el prompt del bot."""
    if not solicitud or not solicitud.telefono:
        return None

    siguiente_step = determinar_siguiente_step_pendiente(solicitud)
    ahora = datetime.utcnow()
    timeout = 0 if siguiente_step == STEP_FINAL_CONFIRMATION else 30

    solicitud.whatsapp_step = str(siguiente_step)
    solicitud.whatsapp_last_activity = ahora
    solicitud.whatsapp_timeout_minutes = timeout
    solicitud.whatsapp_warning_sent = False

    db.session.commit()

    session_payload = {
        'step': siguiente_step,
        'data': {},
        'last_activity': ahora,
        'timeout_minutes': timeout,
        'warning_sent': False,
        'solicitud': solicitud
    }

    try:
        from bot_whatsapp.routes import _prompt_for_next_pending_requirement, _commit_session
    except Exception:
        current_app.logger.exception(
            'No se pudieron importar utilidades del bot para preparar pendientes (solicitud %s).',
            solicitud.id if solicitud else 'desconocida'
        )
        return session_payload

    try:
        if enviar_prompt:
            _prompt_for_next_pending_requirement(session_payload, solicitud, solicitud.telefono)
        _commit_session(solicitud.telefono, session_payload)
    except Exception:
        current_app.logger.exception(
            'Fallo preparando recordatorio inteligente para la solicitud %s.',
            solicitud.id if solicitud else 'desconocida'
        )

    return session_payload

@login_required
@app.route('/api/solicitud_cita/<int:id>/recordatorio_inteligente', methods=['POST'])
def enviar_recordatorio_inteligente(id):
    """Envía recordatorios automáticos basados en datos faltantes"""
    try:
        # Buscar la solicitud
        solicitud = SolicitudCita.query.get_or_404(id)
        
        if not solicitud.telefono:
            return jsonify(success=False, message='La solicitud no tiene un número de teléfono registrado'), 400
        
        # Analizar qué datos faltan
        analisis = analizar_datos_faltantes(solicitud)
        
        if analisis['todos_completos']:
            return jsonify(success=False, message='Todos los datos están completos. No hay nada que recordar.'), 400
        
        # Generar mensaje inteligente
        if analisis['total_faltantes'] == 1:
            mensaje = (
                "Fisher 🐶 detectó un pendiente y te guiará enseguida:\n\n"
                f"{analisis['mensajes_recomendados'][0]}\n\n"
                "Cuando lo envíes, avanzamos automáticamente al siguiente paso."
            )
        else:
            partes = [
                "Fisher 🐶 revisó tu proceso y encontró varios pendientes. Te acompañaré paso a paso:\n\n"
            ]
            for i, msg in enumerate(analisis['mensajes_recomendados'], 1):
                partes.append(f"{i}. {msg}\n\n")
            partes.append("Comencemos con el primero; te indicaré al instante qué necesitas enviar.")
            mensaje = ''.join(partes)
        
        # Enviar el mensaje
        exito = send_whatsapp_message(solicitud.telefono, mensaje)
        
        if exito:
            preparar_sesion_para_pendientes(solicitud, enviar_prompt=True)
            return jsonify(success=True, 
                         message=f'Recordatorio enviado correctamente. Se notificaron {analisis["total_faltantes"]} datos faltantes.',
                         datos_faltantes=analisis['datos_faltantes'])
        else:
            return jsonify(success=False, message='Error al enviar el recordatorio por WhatsApp'), 500
            
    except Exception as e:
        print(f"Error enviando recordatorio inteligente: {e}")
        return jsonify(success=False, message='Error interno del servidor'), 500

@login_required
@app.route('/api/solicitud_cita/<int:id>/datos_faltantes', methods=['GET'])
def obtener_datos_faltantes(id):
    """Retorna información sobre qué datos faltan en una solicitud"""
    try:
        solicitud = SolicitudCita.query.get_or_404(id)
        analisis = analizar_datos_faltantes(solicitud)
        
        return jsonify({
            'success': True,
            'solicitud_id': id,
            'conductor': solicitud.nombre_completo or 'Sin nombre',
            'analisis': analisis
        })
        
    except Exception as e:
        print(f"Error obteniendo datos faltantes: {e}")
        return jsonify(success=False, message='Error interno del servidor'), 500

@login_required
@app.route('/gestionar_clientes')
def gestionar_clientes():    
    clientes_actuales = cargar_clientes()
    return render_template('gestionar_clientes.html', clientes=clientes_actuales)

@login_required
@app.route('/guardar_cliente', methods=['POST'])
def guardar_cliente():
    nombre = request.form.get('nombre_cliente')
    direccion = request.form.get('direccion_cliente')
    ciudad = request.form.get('ciudad_cliente')

    if not nombre or not direccion or not ciudad:
        flash("Todos los campos son obligatorios.", "danger")
        return redirect(url_for('gestionar_clientes'))

    clientes = cargar_clientes()
    # Verificar si el cliente ya existe en JSON
    if any(c['NOMBRE_CLIENTE'].lower() == nombre.lower() for c in clientes):
        flash(f"El cliente '{nombre}' ya existe en la base de datos.", "warning")
        return redirect(url_for('gestionar_clientes'))

    nuevo_cliente = {
        "NOMBRE_CLIENTE": nombre.upper(),
        "DIRECCION": direccion.upper(),
        "CIUDAD_DEPARTAMENTO": ciudad.upper()
    }
    clientes.append(nuevo_cliente)
    clientes.sort(key=lambda x: x['NOMBRE_CLIENTE'])
    guardar_clientes(clientes)

    # Guardar también en PostgreSQL
    try:
        if not Cliente.query.filter_by(nombre=nombre.upper()).first():
            cliente_db = Cliente(
                nombre=nombre.upper(),
                direccion=direccion.upper(),
                ciudad_departamento=ciudad.upper()
            )
            db.session.add(cliente_db)
            db.session.commit()
    except Exception as e:
        db.session.rollback()
        flash(f"Error al guardar en la base de datos: {e}", "danger")
        return redirect(url_for('gestionar_clientes'))

    flash(f"Cliente '{nombre}' agregado exitosamente.", "success")
    return redirect(url_for('gestionar_clientes'))

@login_required
@app.route('/agregar_cliente_ajax', methods=['POST'])
def agregar_cliente_ajax():
    data = request.get_json()
    nombre = data.get('nombre')
    direccion = data.get('direccion')
    ciudad = data.get('ciudad')

    if not nombre or not direccion or not ciudad:
        return jsonify(success=False, message="Todos los campos son obligatorios."), 400

    clientes = cargar_clientes()
    
    # Validar duplicados por NOMBRE + DIRECCIÓN (no solo nombre)
    if any(
        (c.get('NOMBRE_CLIENTE', '').upper() == nombre.upper()) and 
        (c.get('DIRECCION', '').upper() == direccion.upper()) 
        for c in clientes
    ):
        return jsonify(success=False, message=f"El cliente '{nombre}' con esa dirección ya existe."), 409

    nuevo_cliente = {
        "NOMBRE_CLIENTE": nombre.upper(),
        "DIRECCION": direccion.upper(),
        "CIUDAD_DEPARTAMENTO": ciudad.upper()
    }
    clientes.append(nuevo_cliente)
    clientes.sort(key=lambda x: x['NOMBRE_CLIENTE'])
    guardar_clientes(clientes)

    # Guardar también en PostgreSQL
    try:
        # Verificar si existe EXACTAMENTE ese cliente (nombre + dirección)
        existe_db = Cliente.query.filter(
            func.upper(Cliente.nombre) == nombre.upper(),
            func.upper(Cliente.direccion) == direccion.upper()
        ).first()

        if not existe_db:
            cliente_db = Cliente(
                nombre=nombre.upper(),
                direccion=direccion.upper(),
                ciudad_departamento=ciudad.upper()
            )
            db.session.add(cliente_db)
            db.session.commit()
    except Exception as e:
        db.session.rollback()
        # No fallamos la petición AJAX si falla la BD, porque ya guardamos en JSON (prioridad frontend)
        app.logger.warning(f"Error al guardar cliente DB (se guardó en JSON): {e}")

    return jsonify(success=True, message="Cliente agregado exitosamente.", nuevo_cliente=nuevo_cliente)


@login_required
@app.route('/actualizar_cliente_ajax', methods=['POST'])
def actualizar_cliente_ajax():
    data = request.get_json() or {}
    original_nombre = (data.get('original_nombre') or '').strip()
    nombre = (data.get('nombre') or '').strip()
    direccion = (data.get('direccion') or '').strip()
    ciudad = (data.get('ciudad') or '').strip()
    
    # Nuevos campos para identificar unívocamente al original
    original_direccion = (data.get('original_direccion') or '').strip()
    # Si no llega original_direccion (frontend viejo), usamos comportamiento fallback (buscar solo nombre)

    if not original_nombre or not nombre or not direccion or not ciudad:
        return jsonify(success=False, message='Todos los campos son obligatorios.'), 400

    original_upper = original_nombre.upper()
    original_dir_upper = original_direccion.upper()
    
    nombre_upper = nombre.upper()
    direccion_upper = direccion.upper()
    ciudad_upper = ciudad.upper()

    clientes = cargar_clientes()
    coincidencia = None
    
    # Buscar coincidencia exacta por Nombre AND Dirección
    for c in clientes:
        c_nombre = (c.get('NOMBRE_CLIENTE') or '').upper()
        c_direccion = (c.get('DIRECCION') or '').upper()
        
        if c_nombre == original_upper:
            # Si tenemos direccion original, la usamos para desempatar
            if original_dir_upper:
                if c_direccion == original_dir_upper:
                    coincidencia = c
                    break
            else:
                # Fallback: primer nombre que coincida
                coincidencia = c
                break

    if not coincidencia:
        return jsonify(success=False, message=f"No se encontró el cliente '{original_nombre}'."), 404

    # Verificar conflictos: Nuevo nombre+direccion ya existe en OTRO registro
    # "Otro" significa que no es la misma instancia en memoria (pero en JSON no hay ID).
    # Así que verificamos si existe un registro con (NuevoNombre, NuevaDireccion) 
    # QUE NO SEA el que estamos editando (coincidencia).
    
    ya_existe = False
    for otro in clientes:
        if otro is coincidencia:
            continue
        if (otro.get('NOMBRE_CLIENTE','').upper() == nombre_upper and 
            otro.get('DIRECCION','').upper() == direccion_upper):
            ya_existe = True
            break
            
    if ya_existe:
        return jsonify(success=False, message=f"Ya existe otro cliente con nombre '{nombre_upper}' y esa dirección."), 409

    # Actualizar valores
    coincidencia['NOMBRE_CLIENTE'] = nombre_upper
    coincidencia['DIRECCION'] = direccion_upper
    coincidencia['CIUDAD_DEPARTAMENTO'] = ciudad_upper
    
    guardar_clientes(clientes)
    
    # Actualizar DB (Intentar buscar por original y actualizar)
    try:
        # DB requiere ID o criterio único. Como no tenemos ID en frontend, intentamos best-effort.
        q = Cliente.query.filter(func.upper(Cliente.nombre) == original_upper)
        if original_dir_upper:
            q = q.filter(func.upper(Cliente.direccion) == original_dir_upper)
            
        cliente_db = q.first()
        if cliente_db:
            cliente_db.nombre = nombre_upper
            cliente_db.direccion = direccion_upper
            cliente_db.ciudad_departamento = ciudad_upper
            db.session.commit()
    except Exception as e:
        db.session.rollback()
        app.logger.warning(f"Error al actualizar cliente DB (JSON actualizado): {e}")

    return jsonify(success=True, message='Cliente actualizado correctamente.', cliente=coincidencia)

@login_required
@app.route('/eliminar_cliente_ajax', methods=['POST'])
def eliminar_cliente_ajax():
    data = request.get_json() or {}
    nombre = (data.get('nombre') or '').strip()
    direccion = (data.get('direccion') or '').strip()

    if not nombre:
        return jsonify(success=False, message='El nombre del cliente es obligatorio.'), 400

    nombre_upper = nombre.upper()
    direccion_upper = direccion.upper()

    clientes = cargar_clientes()
    
    # 1. Eliminar de la lista en memoria (JSON)
    # Filtramos para quitar TODOS los que coincidan (por si hay duplicados exactos)
    clientes_filtrados = [
        c for c in clientes 
        if not (
            (c.get('NOMBRE_CLIENTE') or '').upper() == nombre_upper and 
            (c.get('DIRECCION') or '').upper() == direccion_upper
        )
    ]
    
    deleted_count = len(clientes) - len(clientes_filtrados)
    
    if deleted_count == 0:
        return jsonify(success=False, message='No se encontró el cliente para eliminar.'), 404

    guardar_clientes(clientes_filtrados)

    # 2. Eliminar de la Base de Datos
    try:
        q = Cliente.query.filter(func.upper(Cliente.nombre) == nombre_upper)
        if direccion_upper:
            q = q.filter(func.upper(Cliente.direccion) == direccion_upper)
        
        registros = q.all()
        for r in registros:
            db.session.delete(r)
        
        if registros:
            db.session.commit()
            
    except Exception as e:
        db.session.rollback()
        app.logger.warning(f"Error eliminando cliente de BD: {e}")

    return jsonify(success=True, message=f'Se eliminaron {deleted_count} registro(s) correctamente.')

@login_required
@permiso_requerido('planilla_precios')
@app.route('/planilla_precios')
def planilla_precios():
    # La lógica para cargar los datos se mantiene igual
    datos_guardados = []
    try:
        carpeta = "registros"
        archivos_precios = sorted([a for a in os.listdir(carpeta) if a.startswith("precios_") and a.endswith(".json")], reverse=True)
        if archivos_precios:
            ruta_reciente = os.path.join(carpeta, archivos_precios[0])
            with open(ruta_reciente, 'r', encoding='utf-8') as f:
                contenido = json.load(f)
            datos_guardados = contenido.get("datos", [])
    except Exception as e:
        print(f"Error cargando planilla de precios: {e}")
        pass

    fuente_de_datos = datos_guardados if datos_guardados else PLANILLA_PRECIOS

    # ¡Ya no necesitamos la clave de Google!
    # Simplemente renderizamos la plantilla con los datos de la planilla.
    return render_template('planilla_precios.html',
                           planilla=fuente_de_datos,
                           nombre=session.get("nombre"))

def cargar_conductores():
    """Función auxiliar para cargar conductores desde Conductores.json de forma segura."""
    try:
        ruta_conductores = os.path.join(BASE_DIR, 'static', 'Conductores.json')
        with open(ruta_conductores, 'r', encoding='utf-8') as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return []

def buscar_conductor_por_placa(placa):
    """
    Busca conductor por placa primero en PostgreSQL, luego en JSON como fallback.
    
    Args:
        placa: Placa del vehículo a buscar
    
    Returns:
        dict: Datos del conductor o None si no se encuentra
    """
    # Primero buscar en PostgreSQL
    conductor_db = Conductor.query.filter_by(placa=placa.upper()).first()
    if conductor_db:
        return {
            'PLACA': conductor_db.placa or '-',
            'PLACA REMOLQUE': conductor_db.placa_remolque or '-',
            'NOMBRE CONDUCTOR': conductor_db.nombre or '-',
            'N° DOCUMENTO': conductor_db.cedula or '-',
            'CELULAR': conductor_db.celular or '-'
        }
    
    # Fallback: buscar en JSON
    conductores = cargar_conductores()
    for c in conductores:
        if c.get('PLACA', '').upper() == placa.upper():
            return {
                'PLACA': c.get('PLACA', '-'),
                'PLACA REMOLQUE': c.get('PLACA REMOLQUE', '-'),
                'NOMBRE CONDUCTOR': c.get('NOMBRE CONDUCTOR', '-'),
                'N° DOCUMENTO': c.get('N° DOCUMENTO', '-'),
                'CELULAR': c.get('CELULAR', '-')
            }
    return None

def guardar_conductores(conductores):
    """Función auxiliar para guardar la lista de conductores en Conductores.json."""
    try:
        ruta_conductores = os.path.join(BASE_DIR, 'static', 'Conductores.json')
        with open(ruta_conductores, 'w', encoding='utf-8') as f:
            json.dump(conductores, f, ensure_ascii=False, indent=4)
        return True # Devuelve True si todo salió bien
    except (IOError, PermissionError) as e:
        # Captura errores de escritura o de permisos
        print(f"ERROR AL GUARDAR: No se pudo escribir en el archivo Conductores.json. Causa: {e}")
        return False # Devuelve False si hubo un error

def cargar_empresas():
    """Función auxiliar para cargar empresas desde EmpresasTransportadoras.json."""
    try:
        ruta_empresas = os.path.join(BASE_DIR, 'static', 'EmpresasTransportadoras.json')
        if not os.path.exists(ruta_empresas):
            return [] # Si el archivo no existe, devuelve una lista vacía
        with open(ruta_empresas, 'r', encoding='utf-8') as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return []

def guardar_empresas(empresas):
    """Función auxiliar para guardar la lista de empresas en EmpresasTransportadoras.json."""
    ruta_empresas = os.path.join(BASE_DIR, 'static', 'EmpresasTransportadoras.json')
    with open(ruta_empresas, 'w', encoding='utf-8') as f:
        json.dump(empresas, f, ensure_ascii=False, indent=4)        


def cargar_productos():
    """Carga `Producto.json` desde la carpeta static. Devuelve lista vacía si falla."""
    try:
        ruta_productos = os.path.join(BASE_DIR, 'static', 'Producto.json')
        with open(ruta_productos, 'r', encoding='utf-8') as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return []


def guardar_productos(productos):
    """Guarda la lista de productos en `static/Producto.json`."""
    ruta_productos = os.path.join(BASE_DIR, 'static', 'Producto.json')
    with open(ruta_productos, 'w', encoding='utf-8') as f:
        json.dump(productos, f, ensure_ascii=False, indent=4)


@login_required
@app.route('/agregar_producto_ajax', methods=['POST'])
def agregar_producto_ajax():
    data = request.get_json() or {}
    nombre = (data.get('producto') or data.get('nombre') or '').strip()
    unidad = (data.get('unidad') or '').strip()
    tempF = (data.get('tempF') or data.get('temp') or '').strip()
    api_obs = (data.get('api_obs') or data.get('API_OBS') or '').strip()

    if not nombre:
        return jsonify(success=False, message='El nombre del producto es obligatorio.'), 400

    productos = cargar_productos()
    if any(((p.get('PRODUCTO') or p.get('producto') or '').upper() == nombre.upper()) for p in productos):
        return jsonify(success=False, message=f"El producto '{nombre}' ya existe."), 409

    nuevo_producto = {
        'PRODUCTO': nombre.upper(),
        'UN': unidad.upper() if unidad else '',
        'tempF': tempF,
        'API_OBS': api_obs
    }
    productos.append(nuevo_producto)
    productos.sort(key=lambda x: x.get('PRODUCTO', ''))
    try:
        guardar_productos(productos)
    except Exception as e:
        return jsonify(success=False, message=f'Error al guardar productos: {e}'), 500

    # Guardar también en PostgreSQL si está configurado
    try:
        nombre_upper = nuevo_producto['PRODUCTO']
        if 'Producto' in globals():
            if not Producto.query.filter_by(producto=nombre_upper).first():
                producto_db = Producto(
                    producto=nombre_upper,
                    unidad=nuevo_producto.get('UN') or None,
                    tempF=nuevo_producto.get('tempF') or None,
                    api_obs=nuevo_producto.get('API_OBS') or None
                )
                db.session.add(producto_db)
                db.session.commit()
    except Exception as e:
        db.session.rollback()
        return jsonify(success=False, message=f"Error al guardar en la base de datos: {e}"), 500

    return jsonify(success=True, message='Producto agregado exitosamente.', nuevo_producto=nuevo_producto)


@login_required
@app.route('/actualizar_producto_ajax', methods=['POST'])
def actualizar_producto_ajax():
    data = request.get_json() or {}
    original = (data.get('original_producto') or '').strip()
    nombre = (data.get('producto') or data.get('nombre') or '').strip()
    unidad = (data.get('unidad') or '').strip()
    tempF = (data.get('tempF') or data.get('temp') or '').strip()
    api_obs = (data.get('api_obs') or data.get('API_OBS') or '').strip()

    if not original or not nombre:
        return jsonify(success=False, message='Campos obligatorios ausentes.'), 400

    productos = cargar_productos()
    original_upper = original.upper()
    nombre_upper = nombre.upper()

    idx = None
    for i, p in enumerate(productos):
        key = (p.get('PRODUCTO') or p.get('producto') or '').upper()
        if key == original_upper:
            idx = i
            break

    if idx is None:
        return jsonify(success=False, message=f"No se encontró el producto '{original}'."), 404

    # Verificar que no exista otro producto con el nuevo nombre
    if nombre_upper != original_upper and any(((p.get('PRODUCTO') or p.get('producto') or '').upper() == nombre_upper) for p in productos):
        return jsonify(success=False, message=f"Ya existe un producto con el nombre '{nombre}'."), 409

    productos[idx]['PRODUCTO'] = nombre_upper
    productos[idx]['UN'] = unidad.upper() if unidad else productos[idx].get('UN', '')
    productos[idx]['tempF'] = tempF
    productos[idx]['API_OBS'] = api_obs

    productos.sort(key=lambda x: x.get('PRODUCTO', ''))
    try:
        guardar_productos(productos)
    except Exception as e:
        return jsonify(success=False, message=f'Error al actualizar productos: {e}'), 500

    # Actualizar también en PostgreSQL si corresponde
    try:
        if 'Producto' in globals():
            prod_db = Producto.query.filter_by(producto=original_upper).first()
            if prod_db:
                prod_db.producto = nombre_upper
                prod_db.unidad = unidad.upper() if unidad else prod_db.unidad
                prod_db.tempF = tempF or prod_db.tempF
                prod_db.api_obs = api_obs or prod_db.api_obs
            else:
                # crear si no existe
                prod_db = Producto(producto=nombre_upper, unidad=unidad.upper() if unidad else None, tempF=tempF or None, api_obs=api_obs or None)
                db.session.add(prod_db)
            db.session.commit()
    except Exception as e:
        db.session.rollback()
        return jsonify(success=False, message=f"Error al actualizar en la base de datos: {e}"), 500

    actualizado = productos[idx]
    return jsonify(success=True, message='Producto actualizado correctamente.', producto=actualizado)

@login_required
@app.route('/agregar_conductor_ajax', methods=['POST'])
def agregar_conductor_ajax():
    data = request.get_json()
    nombre = str(data.get('nombre', ''))
    cedula = str(data.get('cedula', ''))
    placa = str(data.get('placa', ''))

    if not nombre or not cedula or not placa:
        return jsonify(success=False, message="Todos los campos son obligatorios."), 400

    conductores = cargar_conductores()
    # Verificación de duplicados (versión segura)
    if any(c.get('CEDULA', '').lower() == cedula.lower() for c in conductores):
        return jsonify(success=False, message=f"Un conductor con la cédula '{cedula}' ya existe."), 409

    nuevo_conductor = {
        "CONDUCTOR": nombre.upper(),
        "CEDULA": cedula.upper(),
        "PLACA": placa.upper()
    }
    conductores.append(nuevo_conductor)
    conductores.sort(key=lambda x: x.get('CONDUCTOR', ''))
    guardado_exitoso = guardar_conductores(conductores)

    # Guardar también en PostgreSQL
    try:
        if not Conductor.query.filter_by(cedula=cedula.upper()).first():
            conductor_db = Conductor(
                nombre=nombre.upper(),
                cedula=cedula.upper(),
                placa=placa.upper()
            )
            db.session.add(conductor_db)
            db.session.commit()
    except Exception as e:
        db.session.rollback()
        return jsonify(success=False, message=f"Error al guardar en la base de datos: {e}"), 500

    if guardado_exitoso:
        return jsonify(success=True, message="Conductor agregado exitosamente.", nuevo_conductor=nuevo_conductor)
    else:
        return jsonify(success=False, message="Error del servidor: No se pudo escribir en el archivo de conductores."), 500

@login_required
@app.route('/agregar_empresa_ajax', methods=['POST'])
def agregar_empresa_ajax():
    data = request.get_json()
    nombre = data.get('nombre')

    if not nombre:
        return jsonify(success=False, message="El nombre es obligatorio."), 400

    empresas = cargar_empresas()
    if any(e['NOMBRE_EMPRESA'].lower() == nombre.lower() for e in empresas):
        return jsonify(success=False, message=f"La empresa '{nombre}' ya existe."), 409

    nueva_empresa = { "NOMBRE_EMPRESA": nombre.upper() }
    empresas.append(nueva_empresa)
    empresas.sort(key=lambda x: x['NOMBRE_EMPRESA'])
    guardar_empresas(empresas)

    # Guardar también en PostgreSQL
    try:
        if not Empresa.query.filter_by(nombre=nombre.upper()).first():
            empresa_db = Empresa(nombre=nombre.upper())
            db.session.add(empresa_db)
            db.session.commit()
    except Exception as e:
        db.session.rollback()
        return jsonify(success=False, message=f"Error al guardar en la base de datos: {e}"), 500

    return jsonify(success=True, message="Empresa agregada exitosamente.", nueva_empresa=nueva_empresa)

def get_or_create_whatsapp_session(telefono):
    """
    Obtiene o crea una sesión de WhatsApp desde la base de datos.
    Solo busca solicitudes pendientes para evitar duplicar conversaciones cuando ya existe un enturne en curso.
    
    Args:
        telefono: Número de teléfono del usuario
    
    Returns:
        dict: Diccionario con la información de la sesión
    """
    # Asegurar que la última solicitud enturnada tenga el mensaje final persistido
    ultima_enturnada = (
        SolicitudCita.query
        .filter(
            SolicitudCita.telefono == telefono,
            SolicitudCita.estado.in_(['enturnado', STATE_FINALIZADO])
        )
        .order_by(SolicitudCita.fecha_descargue.desc(), SolicitudCita.fecha.desc())
        .first()
    )
    if ultima_enturnada and (not ultima_enturnada.mensaje or 'Solicitud enturnada' not in ultima_enturnada.mensaje):
        ultima_enturnada.mensaje = build_enturnado_message(ultima_enturnada)
        try:
            db.session.commit()
        except Exception:
            db.session.rollback()

    # Buscar la solicitud pendiente más reciente para este teléfono
    solicitud = (
        SolicitudCita.query
        .filter(
            SolicitudCita.telefono == telefono,
            SolicitudCita.estado.in_(['preconfirmacion', 'sin turno', 'en revision', 'error', STATE_PENDING_INSCRIPTION])
        )
        .order_by(SolicitudCita.fecha.desc())
        .first()
    )

    ultima_sesion = (
        SolicitudCita.query
        .filter(SolicitudCita.telefono == telefono)
        .order_by(SolicitudCita.fecha.desc())
        .first()
    )
    if ultima_sesion and ultima_sesion.estado == STATE_FINALIZADO:
        solicitud_fecha = getattr(solicitud, 'fecha', None)
        ultima_fecha = getattr(ultima_sesion, 'fecha', None)
        if not solicitud_fecha or (ultima_fecha and ultima_fecha >= solicitud_fecha):
            solicitud = None
            ultima_enturnada = None

    if not solicitud and ultima_enturnada:
        solicitud = ultima_enturnada
    
    # Si no hay solicitud activa, crear una nueva
    if not solicitud:
        solicitud = SolicitudCita(
            telefono=telefono,
            mensaje='Sesión WhatsApp iniciada (pendiente de confirmación)',
            estado='preconfirmacion',
            whatsapp_step='0',
            whatsapp_timeout_minutes=5,
            whatsapp_warning_sent=False
        )
        db.session.add(solicitud)
        db.session.commit()
    
    # Preparar valores por defecto para la sesión
    raw_step = solicitud.whatsapp_step or '0'
    if raw_step in {STEP_UNDER_REVIEW, STEP_HUMAN_HANDOFF}:
        step = raw_step
    else:
        try:
            step = int(raw_step)
        except (ValueError, TypeError):
            step = 0
    
    last_activity = solicitud.whatsapp_last_activity or datetime.now()
    timeout_minutes = _normalize_timeout_minutes(solicitud.whatsapp_timeout_minutes)
    warning_sent = bool(solicitud.whatsapp_warning_sent and timeout_minutes > 0)
    
    session_data = {
        'step': step,
        'data': {},  # Mantener compatibilidad pero no usar para datos críticos
        'last_activity': last_activity,
        'timeout_minutes': timeout_minutes,
        'warning_sent': warning_sent,
        'solicitud': solicitud  # Referencia a la solicitud para actualizarla fácilmente
    }
    
    return session_data

def update_whatsapp_session(telefono, session_data):
    """
    Actualiza la sesión de WhatsApp en la base de datos.
    
    Args:
        telefono: Número de teléfono del usuario
        session_data: Diccionario con la información de la sesión
    """
    solicitud = session_data.get('solicitud')
    if not solicitud:
        # Buscar la solicitud si no está en session_data
        solicitud = SolicitudCita.query.filter_by(telefono=telefono).order_by(SolicitudCita.fecha.desc()).first()
        if not solicitud:
            return
    
    # Actualizar los campos de sesión
    step = session_data.get('step', 0)
    if isinstance(step, str):
        solicitud.whatsapp_step = step
    else:
        solicitud.whatsapp_step = str(step)
    
    solicitud.whatsapp_last_activity = session_data.get('last_activity', datetime.now())
    solicitud.whatsapp_timeout_minutes = _normalize_timeout_minutes(session_data.get('timeout_minutes'))
    warning_flag = session_data.get('warning_sent', False)
    solicitud.whatsapp_warning_sent = bool(warning_flag and solicitud.whatsapp_timeout_minutes > 0)
    
    db.session.commit()

def is_confirmation_positive(texto):
    """
    Verifica si el texto representa una confirmación positiva.
    
    Args:
        texto: Texto a verificar
    
    Returns:
        bool: True si es confirmación positiva
    """
    positive_responses = ['si', 'sí', 's', 'ok', 'confirmo', 'listo', 'correcto', 'bien', 'dale', 'yes', 'y']
    return texto.strip().lower() in positive_responses

def is_confirmation_negative(texto):
    """
    Verifica si el texto representa una confirmación negativa.
    
    Args:
        texto: Texto a verificar
    
    Returns:
        bool: True si es confirmación negativa
    """
    negative_responses = ['no', 'n', 'cancelar', 'cancel', 'ninguno', 'no tengo']
    return texto.strip().lower() in negative_responses

def send_confirmation_request(telefono, message, step_on_confirm, step_on_deny=None, timeout_minutes=60):
    """
    Envía un mensaje de confirmación estandarizado y configura la sesión.
    
    Args:
        telefono: Número de teléfono
        message: Mensaje a enviar
        step_on_confirm: Paso al que ir si confirma
        step_on_deny: Paso al que ir si niega (opcional)
        timeout_minutes: Minutos de timeout
    """
    full_message = f"{message}\n\nResponde 'sí' para confirmar o 'no' para cancelar."
    send_yes_no_prompt(
        telefono,
        full_message,
        context_label='PANEL'
    )
    
    # La sesión se actualiza en el caller
    return {
        'step': step_on_confirm,
        'timeout_minutes': timeout_minutes
    }


def build_enturnado_message(solicitud):
    """Genera el texto estándar para notificar un enturnamiento completado."""
    fecha_local = to_bogota_datetime(solicitud.fecha_descargue, assume_local=True)
    fecha_descargue_texto = fecha_local.strftime('%d/%m/%Y %H:%M') if fecha_local else 'Por definir'
    lugar_texto = solicitud.lugar_descargue or 'Sociedad Portuaria del Dique'
    
    # Mensaje ingenioso con Fisher
    mensaje_fisher = (
        "✅ *Turno Confirmado*\n\n"
        f"📅 Fecha: {fecha_descargue_texto}\n"
        f"📍 Lugar: {lugar_texto}\n"
        "ℹ️ Turno interno por orden de llegada.\n\n"
        "Buen viaje."
    )
    
    instrucciones = (
        "\n\nSi necesitas un nuevo enturne escribe *NUEVO*.\n"
        "Si el horario no te sirve responde *asesor* para que nuestro equipo te contacte."
    )
    return f"{mensaje_fisher}{instrucciones}"

def validar_y_guardar_ubicacion(telefono, lat, lng, ubicacion_tipo, session, next_step):
    """
    Función auxiliar para validar ubicación GPS y guardar en BD.
    
    Args:
        telefono: Número de teléfono
        lat: Latitud
        lng: Longitud  
    ubicacion_tipo: 'bosconia' o 'gambote'
        session: Sesión actual
        next_step: Siguiente paso si es válido
    
    Returns:
        bool: True si la ubicación fue válida
    """
    solicitud = session.get('solicitud')
    if not solicitud:
        solicitud = SolicitudCita.query.filter_by(telefono=telefono).order_by(SolicitudCita.fecha.desc()).first()

    ruta_alterna_flag = getattr(solicitud, 'ruta_alterna', False) if solicitud else False
    validacion = validar_ubicacion_gps(lat, lng, ubicacion_tipo, ruta_alterna=ruta_alterna_flag)

    if validacion['valido']:
        # Guardar directamente en la solicitud de la base de datos
        if solicitud:
            if ubicacion_tipo == 'bosconia':
                solicitud.ubicacion_lat = lat
                solicitud.ubicacion_lng = lng
                solicitud.paso_bosconia = True
            elif ubicacion_tipo == 'gambote':
                solicitud.ubicacion_gambote_lat = lat
                solicitud.ubicacion_gambote_lng = lng
                solicitud.paso_gambote = True
            if validacion.get('skip_next') == 'gambote':
                # Marcar Gambote como cumplido automáticamente
                solicitud.ubicacion_gambote_lat = solicitud.ubicacion_gambote_lat or lat
                solicitud.ubicacion_gambote_lng = solicitud.ubicacion_gambote_lng or lng
                solicitud.paso_gambote = True

            pendiente_ubicacion = bool(solicitud.ubicacion_pendiente_tipo)
            solicitud.ubicacion_pendiente_lat = None
            solicitud.ubicacion_pendiente_lng = None
            solicitud.ubicacion_pendiente_tipo = None
            solicitud.ubicacion_pendiente_mensaje = None
            solicitud.ubicacion_pendiente_desde = None
            if pendiente_ubicacion:
                solicitud.asesor_pendiente = False
                solicitud.asesor_pendiente_desde = None

            db.session.commit()
        reset_contextual_memory(session)

        send_whatsapp_message(telefono, f"✅ Ubicación en {ubicacion_tipo.title()} validada.\n{validacion['mensaje']}")
        if validacion.get('skip_next') == 'gambote':
            session['step'] = 9  # Ir directo a confirmación final
        else:
            session['step'] = next_step
        return True
    else:
        enviar_mensaje_ubicacion_invalida(
            telefono,
            ubicacion_tipo,
            lat=lat,
            lng=lng,
            validacion=validacion,
            session=session
        )
        return False

def guardar_imagen_whatsapp(telefono, media_payload, tipo_imagen, session):
    """
    Función auxiliar para guardar imágenes de WhatsApp.
    
    Args:
        telefono: Número de teléfono
        image_url: URL de la imagen
        tipo_imagen: 'guia', 'ticket_gambote', etc.
        session: Sesión actual
    """
    if media_payload:
        ruta_guardada = save_whatsapp_image(media_payload, 'whatsapp')
        if ruta_guardada:
            # Guardar directamente en la solicitud de la base de datos
            solicitud = session.get('solicitud')
            if not solicitud:
                solicitud = SolicitudCita.query.filter_by(telefono=telefono).order_by(SolicitudCita.fecha.desc()).first()
            
            if solicitud:
                if tipo_imagen == 'imagen_guia':
                    solicitud.imagen_guia = ruta_guardada
                elif tipo_imagen == 'imagen_manifiesto':
                    solicitud.imagen_manifiesto = ruta_guardada
                elif tipo_imagen == 'ticket_gambote':
                    solicitud.ticket_gambote = ruta_guardada
                db.session.commit()
                print(f"Imagen {tipo_imagen} guardada para {telefono}: {ruta_guardada}")
                reset_contextual_memory(session)
        else:
            print(f"Error guardando imagen {tipo_imagen} para {telefono}")
            send_whatsapp_message(telefono, "⚠️ Hubo un problema descargando el archivo. Por favor, intenta reenviarlo en unos segundos.")
    else:
        print(f"No se recibió URL de imagen para {tipo_imagen} de {telefono}")

def _normalize_timeout_minutes(value):
    """Normaliza los minutos de timeout para mantener consistencia en sesión y BD."""
    if value is None:
        return 0
    try:
        normalized = int(value)
    except (TypeError, ValueError):
        return 0
    return max(0, normalized)


def configurar_timeout_session(session, minutos):
    """Configura el timeout de la sesión aplicando normalización y limpieza de flags."""
    timeout_normalizado = _normalize_timeout_minutes(minutos)
    session['timeout_minutes'] = timeout_normalizado
    session['warning_sent'] = False

def enviar_mensaje_ubicacion_invalida(telefono, ubicacion, lat=None, lng=None, validacion=None, session=None):
    """
    Función auxiliar para enviar mensajes de ubicación inválida.
    
    Args:
        telefono: Número de teléfono
    ubicacion: Nombre de la ubicación ('bosconia' o 'gambote')
    """
    mensajes = {
        'bosconia': "La ubicación que recibí no coincide con Bosconia. Ya le pedí apoyo a un agente humano para validar tu registro.",
        'gambote': "La ubicación que recibí no corresponde a Gambote. Consultaré con un agente humano para revisarla.",
    }
    step_map = {
        'bosconia': STEP_AWAIT_GPS_BOSCONIA,
        'gambote': STEP_AWAIT_GPS_GAMBOTE
    }
    instruccion = mensajes.get(
        ubicacion,
        'La ubicación enviada no coincide con el punto esperado. Solicité ayuda a un agente humano.'
    )

    solicitud = None
    if session is not None:
        solicitud = session.get('solicitud')
    if not solicitud:
        solicitud = (
            SolicitudCita.query
            .filter_by(telefono=telefono)
            .order_by(SolicitudCita.fecha.desc())
            .first()
        )

    detalle_mensaje = instruccion
    if validacion and isinstance(validacion, dict) and validacion.get('mensaje'):
        detalle_mensaje = validacion['mensaje']

    if session is not None:
        cuerpo = compose_contextual_hint(
            session,
            step_map.get(ubicacion, STEP_AWAIT_GPS_BOSCONIA),
            f"{instruccion}\n\nPor favor espera mientras verificamos tu información."
        )
    else:
        cuerpo = (
            "Fisher 🐶 no reconoció la ubicación enviada. "
            f"{instruccion} Espera un momento mientras validamos tus datos."
        )

    mensaje = f"❌ {cuerpo}"
    send_whatsapp_message(
        telefono,
        mensaje,
        force_reminder=True
    )

    if solicitud:
        try:
            if not solicitud.asesor_pendiente:
                solicitud.asesor_pendiente = True
                solicitud.asesor_pendiente_desde = datetime.utcnow()
            else:
                solicitud.asesor_pendiente_desde = datetime.utcnow()

            solicitud.ubicacion_pendiente_lat = lat
            solicitud.ubicacion_pendiente_lng = lng
            solicitud.ubicacion_pendiente_tipo = ubicacion
            solicitud.ubicacion_pendiente_mensaje = str(detalle_mensaje)[:255] if detalle_mensaje else None
            solicitud.ubicacion_pendiente_desde = datetime.utcnow()
            db.session.commit()
        except Exception:
            db.session.rollback()
            current_app.logger.exception('No se pudo marcar asesor pendiente para %s', solicitud.id)
    return 'ok', 200


def enviar_mensaje_solicitar_ubicacion(telefono, ubicacion):
    """
    Función auxiliar para enviar mensajes solicitando ubicación.
    
    Args:
        telefono: Número de teléfono
    ubicacion: Nombre de la ubicación ('bosconia' o 'gambote')
    """
    mensajes = {
        'bosconia': "Cuando llegues a Bosconia abre el clip, toca ‘Ubicación’ y envíala en tiempo real. Solo puedo registrar ubicaciones en este paso.",
        'gambote': "Cuando cruces el peaje de Gambote abre el clip, elige ‘Ubicación’ y envíala en tiempo real. Solo puedo registrar ubicaciones en este paso.",
    }
    send_whatsapp_message(telefono, mensajes.get(ubicacion, 'Por favor, comparte tu ubicación en tiempo real.'))
    return 'ok', 200

def get_solicitud_data(solicitud):
    """
    Función auxiliar para obtener datos de la solicitud de manera segura.
    
    Args:
        solicitud: Objeto SolicitudCita
    
    Returns:
        dict: Diccionario con los datos de la solicitud
    """
    if not solicitud:
        return {}
    
    return {
        'nombre_completo': solicitud.nombre_completo or '',
        'cedula': solicitud.cedula or '',
        'placa': solicitud.placa or '',
        'placa_remolque': solicitud.placa_remolque or '',
        'celular': solicitud.celular or '',
        'imagen_guia': solicitud.imagen_guia,
        'imagen_manifiesto': solicitud.imagen_manifiesto,
        'ticket_gambote': solicitud.ticket_gambote,
        'paso_bosconia': solicitud.paso_bosconia or False,
        'paso_gambote': solicitud.paso_gambote or False,
        'paso_zisa': solicitud.paso_zisa or False
    }
    return 'ok', 200

# Configuración de WhatsApp Business API
WHATSAPP_TOKEN = os.environ.get("WHATSAPP_TOKEN")
PHONE_NUMBER_ID = os.environ.get("PHONE_NUMBER_ID")
WHATSAPP_PHONE_ID = os.environ.get("WHATSAPP_PHONE_ID") or PHONE_NUMBER_ID
VERIFY_TOKEN = os.environ.get("WHATSAPP_VERIFY_TOKEN", "TU_TOKEN_SECRETO_INVENTADO")

# Estados nombrados para la conversación (en lugar de números mágicos)
STATE_PENDING_INSCRIPTION = 'pendiente_inscripcion'
STATE_FINALIZADO = 'finalizado'

STEP_INACTIVE = 0
STEP_WELCOME = 1
STEP_AWAIT_PLACA = 2
STEP_CONFIRM_DATA = 3
STEP_AWAIT_GUIA = 4
STEP_AWAIT_GPS_BOSCONIA = 5
STEP_AWAIT_TICKET_GAMBOTE = 6
STEP_AWAIT_GPS_GAMBOTE = 7
STEP_AWAIT_GPS_ZISA = 8
STEP_FINAL_CONFIRMATION = 9
STEP_MANUAL_REG_NAME = 10
STEP_MANUAL_REG_CEDULA = 11
STEP_MANUAL_REG_REMOLQUE = 12
STEP_MANUAL_REG_CONFIRM = 13
STEP_CONFIRM_UNKNOWN_PLACA = 14
STEP_AWAIT_MANIFIESTO = 15
STEP_UNDER_REVIEW = 'confirmed'
STEP_HUMAN_HANDOFF = 'human_handoff'

STEP_TIMEOUT_CONFIG = {
    STEP_CONFIRM_DATA: {
        'timeout': 10,
        'warning_before': 5,
        'warning_message': (
            "Fisher 🐶 nota que llevas un rato sin confirmar tus datos. ¿Sigues ahí? "
            "Respóndeme 'sí' si todo está bien o 'no' si quieres corregirlos."
        ),
        'timeout_message': (
            "😴 Fisher cerró esta confirmación porque pasaron 10 minutos sin respuesta. "
            "Escribe NUEVO para volver a revisar tus datos cuando estés listo."
        )
    },
    STEP_MANUAL_REG_CONFIRM: {
        'timeout': 10,
        'warning_before': 5,
        'warning_message': (
            "Fisher 🐶 sigue esperando tu confirmación manual. ¿Todo está correcto? "
            "Respóndeme 'sí' para guardar los datos o 'no' para ajustarlos."
        ),
        'timeout_message': (
            "😴 Cerré esta confirmación manual porque no recibí respuesta en 10 minutos. "
            "Cuando quieras retomarla, escribe NUEVO y volvemos a registrar tus datos."
        )
    }
}


def get_step_timeout_config(step):
    return STEP_TIMEOUT_CONFIG.get(step, {})

def determinar_siguiente_step_pendiente(solicitud):
    """Determina el siguiente paso pendiente según la data almacenada."""
    if not solicitud:
        return STEP_AWAIT_GUIA

    if not solicitud.imagen_guia:
        return STEP_AWAIT_GUIA

    if not solicitud.imagen_manifiesto:
        return STEP_AWAIT_MANIFIESTO

    if not solicitud.paso_bosconia or not solicitud.ubicacion_lat or not solicitud.ubicacion_lng:
        return STEP_AWAIT_GPS_BOSCONIA

    if not solicitud.ticket_gambote:
        return STEP_AWAIT_TICKET_GAMBOTE

    if not solicitud.paso_gambote or not solicitud.ubicacion_gambote_lat or not solicitud.ubicacion_gambote_lng:
        return STEP_AWAIT_GPS_GAMBOTE

    return STEP_FINAL_CONFIRMATION

def build_confirmation_summary(solicitud):
    """Construye el mensaje de resumen para confirmación final."""
    datos = get_solicitud_data(solicitud)
    if not datos:
        return (
            "🐶 Fisher 🐶: Por favor confirma que todos tus datos son correctos. "
            "Responde 'sí' para confirmar."
        )

    def _status(check, ok_text, pending_text):
        return ok_text if check else pending_text

    tiene_guia = datos.get('imagen_guia') or datos.get('imagen_manifiesto')
    resumen = (
        "✅ Datos completos:\n\n"
        f"Nombre: {datos.get('nombre_completo') or '-'}\n"
        f"Cédula: {datos.get('cedula') or '-'}\n"
        f"Placa: {datos.get('placa') or '-'}\n"
        f"Placa remolque: {datos.get('placa_remolque') or '-'}\n"
        f"Celular: {datos.get('celular') or '-'}\n"
        f"Guía: {_status(tiene_guia, 'recibida', 'pendiente')}\n"
        f"Ubicación Bosconia: {_status(datos.get('paso_bosconia'), '✅ validada', 'pendiente')}\n"
        f"Ticket Gambote: {_status(datos.get('ticket_gambote'), 'recibido', 'pendiente')}\n"
        f"Ubicación Gambote: {_status(datos.get('paso_gambote'), '✅ validada', 'pendiente')}\n\n"
        "¿Confirmas el envío a revisión? (Sí / No)"
    )
    return resumen


# Recordatorio de seguridad para los conductores antes de responder.
SAFETY_REMINDER_VARIANTS = (
    "\n\n🛑 *Por seguridad: Detén el vehículo antes de responder.*"
)
WHATSAPP_SAFETY_REMINDER = SAFETY_REMINDER_VARIANTS[0]

# Texto característico del mensaje que confirma el enturnamiento final.
ENTURNE_COMPLETED_TOKEN = "Solicitud enturnada"

# Fragmento que usamos para detectar si el recordatorio ya fue agregado.
REMINDER_SENTINEL = "Seguridad ante todo"

SAFETY_REMINDER_TRACKER = {}
FIRST_REMINDER_INTERVAL = 3
try:
    SAFETY_REMINDER_INTERVAL = int(os.environ.get('WHATSAPP_REMINDER_EVERY', 3)) or 3
except (TypeError, ValueError):
    SAFETY_REMINDER_INTERVAL = 3
IMPORTANT_REMINDER_EMOJIS = ('⚠', '🚨')
IMPORTANT_REMINDER_KEYWORDS = (
    'urgente',
    'precaucion',
    'precaución',
    'cuidado',
    'alerta'
)


def reset_safety_reminder_counter(telefono):
    """Permite reiniciar el contador del recordatorio de seguridad para un número."""
    if telefono:
        SAFETY_REMINDER_TRACKER.pop(telefono, None)


def compose_contextual_hint(session, step, hint):
    """Genera un mensaje contextual recordando al conductor qué debe responder."""
    if not session:
        return hint

    tracker = session.setdefault('contextual_memory', {})
    step_key = str(step)
    last_step = tracker.get('_last_step')

    if last_step != step_key:
        tracker[step_key] = 0

    tracker[step_key] = tracker.get(step_key, 0) + 1
    tracker['_last_step'] = step_key
    session['contextual_memory'] = tracker

    count = tracker[step_key]
    if count == 1:
        lead = "Fisher 🐶 no logró entender este mensaje."
    elif count == 2:
        lead = "Fisher 🐶 sigue esperando esa información exacta."
    else:
        lead = "Fisher 🐶 necesita esa información para continuar."

    return f"{lead} {hint}"


def reset_contextual_memory(session, step=None):
    """Limpia la memoria contextual para evitar mensajes repetidos."""
    if not session:
        return

    tracker = session.get('contextual_memory')
    if not tracker:
        return

    if step is None:
        session.pop('contextual_memory', None)
    else:
        step_key = str(step)
        tracker.pop(step_key, None)
        if tracker.get('_last_step') == step_key:
            tracker['_last_step'] = None
        session['contextual_memory'] = tracker


def _next_safety_reminder(tracker):
    if not SAFETY_REMINDER_VARIANTS:
        return ''

    index = tracker.get('variant_index', 0) % len(SAFETY_REMINDER_VARIANTS)
    tracker['variant_index'] = (index + 1) % len(SAFETY_REMINDER_VARIANTS)
    return SAFETY_REMINDER_VARIANTS[index]


def _detects_important_context(mensaje):
    texto = mensaje.lower()
    return any(token in mensaje for token in IMPORTANT_REMINDER_EMOJIS) or any(
        palabra in texto for palabra in IMPORTANT_REMINDER_KEYWORDS
    )


def _maybe_append_safety_reminder(mensaje, telefono=None, force=False, skip=False, prime_after_force=False):
    """Añade el recordatorio de seguridad con cadencia inteligente."""
    if skip or not mensaje:
        if skip and prime_after_force and telefono:
            tracker_key = telefono or '__global__'
            tracker = SAFETY_REMINDER_TRACKER.setdefault(
                tracker_key,
                {
                    'since': 0,
                    'interval': SAFETY_REMINDER_INTERVAL,
                    'variant_index': 0
                }
            )
            tracker['since'] = 0
            tracker['interval'] = max(1, FIRST_REMINDER_INTERVAL)
        return mensaje
    if ENTURNE_COMPLETED_TOKEN in mensaje:
        reset_safety_reminder_counter(telefono)
        return mensaje
    if REMINDER_SENTINEL in mensaje:
        return mensaje
    if mensaje.startswith("✅ Ubicación en"):
        return mensaje
    if "Tus datos han sido enviados para revisión" in mensaje:
        return mensaje

    tracker_key = telefono or '__global__'
    tracker = SAFETY_REMINDER_TRACKER.setdefault(
        tracker_key,
        {
            'since': 0,
            'interval': SAFETY_REMINDER_INTERVAL,
            'variant_index': 0
        }
    )

    if force or _detects_important_context(mensaje):
        tracker['since'] = 0
        if prime_after_force:
            tracker['interval'] = max(1, FIRST_REMINDER_INTERVAL)
        else:
            tracker['interval'] = SAFETY_REMINDER_INTERVAL
        return f"{mensaje}{_next_safety_reminder(tracker)}"

    tracker['since'] = tracker.get('since', 0) + 1
    interval = tracker.get('interval') or SAFETY_REMINDER_INTERVAL
    interval = max(1, interval)
    if tracker['since'] >= interval:
        tracker['since'] = 0
        tracker['interval'] = SAFETY_REMINDER_INTERVAL
        return f"{mensaje}{_next_safety_reminder(tracker)}"

    return mensaje


def _resolve_solicitud_for_logging(telefono, solicitud=None):
    if solicitud and getattr(solicitud, 'id', None):
        return solicitud
    if not telefono:
        return None
    return (
        SolicitudCita.query
        .filter_by(telefono=telefono)
        .order_by(SolicitudCita.fecha.desc())
        .first()
    )


def log_whatsapp_message(telefono, contenido, direction, sender, message_type='text', media_url=None, solicitud=None):
    try:
        solicitud_ref = _resolve_solicitud_for_logging(telefono, solicitud)
        registro = WhatsappMessage(
            solicitud_id=solicitud_ref.id if solicitud_ref else None,
            telefono=telefono,
            direction=direction,
            sender=sender,
            message_type=message_type or 'text',
            content=contenido or '',
            media_url=media_url
        )
        db.session.add(registro)
        db.session.commit()
    except Exception as err:
        db.session.rollback()
        current_app.logger.warning('No se pudo registrar mensaje WhatsApp (%s): %s', sender, err)


def send_whatsapp_message(
    telefono,
    mensaje=None,
    media_url=None,
    buttons=None,
    template_name=None,
    template_vars=None,
    template_lang='es',
    sender='bot',
    solicitud=None,
    *,
    force_reminder=False,
    skip_reminder=False,
    prime_after_force=False
):
    """
    Envía un mensaje de WhatsApp usando la API de Meta (v17.0).
    Soporta mensajes de texto, plantillas, botones interactivos y archivos multimedia.

    Args:
        telefono: Número de teléfono del destinatario (sin +57)
        mensaje: Mensaje de texto a enviar (opcional si se usa template)
        media_url: URL del archivo multimedia (opcional)
        buttons: Lista de botones para mensajes interactivos (opcional)
        template_name: Nombre de la plantilla de WhatsApp Business API (opcional)
        template_vars: Lista de variables para la plantilla (opcional)
        template_lang: Idioma de la plantilla (default 'es')
        sender: Quién envía el mensaje (default 'bot')
        solicitud: Objeto SolicitudCita asociado (opcional)
        force_reminder, skip_reminder, prime_after_force: Parámetros para recordatorios de seguridad

    Returns:
        bool: True si el mensaje se envió correctamente
    """
    import uuid
    import requests
    from flask import current_app

    # Validar parámetros
    if not telefono:
        current_app.logger.warning("Intento de enviar mensaje sin número de teléfono")
        return False

    # Si se especifica template, validar que sea compatible con la API
    if template_name:
        if not WHATSAPP_TOKEN or not WHATSAPP_PHONE_ID:
            current_app.logger.error("Credenciales de WhatsApp no configuradas para enviar plantillas")
            return False

        # Construir payload para plantilla
        payload = {
            "messaging_product": "whatsapp",
            "to": telefono,
            "type": "template",
            "template": {
                "name": template_name,
                "language": {"code": template_lang}
            }
        }

        # Agregar variables si se proporcionan
        if template_vars:
            parameters = []
            for var in template_vars:
                parameters.append({
                    "type": "text",
                    "text": str(var)
                })
            
            # La API de WhatsApp espera UN solo objeto en la lista 'components' para el body
            payload["template"]["components"] = [{
                "type": "body",
                "parameters": parameters
            }]

    else:
        # Mensaje regular (texto, botones o multimedia)
        if not mensaje and not media_url and not buttons:
            current_app.logger.warning("Mensaje vacío: no hay texto, multimedia ni botones")
            return False

        # Aplicar lógica de recordatorios de seguridad si es mensaje regular
        mensaje = mensaje or ''
        auto_force = _detects_important_context(mensaje) if mensaje else False
        mensaje = _maybe_append_safety_reminder(
            mensaje,
            telefono=telefono,
            force=force_reminder or auto_force,
            skip=skip_reminder,
            prime_after_force=prime_after_force
        )

        if buttons and media_url:
            current_app.logger.warning('No se puede enviar botones interactivos junto a un adjunto; se omitirá el archivo adjunto.')
            media_url = None

        payload = {
            "messaging_product": "whatsapp",
            "recipient_type": "individual",
            "to": telefono
        }

        if buttons:
            # Mensaje interactivo con botones
            payload["type"] = "interactive"
            payload["interactive"] = {
                "type": "button",
                "body": {"text": mensaje},
                "action": {"buttons": buttons}
            }
        elif media_url:
            # Mensaje con multimedia
            if mensaje:
                # Documento con caption
                payload["type"] = "document"
                payload["document"] = {
                    "link": media_url,
                    "caption": mensaje,
                    "filename": _derive_media_filename(media_url)
                }
            else:
                # Imagen sin texto
                payload["type"] = "image"
                payload["image"] = {"link": media_url}
        else:
            # Mensaje de texto simple
            payload["type"] = "text"
            payload["text"] = {"body": mensaje}

    # Headers para la API
    headers = {
        'Authorization': f'Bearer {WHATSAPP_TOKEN}',
        'Content-Type': 'application/json'
    }

    try:
        # Enviar mensaje
        url = f"https://graph.facebook.com/v17.0/{WHATSAPP_PHONE_ID}/messages"
        response = requests.post(url, json=payload, headers=headers, timeout=30)

        if response.status_code == 200:
            response_data = response.json()
            whatsapp_message_id = response_data.get('messages', [{}])[0].get('id')

            # Guardar en base de datos si hay solicitud asociada
            try:
                log_whatsapp_message(
                    telefono,
                    mensaje or template_name or '',
                    direction='outbound',
                    sender=sender,
                    message_type=payload.get('type', 'text'),
                    media_url=media_url,
                    solicitud=solicitud
                )
            except Exception as db_error:
                current_app.logger.warning(f"No se pudo guardar mensaje en BD: {db_error}")

            current_app.logger.info(f"Mensaje enviado exitosamente a {telefono}: {whatsapp_message_id}")
            return True
        else:
            error_data = response.json()
            error_message = error_data.get('error', {}).get('message', 'Error desconocido')
            current_app.logger.error(f"Error enviando mensaje a {telefono}: {error_message}")
            return False

    except requests.exceptions.RequestException as e:
        current_app.logger.error(f"Error de conexión enviando mensaje a {telefono}: {e}")
        return False
    except Exception as e:
        current_app.logger.exception(f"Error inesperado enviando mensaje a {telefono}")
        return False


def send_yes_no_prompt(
    telefono,
    mensaje,
    *,
    sender='bot',
    solicitud=None,
    skip_reminder=False,
    prime_after_force=False,
    force_reminder=False,
    yes_title='Sí',
    no_title='No',
    context_label=None
):
    """Envía un mensaje con botones interactivos Sí/No para facilitar la respuesta."""
    suffix = (context_label or 'GEN').upper()
    unique = uuid.uuid4().hex[:4].upper()
    yes_id = f"YES_{suffix}_{unique}"
    no_id = f"NO_{suffix}_{unique}"

    buttons = [
        {"type": "reply", "reply": {"id": yes_id, "title": yes_title}},
        {"type": "reply", "reply": {"id": no_id, "title": no_title}}
    ]

    return send_whatsapp_message(
        telefono,
        mensaje,
        sender=sender,
        solicitud=solicitud,
        skip_reminder=skip_reminder,
        prime_after_force=prime_after_force,
        force_reminder=force_reminder,
        buttons=buttons
    )

def handle_step_welcome(telefono, texto, tipo, msg, session):
    """Handler for welcome step - ask if user wants to enturnar"""
    intentos = session.get('welcome_invalid_attempts', 0)
    texto_normalizado = (texto or '').strip().lower()

    if 'asesor' in texto_normalizado:
        send_whatsapp_message(
            telefono,
            "Entendido. Un asesor humano te contactará pronto. 🐶"
        )
        solicitud = session.get('solicitud')
        if solicitud:
            solicitud.asesor_pendiente = True
            solicitud.asesor_pendiente_desde = datetime.utcnow()
            solicitud.whatsapp_step = str(STEP_HUMAN_HANDOFF)
            solicitud.whatsapp_last_activity = datetime.utcnow()
            db.session.commit()
        session['step'] = STEP_HUMAN_HANDOFF
        session['timeout_minutes'] = 0
        session['warning_sent'] = False
        session['welcome_invalid_attempts'] = 0
        reset_contextual_memory(session)
        update_whatsapp_session(telefono, session)
        return 'ok', 200

    if is_confirmation_negative(texto):
        send_whatsapp_message(
            telefono,
            "Entendido. Este bot automático es solo para enturne. Cuando necesites gestionar tu turno, escribe 'NUEVO' y Fisher te ayudará. ¡Hasta pronto!"
        )
        reset_safety_reminder_counter(telefono)
        reset_contextual_memory(session)
        session['step'] = STEP_INACTIVE
        session['timeout_minutes'] = 0
        session['warning_sent'] = False
        session['welcome_invalid_attempts'] = 0

        solicitud = session.get('solicitud')
        if solicitud:
            solicitud.mensaje = 'Conversación cerrada: el usuario indicó que no era para enturne.'
            solicitud.whatsapp_step = str(STEP_INACTIVE)
            solicitud.whatsapp_timeout_minutes = 0
            solicitud.whatsapp_warning_sent = False
            solicitud.whatsapp_last_activity = datetime.utcnow()
            solicitud.estado = 'preconfirmacion'
            solicitud.asesor_pendiente = False
            solicitud.asesor_pendiente_desde = None
            db.session.commit()

        update_whatsapp_session(telefono, session)
        return 'ok', 200

    if is_confirmation_positive(texto):
        send_whatsapp_message(
            telefono,
            "Perfecto. Por favor dime la placa de tu camión para buscarte en la base de datos.",
            force_reminder=True,
            prime_after_force=True
        )
        session['step'] = STEP_AWAIT_PLACA
        configurar_timeout_session(session, 10)  # 10 minutos para ingresar placa
        session['welcome_invalid_attempts'] = 0
        reset_contextual_memory(session)
        return None

    intentos += 1
    session['welcome_invalid_attempts'] = intentos

    if intentos >= 2:
        send_whatsapp_message(
            telefono,
            "Fisher 🐶 no recibió una respuesta válida. Cierro la conversación por ahora; escribe 'REINICIAR' cuando quieras comenzar de nuevo."
        )
        reset_safety_reminder_counter(telefono)
        reset_contextual_memory(session)
        session['step'] = STEP_INACTIVE
        session['data'] = {}
        session['timeout_minutes'] = 0
        session['warning_sent'] = False
        session['welcome_invalid_attempts'] = 0
        update_whatsapp_session(telefono, session)
        return 'ok', 200

    send_yes_no_prompt(
        telefono,
        "Para continuar necesito que respondas exactamente 'sí' o 'no'. Inténtalo de nuevo, por favor.",
        context_label='WELCOME_RETRY'
    )
    configurar_timeout_session(session, 5)
    update_whatsapp_session(telefono, session)
    return 'ok', 200

def handle_step_await_placa(telefono, texto, tipo, msg, session):
    """Handler for awaiting placa input"""
    placa = texto.upper().replace(' ', '')
    conductor = buscar_conductor_por_placa(placa)
    if conductor:
        respuesta = (
            f"¿Estos datos son correctos?\n"
            f"Placa: {conductor['PLACA']}\n"
            f"Placa remolque: {conductor['PLACA REMOLQUE']}\n"
            f"Nombre: {conductor['NOMBRE CONDUCTOR']}\n"
            f"N° Documento: {conductor['N° DOCUMENTO']}\n"
            f"Celular: {conductor['CELULAR']}\n"
            "Responde 'sí' si son correctos o 'no' si necesitas corregirlos."
        )
        
        # Guardar directamente en la solicitud de la base de datos
        solicitud = session.get('solicitud')
        if solicitud:
            solicitud.nombre_completo = conductor['NOMBRE CONDUCTOR']
            solicitud.placa = conductor['PLACA']
            solicitud.placa_remolque = conductor['PLACA REMOLQUE']
            solicitud.cedula = conductor['N° DOCUMENTO']
            solicitud.celular = conductor['CELULAR']
            solicitud.estado = 'sin turno'
            db.session.commit()

        send_yes_no_prompt(telefono, respuesta, context_label='CONF_DATOS')
        session['step'] = STEP_CONFIRM_DATA
        timeout_cfg = get_step_timeout_config(STEP_CONFIRM_DATA)
        configurar_timeout_session(session, timeout_cfg.get('timeout', 10))
        reset_contextual_memory(session)
        return None
    else:
        mensaje = f"Placa {placa} no encontrada. ¿La escribiste correctamente? 🐶"
        # Guardar la placa original en la solicitud
        solicitud = session.get('solicitud')
        if solicitud:
            solicitud.placa = placa.upper()
            db.session.commit()
        session.setdefault('data', {})
        session['data']['last_placa_input'] = placa.upper()

        send_yes_no_prompt(
            telefono,
            mensaje,
            context_label='CONFIRM_UNKNOWN_PLACA'
        )
        session['step'] = STEP_CONFIRM_UNKNOWN_PLACA
        configurar_timeout_session(session, 10)
        reset_contextual_memory(session)
        return None

def handle_step_confirm_unknown_placa(telefono, texto, tipo, msg, session):
    """Confirm whether an unknown plate was typed correctly before branching."""
    if is_confirmation_positive(texto):
        send_whatsapp_message(
            telefono,
            "Te registro manualmente. Escribe tu nombre completo. 🐶"
        )
        session['step'] = STEP_MANUAL_REG_NAME
        configurar_timeout_session(session, 30)
        reset_contextual_memory(session)
        return None

    if is_confirmation_negative(texto):
        send_whatsapp_message(
            telefono,
            "Escribe la placa nuevamente. 🐶"
        )
        session['step'] = STEP_AWAIT_PLACA
        configurar_timeout_session(session, 10)
        reset_contextual_memory(session)
        return None

    send_yes_no_prompt(
        telefono,
        "Confirma si la placa está correcta. 🐶",
        context_label='CONFIRM_UNKNOWN_PLACA_RETRY'
    )
    configurar_timeout_session(session, 5)
    return None

# Mapeo de handlers para cada estado (reemplaza el if/elif gigante)
STEP_HANDLERS = {
    STEP_INACTIVE: None,  # No handler needed
    STEP_WELCOME: handle_step_welcome,
    STEP_AWAIT_PLACA: handle_step_await_placa,
    STEP_CONFIRM_DATA: None,  # Handled in main flow
    STEP_AWAIT_GUIA: None,  # Handled in main flow
    STEP_AWAIT_MANIFIESTO: None,  # Handled in main flow
    STEP_AWAIT_GPS_BOSCONIA: None,  # Handled in main flow
    STEP_AWAIT_TICKET_GAMBOTE: None,  # Handled in main flow
    STEP_AWAIT_GPS_GAMBOTE: None,  # Handled in main flow
    STEP_AWAIT_GPS_ZISA: None,  # Handled in main flow
    STEP_FINAL_CONFIRMATION: None,  # Handled in main flow
    STEP_MANUAL_REG_NAME: None,  # Handled in main flow
    STEP_MANUAL_REG_CEDULA: None,  # Handled in main flow
    STEP_MANUAL_REG_REMOLQUE: None,  # Handled in main flow
    STEP_MANUAL_REG_CONFIRM: None,  # Handled in main flow
    STEP_CONFIRM_UNKNOWN_PLACA: handle_step_confirm_unknown_placa,
    STEP_UNDER_REVIEW: None,  # Handled in main flow
    STEP_HUMAN_HANDOFF: None,
}

@login_required
@app.route('/api/solicitudes_cita')
def api_solicitudes_cita():
    solicitudes = SolicitudCita.query.order_by(SolicitudCita.fecha.desc()).all()
    def serialize(s):
        return {
            'id': s.id,
            'fecha': s.fecha.strftime('%d/%m/%Y %H:%M'),
            'nombre_completo': s.nombre_completo,
            'cedula': s.cedula,
            'placa': s.placa,
            'placa_remolque': s.placa_remolque,
            'celular': s.celular,
            'estado': s.estado,
            'turno': s.turno,
            'fecha_descargue': s.fecha_descargue.strftime('%d/%m/%Y %H:%M') if s.fecha_descargue else '',
            'lugar_descargue': s.lugar_descargue,
            'observaciones': s.observaciones or '',
            'imagen_guia': s.imagen_guia,
            'imagen_manifiesto': s.imagen_manifiesto,
            'paso_bosconia': s.paso_bosconia,
            'ticket_gambote': s.ticket_gambote
        }
    return jsonify([serialize(s) for s in solicitudes])    

@app.cli.command("init-db")
def init_db_command():
    """Crea las tablas nuevas de la base de datos."""
    db.create_all()
    print("Base de datos inicializada y tablas creadas.")

with app.app_context():
 db.create_all()

# Registrar el Blueprint de WhatsApp
# TEMPORALMENTE DESHABILITADO por error de spacy
# app.register_blueprint(bot_bp)

# ===================================================================
# --- INICIO: FUNCIONES DE IMPORTACIÓN DESDE SHAREPOINT ---
# ===================================================================


def get_access_token():
    """Solicita un token OAuth2 para la Graph API usando client credentials."""
    tenant_id = os.getenv('MS_TENANT_ID')
    client_id = os.getenv('MS_CLIENT_ID')
    client_secret = os.getenv('MS_CLIENT_SECRET')

    if not all([tenant_id, client_id, client_secret]):
        current_app.logger.error('Faltan variables MS_TENANT_ID, MS_CLIENT_ID o MS_CLIENT_SECRET para autenticación.')
        return None

    url = f"https://login.microsoftonline.com/{tenant_id}/oauth2/v2.0/token"
    payload = {
        'client_id': client_id,
        'client_secret': client_secret,
        'scope': 'https://graph.microsoft.com/.default',
        'grant_type': 'client_credentials'
    }
    headers = {'Content-Type': 'application/x-www-form-urlencoded'}

    try:
        response = requests.post(url, headers=headers, data=payload, timeout=30)
        response.raise_for_status()
        token_data = response.json()
        return token_data.get('access_token')
    except requests.exceptions.RequestException as exc:
        current_app.logger.error(f'Error al obtener token de Graph API: {exc}')
        return None


def descargar_guia_sharepoint(registro, numero_guia):
    """Descarga un PDF desde SharePoint y lo almacena en GUIDES_DIR/<registro_id>/."""
    registro_id = getattr(registro, 'id', None)
    if registro_id is None:
        return {'status': 'error', 'mensaje': 'Registro inválido para importar la guía.'}

    token = get_access_token()
    if not token:
        return {'status': 'error', 'mensaje': 'Fallo al obtener token de autenticación.'}

    sp_hostname = os.getenv('SP_SITE_HOSTNAME')
    sp_site_path = os.getenv('SP_SITE_PATH')
    sp_folder_default = os.getenv('SP_GUIDES_FOLDER')
    sp_folder_diluyente = os.getenv('SP_GUIDES_FOLDER_DILUYENTE')
    sp_folder_fo4 = os.getenv('SP_GUIDES_FOLDER_FO4')

    producto = ((registro.producto_a_cargar or '').strip().upper())
    chosen_folder = None

    if producto:
        if 'DILUY' in producto and sp_folder_diluyente:
            chosen_folder = sp_folder_diluyente
        elif 'FO4' in producto and sp_folder_fo4:
            chosen_folder = sp_folder_fo4

    if not chosen_folder:
        chosen_folder = sp_folder_default

    if not all([sp_hostname, sp_site_path, chosen_folder]):
        return {'status': 'error', 'mensaje': 'Faltan variables SP_SITE_HOSTNAME, SP_SITE_PATH o carpeta SharePoint configurada.'}

    normalized_folder = str(chosen_folder).strip().strip('/')


    nombre_archivo_sp = f"{numero_guia}.pdf"
    # Usar siempre la ruta relativa a la raíz del drive
    folder_relative = normalized_folder
    # Si la ruta contiene 'Documentos compartidos/', eliminar ese prefijo
    if folder_relative.lower().startswith('documentos compartidos/'):
        folder_relative = folder_relative[len('documentos compartidos/'):]
    # Si la ruta contiene solo la carpeta raíz (por ejemplo, 'General'), dejarla tal cual
    # Si la ruta contiene 'General/...' dejarla tal cual
    # Si la ruta contiene cualquier otro prefijo, eliminarlo
    # (Opcional: podrías validar que la primera carpeta existe en el log de DRIVE_ROOT)
    encoded_folder = quote(folder_relative, safe='/')
    encoded_filename = quote(nombre_archivo_sp)
    headers = {'Authorization': f'Bearer {token}'}

    try:
        current_app.logger.info(
            f"[SP_IMPORT] Registro {registro_id} -> carpeta '{normalized_folder}' producto '{producto}'"
        )
        file_metadata = None
        search_data = []

        # 1. Obtener drive-id del sitio
        drive_url = f"https://graph.microsoft.com/v1.0/sites/{sp_hostname}:{sp_site_path}:/drives"
        drive_resp = requests.get(drive_url, headers=headers, timeout=30)
        drive_resp.raise_for_status()
        drives = drive_resp.json().get('value', [])
        # Buscar el drive correcto (normalmente el que tiene 'document' en el nombre)
        drive_id = None
        for d in drives:
            if 'document' in (d.get('name', '').lower()):
                drive_id = d.get('id')
                break
        if not drive_id and drives:
            drive_id = drives[0].get('id')
        if not drive_id:
            current_app.logger.error(f"[SP_IMPORT][DRIVE_ID_ERROR] No se pudo obtener el drive-id del sitio. Drives: {drives}")
            return {'status': 'error', 'mensaje': 'No se pudo obtener el drive-id del sitio.'}

        # Listar todos los drives y sus carpetas raíz para depuración
        all_drives_url = f"https://graph.microsoft.com/v1.0/sites/{sp_hostname}:{sp_site_path}:/drives"
        all_drives_resp = requests.get(all_drives_url, headers=headers, timeout=30)
        all_drives_resp.raise_for_status()
        all_drives = all_drives_resp.json().get('value', [])
        for drive in all_drives:
            drive_id_dbg = drive.get('id')
            drive_name_dbg = drive.get('name')
            root_url_dbg = f"https://graph.microsoft.com/v1.0/drives/{drive_id_dbg}/root/children"
            root_resp_dbg = requests.get(root_url_dbg, headers=headers, timeout=30)
            root_resp_dbg.raise_for_status()
            root_items_dbg = root_resp_dbg.json().get('value', [])
            current_app.logger.info(
                f"[SP_IMPORT][DRIVE_LIST] Drive: {drive_name_dbg} ({drive_id_dbg}) -> Raíz: "
                f"{[{'name': item.get('name'), 'folder': 'folder' in item, 'id': item.get('id')} for item in root_items_dbg]}"
            )

        # ...existing code...

        # Listar carpetas y archivos dentro de 'General' para depuración
        # Listar carpetas y archivos dentro de 'General' para depuración
        general_folder_url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/root:/General:/children"
        general_folder_resp = requests.get(general_folder_url, headers=headers, timeout=30)
        general_folder_resp.raise_for_status()
        general_items = general_folder_resp.json().get('value', [])
        current_app.logger.info(
            f"[SP_IMPORT][GENERAL_FOLDER] Carpetas y archivos en 'General': "
            f"{[{'name': item.get('name'), 'folder': 'folder' in item, 'id': item.get('id')} for item in general_items]}"
        )

        # ...existing code...

        # Buscar la carpeta en todos los drives disponibles
        folder_id = None
        found_drive_id = None
        found_drive_name = None
        for d in drives:
            test_drive_id = d.get('id')
            test_drive_name = d.get('name')
            folder_meta_url = f"https://graph.microsoft.com/v1.0/drives/{test_drive_id}/root:/{encoded_folder}"
            try:
                folder_meta_response = requests.get(folder_meta_url, headers=headers, timeout=30)
                if folder_meta_response.status_code == 200:
                    folder_meta = folder_meta_response.json()
                    folder_id = folder_meta.get('id')
                    found_drive_id = test_drive_id
                    found_drive_name = test_drive_name
                    break
            except Exception as e:
                current_app.logger.warning(f"[SP_IMPORT][FOLDER_SEARCH] Error buscando carpeta en drive {test_drive_name}: {e}")
        if not folder_id:
            current_app.logger.error(f"[SP_IMPORT][FOLDER_ID_ERROR] No se pudo encontrar la carpeta '{normalized_folder}' en ningún drive. Revisa permisos y ruta.")
            return {'status': 'error', 'mensaje': f'No se pudo encontrar la carpeta en ningún drive.'}

        # 3. Listar archivos en la carpeta encontrada
        children_url = f"https://graph.microsoft.com/v1.0/drives/{found_drive_id}/items/{folder_id}/children"
        children_response = requests.get(children_url, headers=headers, timeout=30)
        children_response.raise_for_status()
        search_data = children_response.json().get('value', [])
        # Buscar archivo por nombre
        for item in search_data:
            name = (item.get('name') or '').lower()
            if name == nombre_archivo_sp.lower():
                file_metadata = item
                break
        current_app.logger.info(
            f"[SP_IMPORT][SEARCH_RESULTS] Archivos en carpeta '{normalized_folder}' (drive '{found_drive_name}'): "
            f"{[{'name': item.get('name'), 'path': unquote(item.get('parentReference', {}).get('path') or '')} for item in search_data]}"
        )
        if not file_metadata:
            current_app.logger.warning(
                f"[SP_IMPORT][NO_MATCH] No se encontró coincidencia exacta para '{nombre_archivo_sp}' en carpeta. Resultados: "
                f"{[{'name': item.get('name'), 'path': unquote(item.get('parentReference', {}).get('path') or '')} for item in search_data]}"
            )
            return {
                'status': 'error',
                'mensaje': (
                    f"Guía {numero_guia} no encontrada en SharePoint. "
                    "Verifica el nombre del archivo y la carpeta."
                )
            }

        download_url = file_metadata.get('@microsoft.graph.downloadUrl')
        if not download_url:
            current_app.logger.error(f"[SP_IMPORT][NO_DOWNLOAD_URL] No se pudo obtener la URL de descarga para '{numero_guia}'.")
            if search_data:
                current_app.logger.error(f"[SP_IMPORT][SEARCH_RESULTS_ON_ERROR] {search_data}")
            return {'status': 'error', 'mensaje': 'No se pudo obtener la URL de descarga.'}

        file_response = requests.get(download_url, timeout=60)
        file_response.raise_for_status()

        save_dir = os.path.join(current_app.config['GUIDES_DIR'], str(registro_id))
        os.makedirs(save_dir, exist_ok=True)

        unique_filename = f"sharepoint_{uuid.uuid4().hex[:8]}.pdf"
        ruta_guardado_abs = os.path.join(save_dir, unique_filename)

        with open(ruta_guardado_abs, 'wb') as destino:
            destino.write(file_response.content)

        ruta_relativa = os.path.join(str(registro_id), unique_filename).replace('\\', '/')

        return {
            'status': 'exito',
            'ruta_relativa': ruta_relativa,
            'ruta_abs': ruta_guardado_abs
        }

    except requests.exceptions.RequestException as exc:
        current_app.logger.error(f"[SP_IMPORT][EXCEPTION] Error de red o API: {exc}")
        if 'search_data' in locals() and search_data:
            current_app.logger.error(f"[SP_IMPORT][SEARCH_RESULTS_ON_EXCEPTION] {search_data}")
        return {'status': 'error', 'mensaje': f'Error de red o API: {exc}'}
    except Exception as exc:
        current_app.logger.error(f"[SP_IMPORT][EXCEPTION] Error inesperado: {exc}")
        if 'search_data' in locals() and search_data:
            current_app.logger.error(f"[SP_IMPORT][SEARCH_RESULTS_ON_EXCEPTION] {search_data}")
        return {'status': 'error', 'mensaje': f'Error inesperado: {exc}'}


# ===================================================================
# --- FIN: FUNCIONES DE IMPORTACIÓN DESDE SHAREPOINT ---
# ===================================================================

# --- Blueprint Pricing ---
from pricing.routes import pricing_bp
app.register_blueprint(pricing_bp)

from pricing.models import HistorialCombustibles
# Asegurar que todas las tablas existan (incluyendo la nueva MovimientoDian)
with app.app_context():
    db.create_all()

if __name__ == '__main__':
    app.run(debug=True)

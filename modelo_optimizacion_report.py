# -*- coding: utf-8 -*-
import io
from datetime import datetime
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
import pandas as pd

# Claves reales que devuelve modelo_optimizacion_core.ejecutar_modelo() para
# la tabla `ventas`. Estaban mal escritas aqui, asi que el informe mostraba
# precio e ingreso en 0 y el crudo de origen vacio.
COL_PRECIO_VENTA = 'Precio de venta USD/BBL'
COL_INGRESO_BASE = 'Costo Total'
COL_INGRESO_DENSIDAD = 'Ingreso Con Densidad Contractual (USD)'
COL_CRUDO_ORIGEN_VENTA = 'Crudo Origen(SoloRefinados)'


def _ingreso_venta(item):
    """Ingreso de una fila de ventas.

    En las ventas por formula de densidad el ingreso real es el ajustado por la
    densidad contractual; en las demas esa columna viene en 0 y vale el bruto
    (volumen x precio). Es el mismo criterio de la pagina, y asi el total del
    informe cuadra con la utilidad del solver.
    """
    return item.get(COL_INGRESO_DENSIDAD) or item.get(COL_INGRESO_BASE, 0.0)


def _sumar(filas, columna):
    total = 0.0
    for f in filas or []:
        try:
            total += float(f.get(columna) or 0)
        except (TypeError, ValueError):
            pass
    return total


def _costos_por_refineria(resultados):
    """Reparte los costos entre refinerias.

    Lo que se puede atribuir directamente:
      - refinacion : de la tabla throughput, que ya viene por planta
      - transporte : las rutas cuyo destino (entrada) u origen (salida) es la planta
      - crudo      : el que entra por esas rutas, valorado al precio de compra
    """
    refinerias = [str(t.get('Refineria')) for t in resultados.get('throughput', [])]
    if not refinerias:
        return []

    # Precio de compra por crudo, para valorar lo que entra a cada planta
    precio = {}
    for c in resultados.get('compras', []):
        base = str(c.get('Flujo Base Premium') or c.get('Crudo o Producto') or '')
        try:
            precio.setdefault(base, float(c.get('Precio Compra USD/BBL') or 0))
        except (TypeError, ValueError):
            pass

    filas = []
    for t in resultados.get('throughput', []):
        planta = str(t.get('Refineria'))
        entradas = [x for x in resultados.get('transporte', [])
                    if str(x.get('Destino')) == planta]
        salidas = [x for x in resultados.get('transporte', [])
                   if str(x.get('Origen')) == planta]

        vol_crudo = _sumar(entradas, 'Volumen Transportado BPD')
        costo_crudo = 0.0
        for x in entradas:
            try:
                v = float(x.get('Volumen Transportado BPD') or 0)
            except (TypeError, ValueError):
                v = 0.0
            costo_crudo += v * precio.get(str(x.get('Flujo')), 0.0)

        t_entrada = _sumar(entradas, 'Costo Total USD')
        t_salida = _sumar(salidas, 'Costo Total USD')
        c_refinacion = 0.0
        try:
            c_refinacion = float(t.get('Costo Total USD') or 0)
        except (TypeError, ValueError):
            pass

        try:
            throughput = float(t.get('Throughput Refineria BPD') or 0)
        except (TypeError, ValueError):
            throughput = 0.0

        total = costo_crudo + t_entrada + t_salida + c_refinacion
        filas.append({
            'planta': planta,
            'throughput': throughput,
            'vol_crudo': vol_crudo,
            'costo_crudo': costo_crudo,
            'transporte_entrada': t_entrada,
            'transporte_salida': t_salida,
            'refinacion': c_refinacion,
            'total': total,
            'usd_bbl': (total / throughput) if throughput > 0 else 0.0,
        })
    return filas


def _hoja_estructura_costo(wb, resultados, est):
    """Crea la hoja con la cascada de ingresos/egresos y el detalle de costos."""
    ws = wb.create_sheet(title="Estructura de Costo")
    F2 = "#,##0.00"
    M = '"$"#,##0.00'

    def titulo(fila, texto):
        c = ws.cell(row=fila, column=1, value=texto)
        c.font = est['font_title']
        ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=8)
        ws.row_dimensions[fila].height = 26

    def encabezados(fila, cols, ancho_desde=1):
        for i, h in enumerate(cols, ancho_desde):
            c = ws.cell(row=fila, column=i, value=h)
            c.font = est['font_header']
            c.fill = est['fill_header']
            c.alignment = est['align_center']
            c.border = est['thin_border']
        ws.row_dimensions[fila].height = 24

    # ── Datos de la cascada ───────────────────────────────────────────
    ventas = resultados.get('ventas', [])
    # La columna del efecto de densidad es la unica que menciona las DOS
    # variantes. Buscarla por 'Sin Densidad' a secas agarra
    # 'Ingreso TEORICO Sin Densidad Contractual' y duplica los ingresos.
    col_den = next((k for k in (ventas[0].keys() if ventas else [])
                    if 'Ingreso Con Densidad Contractual' in k
                    and 'Ingreso Sin Densidad Contractual' in k), None)
    ingresos = _sumar(ventas, 'Costo Total') + (_sumar(ventas, col_den) if col_den else 0.0)

    partidas = [
        ('Compras de crudo', _sumar(resultados.get('compras'), 'Costo Total USD')),
        ('Transporte', _sumar(resultados.get('transporte'), 'Costo Total USD')),
        ('Refinacion', _sumar(resultados.get('throughput'), 'Costo Total USD')),
        ('Operacion blending', _sumar(resultados.get('costos_operacionales'), 'Costo Total USD')),
        ('Operacion refinados', _sumar(resultados.get('costos_operacionales_ref'), 'Costo Total USD')),
    ]
    utilidad = ingresos - sum(p[1] for p in partidas)

    titulo(1, "ESTRUCTURA DE COSTO - De los ingresos a la utilidad")
    ws.cell(row=2, column=1,
            value="Cada barra es una partida. La barra flota entre el acumulado anterior y el nuevo.").font = est['font_data']

    # Tabla que alimenta el grafico. La columna 'Base' es el piso invisible
    # de cada barra flotante: es el truco para dibujar una cascada en Excel.
    encabezados(4, ["Concepto", "Base", "Ingreso", "Egreso", "Resultado", "Monto USD", "% Ingresos"])
    fila = 5
    acumulado = ingresos

    def escribir(concepto, base, ingreso, egreso, resultado, monto):
        nonlocal fila
        ws.cell(row=fila, column=1, value=concepto).font = est['font_data']
        for col, val in ((2, base), (3, ingreso), (4, egreso), (5, resultado)):
            c = ws.cell(row=fila, column=col, value=val)
            c.number_format = F2
            c.font = est['font_data']
        c = ws.cell(row=fila, column=6, value=monto)
        c.number_format = M
        c.font = est['font_data']
        c.alignment = est['align_right']
        p = ws.cell(row=fila, column=7,
                    value=(monto / ingresos) if ingresos else 0)
        p.number_format = "0.0%"
        p.alignment = est['align_right']
        p.font = est['font_data']
        for col in range(1, 8):
            ws.cell(row=fila, column=col).border = est['thin_border']
        fila += 1

    escribir("Ingresos por ventas", 0, ingresos, 0, 0, ingresos)
    for nombre, monto in partidas:
        nuevo = acumulado - monto
        escribir(nombre, nuevo, 0, monto, 0, -monto)
        acumulado = nuevo
    escribir("UTILIDAD NETA", 0, 0, 0, utilidad, utilidad)
    ultima = fila - 1
    for col in range(1, 8):
        c = ws.cell(row=ultima, column=col)
        c.font = est['font_total']
        c.fill = est['fill_total']
        c.border = est['total_border']

    # ── Grafico de cascada (barras apiladas con base invisible) ───────
    ch = BarChart()
    ch.type = "col"
    ch.grouping = "stacked"
    ch.overlap = 100
    ch.title = "Puente de ingresos a utilidad (USD/dia)"
    ch.y_axis.title = "USD/dia"
    ch.height = 9
    ch.width = 24
    datos = Reference(ws, min_col=2, max_col=5, min_row=4, max_row=ultima)
    cats = Reference(ws, min_col=1, min_row=5, max_row=ultima)
    ch.add_data(datos, titles_from_data=True)
    ch.set_categories(cats)

    # La serie 'Base' sostiene la barra pero no debe verse
    ch.series[0].graphicalProperties.noFill = True
    ch.series[0].graphicalProperties.line.noFill = True
    ch.series[1].graphicalProperties.solidFill = "10B981"   # ingreso
    ch.series[2].graphicalProperties.solidFill = "DC2626"   # egreso
    ch.series[3].graphicalProperties.solidFill = "1E3A5F"   # resultado
    ch.legend.position = 'b'
    ws.add_chart(ch, "I4")

    # ── Detalle por refineria ─────────────────────────────────────────
    fila = ultima + 22
    titulo(fila, "COSTOS POR REFINERIA")
    fila += 1
    ws.cell(row=fila, column=1,
            value="Cada planta con su crudo, su transporte y su costo de proceso, "
                  "para poder compararlas.").font = est['font_data']
    fila += 2

    porref = _costos_por_refineria(resultados)
    encabezados(fila, ["Refineria", "Throughput BPD", "Crudo cargado BPD",
                       "Costo crudo USD", "Transporte entrada USD",
                       "Transporte salida USD", "Refinacion USD",
                       "COSTO TOTAL USD", "USD por barril"])
    cab_ref = fila
    fila += 1
    for r in porref:
        vals = [r['planta'], r['throughput'], r['vol_crudo'], r['costo_crudo'],
                r['transporte_entrada'], r['transporte_salida'], r['refinacion'],
                r['total'], r['usd_bbl']]
        for i, v in enumerate(vals, 1):
            c = ws.cell(row=fila, column=i, value=v)
            c.font = est['font_data']
            c.border = est['thin_border']
            if i in (2, 3):
                c.number_format = F2
                c.alignment = est['align_right']
            elif i >= 4:
                c.number_format = M
                c.alignment = est['align_right']
        ws.row_dimensions[fila].height = 20
        fila += 1
    fin_ref = fila - 1

    if porref:
        # Totales
        ws.cell(row=fila, column=1, value="TOTAL").font = est['font_total']
        ws.cell(row=fila, column=1).fill = est['fill_total']
        ws.cell(row=fila, column=1).border = est['total_border']
        for col in range(2, 10):
            letra = get_column_letter(col)
            if col == 9:
                v = None
            else:
                v = f"=SUM({letra}{cab_ref + 1}:{letra}{fin_ref})"
            c = ws.cell(row=fila, column=col, value=v)
            c.font = est['font_total']
            c.fill = est['fill_total']
            c.border = est['total_border']
            c.number_format = F2 if col in (2, 3) else M
            c.alignment = est['align_right']
        fila += 2

        # Comparativa visual entre plantas
        ch2 = BarChart()
        ch2.type = "col"
        ch2.grouping = "clustered"
        ch2.title = "Costo por refineria (USD/dia)"
        ch2.y_axis.title = "USD/dia"
        ch2.height = 8
        ch2.width = 18
        d2 = Reference(ws, min_col=4, max_col=7, min_row=cab_ref, max_row=fin_ref)
        c2 = Reference(ws, min_col=1, min_row=cab_ref + 1, max_row=fin_ref)
        ch2.add_data(d2, titles_from_data=True)
        ch2.set_categories(c2)
        ch2.series[0].graphicalProperties.solidFill = "1E3A5F"
        ch2.series[1].graphicalProperties.solidFill = "0E7490"
        ch2.series[2].graphicalProperties.solidFill = "64748B"
        ch2.series[3].graphicalProperties.solidFill = "F59E0B"
        ch2.legend.position = 'b'
        ws.add_chart(ch2, f"K{cab_ref}")

    # ── Detalle de produccion por planta ──────────────────────────────
    titulo(fila, "PRODUCCION POR REFINERIA")
    fila += 2
    encabezados(fila, ["Refineria", "Producto Refinado", "Crudo Origen", "Volumen BPD"])
    fila += 1
    for x in sorted(resultados.get('refinacion', []),
                    key=lambda y: (str(y.get('Planta de Refinacion')),
                                   str(y.get('Producto Refinado')))):
        vals = [x.get('Planta de Refinacion'), x.get('Producto Refinado'),
                x.get('Crudo Origen'), x.get('Volumen Producido BPD')]
        for i, v in enumerate(vals, 1):
            c = ws.cell(row=fila, column=i, value=v)
            c.font = est['font_data']
            c.border = est['thin_border']
            if i == 4:
                c.number_format = F2
                c.alignment = est['align_right']
        fila += 1

    anchos = [30, 18, 20, 20, 22, 22, 18, 20, 16]
    for i, w in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    return ws


def generar_informe_excel(nombre_escenario: str, resultados: dict, excel_path_param: str = None) -> bytes:
    """
    Generates a professional corporate Excel report using openpyxl based on optimization results.
    
    Parameters:
        nombre_escenario (str): Name of the scenario (e.g. 'Escenario Base Mayo').
        resultados (dict): Dict of results returned by modelo_optimizacion_core.ejecutar_modelo.
        excel_path_param (str): Optional path to the parameters file to fetch Brent, TRM, etc.
    """
    # 1. Fetch economic parameters
    brent = 0.0
    trm = 0.0
    
    if excel_path_param:
        try:
            # Read from the uploaded Excel file to keep exact parameter numbers
            xls_vals = pd.ExcelFile(excel_path_param)
            # La hoja se llama '0.ECONOMICS' en el archivo real; buscarla por
            # 'ECONOMICOS' a secas hacia que el informe mostrara Brent y TRM en 0.
            hoja_econ = next(
                (h for h in xls_vals.sheet_names
                 if 'ECONOMIC' in str(h).upper().replace('Ó', 'O')), None)
            if hoja_econ:
                df_econ = pd.read_excel(xls_vals, sheet_name=hoja_econ)
                # Find Brent, TRM, etc.
                idx_col = next((c for c in df_econ.columns if str(c).strip().upper() in ("INDEX", "INDICE", "VARIABLE", "NOMBRE")), None)
                val_col = next((c for c in df_econ.columns if str(c).strip().upper() in ("VALUE", "VALOR", "VAL")), None)
                
                if idx_col and val_col:
                    m_brent = df_econ[idx_col].astype(str).str.upper().str.strip() == "BRENT"
                    if m_brent.any():
                        brent = float(df_econ.loc[m_brent, val_col].iloc[0])
                    
                    m_trm = df_econ[idx_col].astype(str).str.upper().str.strip() == "TRM"
                    if m_trm.any():
                        try:
                            # Parse string cop values if necessary
                            raw_trm = df_econ.loc[m_trm, val_col].iloc[0]
                            trm = float(str(raw_trm).replace('$','').replace('.','').replace(',','.').replace(' ','').strip())
                        except:
                            trm = 0.0
                            
                # Antes se leian B4 y B7 a ciegas como "% financiacion" y
                # "premium export". En 0.ECONOMICS esas celdas son DIESEL y una
                # celda vacia, por eso el informe mostraba 847.900 % de
                # financiacion. Si algun dia existen, hay que buscarlas por
                # nombre en la columna INDEX, nunca por posicion.
        except Exception as e:
            print(f"Error loading parameters for report: {e}")
            
    # Calculate margins
    utilidad = resultados.get('utilidad_total', 0.0)
    vol_ventas = resultados.get('vol_ventas_total', 0.0)
    vol_compras = resultados.get('vol_compras_total', 0.0)
    margen_neto = (utilidad / vol_ventas) if vol_ventas > 0 else 0.0
    status = resultados.get('status', 'OPTIMAL')
    
    # 2. Initialize Workbook
    wb = Workbook()
    
    # Setup styles
    font_family = "Segoe UI"
    
    font_title = Font(name=font_family, size=16, bold=True, color="1E3A5F")
    font_subtitle = Font(name=font_family, size=11, italic=True, color="555555")
    font_section = Font(name=font_family, size=12, bold=True, color="1E3A5F")
    font_header = Font(name=font_family, size=10, bold=True, color="FFFFFF")
    font_data = Font(name=font_family, size=10)
    font_total = Font(name=font_family, size=10, bold=True, color="10B981")
    font_kpi_label = Font(name=font_family, size=9, bold=True, color="475569")
    font_kpi_value = Font(name=font_family, size=10, bold=True, color="0F172A")
    
    fill_header = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    fill_kpi = PatternFill(start_color="F0F4F8", end_color="F0F4F8", fill_type="solid")
    fill_total = PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )
    
    total_border = Border(
        top=Side(style='thin', color='10B981'),
        bottom=Side(style='double', color='10B981')
    )
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    
    # ================= Sheet 1: Resumen Ejecutivo =================
    ws = wb.active
    ws.title = "Resumen Ejecutivo"
    ws.views.sheetView[0].showGridLines = True
    
    # Titles
    ws["A1"] = "INFORME EJECUTIVO DE OPTIMIZACIÓN"
    ws["A1"].font = font_title
    
    ws["A2"] = f"Escenario: {nombre_escenario}"
    ws["A2"].font = font_subtitle
    
    ws["A3"] = f"Fecha de reporte: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
    ws["A3"].font = Font(name=font_family, size=9, color="777777")
    
    # 3. Write Economic block in F1:G8
    # Costos abiertos por concepto, que es lo que se quiere ver de un vistazo
    costo_compras = _sumar(resultados.get('compras'), 'Costo Total USD')
    costo_transporte = _sumar(resultados.get('transporte'), 'Costo Total USD')
    costo_refinacion = _sumar(resultados.get('throughput'), 'Costo Total USD')
    costo_blending = (_sumar(resultados.get('costos_operacionales'), 'Costo Total USD')
                      + _sumar(resultados.get('costos_operacionales_ref'), 'Costo Total USD'))
    costo_total = costo_compras + costo_transporte + costo_refinacion + costo_blending

    kpis = [
        ("BRENT (USD/BBL)", brent, "$#,##0.00"),
        ("TRM (COP/USD)", trm, "$#,##0.00"),
        ("Utilidad Total (USD/Día)", utilidad, "$#,##0.00"),
        ("Margen Neto (USD/BBL)", margen_neto, "$#,##0.00"),
        ("Vol. Compras (BPD)", vol_compras, "#,##0.00"),
        ("Vol. Ventas (BPD)", vol_ventas, "#,##0.00"),
        ("Costo Compras (USD/Día)", -costo_compras, "$#,##0.00"),
        ("Costo Transporte (USD/Día)", -costo_transporte, "$#,##0.00"),
        ("Costo Refinación (USD/Día)", -costo_refinacion, "$#,##0.00"),
        ("Costo Blending (USD/Día)", -costo_blending, "$#,##0.00"),
        ("Costo Total (USD/Día)", -costo_total, "$#,##0.00"),
    ]
    
    for idx, (label, val, fmt) in enumerate(kpis):
        row = idx + 1
        ws.cell(row=row, column=6, value=label).font = font_kpi_label
        ws.cell(row=row, column=6).fill = fill_kpi
        ws.cell(row=row, column=6).alignment = align_left
        ws.cell(row=row, column=6).border = thin_border
        
        c_val = ws.cell(row=row, column=7, value=val)
        c_val.font = font_kpi_value
        c_val.fill = fill_kpi
        c_val.alignment = align_right
        c_val.number_format = fmt
        c_val.border = thin_border
        
    # La matriz arranca debajo del bloque de KPIs. Si se agregan KPIs hay que
    # correrla, o los encabezados se pisan con ellos.
    FILA_MATRIZ = len(kpis) + 2
    ws.cell(row=FILA_MATRIZ, column=1, value="MATRIZ DE OPERACIONES OPTIMIZADAS").font = font_section
    
    headers = [
        "Tipo", "Producto/Corriente", "Origen / Planta", "Destino / Pto Venta",
        "Volumen Optimizado (BPD)", "Precio/Tarifa Unit. (USD)", "Valor Bruto (USD/Día)",
        "Tarifa Transporte (USD/BBL)", "Costo Transporte (USD/Día)",
        "Valor Neto (USD/Día)"
    ]
    
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=FILA_MATRIZ + 1, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = thin_border
        ws.row_dimensions[FILA_MATRIZ + 1].height = 25
        
    # Data aggregation
    rows_data = []

    # Indice del transporte, para poder decir de donde a donde viaja cada cosa
    # y cuanto costo moverla. Antes estas dos columnas iban en cero fijo.
    transporte = resultados.get('transporte', [])
    trans_por_clave = {}
    trans_por_flujo_destino = {}
    for t in transporte:
        clave = (str(t.get('Flujo')), str(t.get('Origen')),
                 str(t.get('Destino')), str(t.get('Ruta')))
        trans_por_clave[clave] = t
        trans_por_flujo_destino.setdefault(
            (str(t.get('Flujo')), str(t.get('Destino'))), []).append(t)

    def _num(v):
        try:
            return float(v or 0)
        except (TypeError, ValueError):
            return 0.0

    # Gather Compras
    for item in resultados.get('compras', []):
        vol = _num(item.get('Volumen Comprado BPD'))
        price = _num(item.get('Precio Compra USD/BBL'))
        tot = _num(item.get('Costo Total USD'))

        # 'Variante Transporte' viene como "DESTINO | RUTA"
        destino, ruta = '', ''
        variante = str(item.get('Variante Transporte') or '')
        if '|' in variante:
            destino, ruta = [x.strip() for x in variante.split('|', 1)]

        t = trans_por_clave.get((str(item.get('Flujo Base Premium')),
                                 str(item.get('Origen Base Premium')),
                                 destino, ruta))
        t_rate = _num(t.get('Costo de transporte USD/BBL')) if t else 0.0
        t_tot = _num(t.get('Costo Total USD')) if t else 0.0

        rows_data.append({
            'type': 'Compra',
            'name': item.get('Crudo o Producto', ''),
            'src': item.get('Origen', ''),
            'dst': destino or 'Sin ruta',
            'vol': vol,
            'rate': price,
            # Una compra es salida de caja: va en negativo, para que la
            # columna sume hacia la utilidad en vez de inflar el total.
            'gross': -abs(tot),
            'trans_rate': t_rate,
            'trans_tot': t_tot,
            'net': -abs(tot) - t_tot,
        })

    # Gather Ventas
    for item in resultados.get('ventas', []):
        vol = _num(item.get('Volumen Vendido BPD'))
        price = _num(item.get(COL_PRECIO_VENTA))
        tot = _num(_ingreso_venta(item))

        rutas = trans_por_flujo_destino.get(
            (str(item.get('Corriente Venta')), str(item.get('Punto de Venta'))), [])
        vol_rutas = sum(_num(x.get('Volumen Transportado BPD')) for x in rutas)
        t_tot = sum(_num(x.get('Costo Total USD')) for x in rutas)
        t_rate = (t_tot / vol_rutas) if vol_rutas else 0.0
        origen = ' + '.join(sorted({str(x.get('Origen')) for x in rutas})) if rutas else \
            str(item.get('Punto de Venta', ''))

        rows_data.append({
            'type': 'Venta',
            'name': item.get('Corriente Venta', ''),
            'src': origen,
            'dst': item.get('Punto de Venta', ''),
            'vol': vol,
            'rate': price,
            'gross': tot,
            'trans_rate': t_rate,
            'trans_tot': t_tot,
            'net': tot - t_tot,
        })

    # Costos que no cuelgan de una compra ni de una venta. Sin estas filas la
    # matriz nunca sumaria la utilidad.
    for t in resultados.get('throughput', []):
        costo = _num(t.get('Costo Total USD'))
        if abs(costo) < 0.005:
            continue
        rows_data.append({
            'type': 'Refinacion',
            'name': 'Proceso de refinacion',
            'src': t.get('Refineria', ''),
            'dst': t.get('Refineria', ''),
            'vol': _num(t.get('Throughput Refineria BPD')),
            'rate': _num(t.get('Costo de Refinacion USD/BPD')),
            'gross': -abs(costo),
            'trans_rate': 0.0,
            'trans_tot': 0.0,
            'net': -abs(costo),
        })

    for clave, etiqueta in (('costos_operacionales', 'Operacion blending'),
                            ('costos_operacionales_ref', 'Operacion refinados en blend')):
        for c in resultados.get(clave, []):
            costo = _num(c.get('Costo Total USD'))
            if abs(costo) < 0.005:
                continue
            rows_data.append({
                'type': 'Blending',
                'name': etiqueta,
                'src': c.get('Centro de Operacion', ''),
                'dst': c.get('Centro de Operacion', ''),
                'vol': _num(c.get('Throughput Centro de Operacion BPD')
                            or c.get('Throughput Centro de Operacion Refinados BPD')),
                'rate': _num(c.get('Costos Operacionales USD/BPD')
                             or c.get('Costos Operacionales Refinados USD/BPD')),
                'gross': -abs(costo),
                'trans_rate': 0.0,
                'trans_tot': 0.0,
                'net': -abs(costo),
            })

        
    start_row = FILA_MATRIZ + 2
    current_row = start_row
    
    for r in rows_data:
        ws.cell(row=current_row, column=1, value=r['type']).alignment = align_center
        ws.cell(row=current_row, column=2, value=r['name']).alignment = align_left
        ws.cell(row=current_row, column=3, value=r['src']).alignment = align_left
        ws.cell(row=current_row, column=4, value=r['dst']).alignment = align_left
        
        ws.cell(row=current_row, column=5, value=r['vol']).number_format = "#,##0.00"
        ws.cell(row=current_row, column=6, value=r['rate']).number_format = "$#,##0.00"
        ws.cell(row=current_row, column=7, value=r['gross']).number_format = "$#,##0.00"
        ws.cell(row=current_row, column=8, value=r['trans_rate']).number_format = "$#,##0.00"
        ws.cell(row=current_row, column=9, value=r['trans_tot']).number_format = "$#,##0.00"
        ws.cell(row=current_row, column=10, value=r['net']).number_format = "$#,##0.00"
        
        # Apply font, borders and right alignment for numeric columns
        for c in range(1, 11):
            cell = ws.cell(row=current_row, column=c)
            cell.font = font_data
            cell.border = thin_border
            if c >= 5:
                cell.alignment = align_right
                
        ws.row_dimensions[current_row].height = 20
        current_row += 1
        
    # Totals Row
    ws.cell(row=current_row, column=1, value="TOTALES").font = font_total
    ws.cell(row=current_row, column=1).alignment = align_center
    ws.cell(row=current_row, column=1).fill = fill_total
    ws.cell(row=current_row, column=1).border = total_border
    
    for c in range(2, 5):
        cell = ws.cell(row=current_row, column=c, value="")
        cell.fill = fill_total
        cell.border = total_border
        
    # Formulas for totals
    gross_col_letter = get_column_letter(7)
    trans_col_letter = get_column_letter(9)
    net_col_letter = get_column_letter(10)

    # El volumen NO se totaliza: sumar barriles comprados con barriles vendidos
    # y con throughput de refineria no significa nada. Se leen por fila.
    total_cells = [
        (5, "", ""),
        (6, "", ""),
        (7, f"=SUM({gross_col_letter}{start_row}:{gross_col_letter}{current_row-1})", "$#,##0.00"),
        (8, "", ""),
        (9, f"=SUM({trans_col_letter}{start_row}:{trans_col_letter}{current_row-1})", "$#,##0.00"),
        (10, f"=SUM({net_col_letter}{start_row}:{net_col_letter}{current_row-1})", "$#,##0.00"),
]
    
    for col, formula, fmt in total_cells:
        cell = ws.cell(row=current_row, column=col)
        cell.fill = fill_total
        cell.border = total_border
        cell.font = font_total
        cell.alignment = align_right
        if formula:
            cell.value = formula
            cell.number_format = fmt
            
    ws.row_dimensions[current_row].height = 22
    
    # Auto-adjust column widths for Sheet 1
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        # Avoid title at A1 and scenario at A2 causing massive column width
        for cell in col:
            if cell.row in [1, 2, 3]:
                continue
            val_str = str(cell.value or '')
            if cell.number_format and ('$' in cell.number_format or '%' in cell.number_format):
                val_str += "   " # Add padding for formatting symbols
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
    # ================= Sheet 2: Detalle Compras =================
    ws_c = wb.create_sheet(title="Compras")
    ws_c.views.sheetView[0].showGridLines = True
    
    ws_c["A1"] = "DETALLE DE COMPRAS OPTIMIZADAS"
    ws_c["A1"].font = font_section
    
    c_headers = ["Crudo / Producto", "Origen", "Volumen Comprado (BPD)", "Precio Compra (USD/BBL)", "Costo Total (USD/Día)"]
    for col_idx, h in enumerate(c_headers, 1):
        cell = ws_c.cell(row=3, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = thin_border
    ws_c.row_dimensions[3].height = 24
    
    c_row = 4
    for item in resultados.get('compras', []):
        ws_c.cell(row=c_row, column=1, value=item.get('Crudo o Producto', '')).alignment = align_left
        ws_c.cell(row=c_row, column=2, value=item.get('Origen', '')).alignment = align_left
        ws_c.cell(row=c_row, column=3, value=item.get('Volumen Comprado BPD', 0.0)).number_format = "#,##0.00"
        ws_c.cell(row=c_row, column=4, value=item.get('Precio Compra USD/BBL', 0.0)).number_format = "$#,##0.00"
        ws_c.cell(row=c_row, column=5, value=item.get('Costo Total USD', 0.0)).number_format = "$#,##0.00"
        
        for col in range(1, 6):
            cell = ws_c.cell(row=c_row, column=col)
            cell.font = font_data
            cell.border = thin_border
            if col >= 3:
                cell.alignment = align_right
        ws_c.row_dimensions[c_row].height = 20
        c_row += 1
        
    # Total Compras Row
    ws_c.cell(row=c_row, column=1, value="TOTALES").font = font_total
    ws_c.cell(row=c_row, column=1).alignment = align_center
    ws_c.cell(row=c_row, column=1).fill = fill_total
    ws_c.cell(row=c_row, column=1).border = total_border
    ws_c.cell(row=c_row, column=2, value="").fill = fill_total
    ws_c.cell(row=c_row, column=2).border = total_border
    
    ws_c.cell(row=c_row, column=3, value=f"=SUM(C4:C{c_row-1})").number_format = "#,##0.00"
    ws_c.cell(row=c_row, column=5, value=f"=SUM(E4:E{c_row-1})").number_format = "$#,##0.00"
    
    for c in range(3, 6):
        cell = ws_c.cell(row=c_row, column=c)
        cell.fill = fill_total
        cell.border = total_border
        cell.font = font_total
        cell.alignment = align_right
    ws_c.row_dimensions[c_row].height = 22
    
    # Auto-adjust column widths
    for col in ws_c.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row == 1: continue
            max_len = max(max_len, len(str(cell.value or '')))
        ws_c.column_dimensions[col_letter].width = max(max_len + 4, 15)
        
    # ================= Sheet 3: Detalle Ventas =================
    ws_v = wb.create_sheet(title="Ventas")
    ws_v.views.sheetView[0].showGridLines = True
    
    ws_v["A1"] = "DETALLE DE VENTAS OPTIMIZADAS"
    ws_v["A1"].font = font_section
    
    v_headers = ["Corriente de Venta", "Punto de Venta", "Volumen Vendido (BPD)", "Precio Venta (USD/BBL)", "Ingreso Total (USD/Día)", "Crudo de Origen"]
    for col_idx, h in enumerate(v_headers, 1):
        cell = ws_v.cell(row=3, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = thin_border
    ws_v.row_dimensions[3].height = 24
    
    v_row = 4
    for item in resultados.get('ventas', []):
        ws_v.cell(row=v_row, column=1, value=item.get('Corriente Venta', '')).alignment = align_left
        ws_v.cell(row=v_row, column=2, value=item.get('Punto de Venta', '')).alignment = align_left
        ws_v.cell(row=v_row, column=3, value=item.get('Volumen Vendido BPD', 0.0)).number_format = "#,##0.00"
        ws_v.cell(row=v_row, column=4, value=item.get(COL_PRECIO_VENTA, 0.0)).number_format = "$#,##0.00"
        ws_v.cell(row=v_row, column=5, value=_ingreso_venta(item)).number_format = "$#,##0.00"
        ws_v.cell(row=v_row, column=6, value=item.get(COL_CRUDO_ORIGEN_VENTA, '') or '').alignment = align_left
        
        for col in range(1, 7):
            cell = ws_v.cell(row=v_row, column=col)
            cell.font = font_data
            cell.border = thin_border
            if col in [3, 4, 5]:
                cell.alignment = align_right
        ws_v.row_dimensions[v_row].height = 20
        v_row += 1
        
    # Total Ventas Row
    ws_v.cell(row=v_row, column=1, value="TOTALES").font = font_total
    ws_v.cell(row=v_row, column=1).alignment = align_center
    ws_v.cell(row=v_row, column=1).fill = fill_total
    ws_v.cell(row=v_row, column=1).border = total_border
    for c in [2, 6]:
        ws_v.cell(row=v_row, column=c, value="").fill = fill_total
        ws_v.cell(row=v_row, column=c).border = total_border
        
    ws_v.cell(row=v_row, column=3, value=f"=SUM(C4:C{v_row-1})").number_format = "#,##0.00"
    ws_v.cell(row=v_row, column=5, value=f"=SUM(E4:E{v_row-1})").number_format = "$#,##0.00"
    
    for c in range(3, 6):
        cell = ws_v.cell(row=v_row, column=c)
        cell.fill = fill_total
        cell.border = total_border
        cell.font = font_total
        cell.alignment = align_right
    ws_v.row_dimensions[v_row].height = 22
    
    # Auto-adjust column widths
    for col in ws_v.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row == 1: continue
            max_len = max(max_len, len(str(cell.value or '')))
        ws_v.column_dimensions[col_letter].width = max(max_len + 4, 15)
        
    # ================= Sheet 4: Transporte =================
    ws_t = wb.create_sheet(title="Transporte")
    ws_t.views.sheetView[0].showGridLines = True
    
    ws_t["A1"] = "DETALLE DE TRANSPORTE Y RUTAS LOGÍSTICAS"
    ws_t["A1"].font = font_section
    
    t_headers = ["Flujo/Item", "Origen", "Destino", "Ruta de Transporte", "Volumen (BPD)", "Tarifa (USD/BBL)", "Costo Total (USD/Día)"]
    for col_idx, h in enumerate(t_headers, 1):
        cell = ws_t.cell(row=3, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = thin_border
    ws_t.row_dimensions[3].height = 24
    
    t_row = 4
    for item in resultados.get('transporte', []):
        ws_t.cell(row=t_row, column=1, value=item.get('Flujo', '')).alignment = align_left
        ws_t.cell(row=t_row, column=2, value=item.get('Origen', '')).alignment = align_left
        ws_t.cell(row=t_row, column=3, value=item.get('Destino', '')).alignment = align_left
        ws_t.cell(row=t_row, column=4, value=item.get('Ruta', '')).alignment = align_left
        ws_t.cell(row=t_row, column=5, value=item.get('Volumen Transportado BPD', 0.0)).number_format = "#,##0.00"
        ws_t.cell(row=t_row, column=6, value=item.get('Costo de transporte USD/BBL', 0.0)).number_format = "$#,##0.00"
        ws_t.cell(row=t_row, column=7, value=item.get('Costo Total USD', 0.0)).number_format = "$#,##0.00"
        
        for col in range(1, 8):
            cell = ws_t.cell(row=t_row, column=col)
            cell.font = font_data
            cell.border = thin_border
            if col in [5, 6, 7]:
                cell.alignment = align_right
        ws_t.row_dimensions[t_row].height = 20
        t_row += 1
        
    # Total Transporte Row
    ws_t.cell(row=t_row, column=1, value="TOTALES").font = font_total
    ws_t.cell(row=t_row, column=1).alignment = align_center
    ws_t.cell(row=t_row, column=1).fill = fill_total
    ws_t.cell(row=t_row, column=1).border = total_border
    for c in range(2, 5):
        ws_t.cell(row=t_row, column=c, value="").fill = fill_total
        ws_t.cell(row=t_row, column=c).border = total_border
        
    ws_t.cell(row=t_row, column=5, value=f"=SUM(E4:E{t_row-1})").number_format = "#,##0.00"
    ws_t.cell(row=t_row, column=7, value=f"=SUM(G4:G{t_row-1})").number_format = "$#,##0.00"
    
    for c in [5, 7]:
        cell = ws_t.cell(row=t_row, column=c)
        cell.fill = fill_total
        cell.border = total_border
        cell.font = font_total
        cell.alignment = align_right
    ws_t.row_dimensions[t_row].height = 22
    
    # Auto-adjust column widths
    for col in ws_t.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row == 1: continue
            max_len = max(max_len, len(str(cell.value or '')))
        ws_t.column_dimensions[col_letter].width = max(max_len + 4, 15)
        
    # ================= Sheet 5: Blending & Propiedades =================
    ws_b = wb.create_sheet(title="Blending & Calidades")
    ws_b.views.sheetView[0].showGridLines = True
    
    ws_b["A1"] = "CALIDADES Y PROPIEDADES EN CENTROS DE BLENDING"
    ws_b["A1"].font = font_section
    
    b_headers = [
        "Mezcla", "Centro de Blending", "Volumen (BPD)", "API", "Azufre (% m/m)", 
        "Viscosidad Cst", "Acid Number (mg KOH/g)", "Sedimentos (%)", "Residuo Carbón (%)", "Agua (%)"
    ]
    for col_idx, h in enumerate(b_headers, 1):
        cell = ws_b.cell(row=3, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = thin_border
    ws_b.row_dimensions[3].height = 24
    
    b_row = 4
    for item in resultados.get('blending', []):
        ws_b.cell(row=b_row, column=1, value=item.get('Mezcla', '')).alignment = align_left
        ws_b.cell(row=b_row, column=2, value=item.get('Centro de Blending', '')).alignment = align_left
        ws_b.cell(row=b_row, column=3, value=item.get('Volumen BPD', 0.0)).number_format = "#,##0.00"
        ws_b.cell(row=b_row, column=4, value=item.get('API', 0.0)).number_format = "0.00"
        # 0,48 significa 0,48 % m/m. Con formato de porcentaje Excel lo
        # multiplicaba por 100 y mostraba 48 %.
        ws_b.cell(row=b_row, column=5, value=item.get('%AZUFRE', 0.0)).number_format = "0.000"
        ws_b.cell(row=b_row, column=6, value=item.get('Viscosidad Cst', 0.0)).number_format = "0.00"
        ws_b.cell(row=b_row, column=7, value=item.get('ACID NUMBER mg KOH/g', 0.0)).number_format = "0.00"
        ws_b.cell(row=b_row, column=8, value=item.get('Accelerated Total Sediment by Hot Filtration % m/m', 0.0)).number_format = "0.00%"
        ws_b.cell(row=b_row, column=9, value=item.get('Micro Carbon Residue % m/m', 0.0)).number_format = "0.00%"
        ws_b.cell(row=b_row, column=10, value=item.get('Water by Distillation % v/v', 0.0)).number_format = "0.00%"
        
        for col in range(1, 11):
            cell = ws_b.cell(row=b_row, column=col)
            cell.font = font_data
            cell.border = thin_border
            if col >= 3:
                cell.alignment = align_right
        ws_b.row_dimensions[b_row].height = 20
        b_row += 1
        
    # Auto-adjust column widths
    for col in ws_b.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row == 1: continue
            max_len = max(max_len, len(str(cell.value or '')))
        ws_b.column_dimensions[col_letter].width = max(max_len + 4, 15)
        
    # ================= Estructura de Costo =================
    _hoja_estructura_costo(wb, resultados, {
        'font_title': font_title, 'font_header': font_header,
        'font_data': font_data, 'font_total': font_total,
        'fill_header': fill_header, 'fill_total': fill_total,
        'thin_border': thin_border, 'total_border': total_border,
        'align_center': align_center, 'align_left': align_left,
        'align_right': align_right,
    })

    # ================= Save to memory and return bytes =================
    output_bytes = io.BytesIO()
    wb.save(output_bytes)
    output_bytes.seek(0)
    return output_bytes.read()

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
from datetime import datetime
import io
import matplotlib
import base64
matplotlib.use('Agg')
import matplotlib.pyplot as plt

# Config por defecto (puede ser sobre-escrita desde la vista)
EXCEL_DEFAULT = r"C:\Users\Juan Diego Ayala\OneDrive - conquerstrading\Documentos\Modelo Matematico\Variables.xlsx"

# --------------------------- Utilidades ---------------------------

def vlookup_exact(key, df, col_index_1based):
    try:
        row = df[df.iloc[:, 0].astype(str) == str(key)].iloc[0]
        val = row.iloc[col_index_1based - 1]
        if isinstance(val, float) and np.isnan(val):
            return 0
        return val
    except Exception:
        return 0

def get_named_value(wb, name):
    try:
        dn = wb.defined_names.get(name)
    except Exception:
        dn = None
    if dn is None:
        return None
    try:
        destinations = list(dn.destinations)
    except Exception:
        try:
            destinations = list(dn)
        except Exception:
            destinations = []
    for sheet_name, coord in destinations:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            try:
                return ws[coord].value
            except Exception:
                try:
                    min_c, min_r, max_c, max_r = range_boundaries(coord)
                    return ws.cell(row=min_r, column=min_c).value
                except Exception:
                    pass
    return None

# --------------------------- Core cálculo ---------------------------

def ejecutar_modelo(excel_path: str, generar_excel: bool = False):
    wb_vals = load_workbook(excel_path, data_only=True)
    if "Resultados" not in wb_vals.sheetnames:
        raise RuntimeError("No se encuentra la hoja 'Resultados' en el archivo.")
    ws_res = wb_vals["Resultados"]
    headers = [ws_res.cell(row=1, column=c).value for c in range(1, ws_res.max_column + 1)]

    xls = pd.ExcelFile(excel_path)
    COMPRAS = pd.read_excel(xls, sheet_name="COMPRAS")
    LLEGADA = pd.read_excel(xls, sheet_name="LLEGADA")
    # Cargar hoja económica con flexibilidad en el nombre
    econ_sheet_name = None
    for cand in ["ECONOMICOS", "ECONOMICO", "ECONOMIC", "ECONOMICOS "]:
        if cand in xls.sheet_names:
            econ_sheet_name = cand
            break
    if econ_sheet_name is None:
        # fallback: buscar primera hoja cuyo nombre empiece por 'ECONOM'
        for s in xls.sheet_names:
            if s.upper().startswith('ECONOM'):
                econ_sheet_name = s
                break
    if econ_sheet_name is None:
        raise RuntimeError("No se encontró la hoja ECONOMICOS / ECONOMICO en el archivo.")
    ECONOMICOS = pd.read_excel(xls, sheet_name=econ_sheet_name)

    try:
        PCT_FIN = wb_vals["ECONOMICOS"]["B4"].value or 0.0
    except Exception:
        PCT_FIN = 0.0
    try:
        PREM_EXP = wb_vals["ECONOMICOS"]["B7"].value or 0.0
    except Exception:
        PREM_EXP = 0.0

    BRENT = get_named_value(wb_vals, "BRENT")
    # TRM: puede existir como named range, si no se busca por fila en ECONOMICOS
    TRM = get_named_value(wb_vals, "TRM")
    if BRENT is None:
        idx_col = next((c for c in ECONOMICOS.columns if str(c).strip().upper() in ("INDEX", "INDICE", "VARIABLE", "NOMBRE")), None)
        val_col = next((c for c in ECONOMICOS.columns if str(c).strip().upper() in ("VALUE", "VALOR", "VAL")), None)
        if idx_col and val_col:
            m = ECONOMICOS[idx_col].astype(str).str.upper() == "BRENT"
            if m.any():
                try:
                    BRENT = float(ECONOMICOS.loc[m, val_col].iloc[0])
                except Exception:
                    BRENT = 0.0
    BRENT = BRENT or 0.0

    # Función para normalizar strings numéricos en formatos locales (e.g. 4.100, 4.123,45)
    def parse_num(v):
        if v is None: return None
        if isinstance(v,(int,float)): return float(v)
        s = str(v).strip().replace('$','').replace('USD','').replace('COP','').replace(' ','')
        if not s:
            return None
        # Caso patrón miles con punto y opcional decimales con coma: 4.123,45
        import re
        pattern_miles = re.compile(r'^\d{1,3}(\.\d{3})+(,\d+)?$')
        if pattern_miles.match(s):
            s_norm = s.replace('.','').replace(',','.')
            try: return float(s_norm)
            except: return None
        # Caso sólo miles estilo 4.100 (sin coma decimal pero punto usado como miles)
        if re.match(r'^\d{1,3}(\.\d{3})+$', s):
            try: return float(s.replace('.',''))
            except: return None
        # Si hay coma y no punto: usar coma como decimal
        if ',' in s and '.' not in s:
            try: return float(s.replace(',','.'))
            except: return None
        # Remover miles si ambos presentes asumiendo último es decimal
        if s.count('.')>1 and ',' not in s:
            try: return float(s.replace('.',''))
            except: return None
        # Fallback directo
        try: return float(s)
        except: return None

    if TRM is None:
        idx_col2 = next((c for c in ECONOMICOS.columns if str(c).strip().upper() in ("INDEX", "INDICE", "VARIABLE", "NOMBRE")), None)
        val_col2 = next((c for c in ECONOMICOS.columns if str(c).strip().upper() in ("VALUE", "VALOR", "VAL")), None)
        if idx_col2 and val_col2:
            mask_trm = ECONOMICOS[idx_col2].astype(str).str.upper().str.strip() == 'TRM'
            if mask_trm.any():
                raw_trm = ECONOMICOS.loc[mask_trm, val_col2].iloc[0]
                TRM = parse_num(raw_trm)
    TRM = TRM or 0.0

    result_rows = []
    if COMPRAS.shape[1] == 0:
        raise RuntimeError("La hoja COMPRAS no tiene columnas.")
    ids = COMPRAS.iloc[:, 0].dropna().astype(str).tolist()

    for key in ids:
        contrato = vlookup_exact(key, LLEGADA, 4)
        id_value = key
        producto = vlookup_exact(key, COMPRAS, 3)
        proveedor = vlookup_exact(key, COMPRAS, 4)
        vol_bbl = vlookup_exact(key, COMPRAS, 6)
        val_crudo = vlookup_exact(key, COMPRAS, 7)
        flete_rate = vlookup_exact(key, COMPRAS, 9)
        flete = (flete_rate or 0) * (vol_bbl or 0)
        puerto = vlookup_exact(key, LLEGADA, 8)
        almac1 = vlookup_exact(key, LLEGADA, 13)
        operp1 = vlookup_exact(key, LLEGADA, 18)
        costo_ing_czf = vlookup_exact(key, LLEGADA, 24)
        remol_ing_czf = vlookup_exact(key, LLEGADA, 21)
        transp_czf = vlookup_exact(key, LLEGADA, 23)
        almac2 = vlookup_exact(key, LLEGADA, 14)
        operp2 = vlookup_exact(key, LLEGADA, 19)
        pct_fin = PCT_FIN
        val_fin = (val_crudo or 0) + (flete or 0)
        costo_fin = (val_fin or 0) * (pct_fin or 0)
        if (vol_bbl or 0) != 0:
            usd_bbl_crudo_flete = (val_fin / vol_bbl) - BRENT
            usd_bbl_fin_mes = (BRENT + usd_bbl_crudo_flete) * pct_fin
            usd_bbl_alm_oper1 = ((almac1 or 0) + (operp1 or 0)) / vol_bbl
            remol_a_czf = (remol_ing_czf or 0) / vol_bbl
            transp_ter_bbl = (transp_czf or 0) / vol_bbl
            usd_bbl_ing_czf = ((costo_ing_czf or 0) + (almac2 or 0) + (operp2 or 0)) / vol_bbl
        else:
            usd_bbl_crudo_flete = usd_bbl_fin_mes = usd_bbl_alm_oper1 = 0
            remol_a_czf = transp_ter_bbl = usd_bbl_ing_czf = 0
        premiun_exp = PREM_EXP if str(puerto).strip().upper() == "RETORNO CARGA" else 0
        aseg = (val_fin * 0.005) if premiun_exp == 0 else 0
        aduana = (val_fin * 0.003) if premiun_exp == 0 else 0
        insp = (5500 if (vol_bbl or 0) < 150000 else 8500) if premiun_exp == 0 else 0
        gasto_export = (val_crudo * 0.0035) if premiun_exp > 0 else 0
        if (vol_bbl or 0) != 0:
            nacional_usdbbl = (aseg + aduana + insp + gasto_export) / vol_bbl
        else:
            nacional_usdbbl = 0
        sum_T_Z = (usd_bbl_crudo_flete + usd_bbl_fin_mes + usd_bbl_alm_oper1 + remol_a_czf + transp_ter_bbl + usd_bbl_ing_czf + nacional_usdbbl)
        spread_imp = sum_T_Z if premiun_exp == 0 else 0
        sum_I_O = (almac1 or 0) + (operp1 or 0) + (costo_ing_czf or 0) + (remol_ing_czf or 0) + (transp_czf or 0) + (almac2 or 0) + (operp2 or 0)
        sum_AE_AG = (aseg + aduana + insp)
        costo_total_ingreso = 0 if (flete or 0) == 0 else (val_fin + sum_I_O + costo_fin + sum_AE_AG)
        sum_T_Y = (usd_bbl_crudo_flete + usd_bbl_fin_mes + usd_bbl_alm_oper1 + remol_a_czf + transp_ter_bbl + usd_bbl_ing_czf)
        utilidad_export = 0 if premiun_exp == 0 else (premiun_exp - sum_T_Y)
        usd_bbl_export = (gasto_export / vol_bbl) if (vol_bbl or 0) != 0 else 0
        costo_total_export = (gasto_export + costo_fin + val_fin) if usd_bbl_export > 0 else 0
        spread_export = ((costo_total_export / vol_bbl) - BRENT) if (costo_total_export > 0 and (vol_bbl or 0) != 0) else 0
        row = {
            'CONTRATO': contrato,
            'ID': id_value,
            'Producto': producto,
            'Proveedor': proveedor,
            'Volumen Compra BBL': vol_bbl,
            'Valor Compra Crudo': val_crudo,
            'Flete Marino': flete,
            'Puerto Llegada': puerto,
            'Almacenamiento1': almac1,
            'Oper Portuaria 1': operp1,
            'Costo Ingreso CZF': costo_ing_czf,
            'Remolcador Ingreso CZF': remol_ing_czf,
            'Transporte Terrestre a CZF': transp_czf,
            'Almacenamiento 2': almac2,
            'Oper Portuaria 2': operp2,
            '% Financiación Crudo + Flete Marino': PCT_FIN,
            'Valor a Financiar Crudo + Flete': val_fin,
            'Costo Financiación': costo_fin,
            'Costo Total hasta ingreso TK 109 SPD CZF': costo_total_ingreso,
            'USD/BBL CRUDO + FLETE Marino': usd_bbl_crudo_flete,
            'USD/BBL %FIN mes': usd_bbl_fin_mes,
            'USD/BBL Alm+OperPort 1': usd_bbl_alm_oper1,
            'Remolcador a CZF': remol_a_czf,
            'Transp Terrestre a CZF': transp_ter_bbl,
            'USD/BBL Ingreso a CZF (Alm+OperPort 2)': usd_bbl_ing_czf,
            'Nacionalización USD/BBL': nacional_usdbbl,
            'Spread Total on Brent IMPORTACIONES': spread_imp,
            'Premiun Venta Exportaciones': premiun_exp,
            'Utilidad Exportaciones HC´s ECP u Otros': utilidad_export,
            'Gasto Nacionalización': (aseg + aduana + insp + gasto_export),
            'Aseguranza Crudo + Flete 0.5%': aseg,
            'Agencia de Aduana': aduana,
            'Inspección Recibido': insp,
            'Gasto Exportación': gasto_export,
            'USD/Bbl Exportación': usd_bbl_export,
            'Costo Total Exportación': costo_total_export,
            'Spread Exportaciones': spread_export,
        }
        for h in headers:
            if h not in row:
                row[h] = None
        result_rows.append(row)
    df_result = pd.DataFrame(result_rows)

    # Preparar datos de resumen para la UI
    resumen_rows = []
    for _, r in df_result.iterrows():
        resumen_rows.append({
            'ID': r['ID'],
            'CONTRATO': r['CONTRATO'],
            'Producto': r['Producto'],
            'Volumen': float(r['Volumen Compra BBL'] or 0),
            'SpreadImp': round(float(r['Spread Total on Brent IMPORTACIONES'] or 0), 4),
            'CostoTotalImp': round(float(r['Costo Total hasta ingreso TK 109 SPD CZF'] or 0), 2),
            'SpreadExp': round(float(r['Spread Exportaciones'] or 0), 4),
            'UtilidadExp': round(float(r['Utilidad Exportaciones HC´s ECP u Otros'] or 0), 4),
        })

    grafico_base64 = None
    try:
        df_filtrado = df_result[df_result['Volumen Compra BBL'] > 0].copy()
        if not df_filtrado.empty:
            componentes = [
                'USD/BBL CRUDO + FLETE Marino', 'Remolcador a CZF', 'USD/BBL Ingreso a CZF (Alm+OperPort 2)',
                'USD/BBL %FIN mes', 'USD/BBL Alm+OperPort 1', 'Nacionalización USD/BBL', 'USD/Bbl Exportación', 'Transp Terrestre a CZF'
            ]
            nombres = ['CRUDO+FLETE','REMOLCADOR','INGRESO CZF','FINANCIACIÓN','ALM+OPER','NACIONAL','EXPORT','TRANSP TER']
            etiquetas = [f"{row['ID']}\n{row['Producto']}" for _, row in df_filtrado.iterrows()]
            datos = []
            for _, row in df_filtrado.iterrows():
                datos.append([0 if pd.isna(row[c]) else row[c] for c in componentes])
            plt.style.use('seaborn-v0_8-whitegrid')
            fig, ax = plt.subplots(figsize=(12, 7))
            x_pos = np.arange(len(etiquetas))
            ancho = 0.7
            bottom = np.zeros(len(etiquetas))
            colores = ['#1f77b4','#279e68','#d62728','#aa40fc','#8c564b','#e377c2','#b5bd61','#17becf']
            for i, comp in enumerate(componentes):
                vals = [fila[i] for fila in datos]
                ax.bar(x_pos, vals, ancho, bottom=bottom, label=nombres[i], color=colores[i%len(colores)])
                bottom += vals
            ax.set_ylabel('USD/BBL')
            ax.set_title('Componentes de Costo por Contrato')
            ax.set_xticks(x_pos)
            ax.set_xticklabels(etiquetas, rotation=45, ha='right')
            ax.legend(bbox_to_anchor=(1.05, 1), loc='upper left')
            plt.tight_layout()
            buf = io.BytesIO()
            fig.savefig(buf, format='png', bbox_inches='tight')
            plt.close(fig)
            grafico_base64 = base64.b64encode(buf.getvalue()).decode('utf-8')
    except Exception:
        pass

    output_bytes = None
    if generar_excel:
        from openpyxl import Workbook
        # Guardar en memoria para descarga
        out_buffer = io.BytesIO()
        with pd.ExcelWriter(out_buffer, engine='openpyxl') as w:
            df_result.to_excel(w, sheet_name='Resultados', index=False)
        out_buffer.seek(0)
        output_bytes = out_buffer.read()

    return {
        'df_result': df_result,
        'resumen': resumen_rows,
        'BRENT': BRENT,
    'TRM': TRM,
        'grafico_base64': grafico_base64,
        'excel_bytes': output_bytes
    }
